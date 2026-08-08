import io
import os
import sqlite3
import tempfile
import unittest
from datetime import datetime, timedelta

from openpyxl import Workbook

from services.historical_schema import ensure_historical_energy_schema
from services.import_staging import (
    ImportStagingError,
    claim_staged_import,
    consume_staged_import,
    stage_uploaded_import,
)


BASE_SCHEMA = """
CREATE TABLE users (id INTEGER PRIMARY KEY);
CREATE TABLE members (id INTEGER PRIMARY KEY);
CREATE TABLE import_batches (
    id INTEGER PRIMARY KEY, source_file TEXT NOT NULL,
    period_start TEXT NOT NULL, period_end TEXT NOT NULL
);
CREATE TABLE metering_points (
    id INTEGER PRIMARY KEY, metering_point_id TEXT NOT NULL UNIQUE,
    energy_direction TEXT NOT NULL
);
CREATE TABLE meter_codes (id INTEGER PRIMARY KEY, code TEXT NOT NULL UNIQUE);
CREATE TABLE measurements (
    id INTEGER PRIMARY KEY, batch_id INTEGER NOT NULL,
    metering_point_id TEXT NOT NULL, timestamp_start TEXT NOT NULL,
    timestamp_end TEXT NOT NULL, interval_minutes INTEGER NOT NULL,
    meter_code_id INTEGER NOT NULL, value_kwh REAL NOT NULL,
    quality TEXT NOT NULL, is_estimated INTEGER DEFAULT 0
);
"""


def xlsx_bytes():
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as handle:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Energiedaten"
        sheet.cell(2, 2, "AT001")
        sheet.cell(2, 3, "MM")
        sheet.cell(4, 2, "CONSUMPTION")
        sheet.cell(12, 2, "QH (viertelstündlich)")
        sheet.cell(14, 2, "Gesamtverbrauch lt. Messung (bei Teilnahme gem. Erzeugung) [KWH]")
        start = datetime(2026, 1, 1)
        for index in range(4):
            row = 17 + index
            sheet.cell(row, 1, (start + timedelta(minutes=15 * index)).strftime("%d.%m.%Y %H:%M"))
            sheet.cell(row, 2, 1.0)
            sheet.cell(row, 3, "L1")
        workbook.save(handle.name)
        workbook.close()
        handle.seek(0)
        return handle.read()


class ImportStagingTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tempdir.cleanup)
        self.db = sqlite3.connect(":memory:")
        self.db.row_factory = sqlite3.Row
        self.db.execute("PRAGMA foreign_keys=ON")
        self.db.executescript(BASE_SCHEMA)
        ensure_historical_energy_schema(self.db)
        self.db.executemany("INSERT INTO users(id) VALUES (?)", [(1,), (2,)])
        self.db.commit()

    def tearDown(self):
        self.db.close()

    def stage(self):
        return stage_uploaded_import(
            self.db,
            io.BytesIO(xlsx_bytes()),
            "RC_2026-01-01T00_00-2026-01-01T01_00.xlsx",
            1,
            "provisional",
            False,
            self.tempdir.name,
        )

    def test_stage_is_private_and_token_is_not_stored_in_plaintext(self):
        item = self.stage()
        row = self.db.execute("SELECT * FROM import_staging").fetchone()
        path = os.path.join(self.tempdir.name, row["storage_key"], row["original_filename"])
        self.assertNotEqual(row["token_hash"], item["token"])
        self.assertEqual(os.stat(path).st_mode & 0o777, 0o600)
        self.assertEqual(os.stat(os.path.dirname(path)).st_mode & 0o777, 0o700)
        self.assertEqual(item["preview"]["measurement_count"], 4)

    def test_token_is_bound_to_user_and_single_use(self):
        item = self.stage()
        with self.assertRaises(ImportStagingError):
            claim_staged_import(self.db, item["token"], 2, self.tempdir.name)
        stage = claim_staged_import(self.db, item["token"], 1, self.tempdir.name)
        consume_staged_import(self.db, stage, self.tempdir.name)
        self.assertFalse(os.path.exists(stage["path"]))
        with self.assertRaises(ImportStagingError):
            claim_staged_import(self.db, item["token"], 1, self.tempdir.name)

    def test_tampered_staged_file_is_rejected(self):
        item = self.stage()
        row = self.db.execute("SELECT * FROM import_staging").fetchone()
        path = os.path.join(self.tempdir.name, row["storage_key"], row["original_filename"])
        with open(path, "ab") as target:
            target.write(b"tampered")
        with self.assertRaisesRegex(ImportStagingError, "nicht mehr unverändert"):
            claim_staged_import(self.db, item["token"], 1, self.tempdir.name)
        claimed = self.db.execute("SELECT claimed_at FROM import_staging").fetchone()[0]
        self.assertIsNone(claimed)


if __name__ == "__main__":
    unittest.main()
