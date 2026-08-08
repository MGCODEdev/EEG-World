import os
import tempfile
import unittest
from datetime import datetime, timedelta

from openpyxl import Workbook

from services.eda_preview import EDAPreviewError, preview_eda_xlsx


class EDAPreviewTests(unittest.TestCase):
    def make_file(self, timestamps, values=None):
        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Energiedaten"
        sheet.cell(2, 1, "MeteringPointId")
        sheet.cell(2, 2, "AT001")
        sheet.cell(2, 3, "MM")
        sheet.cell(4, 1, "Energydirection")
        sheet.cell(4, 2, "CONSUMPTION")
        sheet.cell(12, 1, "Metering Interval")
        sheet.cell(12, 2, "QH (viertelstündlich)")
        sheet.cell(14, 1, "MeterCode")
        sheet.cell(14, 2, "Gesamtverbrauch lt. Messung (bei Teilnahme gem. Erzeugung) [KWH]")
        for offset, timestamp in enumerate(timestamps, 17):
            sheet.cell(offset, 1, timestamp.strftime("%d.%m.%Y %H:%M"))
            sheet.cell(offset, 2, 1.0 if values is None else values[offset - 17])
            sheet.cell(offset, 3, "L1")
        workbook.save(handle.name)
        workbook.close()
        self.addCleanup(lambda: os.path.exists(handle.name) and os.unlink(handle.name))
        return handle.name

    def test_preview_is_read_only_and_reports_core_metadata(self):
        start = datetime(2026, 1, 1)
        path = self.make_file([start + timedelta(minutes=15 * index) for index in range(4)])
        preview = preview_eda_xlsx(path)
        self.assertEqual(preview.metering_point_count, 1)
        self.assertEqual(preview.series_count, 1)
        self.assertEqual(preview.interval_count, 4)
        self.assertEqual(preview.measurement_count, 4)
        self.assertEqual(preview.interval_minutes, 15)
        self.assertEqual(preview.quality_counts, {"L1": 4})
        self.assertEqual(preview.warnings, ())

    def test_preview_detects_gaps_duplicates_and_empty_cells(self):
        start = datetime(2026, 1, 1)
        path = self.make_file(
            [start, start, start + timedelta(minutes=30)],
            values=[1.0, None, 2.0],
        )
        preview = preview_eda_xlsx(path)
        self.assertEqual(preview.duplicate_timestamps, 1)
        self.assertEqual(preview.missing_timestamp_intervals, 1)
        self.assertEqual(preview.empty_measurement_cells, 1)
        self.assertEqual(
            {warning.code for warning in preview.warnings},
            {"DUPLICATE_TIMESTAMPS", "MISSING_TIMESTAMPS", "EMPTY_MEASUREMENT_CELLS"},
        )

    def test_preview_rejects_missing_energy_sheet(self):
        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        workbook = Workbook()
        workbook.active.title = "Andere Daten"
        workbook.save(handle.name)
        workbook.close()
        self.addCleanup(lambda: os.path.exists(handle.name) and os.unlink(handle.name))
        with self.assertRaises(EDAPreviewError):
            preview_eda_xlsx(handle.name)


if __name__ == "__main__":
    unittest.main()
