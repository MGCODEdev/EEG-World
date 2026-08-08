import os
import io
import tempfile
import unittest
from datetime import datetime, timedelta

from openpyxl import Workbook

import app as eegapp


class V2UITests(unittest.TestCase):
    def setUp(self):
        # Eigene Datenbank pro Test: init_db() wuerde sonst gegen die
        # Produktivdatenbank laufen.
        self.tmp = tempfile.NamedTemporaryFile(delete=False)
        self.tmp.close()
        self.original_db_path = eegapp.DB_PATH
        self.original_staging_path = eegapp.app.config['IMPORT_STAGING_FOLDER']
        self.staging_tmp = tempfile.TemporaryDirectory()
        eegapp.DB_PATH = self.tmp.name
        eegapp.app.config['IMPORT_STAGING_FOLDER'] = self.staging_tmp.name
        eegapp.init_db()
        eegapp.app.config['WTF_CSRF_ENABLED'] = False

    def tearDown(self):
        eegapp.DB_PATH = self.original_db_path
        eegapp.app.config['IMPORT_STAGING_FOLDER'] = self.original_staging_path
        self.staging_tmp.cleanup()
        for suffix in ('', '-wal', '-shm'):
            try:
                os.unlink(self.tmp.name + suffix)
            except OSError:
                pass

    def _admin_id(self):
        with eegapp.app.app_context():
            row = eegapp.get_db().execute(
                "SELECT id FROM users WHERE is_admin=1 ORDER BY id LIMIT 1"
            ).fetchone()
            return row['id']

    @staticmethod
    def _eda_xlsx_bytes():
        with tempfile.NamedTemporaryFile(suffix='.xlsx') as handle:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Energiedaten'
            sheet.cell(2, 2, 'AT001')
            sheet.cell(2, 3, 'MM')
            sheet.cell(4, 2, 'CONSUMPTION')
            sheet.cell(12, 2, 'QH (viertelstündlich)')
            sheet.cell(14, 2, 'Gesamtverbrauch lt. Messung (bei Teilnahme gem. Erzeugung) [KWH]')
            start = datetime(2026, 1, 1)
            for index in range(4):
                row = 17 + index
                sheet.cell(row, 1, (start + timedelta(minutes=15 * index)).strftime('%d.%m.%Y %H:%M'))
                sheet.cell(row, 2, 1.0)
                sheet.cell(row, 3, 'L1')
            workbook.save(handle.name)
            workbook.close()
            handle.seek(0)
            return handle.read()

    def test_v2_dashboard_renders_react_entry_for_admin(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/v2/')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('window.EEG_V2_DATA', html)
        self.assertIn('/static/v2/assets/index-', html)
        self.assertIn('type="module"', html)
        self.assertIn('id="root"', html)
        self.assertIn('"type": "dashboard"', html)

    def test_release_notes_show_latest_modularization_entry(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            classic = client.get('/release-notes')
            v2 = client.get('/v2/release-notes')

        self.assertEqual(classic.status_code, 200)
        self.assertIn('Modulare Codebasis', classic.get_data(as_text=True))
        self.assertEqual(v2.status_code, 200)
        self.assertIn('Modulare Codebasis', v2.get_data(as_text=True))

    def test_v2_members_renders_native_data_for_admin(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/v2/members')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"current_path": "/members"', html)
        self.assertIn('"type": "members"', html)
        self.assertIn('"counts"', html)

    def test_v2_import_renders_native_data_for_admin(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/v2/import')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"current_path": "/import"', html)
        self.assertIn('"type": "import"', html)
        self.assertIn('"csrf_token"', html)

    def test_v2_import_upload_only_creates_preview_and_keeps_measurements_unchanged(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True
            response = client.post('/v2/import', data={
                'import_action': 'preview',
                'data_status': 'provisional',
                'files': (
                    io.BytesIO(self._eda_xlsx_bytes()),
                    'RC_2026-01-01T00_00-2026-01-01T01_00.xlsx',
                ),
            }, content_type='multipart/form-data')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"previews": [', html)
        with eegapp.app.app_context():
            db = eegapp.get_db()
            self.assertEqual(db.execute('SELECT COUNT(*) FROM measurements').fetchone()[0], 0)
            self.assertEqual(db.execute('SELECT COUNT(*) FROM import_batches').fetchone()[0], 0)
            self.assertEqual(db.execute('SELECT COUNT(*) FROM import_staging').fetchone()[0], 1)

    def test_v2_prices_renders_native_data_for_admin(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/v2/prices')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"current_path": "/prices"', html)
        self.assertIn('"type": "prices"', html)
        self.assertIn('"csrf_token"', html)

    def test_v2_invoices_renders_native_data_for_admin(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/v2/invoices')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"current_path": "/invoices"', html)
        self.assertIn('"type": "invoices"', html)

    def test_v2_payments_renders_native_data_for_admin(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/v2/payments')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"current_path": "/payments"', html)
        self.assertIn('"type": "payments"', html)
        self.assertIn('"csrf_token"', html)

    def test_classic_payments_renders_for_admin(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/payments')

        self.assertEqual(response.status_code, 200)
        self.assertIn('Überweisungen & Forderungen', response.get_data(as_text=True))

    def test_v2_newsletter_renders_native_data_for_admin(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/v2/newsletter')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"current_path": "/newsletter"', html)
        self.assertIn('"type": "newsletter"', html)
        self.assertIn('"csrf_token"', html)

    def test_v2_reports_renders_native_data_for_admin(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/v2/reports')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"current_path": "/reports"', html)
        self.assertIn('"type": "reports"', html)

    def test_v2_users_renders_native_data_for_admin(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/v2/admin/users')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"current_path": "/admin/users"', html)
        self.assertIn('"type": "users"', html)
        self.assertIn('"contracts"', html)
        self.assertIn('"csrf_token"', html)

    def test_classic_users_renders_editable_email_for_superadmin(self):
        admin_id = self._admin_id()
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(admin_id)
                sess['_fresh'] = True
            response = client.get('/admin/users')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn(f'action="/admin/users/{admin_id}/email"', html)
        self.assertIn('type="email" name="email"', html)

    def test_admin_can_edit_email_for_superadmin_and_member_user(self):
        admin_id = self._admin_id()
        with eegapp.app.app_context():
            db = eegapp.get_db()
            member_id = db.execute(
                "INSERT INTO members (name, email, active) VALUES ('Test Mitglied', 'alt@example.org', 1)"
            ).lastrowid
            member_user_id = db.execute("""
                INSERT INTO users (username, password_hash, email, is_admin, role, member_id)
                VALUES ('testmitglied', 'unused', 'alt@example.org', 0, 'member', ?)
            """, (member_id,)).lastrowid
            db.commit()

        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(admin_id)
                sess['_fresh'] = True
            superadmin_response = client.post(
                f'/admin/users/{admin_id}/email',
                data={'email': 'superadmin@example.org'},
            )
            member_response = client.post(
                f'/admin/users/{member_user_id}/email',
                data={'email': 'neu@example.org'},
            )

        self.assertEqual(superadmin_response.status_code, 302)
        self.assertEqual(member_response.status_code, 302)
        with eegapp.app.app_context():
            db = eegapp.get_db()
            self.assertEqual(
                db.execute('SELECT email FROM users WHERE id=?', (admin_id,)).fetchone()['email'],
                'superadmin@example.org',
            )
            self.assertEqual(
                db.execute('SELECT email FROM users WHERE id=?', (member_user_id,)).fetchone()['email'],
                'neu@example.org',
            )
            self.assertEqual(
                db.execute('SELECT email FROM members WHERE id=?', (member_id,)).fetchone()['email'],
                'alt@example.org',
            )

    def test_admin_user_email_rejects_invalid_and_conflicting_addresses(self):
        admin_id = self._admin_id()
        with eegapp.app.app_context():
            db = eegapp.get_db()
            other_id = db.execute("""
                INSERT INTO users (username, password_hash, email, is_admin, role)
                VALUES ('other@example.org', 'unused', 'other-mail@example.org', 0, 'member')
            """).lastrowid
            db.commit()

        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(admin_id)
                sess['_fresh'] = True
            invalid = client.post(
                f'/admin/users/{admin_id}/email', data={'email': 'keine-adresse'}
            )
            conflict = client.post(
                f'/admin/users/{admin_id}/email', data={'email': 'other@example.org'}
            )

        self.assertEqual(invalid.status_code, 302)
        self.assertEqual(conflict.status_code, 302)
        with eegapp.app.app_context():
            db = eegapp.get_db()
            self.assertIsNone(
                db.execute('SELECT email FROM users WHERE id=?', (admin_id,)).fetchone()['email']
            )
            self.assertEqual(
                db.execute('SELECT email FROM users WHERE id=?', (other_id,)).fetchone()['email'],
                'other-mail@example.org',
            )

    def test_v2_audit_renders_native_data_for_admin(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/v2/admin/audit')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"current_path": "/admin/audit"', html)
        self.assertIn('"type": "audit"', html)
        self.assertIn('"pagination"', html)

    def test_v2_backup_renders_native_data_for_admin(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/v2/admin/backup')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"current_path": "/admin/backup"', html)
        self.assertIn('"type": "backup"', html)
        self.assertIn('"local_backups"', html)

    def test_v2_subpage_points_shell_to_embedded_existing_page(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/v2/admin/database')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"current_path": "/admin/database"', html)
        self.assertIn('"content_path": "/admin/database?embed=1"', html)

    def test_embed_mode_can_be_framed_by_same_origin_only(self):
        with eegapp.app.test_client() as client:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(self._admin_id())
                sess['_fresh'] = True

            response = client.get('/import?embed=1')

        self.assertEqual(response.status_code, 200)
        self.assertIn('SAMEORIGIN', response.headers['X-Frame-Options'])
        self.assertIn("frame-ancestors 'self'", response.headers['Content-Security-Policy'])
        self.assertIn('body class="embed-mode"', response.get_data(as_text=True))


if __name__ == '__main__':
    unittest.main()
