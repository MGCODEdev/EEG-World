import io
import os
import tempfile
import unittest
from datetime import date, timedelta
from decimal import Decimal

import app as eegapp

PNG_BYTES = b'\x89PNG\r\n\x1a\n' + b'0' * 64
PDF_BYTES = b'%PDF-1.4\n' + b'0' * 64


class CashbookTests(unittest.TestCase):
    def setUp(self):
        # Eigene Datenbank pro Test, damit init_db() nicht die Produktivdaten trifft.
        self.tmp = tempfile.NamedTemporaryFile(delete=False)
        self.tmp.close()
        self.original_db_path = eegapp.DB_PATH
        eegapp.DB_PATH = self.tmp.name
        eegapp.init_db()
        eegapp.app.config['WTF_CSRF_ENABLED'] = False

    def tearDown(self):
        eegapp.DB_PATH = self.original_db_path
        for suffix in ('', '-wal', '-shm'):
            try:
                os.unlink(self.tmp.name + suffix)
            except OSError:
                pass

    def _admin_id(self):
        with eegapp.app.app_context():
            row = eegapp.get_db().execute(
                "SELECT id FROM users WHERE is_admin=1 ORDER BY id LIMIT 1"
            ).fetchone()
            return row['id']

    def _client(self):
        client = eegapp.app.test_client()
        with client.session_transaction() as sess:
            sess['_user_id'] = str(self._admin_id())
            sess['_fresh'] = True
        return client

    def _post_entry(self, client, **overrides):
        payload = {
            'entry_date': date.today().isoformat(),
            'direction': 'expense',
            'amount_eur': '124,50',
            'payment_method': 'cash',
            'description': 'Bewirtung Generalversammlung',
            'counterparty': 'Gasthaus Muster',
        }
        payload.update(overrides)
        return client.post('/kassabuch/new', data=payload,
                           content_type='multipart/form-data', follow_redirects=True)

    def test_default_categories_are_seeded(self):
        with eegapp.app.app_context():
            categories = eegapp.get_cashbook_categories(eegapp.get_db())
        names = [c['name'] for c in categories]
        self.assertEqual(len(categories), len(eegapp.CASHBOOK_DEFAULT_CATEGORIES))
        self.assertIn('Bewirtung', names)
        self.assertIn('Verwaltungskosten', names)

    def test_entry_is_created_with_document_number(self):
        client = self._client()
        response = self._post_entry(client)
        self.assertEqual(response.status_code, 200)

        with eegapp.app.app_context():
            row = eegapp.get_db().execute("SELECT * FROM cashbook_entries").fetchone()
        self.assertIsNotNone(row)
        self.assertEqual(row['amount_eur'], 124.5)
        self.assertEqual(row['direction'], 'expense')
        self.assertEqual(row['payment_method'], 'cash')
        self.assertEqual(row['document_number'], f'{date.today().year}-0001')

    def test_document_numbers_are_sequential_per_year(self):
        client = self._client()
        self._post_entry(client)
        self._post_entry(client, description='Bankspesen Q1')

        with eegapp.app.app_context():
            numbers = [r['document_number'] for r in eegapp.get_db().execute(
                "SELECT document_number FROM cashbook_entries ORDER BY id").fetchall()]
        year = date.today().year
        self.assertEqual(numbers, [f'{year}-0001', f'{year}-0002'])

    def test_receipt_is_stored_and_downloadable(self):
        client = self._client()
        self._post_entry(client, receipt=(io.BytesIO(PNG_BYTES), 'beleg.png'))

        with eegapp.app.app_context():
            row = eegapp.get_db().execute(
                "SELECT id, receipt_filename, receipt_mimetype FROM cashbook_entries").fetchone()
        self.assertEqual(row['receipt_filename'], 'beleg.png')
        self.assertEqual(row['receipt_mimetype'], 'image/png')

        response = client.get(f'/kassabuch/{row["id"]}/beleg')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, PNG_BYTES)

    def test_pdf_receipt_is_accepted(self):
        client = self._client()
        self._post_entry(client, receipt=(io.BytesIO(PDF_BYTES), 'rechnung.pdf'))
        with eegapp.app.app_context():
            row = eegapp.get_db().execute(
                "SELECT receipt_mimetype FROM cashbook_entries").fetchone()
        self.assertEqual(row['receipt_mimetype'], 'application/pdf')

    def test_receipt_with_wrong_content_is_rejected(self):
        client = self._client()
        response = self._post_entry(client, receipt=(io.BytesIO(b'nur text'), 'beleg.png'))
        self.assertIn('passt nicht zur Dateiendung', response.get_data(as_text=True))
        with eegapp.app.app_context():
            count = eegapp.get_db().execute("SELECT COUNT(*) FROM cashbook_entries").fetchone()[0]
        self.assertEqual(count, 0)

    def test_receipt_with_unsupported_extension_is_rejected(self):
        client = self._client()
        response = self._post_entry(client, receipt=(io.BytesIO(b'MZ'), 'beleg.exe'))
        self.assertIn('nur PDF-, JPG- und PNG-Dateien', response.get_data(as_text=True))

    def test_invalid_amount_is_rejected(self):
        client = self._client()
        response = self._post_entry(client, amount_eur='0')
        self.assertIn('größer als null', response.get_data(as_text=True))

    def test_future_date_is_rejected(self):
        client = self._client()
        response = self._post_entry(client,
                                    entry_date=(date.today() + timedelta(days=1)).isoformat())
        self.assertIn('Zukunft', response.get_data(as_text=True))

    def test_missing_description_is_rejected(self):
        client = self._client()
        response = self._post_entry(client, description='   ')
        self.assertIn('Begründung', response.get_data(as_text=True))

    def _seed_invoice(self, db):
        """Abrechnung mit drei Mitgliedern: gebucht neu, gebucht alt, offen."""
        db.execute("INSERT INTO members (id, name) VALUES (1, 'Zahlendes Mitglied')")
        db.execute("INSERT INTO members (id, name) VALUES (2, 'Einspeisendes Mitglied')")
        db.execute("INSERT INTO members (id, name) VALUES (3, 'Offenes Mitglied')")
        db.execute("""INSERT INTO invoices (id, period_from, period_to, status)
                      VALUES (1, '2026-01-01', '2026-03-31', 'sent')""")
        # Mitglied 1: ueber payment_bookings gebucht (neues Verfahren).
        db.execute("""INSERT INTO invoice_items
                      (invoice_id, member_id, type, kwh, price_per_kwh, amount_eur, paid, paid_at)
                      VALUES (1, 1, 'consumption', 1000, 20, 200.0, 1, '2026-01-10 12:00:00')""")
        db.execute("""INSERT INTO payment_bookings
                      (invoice_id, member_id, amount_eur, direction, booking_date)
                      VALUES (1, 1, 200.0, 'member_to_eeg', '2026-01-10')""")
        # Mitglied 2: nur ueber invoice_items bezahlt (altes Verfahren ohne Buchung).
        db.execute("""INSERT INTO invoice_items
                      (invoice_id, member_id, type, kwh, price_per_kwh, amount_eur, paid, paid_at)
                      VALUES (1, 2, 'generation', 500, 10, 50.0, 1, '2026-01-20 09:30:00')""")
        # Mitglied 3: offen, dazu eine stornierte Buchung.
        db.execute("""INSERT INTO invoice_items
                      (invoice_id, member_id, type, kwh, price_per_kwh, amount_eur, paid)
                      VALUES (1, 3, 'consumption', 5000, 20, 999.0, 0)""")
        db.execute("""INSERT INTO payment_bookings
                      (invoice_id, member_id, amount_eur, direction, booking_date, reversed_at)
                      VALUES (1, 3, 999.0, 'member_to_eeg', '2026-01-25', '2026-01-26')""")

    def test_balance_combines_manual_entries_and_all_invoices(self):
        with eegapp.app.app_context():
            db = eegapp.get_db()
            self._seed_invoice(db)
            db.execute("""INSERT INTO cashbook_entries
                          (entry_date, direction, amount_eur, payment_method, description, document_number)
                          VALUES ('2026-02-01', 'expense', 30.0, 'cash', 'Bewirtung', '2026-0001')""")
            db.commit()

            book = eegapp.build_cashbook(db)

        summary = book['summary']
        self.assertEqual(summary['entry_count'], 3)
        self.assertEqual(summary['income_total'], 200.0)
        self.assertEqual(summary['expense_total'], 80.0)
        self.assertEqual(summary['balance'], 120.0)
        self.assertEqual(summary['cash_balance'], -30.0)
        self.assertEqual(summary['bank_balance'], 150.0)
        categories = {row['category'] for row in book['rows']}
        self.assertIn(eegapp.CASHBOOK_ENERGY_CATEGORIES['income'], categories)
        self.assertIn(eegapp.CASHBOOK_ENERGY_CATEGORIES['expense'], categories)

    def test_invoice_paid_without_payment_booking_is_included(self):
        """Aeltere Abrechnungen wurden ohne payment_bookings gebucht."""
        with eegapp.app.app_context():
            db = eegapp.get_db()
            self._seed_invoice(db)
            db.commit()
            rows = eegapp.build_cashbook(db)['rows']

        by_member = {row['member_id']: row for row in rows}
        self.assertIn(2, by_member)
        self.assertEqual(by_member[2]['entry_date'], '2026-01-20')
        self.assertEqual(by_member[2]['direction'], 'expense')
        self.assertEqual(by_member[2]['amount_eur'], 50.0)

    def test_unpaid_invoice_is_not_in_cashbook(self):
        with eegapp.app.app_context():
            db = eegapp.get_db()
            self._seed_invoice(db)
            db.commit()
            rows = eegapp.build_cashbook(db)['rows']
        self.assertNotIn(3, {row['member_id'] for row in rows})

    def test_carried_over_invoice_is_counted_only_once(self):
        with eegapp.app.app_context():
            db = eegapp.get_db()
            db.execute("INSERT INTO members (id, name) VALUES (1, 'Testmitglied')")
            db.execute("""INSERT INTO invoices (id, period_from, period_to, status)
                          VALUES (1, '2026-01-01', '2026-03-31', 'sent')""")
            db.execute("""INSERT INTO invoices (id, period_from, period_to, status)
                          VALUES (2, '2026-04-01', '2026-06-30', 'sent')""")
            # Abrechnung 1 blieb offen und wurde nach Abrechnung 2 vorgetragen.
            db.execute("""INSERT INTO invoice_items
                          (invoice_id, member_id, type, kwh, price_per_kwh, amount_eur, paid)
                          VALUES (1, 1, 'consumption', 100, 20, 20.0, 0)""")
            db.execute("""INSERT INTO invoice_carryovers
                          (invoice_id, member_id, source_invoice_id, amount_eur)
                          VALUES (2, 1, 1, 20.0)""")
            db.execute("""INSERT INTO invoice_items
                          (invoice_id, member_id, type, kwh, price_per_kwh, amount_eur, paid, paid_at)
                          VALUES (2, 1, 'consumption', 150, 20, 30.0, 1, '2026-07-05 12:00:00')""")
            db.commit()
            book = eegapp.build_cashbook(db)

        self.assertEqual(book['summary']['entry_count'], 1)
        # 30 EUR aus Abrechnung 2 plus 20 EUR Vortrag aus Abrechnung 1.
        self.assertEqual(book['summary']['income_total'], 50.0)
        self.assertIn('inkl. Vortrag', book['rows'][0]['description'])

    def test_filter_keeps_overall_balance(self):
        with eegapp.app.app_context():
            db = eegapp.get_db()
            db.execute("""INSERT INTO cashbook_entries
                          (entry_date, direction, amount_eur, payment_method, description, document_number)
                          VALUES ('2025-05-01', 'income', 100.0, 'transfer', 'Zuschuss', '2025-0001')""")
            db.execute("""INSERT INTO cashbook_entries
                          (entry_date, direction, amount_eur, payment_method, description, document_number)
                          VALUES ('2026-05-01', 'expense', 40.0, 'cash', 'Bewirtung', '2026-0001')""")
            db.commit()

            filtered = eegapp.build_cashbook(db, year='2026')

        self.assertEqual(filtered['summary']['entry_count'], 1)
        self.assertEqual(filtered['summary']['expense_total'], 40.0)
        # Der Saldo bleibt der Gesamtsaldo ueber alle Jahre.
        self.assertEqual(filtered['summary']['balance'], 60.0)
        self.assertEqual([row['year'] for row in filtered['by_year']], ['2026', '2025'])

    def test_amounts_are_summed_in_cents(self):
        """Krumme Betraege duerfen sich nicht zu einer Fliesskomma-Abweichung addieren."""
        values = [0.07, 0.29, 1.13, 2.71, 8.19, 0.01, 33.33, 0.02, 5.55, 0.1, 0.2]
        with eegapp.app.app_context():
            db = eegapp.get_db()
            for index, value in enumerate(values, start=1):
                db.execute("""INSERT INTO cashbook_entries
                              (entry_date, direction, amount_eur, payment_method,
                               description, document_number)
                              VALUES (?, 'income', ?, 'transfer', 'Testbetrag', ?)""",
                           ((date(2026, 1, 1) + timedelta(days=index)).isoformat(),
                            value, f'2026-{index:04d}'))
            db.commit()
            book = eegapp.build_cashbook(db)

        expected = sum(Decimal(str(value)) for value in values)
        self.assertEqual(Decimal(str(book['summary']['balance'])), expected)
        self.assertEqual(Decimal(str(book['summary']['bank_balance'])), expected)
        self.assertEqual(Decimal(str(book['summary']['income_total'])), expected)
        self.assertEqual(Decimal(str(book['summary']['result'])), expected)
        # Der laufende Saldo der letzten Zeile muss demselben Wert entsprechen.
        self.assertEqual(Decimal(str(book['rows'][0]['balance'])), expected)
        self.assertEqual(Decimal(str(book['by_year'][0]['result'])), expected)

    def test_entry_can_be_deleted(self):
        client = self._client()
        self._post_entry(client)
        with eegapp.app.app_context():
            entry_id = eegapp.get_db().execute("SELECT id FROM cashbook_entries").fetchone()['id']

        response = client.post(f'/kassabuch/{entry_id}/delete', follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        with eegapp.app.app_context():
            count = eegapp.get_db().execute("SELECT COUNT(*) FROM cashbook_entries").fetchone()[0]
        self.assertEqual(count, 0)

    def test_invoice_movement_cannot_be_deleted_via_cashbook(self):
        with eegapp.app.app_context():
            db = eegapp.get_db()
            self._seed_invoice(db)
            db.commit()
            rows = eegapp.build_cashbook(db)['rows']
        energy_rows = [row for row in rows if row['source'] == 'energy']
        self.assertTrue(energy_rows)
        self.assertFalse(any(row['deletable'] for row in energy_rows))

    def test_category_can_be_created_and_removed(self):
        client = self._client()
        client.post('/kassabuch/kategorien',
                    data={'name': 'Vereinsausflug', 'direction': 'expense'},
                    follow_redirects=True)
        with eegapp.app.app_context():
            row = eegapp.get_db().execute(
                "SELECT id FROM cashbook_categories WHERE name='Vereinsausflug'").fetchone()
        self.assertIsNotNone(row)

        client.post(f'/kassabuch/kategorien/{row["id"]}/delete', follow_redirects=True)
        with eegapp.app.app_context():
            gone = eegapp.get_db().execute(
                "SELECT id FROM cashbook_categories WHERE name='Vereinsausflug'").fetchone()
        self.assertIsNone(gone)

    def test_used_category_is_deactivated_instead_of_deleted(self):
        client = self._client()
        with eegapp.app.app_context():
            category_id = eegapp.get_db().execute(
                "SELECT id FROM cashbook_categories WHERE name='Bewirtung'").fetchone()['id']
        self._post_entry(client, category_id=str(category_id))

        client.post(f'/kassabuch/kategorien/{category_id}/delete', follow_redirects=True)
        with eegapp.app.app_context():
            row = eegapp.get_db().execute(
                "SELECT active FROM cashbook_categories WHERE id=?", (category_id,)).fetchone()
        self.assertEqual(row['active'], 0)

    def test_csv_export_contains_entry(self):
        client = self._client()
        self._post_entry(client)
        response = client.get('/kassabuch/export.csv')
        self.assertEqual(response.status_code, 200)
        text = response.data.decode('utf-8-sig')
        self.assertIn('Bewirtung Generalversammlung', text)
        self.assertIn('124,50', text)
        self.assertIn('Kassastand bar', text)

    def test_pdf_export_returns_pdf(self):
        client = self._client()
        self._post_entry(client)
        response = client.get('/kassabuch/export.pdf')
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data.startswith(b'%PDF'))

    def _seed_period_entries(self):
        """Je eine Buchung vor, in und nach dem Berichtszeitraum 2026."""
        with eegapp.app.app_context():
            db = eegapp.get_db()
            for entry_date, direction, amount, number in (
                    ('2025-11-10', 'income', 100.0, '2025-0001'),
                    ('2026-03-01', 'income', 60.0, '2026-0001'),
                    ('2026-06-30', 'expense', 25.0, '2026-0002'),
                    ('2027-01-15', 'expense', 10.0, '2027-0001')):
                db.execute("""INSERT INTO cashbook_entries
                              (entry_date, direction, amount_eur, payment_method,
                               description, document_number)
                              VALUES (?, ?, ?, 'transfer', 'Testbuchung', ?)""",
                           (entry_date, direction, amount, number))
            db.commit()

    def test_period_from_year_filter(self):
        self._seed_period_entries()
        with eegapp.app.app_context():
            book = eegapp.build_cashbook(eegapp.get_db(), year='2026')
        self.assertEqual(book['period']['from'], '2026-01-01')
        self.assertEqual(book['period']['to'], '2026-12-31')
        self.assertEqual(book['period']['label'], '01.01.2026 bis 31.12.2026')

    def test_opening_and_closing_balance(self):
        self._seed_period_entries()
        with eegapp.app.app_context():
            summary = eegapp.build_cashbook(eegapp.get_db(), year='2026')['summary']
        # 100 vor dem Zeitraum, im Zeitraum +60 und -25
        self.assertEqual(summary['opening_balance'], 100.0)
        self.assertEqual(summary['income_total'], 60.0)
        self.assertEqual(summary['expense_total'], 25.0)
        self.assertEqual(summary['closing_balance'], 135.0)
        self.assertEqual(summary['entry_count'], 2)
        # Kontrollrechnung des Berichts
        self.assertEqual(summary['opening_balance'] + summary['result'],
                         summary['closing_balance'])

    def test_date_range_beats_year(self):
        self._seed_period_entries()
        with eegapp.app.app_context():
            book = eegapp.build_cashbook(eegapp.get_db(), year='2025',
                                         date_from='2026-01-01', date_to='2026-03-31')
        self.assertEqual(book['period']['from'], '2026-01-01')
        self.assertEqual(book['summary']['entry_count'], 1)
        self.assertEqual(book['summary']['opening_balance'], 100.0)
        self.assertEqual(book['summary']['closing_balance'], 160.0)

    def test_reversed_date_range_is_corrected(self):
        start, end, label = eegapp.resolve_cashbook_period(
            date_from='2026-12-31', date_to='2026-01-01')
        self.assertEqual((start, end), ('2026-01-01', '2026-12-31'))
        self.assertIn('01.01.2026', label)

    def test_open_period_labels(self):
        self.assertEqual(eegapp.resolve_cashbook_period()[2], 'gesamter Zeitraum')
        self.assertEqual(eegapp.resolve_cashbook_period(date_from='2026-05-01')[2],
                         'ab 01.05.2026')
        self.assertEqual(eegapp.resolve_cashbook_period(date_to='2026-05-01')[2],
                         'bis 01.05.2026')

    def test_category_filter_does_not_change_closing_balance(self):
        self._seed_period_entries()
        with eegapp.app.app_context():
            book = eegapp.build_cashbook(eegapp.get_db(), year='2026',
                                         category='Gibt es nicht')
        self.assertEqual(book['summary']['entry_count'], 0)
        # Der Saldo haengt am Zeitraum, nicht an der Kategorieauswahl.
        self.assertEqual(book['summary']['closing_balance'], 135.0)

    def test_xlsx_export_contains_values(self):
        from openpyxl import load_workbook

        self._seed_period_entries()
        response = self._client().get('/kassabuch/export.xlsx?year=2026')
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response.mimetype)

        workbook = load_workbook(io.BytesIO(response.data))
        self.assertEqual(workbook.sheetnames, ['Kassabuch', 'Auswertung'])
        sheet = workbook['Kassabuch']
        self.assertIn('Kassabuch', str(sheet['A1'].value))
        self.assertEqual(sheet['B2'].value, '01.01.2026 bis 31.12.2026')
        labels = [sheet.cell(row=row, column=4).value for row in range(2, 8)]
        self.assertIn('Anfangssaldo', labels)
        self.assertIn('Endsaldo', labels)
        # Betraege muessen Zahlen sein, damit Excel rechnen kann
        amounts = [cell.value for column in ('I', 'J') for cell in sheet[column]
                   if isinstance(cell.value, (int, float))]
        self.assertIn(60.0, amounts)
        self.assertIn(25.0, amounts)

    def test_xlsx_export_has_date_values(self):
        from datetime import datetime as dt

        from openpyxl import load_workbook

        self._seed_period_entries()
        response = self._client().get('/kassabuch/export.xlsx?year=2026')
        sheet = load_workbook(io.BytesIO(response.data))['Kassabuch']
        dates = [cell.value for cell in sheet['C'] if isinstance(cell.value, (dt, date))]
        self.assertTrue(dates)
        self.assertEqual(sheet.freeze_panes, 'A10')

    def test_csv_export_shows_opening_and_closing(self):
        self._seed_period_entries()
        response = self._client().get('/kassabuch/export.csv?year=2026')
        text = response.data.decode('utf-8-sig')
        self.assertIn('01.01.2026 bis 31.12.2026', text)
        self.assertIn('Anfangssaldo', text)
        self.assertIn('Endsaldo', text)
        self.assertIn('100,00', text)
        self.assertIn('135,00', text)

    def test_logo_is_available_as_data_uri(self):
        with eegapp.app.app_context():
            logo = eegapp._logo_data_uri()
        self.assertTrue(logo.startswith('data:image/png;base64,'))
        self.assertGreater(len(logo), 1000)

    def test_pdf_export_contains_logo_and_period(self):
        self._seed_period_entries()
        response = self._client().get('/kassabuch/export.pdf?year=2026')
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data.startswith(b'%PDF'))
        # Das Logo wird als Bildobjekt eingebettet.
        self.assertIn(b'/Image', response.data)
        self.assertIn(b'/XObject', response.data)

    def test_export_filename_uses_period(self):
        self._seed_period_entries()
        response = self._client().get('/kassabuch/export.xlsx?year=2026')
        self.assertIn('kassabuch-2026-01-01_bis_2026-12-31',
                      response.headers['Content-Disposition'])

    def test_sequence_numbers_follow_booking_date(self):
        """Die Belegnummer laeuft chronologisch aufsteigend, je Jahr neu."""
        self._seed_period_entries()
        with eegapp.app.app_context():
            rows = eegapp.build_cashbook(eegapp.get_db())['rows_chronological']
        self.assertEqual([row['sequence_number'] for row in rows],
                         ['2025/001', '2026/001', '2026/002', '2027/001'])
        dates = [row['entry_date'] for row in rows]
        self.assertEqual(dates, sorted(dates))

    def test_sequence_number_is_stable_across_filters(self):
        """Ein Filter darf die Belegnummer nicht verschieben."""
        self._seed_period_entries()
        with eegapp.app.app_context():
            db = eegapp.get_db()
            alle = {row['reference']: row['sequence_number']
                    for row in eegapp.build_cashbook(db)['rows_chronological']}
            gefiltert = {row['reference']: row['sequence_number']
                         for row in eegapp.build_cashbook(db, year='2026')['rows_chronological']}
        for reference, number in gefiltert.items():
            self.assertEqual(alle[reference], number)
        self.assertEqual(gefiltert['2026-0001'], '2026/001')

    def test_backdated_entry_gets_earlier_sequence_number(self):
        """Eine nachtraeglich erfasste, aelter datierte Buchung reiht sich vorne ein."""
        self._seed_period_entries()
        with eegapp.app.app_context():
            db = eegapp.get_db()
            db.execute("""INSERT INTO cashbook_entries
                          (entry_date, direction, amount_eur, payment_method,
                           description, document_number)
                          VALUES ('2026-01-05', 'expense', 5.0, 'cash', 'Nachtrag', '2026-0009')""")
            db.commit()
            rows = eegapp.build_cashbook(db, year='2026')['rows_chronological']
        self.assertEqual([(row['sequence_number'], row['reference']) for row in rows],
                         [('2026/001', '2026-0009'),
                          ('2026/002', '2026-0001'),
                          ('2026/003', '2026-0002')])

    def test_reference_keeps_original_document_number(self):
        self._seed_period_entries()
        with eegapp.app.app_context():
            rows = eegapp.build_cashbook(eegapp.get_db(), year='2026')['rows_chronological']
        self.assertEqual([row['reference'] for row in rows], ['2026-0001', '2026-0002'])

    def test_search_finds_sequence_number(self):
        self._seed_period_entries()
        with eegapp.app.app_context():
            book = eegapp.build_cashbook(eegapp.get_db(), search='2026/002')
        self.assertEqual(book['summary']['entry_count'], 1)
        self.assertEqual(book['rows'][0]['reference'], '2026-0002')

    def test_exports_contain_number_and_reference(self):
        from openpyxl import load_workbook

        self._seed_period_entries()
        client = self._client()
        text = client.get('/kassabuch/export.csv?year=2026').data.decode('utf-8-sig')
        self.assertIn('Beleg-Nr;Referenz;Datum', text)
        self.assertIn('2026/001;2026-0001', text)

        sheet = load_workbook(io.BytesIO(
            client.get('/kassabuch/export.xlsx?year=2026').data))['Kassabuch']
        self.assertEqual([sheet.cell(row=9, column=c).value for c in (1, 2, 3)],
                         ['Beleg-Nr', 'Referenz', 'Datum'])
        self.assertEqual(sheet.cell(row=11, column=1).value, '2026/001')
        self.assertEqual(sheet.cell(row=11, column=2).value, '2026-0001')

    def test_cashbook_page_shows_period_and_balances(self):
        self._seed_period_entries()
        html = self._client().get('/kassabuch?year=2026').get_data(as_text=True)
        self.assertIn('01.01.2026 bis 31.12.2026', html)
        self.assertIn('Anfangssaldo', html)
        self.assertIn('Endsaldo', html)
        self.assertIn('export.xlsx', html)
        self.assertIn('name="date_from"', html)

    def test_cashbook_page_renders(self):
        client = self._client()
        self._post_entry(client)
        response = client.get('/kassabuch')
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('Vereinskassabuch', html)
        self.assertIn('Bewirtung Generalversammlung', html)

    def test_v2_cashbook_data_is_delivered(self):
        client = self._client()
        self._post_entry(client)
        response = client.get('/v2/kassabuch')
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('"type": "cashbook"', html)
        self.assertIn('"current_path": "/kassabuch"', html)

    def test_cashbook_requires_login(self):
        response = eegapp.app.test_client().get('/kassabuch')
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response.headers['Location'])


if __name__ == '__main__':
    unittest.main()
