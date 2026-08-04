#!/usr/bin/env python3
"""EEG Web-Oberfläche - Hauptanwendung."""

import os
import sys
import sqlite3
import secrets
import re
import json
import threading
import time
import base64
import hashlib
import ipaddress
import io
import unicodedata
import mimetypes
from datetime import datetime, date, timezone, timedelta
from functools import wraps
from email.header import Header
from email.utils import formataddr
from html import escape
from urllib.parse import urlparse, urljoin, urlencode
from zoneinfo import ZoneInfo

from flask import (Flask, render_template, request, redirect, url_for,
                   flash, send_file, jsonify, g, abort, get_flashed_messages)
from flask import has_request_context
from flask import session
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                         login_required, current_user)
from flask_wtf.csrf import CSRFProtect, CSRFError, generate_csrf
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix

from core.security import (
    MIN_PASSWORD_LENGTH,
    is_safe_redirect_url as _is_safe_redirect_url,
    sanitize_newsletter_html,
    validate_password,
)
from services.sepa import (
    EPC_MAX_PAYLOAD_BYTES,
    build_epc_payload,
    normalize_iban,
    render_epc_qr_svg,
    sepa_text as _sepa_text,
)
from services.billing import calculate_billing as _calculate_billing

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env'))
except ImportError:
    pass

# App-Pfad setzen
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# EEG_DB_PATH erlaubt es Tests und Wartungsskripten, gezielt auf eine andere
# Datenbank zu zeigen, statt versehentlich die Produktivdatenbank zu treffen.
DB_PATH = os.environ.get('EEG_DB_PATH') or os.path.join(BASE_DIR, '..', 'eeg_data.db')
UPLOAD_FOLDER = os.path.join(BASE_DIR, '..', 'data')
INVOICE_FOLDER = os.path.join(BASE_DIR, 'invoices')
BACKUP_FOLDER = os.path.join(BASE_DIR, '..', 'backups')
INSTANCE_DIR = os.path.join(BASE_DIR, '..', 'instance')
APP_TIMEZONE = ZoneInfo(os.environ.get('EEG_TIMEZONE', 'Europe/Vienna'))

app = Flask(__name__)
_IS_PRODUCTION = os.environ.get('EEG_ENV', '').lower() == 'production' or os.environ.get('FLASK_ENV') == 'production'
_SECRET_KEY = os.environ.get('EEG_SECRET_KEY')
if _IS_PRODUCTION and not _SECRET_KEY:
    raise RuntimeError('EEG_SECRET_KEY muss im Produktivbetrieb gesetzt sein.')
app.config['SECRET_KEY'] = _SECRET_KEY or secrets.token_hex(32)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB
app.config['WTF_CSRF_ENABLED'] = True
app.config['PREFERRED_URL_SCHEME'] = 'https'
app.config['SERVER_NAME_PUBLIC'] = os.environ.get('EEG_SERVER_NAME_PUBLIC', 'localhost')
app.config['PUBLIC_BASE_URL'] = os.environ.get('EEG_PUBLIC_BASE_URL', '').strip().rstrip('/')
app.config['SESSION_COOKIE_SECURE'] = _IS_PRODUCTION
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_PATH'] = '/'
# Absolute Obergrenze einer Sitzung (auch bei durchgehender Aktivitaet).
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(
    hours=int(os.environ.get('EEG_SESSION_MAX_HOURS', '8')))
app.config['SESSION_REFRESH_EACH_REQUEST'] = True
# Zusaetzliche Abmeldung nach Inaktivitaet.
SESSION_IDLE_TIMEOUT_SECONDS = int(os.environ.get('EEG_SESSION_IDLE_MINUTES', '60')) * 60
# Ab dieser Groesse des WAL-Journals warnt der Admin-Bereich.
WAL_WARN_BYTES = int(os.environ.get('EEG_WAL_WARN_MB', '500')) * 1024 * 1024
if os.environ.get('EEG_SESSION_COOKIE_DOMAIN'):
    app.config['SESSION_COOKIE_DOMAIN'] = os.environ['EEG_SESSION_COOKIE_DOMAIN']

DEFAULT_ORG_NAME = os.environ.get('EEG_ORG_NAME', 'EEG Portal')
DEFAULT_ORG_EMAIL = os.environ.get('EEG_ORG_EMAIL', 'office@example.org')
DEFAULT_ORG_WEBSITE = os.environ.get('EEG_ORG_WEBSITE', 'https://example.org/')
DEFAULT_ORG_ADDRESS = os.environ.get('EEG_ORG_ADDRESS', 'Adresse bitte konfigurieren')
DEFAULT_ORG_LEGAL = os.environ.get('EEG_ORG_LEGAL', 'Vereinsdaten bitte konfigurieren')

# Release Notes: Datum und kurze Beschreibung der letzten Änderungen.
# Die neueste Version steht immer an erster Stelle.
RELEASE_NOTES = [
    {
        'date': '2026-07-30',
        'title': 'Fehlende Verträge erkennen + PDF-Vorschau für nachgeladene Links',
        'changes': [
            'Neue Übersicht „Fehlende Verträge“ unter /admin/users zeigt aktive Mitglieder mit Benutzerkonto ohne hochgeladenen Bezieher- bzw. Einspeiser-Vertrag, inkl. Direktlink zum Hochladen.',
            'Einheitlich generierter Dateiname für neu hochgeladene Verträge (Typ, Name, Mitgliedsnummer, Datum).',
            'PDF-Vorschau erkennt jetzt auch Vertragslinks, die nachträglich per JavaScript in die Seite eingefügt werden (V1).',
            'PDF-Vorschau in V2 lädt korrekt die Embed-Ansicht statt der Download-URL.',
            'Fehler beim Download einzelner Verträge (500) behoben.',
        ],
    },
    {
        'date': '2026-07-29',
        'title': 'Release Notes Seite in V1 und V2',
        'changes': [
            'Neue Seite „Release Notes“ mit allen Änderungen chronologisch aufgelistet.',
            'In der klassischen UI unter /release-notes und in der V2-Oberfläche unter /v2/release-notes erreichbar.',
        ],
    },
    {
        'date': '2026-07-29',
        'title': 'V2 Zahlungen: Sortierung + QR-Code',
        'changes': [
            'Gebuchte Zahlungen in V2 sind sortierbar (Mitglied, Zeitraum, Betrag, Datum).',
            'QR-Code für Gutschriften im V2-Design (EPC069-12/GiroCode, Revolut-kompatibel).',
            'Parität der Zahlungsansicht zwischen V1 und V2.',
        ],
    },
    {
        'date': '2026-07-29',
        'title': 'Fortlaufende Belegnummern im Kassabuch',
        'changes': [
            'Belegnummer läuft je Jahr aufsteigend mit dem Buchungsdatum (z.B. 2026/001).',
            'Bisherige Referenznummer bleibt erhalten und verweist auf die manuelle Buchung bzw. Abrechnung.',
            'Suche findet sowohl Belegnummer als auch Referenz.',
            'Neue Spalte Referenz in CSV-, Excel- und PDF-Export.',
        ],
    },
    {
        'date': '2026-07-29',
        'title': 'Kassabuch-Bericht mit Zeitraum, Salden, Logo und Excel',
        'changes': [
            'Freier Berichtszeitraum von/bis mit Schnellwahl für laufendes Jahr, Vorjahr und Quartal.',
            'Anfangs- und Endsaldo je Zeitraum, getrennt nach bar und Bank.',
            'PDF-Export mit Logo, Vereinsdaten, Kennzahlenblock, Seitenzahlen und österreichischer Schreibweise.',
            'Excel-Export zusätzlich zu CSV mit zwei Blättern, fixierter Kopfzeile und Autofilter.',
            'Sortierung innerhalb eines Tages nach Belegnummer.',
            'Dateiname der Exporte enthält den Zeitraum.',
        ],
    },
    {
        'date': '2026-07-29',
        'title': 'Buchungsdaten an Revolut-Kontoauszug angeglichen',
        'changes': [
            'Altbestände und Greimer-Cent-Korrektur nach Kontoauszug korrigiert.',
            'Buchungsdatum nachträglich korrigierbar.',
        ],
    },
    {
        'date': '2026-07-28',
        'title': 'QR-Code zum Auszahlen offener Gutschriften',
        'changes': [
            'QR-Button bei offenen Gutschriften öffnet Popup mit SEPA-Überweisungscode (EPC069-12/GiroCode).',
            'Kompatibel mit Banking-Apps im SEPA-Raum, inklusive Revolut.',
            'Empfänger, IBAN, BIC, Betrag und Verwendungszweck werden kodiert; IBAN inklusive Prüfziffer geprüft.',
        ],
    },
    {
        'date': '2026-07-28',
        'title': 'Gebuchte Zahlungen sortierbar',
        'changes': [
            'Tabelle „Gebuchte Zahlungen“ lässt sich nach Mitglied, Zeitraum, Betrag und Buchungsdatum sortieren.',
            'Buchungen und Altbestände in einer gemeinsamen sortierten Liste.',
            'Aktive Spalte zeigt Sortierrichtung als Pfeil.',
        ],
    },
    {
        'date': '2026-07-28',
        'title': 'Altbestand-Buchungen korrigierbar',
        'changes': [
            'Route /payments/legacy_booking trägt den tatsächlichen Betrag für alte Abrechnungen nach.',
            'Altbestand-Zeilen haben Eingabefelder für Betrag und Buchungsdatum.',
            'Änderungsgrund ist Pflicht; Vorgang erscheint in der Historie.',
        ],
    },
    {
        'date': '2026-07-28',
        'title': 'Sicherheitsabfrage mit Pflicht-Änderungsgrund',
        'changes': [
            'Änderungen an Zahlungsbuchungen laufen über einen Bestätigungsdialog mit Detailübersicht.',
            'Pflichtfeld für Änderungsgrund beim Ändern, Stornieren und bei Abweichung vom Sollbetrag.',
            'Vollständige Änderungshistorie je Buchung als Popup.',
        ],
    },
    {
        'date': '2026-07-28',
        'title': 'Teilzahlungen, Buchungskorrektur und Mitgliedskonten',
        'changes': [
            'Je Abrechnung und Mitglied sind mehrere Buchungen möglich.',
            'Buchungsbetrag frei eingebbar; Differenz bleibt als Restforderung oder Guthaben offen.',
            'Bestehende Buchungen können in Betrag und Datum korrigiert werden.',
            'Einzelne Buchungen stornierbar; Guthaben durch negative Buchung rückerstattbar.',
            'Neue Seite „Mitgliedskonten“ mit Saldo je Mitglied und Kontoauszug.',
        ],
    },
    {
        'date': '2026-07-28',
        'title': 'Vereinskassabuch für Admins',
        'changes': [
            'Neue Tabellen cashbook_categories und cashbook_entries mit Startkategorien.',
            'Manuelle Buchungen (Bar/Überweisung) mit Begründung, Kategorie, Gegenpartei und Beleg-Upload.',
            'Strombewegungen aus Zahlungsbuchungen laufen automatisch mit.',
            'Laufender Saldo sowie Kassastand bar und Bank.',
            'Filter nach Jahr, Kategorie, Art, Zahlungsart und Text.',
            'Klassische Oberfläche und V2-Ansicht.',
        ],
    },
    {
        'date': '2026-07-28',
        'title': 'Sicherheitshärtung',
        'changes': [
            'Zentrale Passwortrichtlinie: mindestens 12 Zeichen, keine Trivialpasswörter, Benutzername nicht enthalten.',
            'Sitzungen laufen nach 60 Minuten Inaktivität bzw. spätestens nach 8 Stunden ab.',
            'Login-Sperre liegt in der Datenbank und überlebt Neustarts.',
            '404-Bremse gegen Scanner: 30 Fehlversuche in 5 Minuten sperren eine IP für 15 Minuten.',
            'Technische Ausnahmetexte erscheinen nicht mehr in der Oberfläche.',
        ],
    },
    {
        'date': '2026-07-28',
        'title': 'V2-Oberfläche (Beta)',
        'changes': [
            'Experimentelle V2-UI parallel zur Jinja-Oberfläche.',
            'Unterstützt Dashboard, Portal, Newsletter, Backup, Audit und Einstellungen.',
            'Zugriffskontrolle: Nicht-Admins erhalten nur Portal-Daten.',
        ],
    },
    {
        'date': '2026-07-28',
        'title': 'WSGI-Einstiegspunkt für gunicorn',
        'changes': [
            'Neue wsgi.py führt Initialisierung (init_db, Mail-Konfiguration, Backup-Scheduler) aus.',
            'Ermöglicht Betrieb unter gunicorn als unprivilegierter Benutzer.',
        ],
    },
    {
        'date': '2026-07-04',
        'title': 'Backup, Reporting und Sicherheit',
        'changes': [
            'Erweiterung von Abrechnung, Backup und Reporting.',
            'Sicherheitskontrollen gehärtet.',
        ],
    },
    {
        'date': '2026-07-03',
        'title': 'Backup- und Datenbank-Verwaltung',
        'changes': [
            'Admin-Backup-Restore-Seite.',
            'Konfigurierbare automatische Backups (lokal standardmäßig aktiviert).',
            'Datenbankwartungsseite mit manueller Löschung.',
            'Google Drive Backup-Upload mit OAuth-PKCE-Flow.',
            'Google Drive Backup-Health-Check.',
        ],
    },
    {
        'date': '2026-05-16',
        'title': 'Initialer Import und Konfiguration',
        'changes': [
            'Initialer Source-Import des EEG Portals.',
            'Öffentliche Einstellungen in allen Templates verfügbar.',
            'Newsletter-Abmeldung und Newsletter-Opt-in für Mitglieder verbessert.',
            'Audit-Log mit österreichischer Zeitdarstellung.',
        ],
    },
]

# Proxy-Fix: Hinter HAProxy/Nginx die echte Client-IP lesen
# x_for=1: Ein Proxy-Level (HAProxy/Nginx) leitet X-Forwarded-For weiter
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1, x_prefix=1)

csrf = CSRFProtect(app)

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(INVOICE_FOLDER, exist_ok=True)
os.makedirs(BACKUP_FOLDER, exist_ok=True)
os.makedirs(INSTANCE_DIR, exist_ok=True)

GOOGLE_DRIVE_SCOPES = ['https://www.googleapis.com/auth/drive.file']
GOOGLE_CLIENT_SECRETS_FILE = os.environ.get(
    'EEG_GOOGLE_CLIENT_SECRETS',
    os.path.join(INSTANCE_DIR, 'google_client_secret.json')
)
GOOGLE_TOKEN_FILE = os.environ.get(
    'EEG_GOOGLE_TOKEN_FILE',
    os.path.join(INSTANCE_DIR, 'google_drive_token.json')
)
GOOGLE_OAUTH_REDIRECT_URI = os.environ.get('EEG_GOOGLE_OAUTH_REDIRECT_URI', '')
RESTORE_MAX_FILES = int(os.environ.get('EEG_RESTORE_MAX_FILES', '1000'))
RESTORE_MAX_UNCOMPRESSED_BYTES = int(os.environ.get('EEG_RESTORE_MAX_UNCOMPRESSED_MB', '512')) * 1024 * 1024

BACKUP_SETTING_DEFAULTS = {
    'backup_auto_enabled': 'true',
    'backup_auto_time': '02:30',
    'backup_retention_daily': '3',
    'backup_retention_weekly': '4',
    'backup_retention_monthly': '6',
    'backup_retention_yearly': '3',
    'backup_email_enabled': 'false',
    'backup_email_weekday': '6',
    'backup_email_time': '03:00',
    'backup_email_to': '',
    'backup_email_max_mb': '20',
    'backup_drive_enabled': 'false',
    'backup_drive_folder_id': '',
    'backup_drive_last_upload': '',
    'backup_drive_last_check': '',
    'backup_drive_last_error': '',
    'backup_auto_last_run_date': '',
    'backup_email_last_attempt_week': '',
    'backup_email_last_sent_week': '',
}
BACKUP_JOB_LOCK = threading.Lock()
BACKUP_SCHEDULER_LOCK = threading.Lock()
BACKUP_SCHEDULER_STARTED = False


def local_now():
    return datetime.now(APP_TIMEZONE)


def utc_now_string():
    return datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')


def form_switch_enabled(name):
    return '1' in request.form.getlist(name)


def to_local_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return None
        if text.endswith('Z'):
            text = text[:-1] + '+00:00'
        try:
            dt = datetime.fromisoformat(text.replace(' ', 'T'))
        except ValueError:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(APP_TIMEZONE)


def local_day_bounds_as_utc_strings(day_text=None):
    """Lokale Tagesgrenzen fuer SQLite-UTC-Zeitstempel."""
    if day_text:
        day = datetime.strptime(day_text, '%Y-%m-%d').date()
    else:
        day = local_now().date()
    start_local = datetime.combine(day, datetime.min.time(), tzinfo=APP_TIMEZONE)
    end_local = datetime.combine(day, datetime.max.time().replace(microsecond=0), tzinfo=APP_TIMEZONE)
    start_utc = start_local.astimezone(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
    end_utc = end_local.astimezone(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
    return start_utc, end_utc


@app.template_filter('localdatetime')
def format_local_datetime(value, fmt='%d.%m.%Y %H:%M'):
    dt = to_local_datetime(value)
    return dt.strftime(fmt) if dt else '—'


@app.template_filter('localdate')
def format_local_date(value, fmt='%d.%m.%Y'):
    dt = to_local_datetime(value)
    return dt.strftime(fmt) if dt else '—'


@app.template_filter('euro')
def format_euro(value):
    """Betrag in oesterreichischer Schreibweise: 1.234,56"""
    try:
        formatted = f'{float(value):,.2f}'
    except (TypeError, ValueError):
        return '0,00'
    return formatted.replace(',', '#').replace('.', ',').replace('#', '.')


@app.context_processor
def inject_template_globals():
    public_cfg = {
        'org_name': DEFAULT_ORG_NAME,
        'org_email': DEFAULT_ORG_EMAIL,
        'org_website': DEFAULT_ORG_WEBSITE,
        'org_address': DEFAULT_ORG_ADDRESS,
        'org_legal': DEFAULT_ORG_LEGAL,
        'payment_bic': '',
        'payment_iban': '',
        'payment_recipient': DEFAULT_ORG_NAME,
    }
    try:
        public_cfg.update(get_public_config(get_db()))
    except Exception:
        pass
    return {
        'now': local_now(),
        'public_cfg': public_cfg,
        'org_name': public_cfg['org_name'],
        'org_email': public_cfg['org_email'],
        'org_website': public_cfg['org_website'],
        'org_address': public_cfg['org_address'],
        'org_legal': public_cfg['org_legal'],
        'timezone_name': getattr(APP_TIMEZONE, 'key', 'Europe/Vienna'),
        'min_password_length': MIN_PASSWORD_LENGTH,
    }


def is_safe_redirect_url(target):
    """Erlaubt nur relative oder gleiche Host-Weiterleitungen."""
    return _is_safe_redirect_url(target, request.host_url)


# === Fehlerausgabe ===
class UserError(Exception):
    """Fachlicher Fehler, dessen Text dem Nutzer gezeigt werden darf."""


# Fachliche Validierungsfehler werden im Projekt bereits als ValueError geworfen.
_USER_SAFE_EXCEPTIONS = (UserError, ValueError)


def flash_exception(exc, fallback='Aktion fehlgeschlagen.', category='danger'):
    """Zeigt fachliche Fehler im Klartext, technische nur mit Referenz-ID.

    Technische Details (Dateipfade, SQL, SMTP-Antworten) landen ausschliesslich
    im Serverlog und koennen ueber die Referenz zugeordnet werden.
    """
    if isinstance(exc, _USER_SAFE_EXCEPTIONS):
        flash(str(exc), category)
        return None
    ref = secrets.token_hex(4)
    app.logger.exception('Fehler %s (%s)', ref, fallback)
    flash(f'{fallback} (Referenz: {ref})', category)
    return ref


def public_base_url():
    """Liefert die oeffentliche Basis-URL fuer Links in E-Mails."""
    configured = (app.config.get('PUBLIC_BASE_URL') or '').strip().rstrip('/')
    if configured:
        parsed = urlparse(configured)
        if parsed.scheme in {'http', 'https'} and parsed.netloc:
            return configured
        app.logger.warning('Ignoring invalid EEG_PUBLIC_BASE_URL: %s', configured)

    public_host = (app.config.get('SERVER_NAME_PUBLIC') or '').strip().rstrip('/')
    if public_host and public_host != 'localhost':
        if '://' in public_host:
            parsed = urlparse(public_host)
            if parsed.scheme in {'http', 'https'} and parsed.netloc:
                return public_host.rstrip('/')
        return f'https://{public_host}'

    if has_request_context():
        return request.url_root.rstrip('/')
    return 'http://localhost'


def public_url_for(endpoint, **values):
    """Erzeugt absolute URLs mit der oeffentlichen Basisadresse."""
    return urljoin(public_base_url() + '/', url_for(endpoint, **values).lstrip('/'))


def _hostname_without_port(host):
    text = (host or '').strip().lower()
    if not text:
        return ''
    if text.startswith('[') and ']' in text:
        return text[1:text.index(']')]
    return text.rsplit(':', 1)[0] if ':' in text else text


def _is_internal_hostname(hostname):
    if hostname in {'localhost', '127.0.0.1', '::1'}:
        return True
    try:
        ip = ipaddress.ip_address(hostname)
        return ip.is_private or ip.is_loopback or ip.is_link_local
    except ValueError:
        return False


def initial_password_hash():
    """Erzeugt sichere Initial-Passwoerter ohne fest codierten Default."""
    password = os.environ.get('EEG_INITIAL_ADMIN_PASSWORD')
    if password:
        return generate_password_hash(password)
    if _IS_PRODUCTION:
        raise RuntimeError('EEG_INITIAL_ADMIN_PASSWORD muss fuer neue Admins im Produktivbetrieb gesetzt sein.')
    app.logger.warning('Kein EEG_INITIAL_ADMIN_PASSWORD gesetzt; neuer Admin erhaelt ein zufaelliges Passwort.')
    return generate_password_hash(secrets.token_urlsafe(32))


def safe_extract_zip_member(zf, member_name, destination):
    """Extrahiert nur Dateien, die im erwarteten Zielverzeichnis bleiben."""
    normalized = os.path.normpath(member_name).replace('\\', '/')
    if normalized.startswith('../') or normalized.startswith('/') or '/..' in normalized:
        raise ValueError(f'Ungueltiger ZIP-Pfad: {member_name}')
    target_path = os.path.abspath(os.path.join(destination, normalized))
    destination_abs = os.path.abspath(destination)
    if not target_path.startswith(destination_abs + os.sep) and target_path != destination_abs:
        raise ValueError(f'Ungueltiger ZIP-Zielpfad: {member_name}')
    os.makedirs(os.path.dirname(target_path), exist_ok=True)
    with zf.open(member_name) as source, open(target_path, 'wb') as target:
        target.write(source.read())
    return target_path


def validate_backup_zip(zf):
    """Prueft Backup-ZIPs vor dem Entpacken gegen unerwartete Dateien und ZIP-Bomben."""
    infos = [info for info in zf.infolist() if not info.is_dir()]
    if len(infos) > RESTORE_MAX_FILES:
        raise ValueError(f'Backup enthaelt zu viele Dateien ({len(infos)} > {RESTORE_MAX_FILES}).')

    total_size = sum(info.file_size for info in infos)
    if total_size > RESTORE_MAX_UNCOMPRESSED_BYTES:
        max_mb = RESTORE_MAX_UNCOMPRESSED_BYTES / 1024 / 1024
        raise ValueError(f'Backup ist entpackt zu gross ({total_size / 1024 / 1024:.1f} MB > {max_mb:.0f} MB).')

    names = []
    for info in infos:
        name = info.filename
        normalized = os.path.normpath(name).replace('\\', '/')
        if normalized.startswith('../') or normalized.startswith('/') or '/..' in normalized:
            raise ValueError(f'Ungueltiger ZIP-Pfad: {name}')
        if normalized not in {'eeg_data.db', 'backup_manifest.txt'} and not normalized.startswith('invoices/'):
            raise ValueError(f'Unerwartete Datei im Backup: {name}')
        names.append(normalized)

    if 'eeg_data.db' not in names:
        raise ValueError('Ungueltiges Backup: eeg_data.db nicht gefunden.')
    return names


@app.before_request
def enforce_allowed_country():
    """Optionaler Laenderblock, gedacht fuer Cloudflare/Reverse-Proxy-Header."""
    allowed = {
        c.strip().upper()
        for c in os.environ.get('EEG_ALLOWED_COUNTRIES', '').split(',')
        if c.strip()
    }
    if not allowed:
        return None
    country = (request.headers.get('CF-IPCountry')
               or request.headers.get('X-Country-Code')
               or '').upper()
    if country not in allowed:
        abort(403)
    return None


@app.before_request
def redirect_internal_host_to_public_url():
    """Verhindert Browser-Sessions ueber interne HTTP-Adressen im Produktivbetrieb."""
    if not _IS_PRODUCTION:
        return None
    redirect_enabled = os.environ.get('EEG_REDIRECT_INTERNAL_HTTP', '').strip().lower() in {'1', 'true', 'yes', 'on'}
    if not redirect_enabled:
        return None
    base_url = public_base_url()
    parsed = urlparse(base_url)
    if not parsed.scheme or not parsed.netloc or parsed.hostname in {None, 'localhost'}:
        return None
    current_hostname = _hostname_without_port(request.host)
    public_hostname = (parsed.hostname or '').lower()
    if current_hostname == public_hostname:
        return None
    if request.is_secure or request.headers.get('X-Forwarded-Proto', '').lower() == 'https':
        return None
    if not _is_internal_hostname(current_hostname):
        return None
    if request.method not in ('GET', 'HEAD', 'OPTIONS'):
        return redirect(urljoin(base_url + '/', 'login?csrf=1'), code=303)
    target = urljoin(base_url + '/', request.full_path.lstrip('/'))
    if target.endswith('?'):
        target = target[:-1]
    return redirect(target, code=302)


@app.errorhandler(CSRFError)
def handle_csrf_error(error):
    app.logger.warning(
        'CSRF validation failed | reason=%s | host=%s | path=%s | secure=%s',
        getattr(error, 'description', str(error)),
        request.host,
        request.path,
        request.is_secure,
    )
    if current_user and current_user.is_authenticated:
        flash('Die Sicherheitsprüfung ist abgelaufen. Bitte Aktion erneut ausführen.', 'warning')
        return redirect(url_for('dashboard' if current_user.is_admin else 'portal_dashboard'))
    return redirect(public_url_for('login', csrf='1'), code=303)


@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    embed_mode = request.args.get('embed') == '1'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN' if embed_mode else 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    frame_ancestors = "'self'" if embed_mode else "'none'"
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "base-uri 'self'; "
        "form-action 'self'; "
        f"frame-ancestors {frame_ancestors}; "
        "img-src 'self' data:; "
        "font-src 'self' https://cdn.jsdelivr.net data:; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://code.highcharts.com; "
        "connect-src 'self'; "
        "object-src 'none'"
    )
    if _IS_PRODUCTION and (request.is_secure or request.headers.get('X-Forwarded-Proto', '').lower() == 'https'):
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Bitte einloggen.'


@login_manager.unauthorized_handler
def unauthorized():
    flash(login_manager.login_message, 'warning')
    if request.path.startswith('/v2'):
        return redirect(url_for('v2_login', next=request.url))
    return redirect(url_for('login', next=request.url))


# === Database ===

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def ensure_import_schema(db):
    """Stellt sicher, dass die EDA-Importtabellen vorhanden sind."""
    import_schema_path = os.path.join(BASE_DIR, '..', 'schema.sql')
    if os.path.exists(import_schema_path):
        with open(import_schema_path) as f:
            db.executescript(f.read())
    for col, coldef in [
        ('data_status', "TEXT NOT NULL DEFAULT 'final'"),
        ('replaced_by_batch_id', 'INTEGER'),
        ('replaced_at', 'TEXT'),
    ]:
        try:
            db.execute(f"ALTER TABLE import_batches ADD COLUMN {col} {coldef}")
        except sqlite3.OperationalError:
            pass
    db.execute("UPDATE import_batches SET data_status='final' WHERE data_status IS NULL OR data_status=''")


def init_db():
    """Schema initialisieren und Admin-User anlegen."""
    db = sqlite3.connect(DB_PATH)
    schema_path = os.path.join(BASE_DIR, 'schema_web.sql')
    with open(schema_path) as f:
        db.executescript(f.read())
    ensure_import_schema(db)
    # Users: member_id, role, invite_token, invite_expires vor Admin-Anlage migrieren
    for col, coldef in [('member_id', 'INTEGER'), ('role', "TEXT DEFAULT 'member'"),
                        ('invite_token', 'TEXT'), ('invite_expires', 'TEXT')]:
        try:
            db.execute(f"ALTER TABLE users ADD COLUMN {col} {coldef}")
        except sqlite3.OperationalError:
            pass
    # Admin-User anlegen falls nicht vorhanden
    existing = db.execute("SELECT id FROM users WHERE username='SuperAdmin'").fetchone()
    if not existing:
        # Auch alten 'admin' User prüfen
        old_admin = db.execute("SELECT id FROM users WHERE username='admin'").fetchone()
        if not old_admin:
            pw_hash = initial_password_hash()
            db.execute("INSERT INTO users (username, password_hash, is_admin, role) VALUES (?, ?, 1, 'admin')",
                       ('SuperAdmin', pw_hash))
        db.commit()
    # Bank-Felder zu members hinzufügen (Migration)
    try:
        db.execute("ALTER TABLE members ADD COLUMN iban TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        db.execute("ALTER TABLE members ADD COLUMN bic TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        db.execute("ALTER TABLE members ADD COLUMN account_holder TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        db.execute("ALTER TABLE members ADD COLUMN phone TEXT")
    except sqlite3.OperationalError:
        pass
    # Users: member_id, role, invite_token, invite_expires
    for col, coldef in [('member_id', 'INTEGER'), ('role', "TEXT DEFAULT 'member'"),
                        ('invite_token', 'TEXT'), ('invite_expires', 'TEXT')]:
        try:
            db.execute(f"ALTER TABLE users ADD COLUMN {col} {coldef}")
        except sqlite3.OperationalError:
            pass
    # Bestehende admins markieren
    db.execute("UPDATE users SET role='admin' WHERE is_admin=1 AND (role IS NULL OR role='')")
    db.execute("UPDATE users SET role='member' WHERE is_admin=0 AND (role IS NULL OR role='')")
    # Erzwungener Passwortwechsel: Bestandsadmins einmalig markieren, weil ihre
    # Passwoerter noch unter der alten Mindestlaenge von 6 Zeichen entstanden sind.
    try:
        db.execute("ALTER TABLE users ADD COLUMN password_change_required INTEGER DEFAULT 0")
        db.execute("UPDATE users SET password_change_required=1 WHERE is_admin=1")
    except sqlite3.OperationalError:
        pass
    # Persistente Login-Sperre (ueberlebt Neustarts)
    db.execute("""CREATE TABLE IF NOT EXISTS login_attempts (
        ip TEXT PRIMARY KEY,
        count INTEGER NOT NULL DEFAULT 0,
        last_attempt REAL NOT NULL DEFAULT 0,
        locked_until REAL NOT NULL DEFAULT 0
    )""")
    # Kassabuch: frei pflegbare Kategorien
    db.execute("""CREATE TABLE IF NOT EXISTS cashbook_categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        direction TEXT NOT NULL DEFAULT 'both',   -- income, expense oder both
        active INTEGER NOT NULL DEFAULT 1,
        sort_order INTEGER NOT NULL DEFAULT 100,
        created_at TEXT NOT NULL DEFAULT (datetime('now'))
    )""")
    for name, cat_direction, sort_order in CASHBOOK_DEFAULT_CATEGORIES:
        db.execute("""INSERT OR IGNORE INTO cashbook_categories (name, direction, sort_order)
                      VALUES (?, ?, ?)""", (name, cat_direction, sort_order))
    # Kassabuch: manuelle Buchungen (Strombewegungen kommen aus payment_bookings)
    db.execute("""CREATE TABLE IF NOT EXISTS cashbook_entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        entry_date TEXT NOT NULL,              -- ISO date
        direction TEXT NOT NULL,               -- income oder expense
        amount_eur REAL NOT NULL,              -- immer positiv
        category_id INTEGER,
        payment_method TEXT NOT NULL,          -- cash oder transfer
        description TEXT NOT NULL,             -- Begruendung
        counterparty TEXT,                     -- Zahler bzw. Empfaenger
        document_number TEXT,                  -- fortlaufende Belegnummer
        receipt_filename TEXT,
        receipt_mimetype TEXT,
        receipt_data BLOB,
        created_at TEXT NOT NULL DEFAULT (datetime('now')),
        created_by TEXT,
        FOREIGN KEY (category_id) REFERENCES cashbook_categories(id)
    )""")
    db.execute("CREATE INDEX IF NOT EXISTS idx_cashbook_entries_date ON cashbook_entries(entry_date)")
    # Aenderungshistorie der Zahlungsbuchungen inklusive Begruendung
    db.execute("""CREATE TABLE IF NOT EXISTS payment_booking_changes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        booking_id INTEGER NOT NULL,
        changed_at TEXT NOT NULL DEFAULT (datetime('now')),
        changed_by_user_id INTEGER,
        changed_by_username TEXT,
        action TEXT NOT NULL,              -- create, edit, reverse
        old_amount_eur REAL,
        new_amount_eur REAL,
        old_booking_date TEXT,
        new_booking_date TEXT,
        reason TEXT NOT NULL DEFAULT '',
        FOREIGN KEY (booking_id) REFERENCES payment_bookings(id)
    )""")
    db.execute("""CREATE INDEX IF NOT EXISTS idx_payment_booking_changes_booking
                  ON payment_booking_changes(booking_id)""")
    # Contracts-Tabelle
    db.execute("""CREATE TABLE IF NOT EXISTS contracts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        member_id INTEGER NOT NULL,
        type TEXT NOT NULL,
        filename TEXT NOT NULL,
        file_data BLOB NOT NULL,
        uploaded_at TEXT NOT NULL DEFAULT (datetime('now')),
        uploaded_by TEXT,
        FOREIGN KEY (member_id) REFERENCES members(id)
    )""")
    # Audit-Log-Tabelle
    db.execute("""CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT NOT NULL DEFAULT (datetime('now')),
        user_id INTEGER,
        username TEXT,
        action TEXT NOT NULL,
        detail TEXT,
        ip TEXT,
        url TEXT,
        method TEXT
    )""")
    db.execute("CREATE INDEX IF NOT EXISTS idx_audit_timestamp ON audit_log(timestamp)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_audit_user ON audit_log(user_id)")
    # Settings-Tabelle für SMTP etc.
    db.execute("""CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )""")
    # Defaults setzen falls leer
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('smtp_host', 'mail.your-server.de')")
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('smtp_port', '587')")
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('smtp_user', '')")
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('smtp_pass', '')")
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('smtp_from', '')")
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('smtp_tls', 'true')")
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('mail_from_address', '')")
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('mail_from_name', ?)", (DEFAULT_ORG_NAME,))
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('mail_reply_to', '')")
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('mail_reply_to_name', ?)", (DEFAULT_ORG_NAME,))
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('org_name', ?)", (DEFAULT_ORG_NAME,))
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('org_email', ?)", (DEFAULT_ORG_EMAIL,))
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('org_website', ?)", (DEFAULT_ORG_WEBSITE,))
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('org_address', ?)", (DEFAULT_ORG_ADDRESS,))
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('org_legal', ?)", (DEFAULT_ORG_LEGAL,))
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('payment_bic', '')")
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('payment_iban', '')")
    db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('payment_recipient', ?)", (DEFAULT_ORG_NAME,))
    for key, value in BACKUP_SETTING_DEFAULTS.items():
        db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)", (key, value))
    db.execute("""CREATE TABLE IF NOT EXISTS oauth_pkce_sessions (
        id TEXT PRIMARY KEY,
        user_id INTEGER,
        state TEXT NOT NULL,
        code_verifier TEXT NOT NULL,
        created_at TEXT NOT NULL,
        expires_at TEXT NOT NULL
    )""")
    # Zahlungsstatus und Buchungsjournal
    for col, coldef in [('paid', 'INTEGER DEFAULT 0'), ('paid_at', 'TEXT')]:
        try:
            db.execute(f"ALTER TABLE invoice_items ADD COLUMN {col} {coldef}")
        except sqlite3.OperationalError:
            pass
    for table, col, coldef in [
        ('invoices', 'data_status', "TEXT NOT NULL DEFAULT 'final'"),
        ('import_log', 'data_status', "TEXT NOT NULL DEFAULT 'final'"),
    ]:
        try:
            db.execute(f"ALTER TABLE {table} ADD COLUMN {col} {coldef}")
        except sqlite3.OperationalError:
            pass
    db.execute("UPDATE invoices SET data_status='final' WHERE data_status IS NULL OR data_status=''")
    db.execute("UPDATE import_log SET data_status='final' WHERE data_status IS NULL OR data_status=''")
    db.execute("""CREATE TABLE IF NOT EXISTS payment_bookings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_id INTEGER NOT NULL,
        member_id INTEGER NOT NULL,
        amount_eur REAL NOT NULL,
        direction TEXT NOT NULL,
        booking_date TEXT NOT NULL,
        recorded_at TEXT NOT NULL DEFAULT (datetime('now')),
        recorded_by_user_id INTEGER,
        recorded_by_username TEXT,
        note TEXT,
        reversed_at TEXT,
        reversed_by_user_id INTEGER,
        reversed_by_username TEXT,
        reverse_note TEXT,
        FOREIGN KEY (invoice_id) REFERENCES invoices(id),
        FOREIGN KEY (member_id) REFERENCES members(id),
        FOREIGN KEY (recorded_by_user_id) REFERENCES users(id),
        FOREIGN KEY (reversed_by_user_id) REFERENCES users(id)
    )""")
    db.execute("CREATE INDEX IF NOT EXISTS idx_payment_bookings_member ON payment_bookings(member_id, booking_date)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_payment_bookings_invoice_member ON payment_bookings(invoice_id, member_id)")
    db.execute("""CREATE TABLE IF NOT EXISTS invoice_carryovers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_id INTEGER NOT NULL,
        member_id INTEGER NOT NULL,
        source_invoice_id INTEGER NOT NULL,
        amount_eur REAL NOT NULL,
        description TEXT,
        created_at TEXT NOT NULL DEFAULT (datetime('now')),
        FOREIGN KEY (invoice_id) REFERENCES invoices(id),
        FOREIGN KEY (member_id) REFERENCES members(id),
        FOREIGN KEY (source_invoice_id) REFERENCES invoices(id),
        UNIQUE(invoice_id, member_id, source_invoice_id)
    )""")
    db.execute("CREATE INDEX IF NOT EXISTS idx_invoice_carryovers_invoice_member ON invoice_carryovers(invoice_id, member_id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_invoice_carryovers_source ON invoice_carryovers(source_invoice_id, member_id)")
    # Newsletter-Tabellen
    db.execute("""CREATE TABLE IF NOT EXISTS newsletters (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        subject TEXT NOT NULL,
        body_html TEXT NOT NULL,
        created_by TEXT,
        created_at TEXT NOT NULL DEFAULT (datetime('now')),
        sent_at TEXT,
        recipients_count INTEGER DEFAULT 0
    )""")
    db.execute("""CREATE TABLE IF NOT EXISTS newsletter_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        newsletter_id INTEGER NOT NULL,
        member_id INTEGER NOT NULL,
        email TEXT NOT NULL,
        status TEXT NOT NULL,
        error_message TEXT,
        sent_at TEXT NOT NULL DEFAULT (datetime('now')),
        FOREIGN KEY (newsletter_id) REFERENCES newsletters(id),
        FOREIGN KEY (member_id) REFERENCES members(id)
    )""")
    # Newsletter-Opt-out Spalte in members
    try:
        db.execute("ALTER TABLE members ADD COLUMN newsletter_optout INTEGER DEFAULT 0")
    except sqlite3.OperationalError:
        pass
    # Unsubscribe-Token in members
    try:
        db.execute("ALTER TABLE members ADD COLUMN unsubscribe_token TEXT")
    except sqlite3.OperationalError:
        pass
    db.commit()
    db.close()


def _create_named_admin(db, username, member_id, email):
    """Erstellt einen Admin-User falls noch nicht vorhanden."""
    existing = db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
    if not existing:
        pw_hash = initial_password_hash()
        db.execute("""INSERT INTO users (username, password_hash, email, is_admin, role, member_id)
                      VALUES (?, ?, ?, 1, 'admin', ?)""",
                   (username, pw_hash, email, member_id))


def _is_valid_email(address):
    """Einfache E-Mail-Validierung für Header/SMTP-Konfiguration."""
    if not address:
        return False
    return bool(re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', address.strip()))


def _mail_header(name, address):
    """Erzeugt RFC-konformen Address-Header mit UTF-8 Anzeigename."""
    return formataddr((str(Header(name or '', 'utf-8')), address))


def _load_mail_config(db):
    """Lädt SMTP- und Mail-Absenderkonfiguration aus settings."""
    rows = db.execute("SELECT key, value FROM settings").fetchall()
    cfg = {r['key']: r['value'] for r in rows}

    smtp_user = (cfg.get('smtp_user') or '').strip()
    from_address = (cfg.get('mail_from_address') or cfg.get('smtp_from') or smtp_user).strip()
    from_name = (cfg.get('mail_from_name') or DEFAULT_ORG_NAME).strip()
    reply_to_address = (cfg.get('mail_reply_to') or from_address).strip()
    reply_to_name = (cfg.get('mail_reply_to_name') or DEFAULT_ORG_NAME).strip()
    smtp_tls = (cfg.get('smtp_tls') or 'true').strip().lower() in ('1', 'true', 'yes', 'on')

    return {
        'smtp_host': (cfg.get('smtp_host') or '').strip(),
        'smtp_port': int((cfg.get('smtp_port') or '587').strip() or '587'),
        'smtp_user': smtp_user,
        'smtp_pass': cfg.get('smtp_pass') or '',
        'smtp_tls': smtp_tls,
        'from_address': from_address,
        'from_name': from_name,
        'reply_to_address': reply_to_address,
        'reply_to_name': reply_to_name,
        'from_header': _mail_header(from_name, from_address) if from_address else '',
        'reply_to_header': _mail_header(reply_to_name, reply_to_address) if reply_to_address else '',
    }


def _validate_mail_config(mail_cfg):
    """Validiert Mail-Konfiguration gemäß RFC/Anwendungsanforderungen."""
    if not mail_cfg.get('smtp_user'):
        return False, 'SMTP-Benutzername fehlt.'
    if not mail_cfg.get('smtp_host'):
        return False, 'SMTP-Server fehlt.'
    if not mail_cfg.get('smtp_pass'):
        return False, 'SMTP-Passwort fehlt.'
    if not mail_cfg.get('from_address'):
        return False, 'Absenderadresse für E-Mails fehlt.'
    if not _is_valid_email(mail_cfg.get('from_address')):
        return False, 'Absenderadresse für E-Mails ist ungültig.'
    if not _is_valid_email(mail_cfg.get('reply_to_address')):
        return False, 'Antwortadresse ist ungültig.'

    smtp_user = mail_cfg.get('smtp_user').lower()
    from_addr = mail_cfg.get('from_address').lower()
    if from_addr != smtp_user:
        smtp_domain = smtp_user.split('@')[-1] if '@' in smtp_user else ''
        from_domain = from_addr.split('@')[-1] if '@' in from_addr else ''
        if not smtp_domain or smtp_domain != from_domain:
            return False, 'Absenderadresse muss dem SMTP-Benutzernamen oder einer Alias-Adresse derselben Domain entsprechen.'

    return True, ''


def _get_valid_mail_config(db):
    """Lädt und validiert Mail-Konfiguration; wirft RuntimeError bei Fehlern."""
    mail_cfg = _load_mail_config(db)
    ok, error = _validate_mail_config(mail_cfg)
    if not ok:
        raise RuntimeError(error)
    return mail_cfg


def get_public_config(db):
    rows = db.execute("""SELECT key, value FROM settings WHERE key IN (
        'org_name', 'org_email', 'org_website', 'org_address', 'org_legal',
        'payment_bic', 'payment_iban', 'payment_recipient'
    )""").fetchall()
    cfg = {r['key']: r['value'] for r in rows}
    return {
        'org_name': cfg.get('org_name') or DEFAULT_ORG_NAME,
        'org_email': cfg.get('org_email') or DEFAULT_ORG_EMAIL,
        'org_website': cfg.get('org_website') or DEFAULT_ORG_WEBSITE,
        'org_address': cfg.get('org_address') or DEFAULT_ORG_ADDRESS,
        'org_legal': cfg.get('org_legal') or DEFAULT_ORG_LEGAL,
        'payment_bic': cfg.get('payment_bic') or '',
        'payment_iban': cfg.get('payment_iban') or '',
        'payment_recipient': cfg.get('payment_recipient') or cfg.get('org_name') or DEFAULT_ORG_NAME,
    }


def _log_mail_send(mail_cfg, recipient, subject):
    """Loggt Versandparameter ohne sensitive Daten (kein Passwort)."""
    app.logger.info(
        'Sending mail | SMTP host: %s | SMTP user: %s | From: %s | Reply-To: %s | To: %s | Subject: %s',
        mail_cfg.get('smtp_host'),
        mail_cfg.get('smtp_user'),
        mail_cfg.get('from_header'),
        mail_cfg.get('reply_to_header'),
        recipient,
        subject,
    )


def _format_invite_expires(invite_expires):
    """Formatiert das Ablaufdatum für Einladungs-Mails."""
    if not invite_expires:
        return 'in 14 Tagen'
    try:
        return datetime.fromisoformat(str(invite_expires)).strftime('%d.%m.%Y um %H:%M Uhr')
    except ValueError:
        return str(invite_expires)


def _build_invitation_email(member_name, username, role, invite_url, invite_expires, public_cfg, logo_src=None):
    """Erzeugt Betreff, Text- und HTML-Teil für Portal-Einladungen."""
    display_name = (member_name or username or 'Mitglied').strip()
    role_label = 'Administrator' if role == 'admin' else 'Teilnehmer'
    expires_text = _format_invite_expires(invite_expires)
    org_name = public_cfg.get('org_name') or DEFAULT_ORG_NAME
    org_email = public_cfg.get('org_email') or DEFAULT_ORG_EMAIL
    org_address = public_cfg.get('org_address') or ''
    org_website = public_cfg.get('org_website') or ''

    subject = f'Einladung zum {org_name}'
    body_text = f"""Hallo {display_name},

Christian und Markus von der EEG haben für Sie einen Zugang zum {org_name} eingerichtet.

So starten Sie:
1. Öffnen Sie den folgenden Einladungslink:
{invite_url}
2. Legen Sie Ihr eigenes Passwort fest.
3. Melden Sie sich danach mit Ihrem Benutzernamen an: {username}

Der Link ist bis {expires_text} gültig. Falls der Link abgelaufen ist, antworten Sie bitte auf diese E-Mail oder wenden Sie sich an {org_email}.

Ihre Rolle im Portal: {role_label}

Viele Grüße
Christian und Markus
von der EEG
"""

    safe_name = escape(display_name)
    safe_username = escape(username or '')
    safe_role = escape(role_label)
    safe_url = escape(invite_url)
    safe_expires = escape(expires_text)
    safe_org_name = escape(org_name)
    safe_org_email = escape(org_email)
    safe_org_address = escape(org_address)
    safe_org_website = escape(org_website)
    safe_logo_src = escape(logo_src) if logo_src else ''

    html_footer_website = (
        f'<br><a href="{safe_org_website}" style="color:#2b5e3a;text-decoration:none;">{safe_org_website}</a>'
        if org_website else ''
    )
    logo_html = (
        f'<img src="{safe_logo_src}" width="58" height="58" alt="{safe_org_name}" '
        'style="display:block;border-radius:8px;margin:0 0 12px 0;background:#ffffff;">'
        if logo_src else ''
    )
    body_html = f"""<!doctype html>
<html lang="de">
<body style="margin:0;padding:0;background:#f5f7f4;font-family:Arial,Helvetica,sans-serif;color:#1f2a24;">
  <div style="display:none;max-height:0;overflow:hidden;color:transparent;">
    Christian und Markus von der EEG laden Sie zum Portal ein.
  </div>
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#f5f7f4;padding:28px 12px;">
    <tr>
      <td align="center">
        <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="max-width:640px;background:#ffffff;border-radius:10px;overflow:hidden;border:1px solid #dfe7df;">
          <tr>
            <td style="background:#2b5e3a;padding:24px 28px;color:#ffffff;">
              {logo_html}
              <div style="font-size:13px;letter-spacing:.04em;text-transform:uppercase;opacity:.85;">Einladung zum Portal</div>
              <h1 style="margin:8px 0 0 0;font-size:24px;line-height:1.25;font-weight:700;">{safe_org_name}</h1>
            </td>
          </tr>
          <tr>
            <td style="padding:28px;">
              <p style="margin:0 0 16px 0;font-size:16px;line-height:1.6;">Hallo {safe_name},</p>
              <p style="margin:0 0 16px 0;font-size:16px;line-height:1.6;">
                Christian und Markus von der EEG haben für Sie einen Zugang zum <strong>{safe_org_name}</strong> eingerichtet.
              </p>
              <div style="background:#eef6ef;border-left:4px solid #2b5e3a;padding:16px 18px;margin:22px 0;border-radius:6px;">
                <p style="margin:0 0 8px 0;font-size:15px;line-height:1.5;"><strong>Ihre Zugangsdaten</strong></p>
                <p style="margin:0;font-size:15px;line-height:1.6;">Benutzername: <strong>{safe_username}</strong><br>Rolle: <strong>{safe_role}</strong></p>
              </div>
              <p style="margin:0 0 12px 0;font-size:16px;line-height:1.6;"><strong>So starten Sie:</strong></p>
              <ol style="margin:0 0 22px 20px;padding:0;font-size:16px;line-height:1.7;">
                <li>Einladungslink öffnen.</li>
                <li>Eigenes Passwort festlegen.</li>
                <li>Danach mit Ihrem Benutzernamen anmelden.</li>
              </ol>
              <p style="margin:0 0 24px 0;text-align:center;">
                <a href="{safe_url}" style="display:inline-block;background:#2b5e3a;color:#ffffff;text-decoration:none;font-weight:700;padding:13px 22px;border-radius:6px;font-size:16px;">Einladung annehmen</a>
              </p>
              <p style="margin:0 0 16px 0;font-size:14px;line-height:1.6;color:#58665e;">
                Der Link ist bis <strong>{safe_expires}</strong> gültig. Falls der Button nicht funktioniert, kopieren Sie diesen Link in Ihren Browser:
              </p>
              <p style="margin:0 0 22px 0;word-break:break-all;font-size:13px;line-height:1.5;color:#2b5e3a;">
                <a href="{safe_url}" style="color:#2b5e3a;">{safe_url}</a>
              </p>
              <p style="margin:0;font-size:16px;line-height:1.6;">
                Viele Grüße<br>
                <strong>Christian und Markus</strong><br>
                von der EEG
              </p>
            </td>
          </tr>
          <tr>
            <td style="padding:18px 28px;background:#f0f4ef;border-top:1px solid #dfe7df;color:#6a766e;font-size:12px;line-height:1.5;text-align:center;">
              <strong>{safe_org_name}</strong><br>
              {safe_org_address}<br>
              <a href="mailto:{safe_org_email}" style="color:#2b5e3a;text-decoration:none;">{safe_org_email}</a>{html_footer_website}
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>"""

    return subject, body_text, body_html


def send_invitation_email(db, user_row, invite_url, invite_expires):
    """Sendet eine Portal-Einladung als HTML-Mail mit Plaintext-Fallback."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.image import MIMEImage

    recipient = (user_row['email'] or '').strip()
    if not _is_valid_email(recipient):
        raise RuntimeError('Keine gültige E-Mail-Adresse für diesen Benutzer hinterlegt.')

    mail_cfg = _get_valid_mail_config(db)
    public_cfg = get_public_config(db)
    member_name = user_row['member_name'] if 'member_name' in user_row.keys() else ''
    logo_cid = 'eeg-logo'
    subject, body_text, body_html = _build_invitation_email(
        member_name,
        user_row['username'],
        user_row['role'],
        invite_url,
        invite_expires,
        public_cfg,
        logo_src=f'cid:{logo_cid}',
    )

    msg = MIMEMultipart('related')
    msg['From'] = mail_cfg['from_header']
    msg['Reply-To'] = mail_cfg['reply_to_header']
    msg['To'] = recipient
    msg['Subject'] = subject

    msg_alt = MIMEMultipart('alternative')
    msg_alt.attach(MIMEText(body_text, 'plain', 'utf-8'))
    msg_alt.attach(MIMEText(body_html, 'html', 'utf-8'))
    msg.attach(msg_alt)

    logo_path = os.path.join(BASE_DIR, 'static', 'logo_small.png')
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo = MIMEImage(f.read(), _subtype='png')
        logo.add_header('Content-ID', f'<{logo_cid}>')
        logo.add_header('Content-Disposition', 'inline', filename='eeg-logo.png')
        msg.attach(logo)

    _log_mail_send(mail_cfg, recipient, subject)
    with smtplib.SMTP(mail_cfg['smtp_host'], mail_cfg['smtp_port']) as server:
        if mail_cfg['smtp_tls']:
            server.starttls()
        server.login(mail_cfg['smtp_user'], mail_cfg['smtp_pass'])
        server.send_message(msg, from_addr=mail_cfg['from_address'], to_addrs=[recipient])


def _startup_mail_config_check():
    """Prüft Mail-Konfiguration beim Start und loggt das Ergebnis."""
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    try:
        cfg = _load_mail_config(db)
        ok, error = _validate_mail_config(cfg)
        if ok:
            app.logger.info('Mail config check passed on startup. SMTP user=%s, From=%s',
                            cfg.get('smtp_user'), cfg.get('from_header'))
        else:
            app.logger.error('Mail config invalid on startup: %s', error)
    finally:
        db.close()


# === User Model ===

class User(UserMixin):
    def __init__(self, id, username, is_admin=False, member_id=None, role='member'):
        self.id = id
        self.username = username
        self.is_admin = is_admin
        self.member_id = member_id
        self.role = role or ('admin' if is_admin else 'member')


@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    row = db.execute("SELECT id, username, is_admin, member_id, role FROM users WHERE id=?",
                     (user_id,)).fetchone()
    if row:
        return User(row['id'], row['username'], row['is_admin'],
                    row['member_id'], row['role'])
    return None


def admin_required(f):
    """Decorator: Route nur für Admins zugänglich."""
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_admin:
            flash('Zugriff nur für Administratoren.', 'danger')
            return redirect(url_for('portal_dashboard'))
        return f(*args, **kwargs)
    return decorated


# === Audit Logging ===

# Seitenaufrufe, die automatisch geloggt werden (GET-Requests)
_AUDIT_PAGE_ENDPOINTS = {
    'dashboard': 'Dashboard',
    'import_data': 'Import',
    'members_list': 'Mitglieder',
    'member_new': 'Neues Mitglied',
    'member_edit': 'Mitglied bearbeiten',
    'prices': 'Preise',
    'invoices_list': 'Abrechnungen',
    'invoice_new': 'Neue Abrechnung',
    'invoice_detail': 'Abrechnungsdetail',
    'reports': 'Reports',
    'settings': 'Einstellungen',
    'admin_backup': 'Backup',
    'admin_database': 'Datenbank-Wartung',
    'admin_users': 'Benutzerverwaltung',
    'payments': 'Überweisungen',
    'portal_dashboard': 'Portal: Übersicht',
    'portal_data': 'Portal: Meine Daten',
    'portal_invoices': 'Portal: Abrechnungen',
    'portal_contracts': 'Portal: Verträge',
}


def get_real_ip():
    """Echte Client-IP ermitteln (hinter Reverse-Proxy)."""
    # ProxyFix setzt remote_addr bereits korrekt, aber als Fallback:
    return request.remote_addr


def audit_log(action, detail=None, user_id=None, username=None):
    """Schreibt einen Eintrag ins Audit-Log."""
    try:
        db = get_db()
        uid = user_id
        uname = username
        if uid is None and has_request_context() and current_user and current_user.is_authenticated:
            uid = current_user.id
            uname = current_user.username
        db.execute(
            """INSERT INTO audit_log
               (timestamp, user_id, username, action, detail, ip, url, method)
               VALUES (?,?,?,?,?,?,?,?)""",
            (utc_now_string(), uid, uname, action, detail,
             get_real_ip() if has_request_context() else None,
             request.url if has_request_context() else None,
             request.method if has_request_context() else None))
        db.commit()
    except Exception:
        pass  # Audit-Log darf nie die App blockieren


@app.after_request
def audit_page_views(response):
    """Loggt Seitenaufrufe automatisch für authentifizierte User."""
    try:
        if (request.method == 'GET'
                and response.status_code == 200
                and current_user
                and current_user.is_authenticated
                and request.endpoint in _AUDIT_PAGE_ENDPOINTS):
            label = _AUDIT_PAGE_ENDPOINTS[request.endpoint]
            audit_log('page_view', label)
    except Exception:
        pass
    return response


# === Login Security ===
# Die Zaehler liegen in der Datenbank, damit eine Sperre einen Neustart oder ein
# Deployment ueberlebt (frueher: Prozessspeicher, nach jedem Restart zurueckgesetzt).
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_SECONDS = 300  # 5 Minuten
LOGIN_ATTEMPT_RESET_SECONDS = 900  # Zaehler nach 15 Minuten Ruhe zuruecksetzen


def _login_attempt_row(ip):
    try:
        return get_db().execute(
            "SELECT count, last_attempt, locked_until FROM login_attempts WHERE ip=?",
            (ip,)).fetchone()
    except sqlite3.Error:
        app.logger.exception('Login-Sperrtabelle nicht lesbar')
        return None


def _check_login_rate(ip):
    """Prüft ob eine IP gesperrt ist. Gibt verbleibende Sekunden zurück, oder 0."""
    row = _login_attempt_row(ip)
    if not row:
        return 0
    remaining = (row['locked_until'] or 0) - time.time()
    return int(remaining) if remaining > 0 else 0


def _login_attempt_count(ip):
    """Aktuelle Anzahl Fehlversuche innerhalb des Beobachtungsfensters."""
    row = _login_attempt_row(ip)
    if not row:
        return 0
    if time.time() - (row['last_attempt'] or 0) > LOGIN_ATTEMPT_RESET_SECONDS:
        return 0
    return row['count'] or 0


def _record_failed_login(ip):
    """Zählt fehlgeschlagene Login-Versuche und sperrt ggf."""
    now = time.time()
    count = _login_attempt_count(ip) + 1
    locked_until = now + LOCKOUT_SECONDS if count >= MAX_LOGIN_ATTEMPTS else 0
    try:
        db = get_db()
        db.execute("""
            INSERT INTO login_attempts (ip, count, last_attempt, locked_until)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(ip) DO UPDATE SET
                count = excluded.count,
                last_attempt = excluded.last_attempt,
                locked_until = excluded.locked_until
        """, (ip, count, now, locked_until))
        db.commit()
    except sqlite3.Error:
        app.logger.exception('Fehlversuch konnte nicht gespeichert werden')


def _reset_login_attempts(ip):
    try:
        db = get_db()
        db.execute("DELETE FROM login_attempts WHERE ip=?", (ip,))
        db.commit()
    except sqlite3.Error:
        app.logger.exception('Login-Sperre konnte nicht aufgehoben werden')


# === Schutz gegen Scanner-Fluten ===
# Reine Prozess-Statistik: 404-Serien stammen von Bots, ein Neustart darf den
# Zaehler folgenlos leeren. Deshalb bewusst keine Datenbankschreiblast.
NOT_FOUND_LIMIT = int(os.environ.get('EEG_NOT_FOUND_LIMIT', '30'))
NOT_FOUND_WINDOW_SECONDS = 300
NOT_FOUND_BLOCK_SECONDS = 900
_NOT_FOUND_HITS = {}  # {ip: [count, window_start, blocked_until]}
_NOT_FOUND_LOCK = threading.Lock()


def _not_found_block_seconds(ip):
    with _NOT_FOUND_LOCK:
        entry = _NOT_FOUND_HITS.get(ip)
        if not entry:
            return 0
        remaining = entry[2] - time.time()
        return int(remaining) if remaining > 0 else 0


def _record_not_found(ip):
    now = time.time()
    with _NOT_FOUND_LOCK:
        if len(_NOT_FOUND_HITS) > 5000:
            for key, value in list(_NOT_FOUND_HITS.items()):
                if value[2] < now and now - value[1] > NOT_FOUND_WINDOW_SECONDS:
                    _NOT_FOUND_HITS.pop(key, None)
        entry = _NOT_FOUND_HITS.get(ip)
        if not entry or now - entry[1] > NOT_FOUND_WINDOW_SECONDS:
            entry = [0, now, entry[2] if entry else 0]
        entry[0] += 1
        if entry[0] >= NOT_FOUND_LIMIT:
            entry = [0, now, now + NOT_FOUND_BLOCK_SECONDS]
            app.logger.warning('IP %s wegen 404-Flut fuer %s Sekunden gesperrt',
                               ip, NOT_FOUND_BLOCK_SECONDS)
        _NOT_FOUND_HITS[ip] = entry


def _password_change_allowed_endpoint():
    """Seiten, die trotz erzwungenem Passwortwechsel erreichbar bleiben muessen."""
    if (request.endpoint or '') in {'change_password', 'logout', 'static'}:
        return True
    # V2-Oberflaeche: /v2/change-password laeuft ueber die SPA-Sammelroute.
    return request.path.rstrip('/').endswith('/change-password')


@app.before_request
def enforce_request_policies():
    """Scanner-Sperre, Sitzungs-Timeout und erzwungener Passwortwechsel."""
    blocked = _not_found_block_seconds(get_real_ip())
    if blocked > 0:
        return ('Zu viele ungueltige Anfragen.', 429, {'Retry-After': str(blocked)})

    if not (current_user and current_user.is_authenticated):
        return None

    session.permanent = True
    now = time.time()
    last_seen = session.get('last_seen')
    if isinstance(last_seen, (int, float)) and now - last_seen > SESSION_IDLE_TIMEOUT_SECONDS:
        audit_log('session_timeout', 'Sitzung wegen Inaktivität beendet')
        logout_user()
        session.clear()
        flash('Sitzung wegen Inaktivität beendet. Bitte erneut anmelden.', 'warning')
        return redirect('/v2/login' if request.path.startswith('/v2') else url_for('login'))
    session['last_seen'] = now

    if session.get('must_change_password') and not _password_change_allowed_endpoint():
        flash('Bitte vergeben Sie zuerst ein neues, sicheres Passwort.', 'warning')
        return redirect('/v2/change-password' if request.path.startswith('/v2')
                        else url_for('change_password'))
    return None


@app.after_request
def track_not_found_flood(response):
    """Zaehlt 404er nur fuer nicht angemeldete Aufrufer (typische Scanner)."""
    if response.status_code == 404 and not (current_user and current_user.is_authenticated):
        _record_not_found(get_real_ip())
    return response


def _v2_assets():
    manifest_path = os.path.join(BASE_DIR, 'static', 'v2', '.vite', 'manifest.json')
    if not os.path.exists(manifest_path):
        raise RuntimeError('V2-Assets fehlen. Bitte in webapp/v2_src "npm install" und "npm run build" ausführen.')
    with open(manifest_path, encoding='utf-8') as f:
        manifest = json.load(f)
    entry = manifest.get('index.html')
    if not entry or not entry.get('file'):
        raise RuntimeError('V2-Manifest ist unvollständig.')
    return {
        'js': entry['file'],
        'css': entry.get('css', []),
    }


def _v2_public_dict(row, fields):
    if not row:
        return {}
    return {field: row[field] for field in fields if field in row.keys()}


def _v2_dashboard_data(db):
    stats = {
        'members': db.execute("SELECT COUNT(*) FROM members WHERE active=1").fetchone()[0],
        'measurements': db.execute("SELECT COUNT(*) FROM measurements").fetchone()[0],
        'batches': db.execute("SELECT COUNT(*) FROM import_batches").fetchone()[0],
        'invoices': db.execute("SELECT COUNT(*) FROM invoices").fetchone()[0],
    }
    imports = db.execute("""
        SELECT source_file, period_start, period_end, imported_at, data_status
        FROM import_batches
        WHERE replaced_at IS NULL
        ORDER BY imported_at DESC, id DESC
        LIMIT 6
    """).fetchall()
    invoices = db.execute("""
        SELECT id, period_from, period_to, status, data_status, created_at, finalized_at
        FROM invoices
        ORDER BY period_from DESC, id DESC
        LIMIT 6
    """).fetchall()
    monthly = db.execute("""
        SELECT b.period_start, ROUND(SUM(m.value_kwh), 1) as kwh, COUNT(*) as records
        FROM measurements m
        JOIN import_batches b ON b.id = m.batch_id
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE mc.code = '1-1:2.9.0 G.03'
        GROUP BY b.period_start
        ORDER BY b.period_start DESC
        LIMIT 6
    """).fetchall()
    return {
        'type': 'dashboard',
        'stats': stats,
        'imports': [_v2_public_dict(row, ('source_file', 'period_start', 'period_end', 'imported_at', 'data_status')) for row in imports],
        'invoices': [_v2_public_dict(row, ('id', 'period_from', 'period_to', 'status', 'data_status', 'created_at', 'finalized_at')) for row in invoices],
        'monthly': [_v2_public_dict(row, ('period_start', 'kwh', 'records')) for row in reversed(monthly)],
    }


def _v2_import_data(db, results=None):
    imports = db.execute("""
        SELECT id, source_file, period_start, period_end, data_status,
               replaced_by_batch_id, replaced_at, imported_at
        FROM import_batches
        ORDER BY imported_at DESC, id DESC
        LIMIT 100
    """).fetchall()
    import_values = db.execute("""
        SELECT id, filename, records_imported, records_overwritten, status,
               data_status, error_message, imported_by, imported_at
        FROM import_log
        ORDER BY imported_at DESC, id DESC
        LIMIT 100
    """).fetchall()
    return {
        'type': 'import',
        'results': results or [],
        'imports': [_v2_public_dict(row, (
            'id', 'source_file', 'period_start', 'period_end', 'data_status',
            'replaced_by_batch_id', 'replaced_at', 'imported_at',
        )) for row in imports],
        'history': [_v2_public_dict(row, (
            'id', 'filename', 'records_imported', 'records_overwritten', 'status',
            'data_status', 'error_message', 'imported_by', 'imported_at',
        )) for row in import_values],
    }


def _v2_prices_data(db):
    prices = db.execute("SELECT * FROM prices ORDER BY valid_from DESC").fetchall()
    invoice_map = {}
    for price in prices:
        invoice = db.execute("""
            SELECT id, period_from, period_to
            FROM invoices
            WHERE period_from <= ? AND period_to >= ?
            ORDER BY id DESC
            LIMIT 1
        """, (price['valid_to'], price['valid_from'])).fetchone()
        if invoice:
            invoice_map[price['id']] = _v2_public_dict(invoice, ('id', 'period_from', 'period_to'))
    fields = ('id', 'valid_from', 'valid_to', 'price_consumption',
              'price_generation', 'description', 'created_at')
    rows = []
    for price in prices:
        item = _v2_public_dict(price, fields)
        item['invoice'] = invoice_map.get(price['id'])
        rows.append(item)
    return {
        'type': 'prices',
        'prices': rows,
    }


def _v2_invoices_data(db):
    rows = db.execute("""
        SELECT id, period_from, period_to, status, total_kwh_traded, total_income,
               total_expense, total_margin, created_at, finalized_at, data_status
        FROM invoices
        ORDER BY period_from DESC, id DESC
        LIMIT 200
    """).fetchall()
    fields = ('id', 'period_from', 'period_to', 'status', 'total_kwh_traded',
              'total_income', 'total_expense', 'total_margin', 'created_at',
              'finalized_at', 'data_status')
    return {
        'type': 'invoices',
        'invoices': [_v2_public_dict(row, fields) for row in rows],
    }


def _invoice_period_suggestion():
    today = date.today()
    q_month = ((today.month - 1) // 3) * 3 + 1
    q_start = date(today.year, q_month, 1)
    if q_month > 3:
        prev_q_start = date(today.year, q_month - 3, 1)
    else:
        prev_q_start = date(today.year - 1, 10, 1)
    from calendar import monthrange
    prev_end_month = q_month - 1 if q_month > 1 else 12
    prev_end_year = today.year if q_month > 1 else today.year - 1
    _, last_day = monthrange(prev_end_year, prev_end_month)
    prev_q_end = date(prev_end_year, prev_end_month, last_day)
    return prev_q_start.isoformat(), prev_q_end.isoformat()


def _v2_invoice_new_data(db):
    suggested_from, suggested_to = _invoice_period_suggestion()
    import_status = get_import_status_for_period(db, suggested_from, suggested_to)
    price_cons, price_gen = get_price_for_date(db, suggested_from)
    return {
        'type': 'invoice_new',
        'suggested_from': suggested_from,
        'suggested_to': suggested_to,
        'import_status': {
            'data_status': import_status['data_status'],
            'is_final': import_status['is_final'],
            'reason': import_status.get('reason', ''),
        },
        'price': {
            'consumption': price_cons,
            'generation': price_gen,
        },
    }


def _create_invoice_from_request(db):
    period_from = request.form['period_from']
    period_to = request.form['period_to']
    existing = db.execute("""
        SELECT id, period_from, period_to FROM invoices
        WHERE period_from <= ? AND period_to >= ?
    """, (period_to, period_from)).fetchone()
    if existing:
        raise ValueError(
            f'Es existiert bereits eine Abrechnung für diesen Zeitraum '
            f'(Nr. {existing["id"]}: {existing["period_from"]} - {existing["period_to"]}). '
            f'Pro Quartal ist nur eine Abrechnung zulässig.'
        )

    price_cons, price_gen = get_price_for_date(db, period_from)
    import_status = get_import_status_for_period(db, period_from, period_to)
    result = calculate_billing(db, period_from, period_to, price_cons, price_gen)
    cur = db.execute("""INSERT INTO invoices (period_from, period_to, total_kwh_traded,
                        total_income, total_expense, total_margin, data_status)
                        VALUES (?, ?, ?, ?, ?, ?, ?)""",
                     (period_from, period_to, result['total_kwh'],
                      result['total_income'], result['total_expense'], result['total_margin'],
                      import_status['data_status']))
    invoice_id = cur.lastrowid
    for item in result['items']:
        db.execute("""INSERT INTO invoice_items (invoice_id, member_id, type, kwh, price_per_kwh, amount_eur)
                      VALUES (?, ?, ?, ?, ?, ?)""",
                   (invoice_id, item['member_id'], item['type'],
                    item['kwh'], item['price'], item['amount']))
    save_invoice_carryovers(db, invoice_id, result['carryovers'])
    db.commit()
    audit_log('invoice_create', f'Abrechnung #{invoice_id} erstellt: {period_from} - {period_to} ({result["total_kwh"]:.1f} kWh)')
    carryover_total = round(sum(item['amount'] for item in result['carryovers']), 2)
    carryover_info = f' Finanzvortrag: {carryover_total:.2f} EUR.' if result['carryovers'] else ''
    if import_status['is_final']:
        flash(f'Abrechnung #{invoice_id} erstellt ({result["total_kwh"]:.1f} kWh).{carryover_info}', 'success')
    else:
        flash(f'Vorläufige Abrechnung #{invoice_id} erstellt ({result["total_kwh"]:.1f} kWh).{carryover_info} Versand und Abschluss sind gesperrt, bis endgültige Daten importiert und die Abrechnung neu berechnet wurde.', 'warning')
    return invoice_id


def _v2_invoice_detail_data(db, invoice_id):
    invoice = db.execute("SELECT * FROM invoices WHERE id=?", (invoice_id,)).fetchone()
    if not invoice:
        return None
    items = db.execute("""
        SELECT ii.*, m.name as member_name, m.email as member_email
        FROM invoice_items ii
        JOIN members m ON m.id = ii.member_id
        WHERE ii.invoice_id = ?
        ORDER BY m.name, ii.type
    """, (invoice_id,)).fetchall()
    members_map = {}
    for item in items:
        mid = item['member_id']
        if mid not in members_map:
            members_map[mid] = {
                'member_id': mid,
                'member_name': item['member_name'],
                'member_email': item['member_email'],
                'cons_kwh': 0,
                'cons_eur': 0,
                'cons_price': 0,
                'gen_kwh': 0,
                'gen_eur': 0,
                'gen_price': 0,
            }
        if item['type'] == 'consumption':
            members_map[mid]['cons_kwh'] = item['kwh']
            members_map[mid]['cons_eur'] = item['amount_eur']
            members_map[mid]['cons_price'] = item['price_per_kwh']
        else:
            members_map[mid]['gen_kwh'] = item['kwh']
            members_map[mid]['gen_eur'] = item['amount_eur']
            members_map[mid]['gen_price'] = item['price_per_kwh']
    for member in members_map.values():
        member['energy_net_eur'] = round(member['cons_eur'] - member['gen_eur'], 2)
        member['carryover_eur'] = 0.0
        member['carryovers'] = []

    carryover_map = get_invoice_carryover_map(db, invoice_id)
    for mid, data in carryover_map.items():
        if mid not in members_map:
            first = data['rows'][0]
            members_map[mid] = {
                'member_id': mid,
                'member_name': first['member_name'],
                'member_email': first['member_email'],
                'cons_kwh': 0,
                'cons_eur': 0,
                'cons_price': 0,
                'gen_kwh': 0,
                'gen_eur': 0,
                'gen_price': 0,
                'energy_net_eur': 0.0,
            }
        members_map[mid]['carryover_eur'] = data['total']
        members_map[mid]['carryovers'] = [
            _v2_public_dict(row, ('source_invoice_id', 'source_period_from', 'source_period_to', 'amount_eur', 'description'))
            for row in data['rows']
        ]

    emails = db.execute("""
        SELECT el.*, m.name as member_name
        FROM email_log el
        LEFT JOIN members m ON m.id = el.member_id
        WHERE el.invoice_id=? ORDER BY el.sent_at DESC
        LIMIT 100
    """, (invoice_id,)).fetchall()
    sent_members = {row['member_id'] for row in emails if row['status'] == 'sent' and row['member_id']}
    member_rows = []
    for member in members_map.values():
        member.setdefault('energy_net_eur', round(member['cons_eur'] - member['gen_eur'], 2))
        member.setdefault('carryover_eur', 0.0)
        member.setdefault('carryovers', [])
        member['net_eur'] = round(member['energy_net_eur'] + member['carryover_eur'], 2)
        member['email_sent'] = member['member_id'] in sent_members
        member_rows.append(member)
    member_rows = sorted(member_rows, key=lambda item: item['member_name'])

    import_status = get_import_status_for_period(db, invoice['period_from'], invoice['period_to'])
    return {
        'type': 'invoice_detail',
        'invoice': _v2_public_dict(invoice, (
            'id', 'period_from', 'period_to', 'status', 'data_status',
            'total_kwh_traded', 'total_income', 'total_expense', 'total_margin',
            'created_at', 'sent_at', 'finalized_at',
        )),
        'members': member_rows,
        'emails': [
            {
                **_v2_public_dict(row, ('id', 'member_id', 'member_name', 'email', 'subject', 'status', 'error_message')),
                'sent_at': format_local_datetime(row['sent_at']),
            }
            for row in emails
        ],
        'import_status': {
            'data_status': import_status['data_status'],
            'is_final': import_status['is_final'],
            'reason': import_status.get('reason', ''),
            'provisional_count': len(import_status.get('provisional_batches') or []),
        },
        'finalization_blocker': invoice_finalization_blocker(db, invoice),
    }


def _v2_payments_data(db):
    rows = get_payment_rows(db)
    fields = (
        'invoice_id', 'member_id', 'member_name', 'iban', 'bic', 'account_holder',
        'period_from', 'period_to', 'invoice_status', 'net_total', 'energy_total',
        'carryover_total', 'booked_total', 'open_amount', 'is_partially_booked',
        'paid', 'paid_at', 'booking_date', 'booking_note', 'booking_id',
        'direction', 'open_direction', 'is_settled_by_carryover', 'carried_forward_to_invoice_id',
        'is_overdue', 'is_previous_period_open',
    )

    def serialize(row):
        item = _v2_public_dict(row, fields)
        for key in ('reference_date', 'due_on'):
            value = row.get(key)
            item[key] = value.isoformat() if hasattr(value, 'isoformat') else value
        item['bookings'] = [{
            'id': booking['id'],
            'amount_eur': booking['amount_eur'],
            'booking_date': booking['booking_date'],
            'note': booking['note'] or '',
            'recorded_by': booking['recorded_by_username'] or '',
            'changes': [{
                'changed_at': change['changed_at'],
                'changed_by': change['changed_by_username'] or '',
                'action': change['action'],
                'old_amount_eur': change['old_amount_eur'],
                'new_amount_eur': change['new_amount_eur'],
                'old_booking_date': change['old_booking_date'],
                'new_booking_date': change['new_booking_date'],
                'reason': change['reason'] or '',
            } for change in booking['changes']],
        } for booking in row['bookings']]
        item['sepa'] = {
            'name': _sepa_text(row.get('account_holder') or row['member_name'], 70),
            'iban': row.get('iban') or '',
            'bic': _sepa_text(row.get('bic') or '', 11),
            'reference': _sepa_text(payment_transfer_reference(row), 140),
            'qr_url': url_for('payment_transfer_qr', invoice_id=row['invoice_id'],
                              member_id=row['member_id']) if row.get('iban') else '',
        }
        return item

    active_open = [row for row in rows if not row['paid'] and not row['is_settled_by_carryover']]

    booked_sort = request.args.get('booked_sort') or 'date'
    booked_dir = request.args.get('booked_dir') or 'desc'
    if booked_sort not in BOOKED_SORT_KEYS:
        booked_sort = 'date'
    if booked_dir not in {'asc', 'desc'}:
        booked_dir = 'desc'
    paid_entries_raw = booked_payment_entries(rows, booked_sort, booked_dir)
    paid_entries = []
    for entry in paid_entries_raw:
        row = entry['row']
        booking = entry['booking']
        paid_entries.append({
            'kind': entry['kind'],
            'position': entry['position'],
            'total_bookings': entry['total_bookings'],
            'amount': entry['amount'],
            'booking_date': entry['booking_date'],
            'row': serialize(row),
            'booking': {
                'id': booking['id'],
                'amount_eur': booking['amount_eur'],
                'booking_date': booking['booking_date'],
                'note': booking.get('note') or '',
            } if booking else None,
        })

    return {
        'type': 'payments',
        'today': local_now().date().isoformat(),
        'booked_sort': booked_sort,
        'booked_dir': booked_dir,
        'summary': {
            'open_claims_count': len([row for row in active_open if row['open_amount'] > 0]),
            'open_credits_count': len([row for row in active_open if row['open_amount'] < 0]),
            'overdue_count': len([row for row in rows if row['is_overdue']]),
            'paid_count': len([row for row in rows if row['paid']]),
            'open_claims_total': round(sum(row['open_amount'] for row in active_open if row['open_amount'] > 0), 2),
            'open_credits_total': round(sum(row['open_amount'] for row in active_open if row['open_amount'] < 0), 2),
        },
        'booked_entries': paid_entries,
        'payments': [serialize(row) for row in rows],
    }


def _v2_member_accounts_data(db):
    accounts = get_member_account_overview(db)
    return {
        'type': 'member_accounts',
        'accounts': accounts,
        'summary': {
            'claims': round(sum(a['balance'] for a in accounts if a['balance'] > 0), 2),
            'credits': round(sum(a['balance'] for a in accounts if a['balance'] < 0), 2),
            'balance': round(sum(a['balance'] for a in accounts), 2),
            'overdue': round(sum(a['overdue'] for a in accounts), 2),
        },
    }


def _save_settings_from_request(db):
    existing_settings = {
        row['key']: row['value']
        for row in db.execute("SELECT key, value FROM settings").fetchall()
    }
    for key in (
        'smtp_host', 'smtp_port', 'smtp_user', 'smtp_pass', 'smtp_from', 'smtp_tls',
        'mail_from_address', 'mail_from_name', 'mail_reply_to', 'mail_reply_to_name',
        'email_subject', 'email_body',
        'org_name', 'org_email', 'org_website', 'org_address', 'org_legal',
        'payment_recipient', 'payment_iban', 'payment_bic'
    ):
        value = request.form.get(key, '')
        if key == 'smtp_pass' and not value:
            value = existing_settings.get('smtp_pass', '')
        db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
    db.commit()
    audit_log('settings_update', 'Einstellungen geändert')


def _v2_settings_data(db):
    rows = db.execute("SELECT key, value FROM settings").fetchall()
    settings = {row['key']: row['value'] for row in rows}
    smtp_configured, mail_errors = _validate_mail_config(_load_mail_config(db))
    return {
        'type': 'settings',
        'settings': settings,
        'smtp_configured': smtp_configured,
        'mail_errors': mail_errors,
        'db_path': DB_PATH,
    }


def _v2_database_data(check_result=None, maintenance_result=None):
    stats = get_database_stats()
    return {
        'type': 'database',
        'stats': {
            'db_path': stats.get('db_path', ''),
            'db_size': stats.get('db_size', 0),
            'wal_size': stats.get('wal_size', 0),
            'shm_size': stats.get('shm_size', 0),
            'page_count': stats.get('page_count', 0),
            'page_size': stats.get('page_size', 0),
            'freelist_count': stats.get('freelist_count', 0),
            'fragmentation_mb': round(stats.get('fragmentation_mb', 0), 2),
            'tables': stats.get('tables', []),
        },
        'check_result': check_result,
        'maintenance_result': maintenance_result,
    }


def _v2_member_portal_data(db):
    if not current_user.member_id:
        return None
    return db.execute("SELECT * FROM members WHERE id=?", (current_user.member_id,)).fetchone()


def _v2_portal_dashboard_data(db):
    member = _v2_member_portal_data(db)
    if not member:
        return {'type': 'portal_dashboard', 'member': None, 'invoices': [], 'stats': None, 'account': None}
    account = get_member_account_summary(db, current_user.member_id)
    invoices = db.execute("""
        SELECT DISTINCT i.* FROM invoices i
        WHERE i.id IN (
            SELECT invoice_id FROM invoice_items WHERE member_id=?
            UNION
            SELECT invoice_id FROM invoice_carryovers WHERE member_id=?
        )
        ORDER BY i.period_from DESC
    """, (current_user.member_id, current_user.member_id)).fetchall()
    stats = None
    if invoices:
        latest = invoices[0]
        stats = get_member_stats(db, member, latest['period_from'], latest['period_to'])
        items = db.execute("SELECT * FROM invoice_items WHERE invoice_id=? AND member_id=?",
                           (latest['id'], current_user.member_id)).fetchall()
        carryovers = get_invoice_carryovers(db, latest['id'], current_user.member_id)
        net = sum(item['amount_eur'] if item['type'] == 'consumption' else -item['amount_eur'] for item in items)
        net += sum(row['amount_eur'] for row in carryovers)
        stats['net_total'] = round(net, 2)
        stats['invoice_id'] = latest['id']
    return {
        'type': 'portal_dashboard',
        'member': _v2_public_dict(member, ('id', 'name', 'email', 'phone', 'address_street', 'address_zip', 'address_city')),
        'invoices': [_v2_public_dict(row, ('id', 'period_from', 'period_to', 'status', 'data_status', 'created_at')) for row in invoices[:8]],
        'stats': stats,
        'account': _v2_account_summary(account),
    }


def _v2_account_summary(account):
    if not account:
        return None
    def serialize_value(value):
        return value.isoformat() if hasattr(value, 'isoformat') else value

    def serialize_mapping(row):
        return {key: serialize_value(value) for key, value in dict(row).items()}

    return {
        'balance': account.get('balance', 0),
        'open_claims': account.get('open_claims', 0),
        'open_credits': account.get('open_credits', 0),
        'overdue_claims': account.get('overdue_claims', 0),
        'previous_open': [serialize_mapping(row) for row in account.get('previous_open', [])[:100]],
        'history': [serialize_mapping(row) for row in account.get('history', [])[:100]],
        'rows': [serialize_mapping(row) for row in account.get('rows', [])[:100]],
    }


def _v2_portal_data_data(db):
    member = _v2_member_portal_data(db)
    return {
        'type': 'portal_data',
        'member': _v2_public_dict(member, (
            'id', 'name', 'email', 'phone', 'address_street', 'address_zip', 'address_city',
            'account_holder', 'iban', 'bic', 'bezug_zp', 'einspeiser_zp', 'newsletter_optout',
        )) if member else None,
    }


def _update_portal_member_from_request(db):
    member = _v2_member_portal_data(db)
    if not member:
        flash('Kein Mitglied zugeordnet.', 'warning')
        return
    newsletter_optout = 0 if form_switch_enabled('newsletter_enabled') else 1
    db.execute("""UPDATE members SET
        name=?, email=?, phone=?,
        address_street=?, address_zip=?, address_city=?,
        iban=?, bic=?, account_holder=?, newsletter_optout=?,
        updated_at=datetime('now')
        WHERE id=?""", (
        request.form.get('name', member['name']),
        request.form.get('email', member['email']),
        request.form.get('phone', member['phone']),
        request.form.get('address_street', member['address_street']),
        request.form.get('address_zip', member['address_zip']),
        request.form.get('address_city', member['address_city']),
        request.form.get('iban', member['iban']),
        request.form.get('bic', member['bic']),
        request.form.get('account_holder', member['account_holder']),
        newsletter_optout,
        current_user.member_id,
    ))
    db.commit()
    audit_log('portal_data_update', 'Eigene Stammdaten aktualisiert')
    flash('Daten aktualisiert.', 'success')


def _v2_portal_invoices_data(db):
    if not current_user.member_id:
        return {'type': 'portal_invoices', 'invoices': [], 'account': None}
    account = get_member_account_summary(db, current_user.member_id)
    rows = db.execute("""
        SELECT i.id, i.period_from, i.period_to, i.status, i.data_status, i.created_at,
               COALESCE(SUM(CASE WHEN ii.type='consumption' THEN ii.amount_eur ELSE 0 END), 0) as total_cons,
               COALESCE(SUM(CASE WHEN ii.type='generation' THEN ii.amount_eur ELSE 0 END), 0) as total_gen,
               COALESCE(SUM(ii.kwh), 0) as total_kwh
        FROM invoices i
        LEFT JOIN invoice_items ii ON ii.invoice_id = i.id AND ii.member_id = ?
        WHERE i.id IN (
            SELECT invoice_id FROM invoice_items WHERE member_id=?
            UNION
            SELECT invoice_id FROM invoice_carryovers WHERE member_id=?
        )
        GROUP BY i.id
        ORDER BY i.period_from DESC
    """, (current_user.member_id, current_user.member_id, current_user.member_id)).fetchall()
    payment_by_invoice = {row['invoice_id']: row for row in account['rows']}
    invoices = []
    for row in rows:
        item = _v2_public_dict(row, ('id', 'period_from', 'period_to', 'status', 'data_status', 'created_at', 'total_cons', 'total_gen', 'total_kwh'))
        payment = payment_by_invoice.get(row['id'])
        item['net_total'] = payment['net_total'] if payment else round((row['total_cons'] or 0) - (row['total_gen'] or 0), 2)
        item['paid'] = bool(payment['paid']) if payment else False
        item['booking_date'] = payment['booking_date'] if payment else ''
        invoices.append(item)
    return {'type': 'portal_invoices', 'invoices': invoices, 'account': _v2_account_summary(account), 'member_id': current_user.member_id}


def _v2_portal_contracts_data(db):
    if not current_user.member_id:
        return {'type': 'portal_contracts', 'contracts': []}
    contracts = db.execute("SELECT * FROM contracts WHERE member_id=? ORDER BY uploaded_at DESC",
                           (current_user.member_id,)).fetchall()
    return {
        'type': 'portal_contracts',
        'contracts': [_v2_public_dict(row, ('id', 'type', 'filename', 'uploaded_at', 'uploaded_by')) for row in contracts],
    }


def _v2_newsletter_data(db):
    rows = db.execute("""
        SELECT id, subject, created_by, created_at, sent_at, recipients_count
        FROM newsletters
        ORDER BY created_at DESC, id DESC
        LIMIT 200
    """).fetchall()
    fields = ('id', 'subject', 'created_by', 'created_at', 'sent_at', 'recipients_count')
    recipient_count = db.execute("""
        SELECT COUNT(*)
        FROM members
        WHERE active=1 AND email IS NOT NULL AND email != ''
          AND (newsletter_optout IS NULL OR newsletter_optout=0)
    """).fetchone()[0]
    return {
        'type': 'newsletter',
        'recipient_count': recipient_count,
        'newsletters': [_v2_public_dict(row, fields) for row in rows],
    }


def _v2_reports_data(db):
    min_date, max_date = _report_ts_bounds(db)
    aggregation = request.args.get('aggregation', 'month')
    if aggregation not in REPORT_AGGREGATIONS:
        aggregation = 'month'
    period_from = _parse_report_date(request.args.get('date_from'), min_date)
    period_to = _parse_report_date(request.args.get('date_to'), max_date)
    if period_from > period_to:
        period_from, period_to = period_to, period_from

    if current_user.is_admin:
        members = db.execute("""
            SELECT id, name, bezug_zp, einspeiser_zp
            FROM members
            WHERE active=1
            ORDER BY name
        """).fetchall()
        member_id = request.args.get('member_id', type=int)
        if not member_id and members:
            member_id = members[0]['id']
    else:
        members = db.execute("""
            SELECT id, name, bezug_zp, einspeiser_zp
            FROM members
            WHERE active=1 AND id=?
        """, (current_user.member_id,)).fetchall() if current_user.member_id else []
        member_id = current_user.member_id if members else None
    member = db.execute("SELECT * FROM members WHERE id=? AND active=1", (member_id,)).fetchone() if member_id else None
    report = _build_member_report(db, member, period_from, period_to, aggregation) if member else None

    return {
        'type': 'reports',
        'members': [_v2_public_dict(row, ('id', 'name', 'bezug_zp', 'einspeiser_zp')) for row in members],
        'selected_member': _v2_public_dict(member, ('id', 'name')) if member else None,
        'report': report,
        'period_from': period_from,
        'period_to': period_to,
        'aggregation': aggregation,
        'aggregations': {
            key: {'label': cfg['label']}
            for key, cfg in REPORT_AGGREGATIONS.items()
        },
        'min_date': min_date,
        'max_date': max_date,
    }


def _v2_users_data(db):
    users = db.execute("""
        SELECT u.id, u.username, u.email, u.is_admin, u.role, u.invite_token,
               u.invite_expires, u.created_at, u.member_id,
               m.name as member_name, m.email as member_email
        FROM users u LEFT JOIN members m ON u.member_id = m.id
        WHERE NOT EXISTS (
            SELECT 1
            FROM users other
            WHERE other.id != u.id
              AND (
                  (u.member_id IS NOT NULL AND other.member_id = u.member_id)
                  OR (
                      u.email IS NOT NULL AND u.email != ''
                      AND other.email IS NOT NULL AND other.email != ''
                      AND LOWER(other.email) = LOWER(u.email)
                  )
              )
              AND (
                  (other.invite_token IS NULL AND u.invite_token IS NOT NULL)
                  OR (
                      (other.invite_token IS NULL) = (u.invite_token IS NULL)
                      AND other.id < u.id
                  )
              )
        )
        ORDER BY u.is_admin DESC, u.username
    """).fetchall()
    members = db.execute("""
        SELECT id, name, email
        FROM members
        WHERE active=1
          AND id NOT IN (SELECT member_id FROM users WHERE member_id IS NOT NULL)
        ORDER BY name
    """).fetchall()
    contract_members = db.execute("""
        SELECT id, name, email
        FROM members
        WHERE active=1
        ORDER BY name
    """).fetchall()
    contracts = db.execute("""
        SELECT c.id, c.member_id, c.type, c.filename, c.uploaded_at,
               c.uploaded_by, m.name as member_name
        FROM contracts c JOIN members m ON m.id = c.member_id
        ORDER BY m.name, c.type, c.uploaded_at DESC
    """).fetchall()
    user_fields = ('id', 'username', 'email', 'is_admin', 'role', 'invite_expires',
                   'created_at', 'member_id', 'member_name', 'member_email')
    member_fields = ('id', 'name', 'email')
    contract_fields = ('id', 'member_id', 'type', 'filename', 'uploaded_at',
                       'uploaded_by', 'member_name')
    rows = []
    for user in users:
        item = _v2_public_dict(user, user_fields)
        item['invite_open'] = bool(user['invite_token'])
        rows.append(item)
    contract_rows = []
    for contract in contracts:
        item = _v2_public_dict(contract, contract_fields)
        item['uploaded_at'] = format_local_date(item.get('uploaded_at'))
        contract_rows.append(item)
    return {
        'type': 'users',
        'current_user_id': current_user.id,
        'users': rows,
        'members': [_v2_public_dict(row, member_fields) for row in members],
        'contract_members': [_v2_public_dict(row, member_fields) for row in contract_members],
        'contracts': contract_rows,
    }


def _v2_audit_data(db):
    page = max(request.args.get('page', 1, type=int), 1)
    per_page = 50
    offset = (page - 1) * per_page
    action_filter = request.args.get('action', '')
    user_filter = request.args.get('user', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    where_clauses = []
    params = []
    if action_filter:
        where_clauses.append("a.action = ?")
        params.append(action_filter)
    if user_filter:
        where_clauses.append("a.username LIKE ?")
        params.append(f'%{user_filter}%')
    if date_from:
        date_from_utc, _ = local_day_bounds_as_utc_strings(date_from)
        where_clauses.append("a.timestamp >= ?")
        params.append(date_from_utc)
    if date_to:
        _, date_to_utc = local_day_bounds_as_utc_strings(date_to)
        where_clauses.append("a.timestamp <= ?")
        params.append(date_to_utc)

    where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    total = db.execute(f"SELECT COUNT(*) FROM audit_log a{where_sql}", params).fetchone()[0]
    logs = db.execute(f"""
        SELECT a.id, a.timestamp, a.user_id, a.username, a.action, a.detail, a.ip
        FROM audit_log a{where_sql}
        ORDER BY a.timestamp DESC LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()
    action_list = [
        row['action']
        for row in db.execute("SELECT DISTINCT action FROM audit_log ORDER BY action").fetchall()
    ]
    today_from_utc, today_to_utc = local_day_bounds_as_utc_strings()
    stats = {
        'total_entries': db.execute("SELECT COUNT(*) FROM audit_log").fetchone()[0],
        'today_entries': db.execute(
            "SELECT COUNT(*) FROM audit_log WHERE timestamp >= ? AND timestamp <= ?",
            (today_from_utc, today_to_utc)
        ).fetchone()[0],
        'active_users': db.execute(
            "SELECT COUNT(DISTINCT username) FROM audit_log WHERE timestamp >= ? AND timestamp <= ?",
            (today_from_utc, today_to_utc)
        ).fetchone()[0],
    }
    fields = ('id', 'timestamp', 'user_id', 'username', 'action', 'detail', 'ip')
    log_rows = []
    for row in logs:
        item = _v2_public_dict(row, fields)
        item['timestamp_display'] = format_local_datetime(item.get('timestamp'))
        log_rows.append(item)
    return {
        'type': 'audit',
        'logs': log_rows,
        'actions': action_list,
        'stats': stats,
        'filters': {
            'action': action_filter,
            'user': user_filter,
            'date_from': date_from,
            'date_to': date_to,
        },
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': total,
            'total_pages': (total + per_page - 1) // per_page,
        },
        'timezone': getattr(APP_TIMEZONE, 'key', 'Europe/Vienna'),
    }


def _v2_backup_data(db):
    info = get_backup_info()
    settings = get_backup_settings(db)
    google_drive = get_google_drive_status()
    drive_backups = []
    drive_backups_error = ''
    if google_drive['connected']:
        try:
            drive_backups = list_google_drive_backups(db)
        except Exception as e:
            drive_backups_error = str(e)
            app.logger.warning('Could not list Google Drive backups for V2: %s', e, exc_info=True)

    def local_backup_row(item):
        created_at = item.get('created_at')
        return {
            'name': item.get('name', ''),
            'size': item.get('size', 0),
            'created_at': created_at.isoformat() if hasattr(created_at, 'isoformat') else created_at,
            'created_at_display': format_local_datetime(created_at),
            'kind': item.get('kind', ''),
        }

    def drive_backup_row(item):
        return {
            'id': item.get('id', ''),
            'name': item.get('name', ''),
            'size': item.get('size', 0),
            'created_at': item.get('created_at', ''),
            'created_at_display': format_local_datetime(item.get('created_at')),
            'modified_at': item.get('modified_at', ''),
            'web_view_link': item.get('web_view_link', ''),
            'mime_type': item.get('mime_type', ''),
        }

    return {
        'type': 'backup',
        'info': {
            'db_size': info['db_size'],
            'invoice_count': info['invoice_count'],
            'invoice_size': info['invoice_size'],
            'backup_folder': info['backup_folder'],
        },
        'settings': {
            'auto_enabled': settings['auto_enabled'],
            'auto_time': settings['auto_time'],
            'retention_daily': settings['retention_daily'],
            'retention_weekly': settings['retention_weekly'],
            'retention_monthly': settings['retention_monthly'],
            'retention_yearly': settings['retention_yearly'],
            'drive_enabled': settings['drive_enabled'],
            'drive_folder_id': settings['drive_folder_id'],
            'drive_last_upload': settings['drive_last_upload'],
            'drive_last_check': settings['drive_last_check'],
            'drive_last_error': settings['drive_last_error'],
            'email_enabled': settings['email_enabled'],
            'email_weekday': settings['email_weekday'],
            'email_time': settings['email_time'],
            'email_to': settings['email_to'],
            'email_max_mb': settings['email_max_mb'],
        },
        'google_drive': {
            'libs_available': google_drive['libs_available'],
            'client_configured': google_drive['client_configured'],
            'connected': google_drive['connected'],
            'error': google_drive['error'],
            'redirect_uri': google_drive['redirect_uri'],
        },
        'local_backups': [local_backup_row(item) for item in list_local_backups()[:50]],
        'drive_backups': [drive_backup_row(item) for item in drive_backups],
        'drive_backups_error': drive_backups_error,
    }


def _v2_members_data(db):
    rows = db.execute("""
        SELECT id, name, email, phone, address_street, address_zip, address_city,
               active, teilnahme, bezug_zp, einspeiser_zp, newsletter_optout,
               CASE WHEN iban IS NOT NULL AND TRIM(iban) <> '' THEN 1 ELSE 0 END AS has_bank
        FROM members
        ORDER BY active DESC, name COLLATE NOCASE ASC
        LIMIT 300
    """).fetchall()
    counts = db.execute("""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN active=1 THEN 1 ELSE 0 END) AS active,
            SUM(CASE WHEN active=0 THEN 1 ELSE 0 END) AS inactive,
            SUM(CASE WHEN newsletter_optout=0 THEN 1 ELSE 0 END) AS newsletter
        FROM members
    """).fetchone()
    fields = ('id', 'name', 'email', 'phone', 'address_street', 'address_zip',
              'address_city', 'active', 'teilnahme', 'bezug_zp', 'einspeiser_zp',
              'newsletter_optout', 'has_bank')
    return {
        'type': 'members',
        'counts': {
            'total': counts['total'] or 0,
            'active': counts['active'] or 0,
            'inactive': counts['inactive'] or 0,
            'newsletter': counts['newsletter'] or 0,
        },
        'members': [_v2_public_dict(row, fields) for row in rows],
    }


def _v2_member_form_data(db, member_id=None):
    member = None
    if member_id is not None:
        member = db.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
        if not member:
            return None
    fields = (
        'id', 'name', 'email', 'phone', 'address_street', 'address_zip',
        'address_city', 'einspeiser_zp', 'einspeiser_ab', 'bezug_zp',
        'bezug_ab', 'teilnahme', 'active', 'iban', 'bic', 'account_holder',
        'newsletter_optout',
    )
    return {
        'type': 'member_form',
        'mode': 'edit' if member else 'new',
        'member': _v2_public_dict(member, fields) if member else {},
    }


def _save_member_from_request(db, member_id=None):
    newsletter_optout = 0 if form_switch_enabled('newsletter_enabled') else 1
    if member_id is None:
        db.execute("""INSERT INTO members (name, email, phone, address_street, address_zip, address_city,
                      einspeiser_zp, einspeiser_ab, bezug_zp, bezug_ab, teilnahme,
                      iban, bic, account_holder, newsletter_optout, updated_at)
                      VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))""",
                   (request.form['name'], request.form.get('email'),
                    request.form.get('phone'),
                    request.form.get('address_street'), request.form.get('address_zip'),
                    request.form.get('address_city'),
                    request.form.get('einspeiser_zp') or None,
                    request.form.get('einspeiser_ab') or None,
                    request.form.get('bezug_zp') or None,
                    request.form.get('bezug_ab') or None,
                    float(request.form.get('teilnahme', 1.0)),
                    request.form.get('iban') or None,
                    request.form.get('bic') or None,
                    request.form.get('account_holder') or None,
                    newsletter_optout))
        db.commit()
        audit_log('member_create', f'Mitglied angelegt: {request.form["name"]}')
        flash('Mitglied angelegt.', 'success')
        return db.execute("SELECT last_insert_rowid()").fetchone()[0]

    db.execute("""UPDATE members SET name=?, email=?, phone=?, address_street=?, address_zip=?,
                  address_city=?, einspeiser_zp=?, einspeiser_ab=?, bezug_zp=?,
                  bezug_ab=?, teilnahme=?, active=?, iban=?, bic=?, account_holder=?,
                  newsletter_optout=?, updated_at=datetime('now')
                  WHERE id=?""",
               (request.form['name'], request.form.get('email'),
                request.form.get('phone'),
                request.form.get('address_street'), request.form.get('address_zip'),
                request.form.get('address_city'),
                request.form.get('einspeiser_zp') or None,
                request.form.get('einspeiser_ab') or None,
                request.form.get('bezug_zp') or None,
                request.form.get('bezug_ab') or None,
                float(request.form.get('teilnahme', 1.0)),
                1 if request.form.get('active') else 0,
                request.form.get('iban') or None,
                request.form.get('bic') or None,
                request.form.get('account_holder') or None,
                newsletter_optout,
                member_id))
    db.commit()
    audit_log('member_edit', f'Mitglied bearbeitet: {request.form["name"]} (ID {member_id})')
    flash('Mitglied aktualisiert.', 'success')
    return member_id


def _v2_price_edit_data(db, price_id):
    price = db.execute("SELECT * FROM prices WHERE id=?", (price_id,)).fetchone()
    if not price:
        return None
    invoice = db.execute("""
        SELECT id, period_from, period_to
        FROM invoices
        WHERE period_from <= ? AND period_to >= ?
        ORDER BY id DESC
        LIMIT 1
    """, (price['valid_to'], price['valid_from'])).fetchone()
    return {
        'type': 'price_edit',
        'price': _v2_public_dict(price, ('id', 'valid_from', 'valid_to', 'price_consumption', 'price_generation', 'description', 'created_at')),
        'invoice': _v2_public_dict(invoice, ('id', 'period_from', 'period_to')) if invoice else None,
    }


def _update_price_from_request(db, price_id):
    price = db.execute("SELECT * FROM prices WHERE id=?", (price_id,)).fetchone()
    if not price:
        flash('Preis nicht gefunden.', 'danger')
        return
    db.execute("""UPDATE prices SET valid_from=?, valid_to=?, price_consumption=?,
                  price_generation=?, description=? WHERE id=?""",
               (request.form['valid_from'], request.form['valid_to'],
                float(request.form['price_consumption']),
                float(request.form['price_generation']),
                request.form.get('description', ''), price_id))
    db.commit()
    audit_log('price_edit', f'Preis bearbeitet: {request.form["valid_from"]} - {request.form["valid_to"]} (ID {price_id})')
    inv = db.execute("""SELECT id FROM invoices
                       WHERE period_from <= ? AND period_to >= ?""",
                    (request.form['valid_to'], request.form['valid_from'])).fetchone()
    if inv:
        flash(f'Achtung: Für diesen Zeitraum existiert bereits Abrechnung #{inv["id"]}. Es muss eine neue Abrechnung erstellt werden, damit die Preisänderung wirksam wird!', 'warning')
    else:
        flash('Preis aktualisiert.', 'success')


def _v2_newsletter_form_data(db, newsletter_id=None):
    newsletter = None
    if newsletter_id is not None:
        newsletter = db.execute("SELECT * FROM newsletters WHERE id=?", (newsletter_id,)).fetchone()
        if not newsletter:
            return None
    return {
        'type': 'newsletter_form',
        'mode': 'edit' if newsletter else 'new',
        'newsletter': _v2_public_dict(newsletter, ('id', 'subject', 'body_html', 'created_by', 'created_at', 'sent_at', 'recipients_count')) if newsletter else {},
    }


def _save_newsletter_from_request(db, newsletter_id=None):
    subject = request.form.get('subject', '').strip()
    body_html = sanitize_newsletter_html(request.form.get('body_html', '').strip())
    if not subject or not body_html:
        raise ValueError('Betreff und Inhalt sind erforderlich.')
    if newsletter_id is None:
        db.execute("INSERT INTO newsletters (subject, body_html, created_by) VALUES (?,?,?)",
                   (subject, body_html, current_user.username))
        db.commit()
        audit_log('newsletter_create', f'Newsletter erstellt: {subject}')
        flash('Newsletter gespeichert.', 'success')
        return db.execute("SELECT last_insert_rowid()").fetchone()[0]

    newsletter = db.execute("SELECT * FROM newsletters WHERE id=?", (newsletter_id,)).fetchone()
    if not newsletter:
        raise ValueError('Newsletter nicht gefunden.')
    if newsletter['sent_at']:
        raise ValueError('Bereits versendeter Newsletter kann nicht bearbeitet werden.')
    db.execute("UPDATE newsletters SET subject=?, body_html=? WHERE id=?", (subject, body_html, newsletter_id))
    db.commit()
    audit_log('newsletter_edit', f'Newsletter bearbeitet: {subject} (ID {newsletter_id})')
    flash('Newsletter aktualisiert.', 'success')
    return newsletter_id


def _v2_newsletter_preview_data(db, newsletter_id):
    newsletter = db.execute("SELECT * FROM newsletters WHERE id=?", (newsletter_id,)).fetchone()
    if not newsletter:
        return None
    logo_url = public_url_for('static', filename='logo.png')
    html = render_template('newsletter_email.html',
        subject=newsletter['subject'],
        preview_text=newsletter['subject'],
        logo_url=logo_url,
        edition_label=newsletter['subject'].split('–')[0].strip() if '\u2013' in newsletter['subject'] else newsletter['subject'],
        headline=newsletter['subject'],
        subtitle='',
        body_html=sanitize_newsletter_html(newsletter['body_html']),
        unsubscribe_url='#',
    )
    return {
        'type': 'newsletter_preview',
        'newsletter': _v2_public_dict(newsletter, ('id', 'subject', 'created_by', 'created_at', 'sent_at', 'recipients_count')),
        'html': html,
    }


def _v2_change_password_data():
    return {'type': 'change_password'}


def _v2_release_notes_data():
    return {
        'type': 'release_notes',
        'release_notes': RELEASE_NOTES,
    }


def _v2_cashbook_data(db):
    filters = _cashbook_filters()
    book = build_cashbook(db, **filters)
    return {
        'type': 'cashbook',
        'today': local_now().date().isoformat(),
        'filters': filters,
        'period': book['period'],
        'entries': book['rows'],
        'categories': book['categories'],
        'years': book['years'],
        'summary': book['summary'],
        'by_category': book['by_category'],
        'by_year': book['by_year'],
        'directions': CASHBOOK_DIRECTIONS,
        'methods': CASHBOOK_PAYMENT_METHODS,
    }


def _change_password_from_request(db):
    old_pw = request.form.get('old_password', '')
    new_pw = request.form.get('new_password', '')
    confirm = request.form.get('confirm_password', '')
    row = db.execute("SELECT password_hash FROM users WHERE id=?", (current_user.id,)).fetchone()
    if not check_password_hash(row['password_hash'], old_pw):
        flash('Altes Passwort falsch.', 'danger')
        return False
    if new_pw != confirm:
        flash('Neue Passwörter stimmen nicht überein.', 'danger')
        return False
    policy_error = validate_password(new_pw, current_user.username)
    if policy_error:
        flash(policy_error, 'danger')
        return False
    db.execute("UPDATE users SET password_hash=?, password_change_required=0 WHERE id=?",
               (generate_password_hash(new_pw), current_user.id))
    db.commit()
    session.pop('must_change_password', None)
    audit_log('password_change', 'Passwort geändert')
    flash('Passwort geändert.', 'success')
    return True


def _v2_native_data(db, current_path):
    if current_path == '/change-password':
        return _v2_change_password_data()
    if not current_user.is_admin:
        if current_path == '/portal':
            return _v2_portal_dashboard_data(db)
        if current_path == '/portal/data':
            return _v2_portal_data_data(db)
        if current_path == '/portal/invoices':
            return _v2_portal_invoices_data(db)
        if current_path == '/portal/contracts':
            return _v2_portal_contracts_data(db)
        if current_path == '/portal/reports':
            return _v2_reports_data(db)
        return None
    if current_path == '/':
        return _v2_dashboard_data(db)
    if current_path == '/import':
        return _v2_import_data(db)
    if current_path == '/members':
        return _v2_members_data(db)
    if current_path == '/members/new':
        return _v2_member_form_data(db)
    member_match = re.fullmatch(r'/members/(\d+)/edit', current_path)
    if member_match:
        return _v2_member_form_data(db, int(member_match.group(1)))
    if current_path == '/prices':
        return _v2_prices_data(db)
    price_match = re.fullmatch(r'/prices/(\d+)/edit', current_path)
    if price_match:
        return _v2_price_edit_data(db, int(price_match.group(1)))
    if current_path == '/invoices':
        return _v2_invoices_data(db)
    if current_path == '/invoices/new':
        return _v2_invoice_new_data(db)
    invoice_match = re.fullmatch(r'/invoices/(\d+)', current_path)
    if invoice_match:
        return _v2_invoice_detail_data(db, int(invoice_match.group(1)))
    if current_path == '/payments':
        return _v2_payments_data(db)
    if current_path == '/kassabuch':
        return _v2_cashbook_data(db)
    if current_path == '/mitgliederkonten':
        return _v2_member_accounts_data(db)
    if current_path == '/newsletter':
        return _v2_newsletter_data(db)
    if current_path == '/newsletter/new':
        return _v2_newsletter_form_data(db)
    newsletter_edit_match = re.fullmatch(r'/newsletter/(\d+)/edit', current_path)
    if newsletter_edit_match:
        return _v2_newsletter_form_data(db, int(newsletter_edit_match.group(1)))
    newsletter_preview_match = re.fullmatch(r'/newsletter/(\d+)/preview', current_path)
    if newsletter_preview_match:
        return _v2_newsletter_preview_data(db, int(newsletter_preview_match.group(1)))
    if current_path == '/reports':
        return _v2_reports_data(db)
    if current_path == '/admin/users':
        return _v2_users_data(db)
    if current_path == '/admin/audit':
        return _v2_audit_data(db)
    if current_path == '/admin/backup':
        return _v2_backup_data(db)
    if current_path == '/admin/database':
        return _v2_database_data()
    if current_path == '/settings':
        return _v2_settings_data(db)
    if current_path == '/release-notes':
        return _v2_release_notes_data()
    return None


def _v2_shell_data(db, current_path=None):
    public_cfg = get_public_config(db)
    org_legal = public_cfg.get('org_legal') or DEFAULT_ORG_LEGAL
    zvr_match = re.search(r'ZVR\D*(\d+)', org_legal, re.IGNORECASE)
    user_row = db.execute("SELECT email FROM users WHERE id=?", (current_user.id,)).fetchone()
    data = {
        'user': {
            'id': current_user.id,
            'username': current_user.username,
            'role': current_user.role,
            'email': user_row['email'] if user_row else '',
        },
        'org': {
            'name': public_cfg.get('org_name') or DEFAULT_ORG_NAME,
            'legal': org_legal,
            'zvr': f'ZVR {zvr_match.group(1)}' if zvr_match else org_legal,
        },
        'security': {
            'csrf_token': generate_csrf(),
        },
        'messages': [
            {'category': category, 'text': message}
            for category, message in get_flashed_messages(with_categories=True)
        ],
    }
    if current_path:
        data['native'] = _v2_native_data(db, current_path)
    return data


def _v2_content_path(subpath=None):
    old_path = '/' + (subpath or '').strip('/')
    if old_path == '/':
        old_path = '/' if current_user.is_admin else '/portal'
    query = {
        key: values
        for key, values in request.args.lists()
        if key != 'embed'
    }
    query['embed'] = '1'
    return f'{old_path}?{urlencode(query, doseq=True)}'


# === Auth Routes ===

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        if current_user.is_admin:
            return redirect(url_for('dashboard'))
        return redirect(url_for('portal_dashboard'))

    if request.method == 'GET' and request.args.get('csrf'):
        flash('Die Sitzung war nicht mehr gültig. Bitte erneut anmelden.', 'warning')

    ip = get_real_ip()
    locked_secs = _check_login_rate(ip)

    if request.method == 'POST':
        if locked_secs > 0:
            flash(f'Zu viele Fehlversuche. Bitte warten Sie {locked_secs} Sekunden.', 'danger')
            return render_template('login.html', locked_until=locked_secs)

        login_identifier = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        db = get_db()
        candidates = db.execute("""
            SELECT id, username, password_hash, is_admin, member_id, role,
                   password_change_required
            FROM users
            WHERE LOWER(username)=? OR LOWER(email)=?
            ORDER BY
                CASE
                    WHEN invite_token IS NULL THEN 0
                    WHEN LOWER(username)=? THEN 1
                    ELSE 2
                END,
                id
        """, (login_identifier, login_identifier, login_identifier)).fetchall()
        row = None
        for candidate in candidates:
            if check_password_hash(candidate['password_hash'], password):
                row = candidate
                break
        if row:
            _reset_login_attempts(ip)
            user = User(row['id'], row['username'], row['is_admin'],
                        row['member_id'], row['role'])
            login_user(user)
            session['last_seen'] = time.time()
            audit_log('login', f'Anmeldung erfolgreich (Rolle: {user.role})')
            if row['password_change_required']:
                session['must_change_password'] = True
                flash('Bitte vergeben Sie zuerst ein neues, sicheres Passwort.', 'warning')
                return redirect(url_for('change_password'))
            next_page = request.args.get('next')
            if next_page and is_safe_redirect_url(next_page):
                return redirect(next_page)
            if user.is_admin:
                return redirect(url_for('dashboard'))
            return redirect(url_for('portal_dashboard'))
        _record_failed_login(ip)
        audit_log('login_failed', f'Fehlgeschlagener Login für "{login_identifier}"', user_id=0, username=login_identifier)
        remaining = MAX_LOGIN_ATTEMPTS - _login_attempt_count(ip)
        if remaining > 0:
            flash(f'Ungültiger Benutzername oder Passwort. Noch {remaining} Versuche.', 'danger')
        else:
            flash(f'Konto gesperrt für {LOCKOUT_SECONDS // 60} Minuten.', 'danger')
        locked_secs = _check_login_rate(ip)

    return render_template('login.html', locked_until=locked_secs if locked_secs > 0 else None)


@app.route('/v2/login', methods=['GET', 'POST'])
def v2_login():
    if current_user.is_authenticated:
        return redirect('/v2/' if current_user.is_admin else '/v2/portal')

    if request.method == 'GET' and request.args.get('csrf'):
        flash('Die Sitzung war nicht mehr gültig. Bitte erneut anmelden.', 'warning')

    ip = get_real_ip()
    locked_secs = _check_login_rate(ip)

    if request.method == 'POST':
        if locked_secs > 0:
            flash(f'Zu viele Fehlversuche. Bitte warten Sie {locked_secs} Sekunden.', 'danger')
            return render_template('v2_public.html', page='login', locked_until=locked_secs)

        login_identifier = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        db = get_db()
        candidates = db.execute("""
            SELECT id, username, password_hash, is_admin, member_id, role,
                   password_change_required
            FROM users
            WHERE LOWER(username)=? OR LOWER(email)=?
            ORDER BY
                CASE
                    WHEN invite_token IS NULL THEN 0
                    WHEN LOWER(username)=? THEN 1
                    ELSE 2
                END,
                id
        """, (login_identifier, login_identifier, login_identifier)).fetchall()
        row = None
        for candidate in candidates:
            if check_password_hash(candidate['password_hash'], password):
                row = candidate
                break
        if row:
            _reset_login_attempts(ip)
            user = User(row['id'], row['username'], row['is_admin'], row['member_id'], row['role'])
            login_user(user)
            session['last_seen'] = time.time()
            audit_log('login', f'Anmeldung erfolgreich (Rolle: {user.role})')
            if row['password_change_required']:
                session['must_change_password'] = True
                flash('Bitte vergeben Sie zuerst ein neues, sicheres Passwort.', 'warning')
                return redirect('/v2/change-password')
            next_page = request.args.get('next') or request.form.get('next')
            if next_page and is_safe_redirect_url(next_page):
                return redirect(next_page)
            return redirect('/v2/' if user.is_admin else '/v2/portal')
        _record_failed_login(ip)
        audit_log('login_failed', f'Fehlgeschlagener Login für "{login_identifier}"', user_id=0, username=login_identifier)
        remaining = MAX_LOGIN_ATTEMPTS - _login_attempt_count(ip)
        if remaining > 0:
            flash(f'Ungültiger Benutzername oder Passwort. Noch {remaining} Versuche.', 'danger')
        else:
            flash(f'Konto gesperrt für {LOCKOUT_SECONDS // 60} Minuten.', 'danger')
        locked_secs = _check_login_rate(ip)

    return render_template('v2_public.html', page='login', locked_until=locked_secs if locked_secs > 0 else None)


@app.route('/logout')
@login_required
def logout():
    audit_log('logout', 'Abmeldung')
    logout_user()
    return redirect(url_for('login'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        old_pw = request.form.get('old_password', '')
        new_pw = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        db = get_db()
        row = db.execute("SELECT password_hash FROM users WHERE id=?",
                         (current_user.id,)).fetchone()
        policy_error = validate_password(new_pw, current_user.username)
        if not check_password_hash(row['password_hash'], old_pw):
            flash('Altes Passwort falsch.', 'danger')
        elif new_pw != confirm:
            flash('Neue Passwörter stimmen nicht überein.', 'danger')
        elif policy_error:
            flash(policy_error, 'danger')
        else:
            db.execute("UPDATE users SET password_hash=?, password_change_required=0 WHERE id=?",
                       (generate_password_hash(new_pw), current_user.id))
            db.commit()
            session.pop('must_change_password', None)
            audit_log('password_change', 'Passwort geändert')
            flash('Passwort geändert.', 'success')
            return redirect(url_for('dashboard'))
    return render_template('change_password.html')


# === Dashboard ===

@app.route('/')
@admin_required
def dashboard():
    db = get_db()
    stats = {}
    stats['members'] = db.execute("SELECT COUNT(*) FROM members WHERE active=1").fetchone()[0]
    stats['measurements'] = db.execute("SELECT COUNT(*) FROM measurements").fetchone()[0]
    stats['batches'] = db.execute("SELECT COUNT(*) FROM import_batches").fetchone()[0]
    stats['invoices'] = db.execute("SELECT COUNT(*) FROM invoices").fetchone()[0]

    # Sortierung Dashboard
    imp_sort = request.args.get('imp_sort', 'imported_at')
    imp_dir = request.args.get('imp_dir', 'desc').lower()
    mon_sort = request.args.get('mon_sort', 'period_start')
    mon_dir = request.args.get('mon_dir', 'desc').lower()

    allowed_imp_sort = {'imported_at', 'source_file', 'period_start'}
    allowed_mon_sort = {'period_start', 'kwh'}
    if imp_sort not in allowed_imp_sort:
        imp_sort = 'imported_at'
    if mon_sort not in allowed_mon_sort:
        mon_sort = 'period_start'
    if imp_dir not in {'asc', 'desc'}:
        imp_dir = 'desc'
    if mon_dir not in {'asc', 'desc'}:
        mon_dir = 'asc'

    # Importe
    stats['last_imports'] = db.execute(f"""
        SELECT source_file, period_start, period_end, imported_at
        FROM import_batches ORDER BY {imp_sort} {imp_dir.upper()}
    """).fetchall()

    # Monatssummen
    order_col = 'kwh' if mon_sort == 'kwh' else 'b.period_start'
    stats['monthly'] = db.execute(f"""
        SELECT b.period_start, ROUND(SUM(m.value_kwh), 1) as kwh, COUNT(*) as cnt
        FROM measurements m
        JOIN import_batches b ON b.id = m.batch_id
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE mc.code = '1-1:2.9.0 G.03'
        GROUP BY b.period_start
        ORDER BY {order_col} {mon_dir.upper()}
    """).fetchall()

    return render_template('dashboard.html', stats=stats,
                           imp_sort=imp_sort, imp_dir=imp_dir,
                           mon_sort=mon_sort, mon_dir=mon_dir)


def _run_import_uploads():
    files = request.files.getlist('files')
    overwrite = request.form.get('overwrite') == '1'
    data_status = _valid_import_data_status(request.form.get('data_status'))
    results = []
    for f in files:
        if f and f.filename.lower().endswith('.xlsx'):
            filename = secure_filename(f.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            try:
                f.save(filepath)
                result = run_import(filepath, overwrite, data_status)
            except Exception as e:
                app.logger.exception('Import upload handling failed for %s', filename)
                result = {
                    'filename': filename,
                    'status': 'error',
                    'data_status': data_status,
                    'records': 0,
                    'overwritten': 0,
                    'error': str(e),
                    'imported_at': None,
                }
            results.append(result)
            audit_log('import', f'Datei importiert: {filename} ({result["records"]} Datensätze, Status: {result["status"]})')
    return results


def _create_price_from_request(db):
    db.execute("""INSERT INTO prices (valid_from, valid_to, price_consumption, price_generation, description)
                  VALUES (?, ?, ?, ?, ?)""",
               (request.form['valid_from'], request.form['valid_to'],
                float(request.form['price_consumption']),
                float(request.form['price_generation']),
                request.form.get('description', '')))
    db.commit()
    audit_log('price_create', f'Preis angelegt: {request.form["valid_from"]} - {request.form["valid_to"]}')


@app.route('/v2/', methods=['GET', 'POST'])
@app.route('/v2', methods=['GET', 'POST'])
@app.route('/v2/<path:subpath>', methods=['GET', 'POST'])
@login_required
def v2_dashboard(subpath=None):
    """Experimentelle V2-Oberflaeche, getrennt von der bestehenden Jinja-UI."""
    current_path = '/' + (subpath or '').strip('/')
    if current_path == '/':
        current_path = '/' if current_user.is_admin else '/portal'
    writable_paths = {'/import', '/prices', '/settings', '/admin/database', '/invoices/new', '/portal/data', '/members/new', '/newsletter/new', '/change-password'}
    writable_patterns = (
        r'/members/\d+/edit',
        r'/prices/\d+/edit',
        r'/newsletter/\d+/edit',
    )
    is_writable_dynamic = any(re.fullmatch(pattern, current_path) for pattern in writable_patterns)
    if request.method == 'POST' and current_path not in writable_paths and not is_writable_dynamic:
        abort(405)
    results = None
    check_result = None
    maintenance_result = None
    if request.method == 'POST':
        if current_path not in {'/portal/data', '/change-password'} and not current_user.is_admin:
            abort(403)
        if current_path == '/import':
            results = _run_import_uploads()
        elif current_path == '/prices':
            _create_price_from_request(get_db())
            flash('Preis angelegt.', 'success')
            return redirect('/v2/prices')
        elif re.fullmatch(r'/prices/\d+/edit', current_path):
            price_id = int(current_path.split('/')[2])
            _update_price_from_request(get_db(), price_id)
            return redirect('/v2/prices')
        elif current_path == '/members/new':
            _save_member_from_request(get_db())
            return redirect('/v2/members')
        elif re.fullmatch(r'/members/\d+/edit', current_path):
            member_id = int(current_path.split('/')[2])
            _save_member_from_request(get_db(), member_id)
            return redirect('/v2/members')
        elif current_path == '/settings':
            _save_settings_from_request(get_db())
            flash('Einstellungen gespeichert.', 'success')
            return redirect('/v2/settings')
        elif current_path == '/admin/database':
            action = request.form.get('database_action', '')
            try:
                if action == 'check':
                    check_result = run_database_quality_check()
                    audit_log('database_quality_check', check_result['summary'])
                    flash(f'Datenbank-Qualitätscheck abgeschlossen: {check_result["summary"]}.',
                          'success' if check_result['status'] == 'ok' else 'warning')
                elif action == 'maintenance':
                    maintenance_result = run_database_maintenance(request.form.get('maintenance_action', ''))
                    audit_log('database_maintenance', f'{maintenance_result["label"]} ausgeführt')
                    flash(f'{maintenance_result["label"]} erfolgreich abgeschlossen.', 'success')
                else:
                    flash('Unbekannte Datenbank-Aktion.', 'danger')
            except Exception as e:
                flash_exception(e, 'Datenbank-Aktion fehlgeschlagen.')
        elif current_path == '/invoices/new':
            try:
                invoice_id = _create_invoice_from_request(get_db())
                return redirect(f'/v2/invoices/{invoice_id}')
            except Exception as e:
                flash_exception(e, 'Rechnung konnte nicht erstellt werden.')
        elif current_path == '/newsletter/new':
            try:
                _save_newsletter_from_request(get_db())
                return redirect('/v2/newsletter')
            except Exception as e:
                flash_exception(e, 'Newsletter konnte nicht gespeichert werden.')
        elif re.fullmatch(r'/newsletter/\d+/edit', current_path):
            newsletter_id = int(current_path.split('/')[2])
            try:
                _save_newsletter_from_request(get_db(), newsletter_id)
                return redirect('/v2/newsletter')
            except Exception as e:
                flash_exception(e, 'Newsletter konnte nicht gespeichert werden.')
        elif current_path == '/change-password':
            target = '/v2/' if current_user.is_admin else '/v2/portal'
            if _change_password_from_request(get_db()):
                return redirect(target)
        elif current_path == '/portal/data':
            if not current_user.member_id:
                abort(403)
            _update_portal_member_from_request(get_db())
            return redirect('/v2/portal/data')
    db = get_db()
    data = _v2_shell_data(db, current_path)
    if current_path == '/import':
        data['native'] = _v2_import_data(db, results)
    elif current_path == '/admin/database':
        data['native'] = _v2_database_data(check_result, maintenance_result)
    data['content_path'] = _v2_content_path(subpath)
    data['current_path'] = current_path
    return render_template(
        'v2_index.html',
        v2_assets=_v2_assets(),
        v2_data=data,
    )


# === Import ===

@app.route('/import', methods=['GET', 'POST'])
@admin_required
def import_data():
    files_sort = request.args.get('files_sort', 'imported_at')
    files_dir = request.args.get('files_dir', 'asc').lower()
    values_sort = request.args.get('values_sort', 'imported_at')
    values_dir = request.args.get('values_dir', 'asc').lower()

    allowed_files_sort = {'period_start', 'source_file', 'imported_at'}
    allowed_values_sort = {'imported_at', 'filename', 'records_imported', 'status'}
    if files_sort not in allowed_files_sort:
        files_sort = 'imported_at'
    if values_sort not in allowed_values_sort:
        values_sort = 'imported_at'
    if files_dir not in {'asc', 'desc'}:
        files_dir = 'asc'
    if values_dir not in {'asc', 'desc'}:
        values_dir = 'asc'

    if request.method == 'POST':
        files = request.files.getlist('files')
        overwrite = request.form.get('overwrite') == '1'
        data_status = _valid_import_data_status(request.form.get('data_status'))
        results = []
        for f in files:
            if f and f.filename.lower().endswith('.xlsx'):
                filename = secure_filename(f.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                try:
                    f.save(filepath)
                    result = run_import(filepath, overwrite, data_status)
                except Exception as e:
                    app.logger.exception('Import upload handling failed for %s', filename)
                    result = {
                        'filename': filename,
                        'status': 'error',
                        'data_status': data_status,
                        'records': 0,
                        'overwritten': 0,
                        'error': str(e),
                        'imported_at': None,
                    }
                results.append(result)
                audit_log('import', f'Datei importiert: {filename} ({result["records"]} Datensätze, Status: {result["status"]})')
        db = get_db()
        imports = db.execute(f"""
            SELECT id, source_file, period_start, period_end, data_status, replaced_by_batch_id, replaced_at, imported_at
            FROM import_batches ORDER BY {files_sort} {files_dir.upper()}
        """).fetchall()
        import_values = db.execute(f"""
            SELECT id, filename, records_imported, records_overwritten, status, data_status, error_message, imported_by, imported_at
            FROM import_log ORDER BY {values_sort} {values_dir.upper()}
        """).fetchall()
        return render_template('import.html', results=results, imports=imports,
                               import_values=import_values,
                               files_sort=files_sort, files_dir=files_dir,
                               values_sort=values_sort, values_dir=values_dir)

    # Vorhandene Importe zeigen
    db = get_db()
    imports = db.execute(f"""
        SELECT id, source_file, period_start, period_end, data_status, replaced_by_batch_id, replaced_at, imported_at
        FROM import_batches ORDER BY {files_sort} {files_dir.upper()}
    """).fetchall()
    import_values = db.execute(f"""
        SELECT id, filename, records_imported, records_overwritten, status, data_status, error_message, imported_by, imported_at
        FROM import_log ORDER BY {values_sort} {values_dir.upper()}
    """).fetchall()
    return render_template('import.html', imports=imports, import_values=import_values,
                           files_sort=files_sort, files_dir=files_dir,
                           values_sort=values_sort, values_dir=values_dir)


def _valid_import_data_status(value):
    return 'final' if value == 'final' else 'provisional'


def _format_import_status_label(data_status):
    return 'Endgültig' if data_status == 'final' else 'Vorläufig'


def safe_invoice_pdf_filename(invoice_id, member_id, member_name):
    """Erzeugt einen Download-Dateinamen ohne daraus einen Serverpfad zu bauen."""
    safe_name = secure_filename(str(member_name or '').replace(' ', '_')) or f'mitglied_{member_id}'
    safe_name = safe_name[:80]
    return f'abrechnung_{int(invoice_id)}_{int(member_id)}_{safe_name}.pdf'


def _active_batches_for_period(db, period_start, period_end, data_status=None):
    params = [period_start, period_end]
    status_filter = ''
    if data_status:
        status_filter = ' AND data_status=?'
        params.append(data_status)
    return db.execute(f"""
        SELECT *
        FROM import_batches
        WHERE replaced_at IS NULL
          AND period_start = ?
          AND period_end = ?
          {status_filter}
        ORDER BY id
    """, params).fetchall()


def _delete_batch_measurements(db, batch_id):
    count = db.execute("SELECT COUNT(*) FROM measurements WHERE batch_id=?", (batch_id,)).fetchone()[0]
    db.execute("DELETE FROM measurements WHERE batch_id=?", (batch_id,))
    return count


def _mark_batches_replaced(db, batches, replacement_batch_id):
    for batch in batches:
        _delete_batch_measurements(db, batch['id'])
        db.execute("""
            UPDATE import_batches
            SET replaced_by_batch_id=?, replaced_at=datetime('now')
            WHERE id=?
        """, (replacement_batch_id, batch['id']))


def run_import(filepath, overwrite=False, data_status='final'):
    """Importiert eine Excel-Datei. Bei overwrite=True werden bestehende Daten überschrieben."""
    sys.path.insert(0, os.path.join(BASE_DIR, '..'))
    from import_eda import import_file, parse_filename

    data_status = _valid_import_data_status(data_status)
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    ensure_import_schema(db)

    filename = os.path.basename(filepath)
    records_overwritten = 0
    new_batch_id = None
    replacement_candidates = []

    try:
        info = parse_filename(filename)
        period_start = info.get('period_start')
        period_end = info.get('period_end')
        if not period_start or not period_end:
            raise ValueError('Der Zeitraum konnte aus dem Dateinamen nicht erkannt werden.')

        active_final = _active_batches_for_period(db, period_start, period_end, 'final')
        active_provisional = _active_batches_for_period(db, period_start, period_end, 'provisional')

        if data_status == 'provisional' and active_final and not overwrite:
            raise ValueError('Für diesen Zeitraum sind bereits endgültige Daten vorhanden. Vorläufige Daten werden nicht darüber importiert.')
        if data_status == 'final' and active_final and not overwrite:
            raise ValueError('Für diesen Zeitraum sind bereits endgültige Daten vorhanden. Zum Ersetzen bitte Überschreiben aktivieren.')

        if data_status == 'final':
            replacement_candidates = list(active_provisional)
            if overwrite:
                replacement_candidates.extend(active_final)
        elif overwrite:
            replacement_candidates = list(active_provisional)

        count = import_file(filepath, db, allow_duplicate=bool(replacement_candidates or overwrite))
        new_batch = db.execute("""
            SELECT id
            FROM import_batches
            WHERE source_file=? AND period_start=? AND period_end=? AND replaced_at IS NULL
            ORDER BY id DESC
            LIMIT 1
        """, (filename, period_start, period_end)).fetchone()
        if new_batch:
            new_batch_id = new_batch['id']
            db.execute("UPDATE import_batches SET data_status=? WHERE id=?", (data_status, new_batch_id))

        if new_batch_id and replacement_candidates:
            records_overwritten = sum(_delete_batch_measurements(db, batch['id']) for batch in replacement_candidates)
            for batch in replacement_candidates:
                db.execute("""
                    UPDATE import_batches
                    SET replaced_by_batch_id=?, replaced_at=datetime('now')
                    WHERE id=?
                """, (new_batch_id, batch['id']))

        db.commit()
        status = 'success'
        error = None
    except Exception as e:
        status = 'error'
        count = 0
        error = str(e)
        app.logger.exception('Import failed for %s', filename)
    finally:
        # Log
        cur = db.execute("""INSERT INTO import_log (filename, records_imported, records_overwritten, status, data_status, error_message, imported_by)
                      VALUES (?, ?, ?, ?, ?, ?, ?)""",
                         (filename, count, records_overwritten, status, data_status, error,
                          current_user.username if has_request_context() and current_user and current_user.is_authenticated else 'system'))
        log_row = db.execute("SELECT imported_at FROM import_log WHERE id=?", (cur.lastrowid,)).fetchone()
        db.commit()
        db.close()

    return {'filename': filename, 'status': status, 'records': count,
            'data_status': data_status,
            'overwritten': records_overwritten, 'error': error,
            'imported_at': (log_row['imported_at'] if log_row else None)}


# === Mitglieder ===

@app.route('/members')
@admin_required
def members_list():
    db = get_db()
    members = db.execute("""
        SELECT * FROM members ORDER BY name
    """).fetchall()
    return render_template('members.html', members=members)


@app.route('/members/new', methods=['GET', 'POST'])
@admin_required
def member_new():
    if request.method == 'POST':
        db = get_db()
        newsletter_optout = 0 if form_switch_enabled('newsletter_enabled') else 1
        db.execute("""INSERT INTO members (name, email, phone, address_street, address_zip, address_city,
                      einspeiser_zp, einspeiser_ab, bezug_zp, bezug_ab, teilnahme,
                      iban, bic, account_holder, newsletter_optout, updated_at)
                      VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))""",
                   (request.form['name'], request.form.get('email'),
                    request.form.get('phone'),
                    request.form.get('address_street'), request.form.get('address_zip'),
                    request.form.get('address_city'),
                    request.form.get('einspeiser_zp') or None,
                    request.form.get('einspeiser_ab') or None,
                    request.form.get('bezug_zp') or None,
                    request.form.get('bezug_ab') or None,
                    float(request.form.get('teilnahme', 1.0)),
                    request.form.get('iban') or None,
                    request.form.get('bic') or None,
                    request.form.get('account_holder') or None,
                    newsletter_optout))
        db.commit()
        audit_log('member_create', f'Mitglied angelegt: {request.form["name"]}')
        flash('Mitglied angelegt.', 'success')
        return redirect(url_for('members_list'))
    return render_template('member_edit.html', member=None)


@app.route('/members/<int:id>/edit', methods=['GET', 'POST'])
@admin_required
def member_edit(id):
    db = get_db()
    if request.method == 'POST':
        newsletter_optout = 0 if form_switch_enabled('newsletter_enabled') else 1
        db.execute("""UPDATE members SET name=?, email=?, phone=?, address_street=?, address_zip=?,
                      address_city=?, einspeiser_zp=?, einspeiser_ab=?, bezug_zp=?,
                      bezug_ab=?, teilnahme=?, active=?, iban=?, bic=?, account_holder=?,
                      newsletter_optout=?,
                      updated_at=datetime('now')
                      WHERE id=?""",
                   (request.form['name'], request.form.get('email'),
                    request.form.get('phone'),
                    request.form.get('address_street'), request.form.get('address_zip'),
                    request.form.get('address_city'),
                    request.form.get('einspeiser_zp') or None,
                    request.form.get('einspeiser_ab') or None,
                    request.form.get('bezug_zp') or None,
                    request.form.get('bezug_ab') or None,
                    float(request.form.get('teilnahme', 1.0)),
                    1 if request.form.get('active') else 0,
                    request.form.get('iban') or None,
                    request.form.get('bic') or None,
                    request.form.get('account_holder') or None,
                    newsletter_optout,
                    id))
        db.commit()
        audit_log('member_edit', f'Mitglied bearbeitet: {request.form["name"]} (ID {id})')
        flash('Mitglied aktualisiert.', 'success')
        return redirect(url_for('members_list'))
    member = db.execute("SELECT * FROM members WHERE id=?", (id,)).fetchone()
    return render_template('member_edit.html', member=member)


@app.route('/members/<int:id>/delete', methods=['POST'])
@admin_required
def member_delete(id):
    db = get_db()
    member = db.execute("SELECT name FROM members WHERE id=?", (id,)).fetchone()
    db.execute("UPDATE members SET active=0, updated_at=datetime('now') WHERE id=?", (id,))
    db.commit()
    audit_log('member_delete', f'Mitglied deaktiviert: {member["name"]} (ID {id})')
    flash('Mitglied deaktiviert.', 'success')
    next_url = request.form.get('next') or request.args.get('next')
    if next_url and is_safe_redirect_url(next_url):
        return redirect(next_url)
    return redirect(url_for('members_list'))


# === Preise ===

@app.route('/prices', methods=['GET', 'POST'])
@admin_required
def prices():
    db = get_db()
    if request.method == 'POST':
        db.execute("""INSERT INTO prices (valid_from, valid_to, price_consumption, price_generation, description)
                      VALUES (?, ?, ?, ?, ?)""",
                   (request.form['valid_from'], request.form['valid_to'],
                    float(request.form['price_consumption']),
                    float(request.form['price_generation']),
                    request.form.get('description', '')))
        db.commit()
        audit_log('price_create', f'Preis angelegt: {request.form["valid_from"]} - {request.form["valid_to"]}')
        flash('Preis angelegt.', 'success')
        return redirect(url_for('prices'))
    all_prices = db.execute("SELECT * FROM prices ORDER BY valid_from DESC").fetchall()
    # Prüfe ob es Abrechnungen für die Preis-Zeiträume gibt
    invoices_for_prices = {}
    for p in all_prices:
        inv = db.execute("""SELECT id, period_from, period_to FROM invoices
                           WHERE period_from <= ? AND period_to >= ?""",
                        (p['valid_to'], p['valid_from'])).fetchone()
        if inv:
            invoices_for_prices[p['id']] = inv
    return render_template('prices.html', prices=all_prices, invoices_for_prices=invoices_for_prices)


@app.route('/prices/<int:id>/edit', methods=['POST'])
@admin_required
def price_edit(id):
    db = get_db()
    price = db.execute("SELECT * FROM prices WHERE id=?", (id,)).fetchone()
    if not price:
        flash('Preis nicht gefunden.', 'danger')
        return redirect(url_for('prices'))
    db.execute("""UPDATE prices SET valid_from=?, valid_to=?, price_consumption=?,
                  price_generation=?, description=? WHERE id=?""",
               (request.form['valid_from'], request.form['valid_to'],
                float(request.form['price_consumption']),
                float(request.form['price_generation']),
                request.form.get('description', ''), id))
    db.commit()
    audit_log('price_edit', f'Preis bearbeitet: {request.form["valid_from"]} - {request.form["valid_to"]} (ID {id})')
    # Warnung wenn Abrechnung existiert
    inv = db.execute("""SELECT id FROM invoices
                       WHERE period_from <= ? AND period_to >= ?""",
                    (request.form['valid_to'], request.form['valid_from'])).fetchone()
    if inv:
        flash(f'Achtung: Für diesen Zeitraum existiert bereits Abrechnung #{inv["id"]}. '
              f'Es muss eine neue Abrechnung erstellt werden, damit die Preisänderung wirksam wird!', 'warning')
    else:
        flash('Preis aktualisiert.', 'success')
    return redirect(url_for('prices'))


@app.route('/prices/<int:id>/duplicate', methods=['POST'])
@admin_required
def price_duplicate(id):
    """Preis in die nächste Periode (Quartal) duplizieren."""
    from datetime import timedelta
    db = get_db()
    price = db.execute("SELECT * FROM prices WHERE id=?", (id,)).fetchone()
    if not price:
        flash('Preis nicht gefunden.', 'danger')
        return redirect(url_for('prices'))
    # Nächstes Quartal berechnen
    old_from = datetime.strptime(price['valid_from'], '%Y-%m-%d').date()
    old_to = datetime.strptime(price['valid_to'], '%Y-%m-%d').date()
    duration = (old_to - old_from).days + 1
    new_from = old_to + timedelta(days=1)
    new_to = new_from + timedelta(days=duration - 1)
    # Duplikat prüfen
    existing = db.execute("SELECT id FROM prices WHERE valid_from=? AND valid_to=?",
                          (new_from.isoformat(), new_to.isoformat())).fetchone()
    if existing:
        flash(f'Für den Zeitraum {new_from} – {new_to} existiert bereits ein Preis.', 'warning')
        return redirect(url_for('prices'))
    # Neue Periode: Beschreibung anpassen
    new_desc = price['description'] or ''
    # Versuche Q-Nummer hochzuzählen
    import re
    q_match = re.search(r'Q(\d)/(\d{4})', new_desc)
    if q_match:
        q_num = int(q_match.group(1))
        q_year = int(q_match.group(2))
        if q_num < 4:
            new_desc = new_desc.replace(q_match.group(0), f'Q{q_num+1}/{q_year}')
        else:
            new_desc = new_desc.replace(q_match.group(0), f'Q1/{q_year+1}')
    db.execute("""INSERT INTO prices (valid_from, valid_to, price_consumption, price_generation, description)
                  VALUES (?, ?, ?, ?, ?)""",
               (new_from.isoformat(), new_to.isoformat(),
                price['price_consumption'], price['price_generation'], new_desc))
    db.commit()
    audit_log('price_duplicate', f'Preis dupliziert: {new_from} - {new_to} (von ID {id})')
    flash(f'Preis in nächste Periode kopiert: {new_from} – {new_to}', 'success')
    return redirect(url_for('prices'))


@app.route('/prices/<int:id>/delete', methods=['POST'])
@admin_required
def price_delete(id):
    db = get_db()
    price = db.execute("SELECT valid_from, valid_to FROM prices WHERE id=?", (id,)).fetchone()
    db.execute("DELETE FROM prices WHERE id=?", (id,))
    db.commit()
    audit_log('price_delete', f'Preis gelöscht: {price["valid_from"]} - {price["valid_to"]}' if price else f'Preis ID {id} gelöscht')
    flash('Preis gelöscht.', 'success')
    return redirect(url_for('prices'))


def get_price_for_date(db, target_date):
    """Ermittelt den gültigen Preis für ein Datum."""
    row = db.execute("""
        SELECT price_consumption, price_generation FROM prices
        WHERE valid_from <= ? AND valid_to >= ?
        ORDER BY valid_from DESC LIMIT 1
    """, (target_date, target_date)).fetchone()
    if row:
        return row['price_consumption'], row['price_generation']
    # Fallback: Letzten eingetragenen Preis verwenden
    last = db.execute("SELECT price_consumption, price_generation FROM prices ORDER BY valid_from DESC LIMIT 1").fetchone()
    if last:
        return last['price_consumption'], last['price_generation']
    return 12.0, 10.0  # Absoluter Fallback (aktueller Standardpreis 2026)


def get_import_status_for_period(db, period_from, period_to):
    """Prüft, ob fuer einen Abrechnungszeitraum nur finale aktive Importdaten vorliegen."""
    ts_from = period_from + "T00:00:00" if "T" not in period_from else period_from
    ts_to = period_to + "T23:45:00" if "T" not in period_to else period_to
    batches = db.execute("""
        SELECT id, source_file, period_start, period_end, data_status
        FROM import_batches
        WHERE replaced_at IS NULL
          AND period_start <= ?
          AND period_end >= ?
        ORDER BY period_start, id
    """, (ts_to, ts_from)).fetchall()
    has_final = any(batch['data_status'] == 'final' for batch in batches)
    provisional = [batch for batch in batches if batch['data_status'] != 'final']
    if not batches:
        return {
            'data_status': 'provisional',
            'is_final': False,
            'reason': 'Für diesen Zeitraum wurden noch keine Messdaten importiert.',
            'batches': [],
            'provisional_batches': [],
        }
    if provisional:
        return {
            'data_status': 'provisional',
            'is_final': False,
            'reason': 'Im Zeitraum sind noch vorläufige Messdaten vorhanden.',
            'batches': batches,
            'provisional_batches': provisional,
        }
    if not has_final:
        return {
            'data_status': 'provisional',
            'is_final': False,
            'reason': 'Es wurden keine endgültigen Messdaten für diesen Zeitraum gefunden.',
            'batches': batches,
            'provisional_batches': [],
        }
    return {
        'data_status': 'final',
        'is_final': True,
        'reason': '',
        'batches': batches,
        'provisional_batches': [],
    }


def invoice_finalization_blocker(db, invoice):
    """Liefert eine Fehlermeldung, wenn eine Abrechnung nicht finalisiert/versendet werden darf."""
    import_status = get_import_status_for_period(db, invoice['period_from'], invoice['period_to'])
    invoice_data_status = invoice['data_status'] if 'data_status' in invoice.keys() else 'final'
    if invoice_data_status != 'final':
        if import_status['is_final']:
            return 'Diese Abrechnung wurde mit vorläufigen Daten berechnet. Bitte zuerst neu berechnen, danach kann sie versendet werden.'
        return 'Diese Abrechnung basiert auf vorläufigen Messdaten. Versand und Abschluss sind erst mit endgültigen Daten möglich.'
    if not import_status['is_final']:
        return import_status['reason'] or 'Für diesen Zeitraum liegen noch keine endgültigen Messdaten vor.'
    return ''


def get_invoice_carryovers(db, invoice_id, member_id=None):
    """Liefert Finanzvortraege einer Abrechnung, optional fuer ein Mitglied."""
    params = [invoice_id]
    member_filter = ''
    if member_id is not None:
        member_filter = 'AND c.member_id=?'
        params.append(member_id)
    return db.execute(f"""
        SELECT c.*, m.name AS member_name, m.email AS member_email,
               src.period_from AS source_period_from,
               src.period_to AS source_period_to
        FROM invoice_carryovers c
        JOIN members m ON m.id = c.member_id
        JOIN invoices src ON src.id = c.source_invoice_id
        WHERE c.invoice_id=?
          {member_filter}
        ORDER BY m.name, src.period_from, c.source_invoice_id
    """, params).fetchall()


def get_invoice_carryover_map(db, invoice_id):
    """Gruppiert Vortraege nach Mitglied fuer Detailansichten und Zahlungslogik."""
    carryovers = get_invoice_carryovers(db, invoice_id)
    grouped = {}
    for row in carryovers:
        member_id = row['member_id']
        bucket = grouped.setdefault(member_id, {'total': 0.0, 'rows': []})
        bucket['total'] = round(bucket['total'] + row['amount_eur'], 2)
        bucket['rows'].append(row)
    return grouped


def calculate_carryovers_for_period(db, period_from):
    """Berechnet offene Vorperioden, die in eine neue Abrechnung uebernommen werden.

    Uebernommen wird der offene Restbetrag. Bei einer Unterzahlung bleibt die
    Restforderung offen, bei einer Ueberzahlung wird das Guthaben vorgetragen.
    """
    carryovers = []
    for row in get_payment_rows(db):
        if row['period_to'] >= period_from:
            continue
        if row.get('invoice_status') not in {'sent', 'finalized'}:
            continue
        if row['paid'] or row.get('is_settled_by_carryover'):
            continue
        if abs(row['open_amount']) < 0.005:
            continue
        if row['booked_total']:
            description = ('Restforderung aus Vorperioden' if row['open_amount'] > 0
                           else 'Überzahlung aus Vorperioden')
        else:
            description = ('Buchungsrückstand aus Vorperioden' if row['open_amount'] > 0
                           else 'Guthaben aus Vorperioden')
        carryovers.append({
            'member_id': row['member_id'],
            'source_invoice_id': row['invoice_id'],
            'amount': round(row['open_amount'], 2),
            'description': description,
        })
    return carryovers


def save_invoice_carryovers(db, invoice_id, carryovers):
    for carryover in carryovers:
        db.execute("""
            INSERT OR REPLACE INTO invoice_carryovers (
                invoice_id, member_id, source_invoice_id, amount_eur, description
            ) VALUES (?, ?, ?, ?, ?)
        """, (
            invoice_id,
            carryover['member_id'],
            carryover['source_invoice_id'],
            carryover['amount'],
            carryover.get('description') or '',
        ))


def invoice_recipient_rows(db, invoice_id):
    """Mitglieder, die Positionen oder einen Finanzvortrag in der Abrechnung haben."""
    return db.execute("""
        SELECT m.id AS member_id, m.name, m.email
        FROM members m
        WHERE m.id IN (
            SELECT member_id FROM invoice_items WHERE invoice_id=?
            UNION
            SELECT member_id FROM invoice_carryovers WHERE invoice_id=?
        )
        ORDER BY m.name
    """, (invoice_id, invoice_id)).fetchall()


# === Abrechnung ===

@app.route('/invoices')
@admin_required
def invoices_list():
    db = get_db()
    invoices = db.execute("SELECT * FROM invoices ORDER BY period_from DESC").fetchall()
    return render_template('invoices.html', invoices=invoices)


@app.route('/invoices/new', methods=['GET', 'POST'])
@admin_required
def invoice_new():
    if request.method == 'POST':
        period_from = request.form['period_from']
        period_to = request.form['period_to']
        db = get_db()

        # Duplikat-Prüfung: Keine überlappenden Abrechnungen erlauben
        existing = db.execute("""
            SELECT id, period_from, period_to FROM invoices
            WHERE period_from <= ? AND period_to >= ?
        """, (period_to, period_from)).fetchone()
        if existing:
            flash(f'Es existiert bereits eine Abrechnung für diesen Zeitraum '
                  f'(Nr. {existing["id"]}: {existing["period_from"]} – {existing["period_to"]}). '
                  f'Pro Quartal ist nur eine Abrechnung zulässig.', 'danger')
            return redirect(url_for('invoice_new'))

        # Preise für Zeitraum
        price_cons, price_gen = get_price_for_date(db, period_from)
        import_status = get_import_status_for_period(db, period_from, period_to)

        # Abrechnung berechnen
        result = calculate_billing(db, period_from, period_to, price_cons, price_gen)

        # Speichern
        cur = db.execute("""INSERT INTO invoices (period_from, period_to, total_kwh_traded,
                            total_income, total_expense, total_margin, data_status)
                            VALUES (?, ?, ?, ?, ?, ?, ?)""",
                         (period_from, period_to, result['total_kwh'],
                          result['total_income'], result['total_expense'], result['total_margin'],
                          import_status['data_status']))
        invoice_id = cur.lastrowid

        # Einzelpositionen
        for item in result['items']:
            db.execute("""INSERT INTO invoice_items (invoice_id, member_id, type, kwh, price_per_kwh, amount_eur)
                          VALUES (?, ?, ?, ?, ?, ?)""",
                       (invoice_id, item['member_id'], item['type'],
                        item['kwh'], item['price'], item['amount']))
        save_invoice_carryovers(db, invoice_id, result['carryovers'])
        db.commit()
        audit_log('invoice_create', f'Abrechnung #{invoice_id} erstellt: {period_from} - {period_to} ({result["total_kwh"]:.1f} kWh)')
        carryover_total = round(sum(item['amount'] for item in result['carryovers']), 2)
        carryover_info = f' Finanzvortrag: {carryover_total:.2f} EUR.' if result['carryovers'] else ''
        if import_status['is_final']:
            flash(f'Abrechnung #{invoice_id} erstellt ({result["total_kwh"]:.1f} kWh).{carryover_info}', 'success')
        else:
            flash(f'Vorläufige Abrechnung #{invoice_id} erstellt ({result["total_kwh"]:.1f} kWh).{carryover_info} Versand und Abschluss sind gesperrt, bis endgültige Daten importiert und die Abrechnung neu berechnet wurde.', 'warning')
        return redirect(url_for('invoice_detail', id=invoice_id))

    # Quartalsvorschläge
    today = date.today()
    q_month = ((today.month - 1) // 3) * 3 + 1
    q_start = date(today.year, q_month, 1)
    if q_month > 3:
        prev_q_start = date(today.year, q_month - 3, 1)
    else:
        prev_q_start = date(today.year - 1, 10, 1)
    prev_q_end = date(q_start.year, q_start.month, 1)
    from calendar import monthrange
    prev_end_month = q_month - 1 if q_month > 1 else 12
    prev_end_year = today.year if q_month > 1 else today.year - 1
    _, last_day = monthrange(prev_end_year, prev_end_month)
    prev_q_end = date(prev_end_year, prev_end_month, last_day)

    return render_template('invoice_new.html',
                           suggested_from=prev_q_start.isoformat(),
                           suggested_to=prev_q_end.isoformat())


@app.route('/invoices/<int:id>')
@admin_required
def invoice_detail(id):
    db = get_db()
    invoice = db.execute("SELECT * FROM invoices WHERE id=?", (id,)).fetchone()
    items = db.execute("""
        SELECT ii.*, m.name as member_name, m.email as member_email
        FROM invoice_items ii
        JOIN members m ON m.id = ii.member_id
        WHERE ii.invoice_id = ?
        ORDER BY m.name, ii.type
    """, (id,)).fetchall()
    # Pro Mitglied zusammenfassen
    members_map = {}
    for item in items:
        mid = item['member_id']
        if mid not in members_map:
            members_map[mid] = {
                'member_id': mid,
                'member_name': item['member_name'],
                'member_email': item['member_email'],
                'cons_kwh': 0, 'cons_eur': 0, 'cons_price': 0,
                'gen_kwh': 0, 'gen_eur': 0, 'gen_price': 0,
            }
        if item['type'] == 'consumption':
            members_map[mid]['cons_kwh'] = item['kwh']
            members_map[mid]['cons_eur'] = item['amount_eur']
            members_map[mid]['cons_price'] = item['price_per_kwh']
        else:
            members_map[mid]['gen_kwh'] = item['kwh']
            members_map[mid]['gen_eur'] = item['amount_eur']
            members_map[mid]['gen_price'] = item['price_per_kwh']
    for m in members_map.values():
        m['energy_net_eur'] = round(m['cons_eur'] - m['gen_eur'], 2)
        m['carryover_eur'] = 0.0
        m['carryovers'] = []

    carryover_map = get_invoice_carryover_map(db, id)
    for mid, data in carryover_map.items():
        if mid not in members_map:
            first = data['rows'][0]
            members_map[mid] = {
                'member_id': mid,
                'member_name': first['member_name'],
                'member_email': first['member_email'],
                'cons_kwh': 0, 'cons_eur': 0, 'cons_price': 0,
                'gen_kwh': 0, 'gen_eur': 0, 'gen_price': 0,
                'energy_net_eur': 0.0,
            }
        members_map[mid]['carryover_eur'] = data['total']
        members_map[mid]['carryovers'] = data['rows']

    for m in members_map.values():
        m.setdefault('energy_net_eur', round(m['cons_eur'] - m['gen_eur'], 2))
        m.setdefault('carryover_eur', 0.0)
        m.setdefault('carryovers', [])
        m['net_eur'] = round(m['energy_net_eur'] + m['carryover_eur'], 2)
    member_rows = sorted(members_map.values(), key=lambda x: x['member_name'])

    emails = db.execute("""
        SELECT el.*, m.name as member_name
        FROM email_log el
        LEFT JOIN members m ON m.id = el.member_id
        WHERE el.invoice_id=? ORDER BY el.sent_at DESC
    """, (id,)).fetchall()

    # E-Mail-Status pro Mitglied ermitteln
    sent_members = set()
    for e in emails:
        if e['status'] == 'sent' and e['member_id']:
            sent_members.add(e['member_id'])
    for m in member_rows:
        m['email_sent'] = m['member_id'] in sent_members

    import_status = get_import_status_for_period(db, invoice['period_from'], invoice['period_to'])
    finalization_blocker = invoice_finalization_blocker(db, invoice)
    return render_template('invoice_detail.html', invoice=invoice, items=items,
                           member_rows=member_rows, emails=emails,
                           import_status=import_status,
                           finalization_blocker=finalization_blocker)


@app.route('/invoices/<int:id>/regenerate', methods=['POST'])
@admin_required
def invoice_regenerate(id):
    """Abrechnung neu berechnen (z.B. nach Preisänderung)."""
    db = get_db()
    invoice = db.execute("SELECT * FROM invoices WHERE id=?", (id,)).fetchone()
    if not invoice:
        flash('Abrechnung nicht gefunden.', 'danger')
        return redirect(url_for('invoices_list'))

    period_from = invoice['period_from']
    period_to = invoice['period_to']

    # Aktuelle Preise für Zeitraum laden
    price_cons, price_gen = get_price_for_date(db, period_from)
    import_status = get_import_status_for_period(db, period_from, period_to)

    # Alte Items und Finanzvortraege löschen
    db.execute("DELETE FROM invoice_items WHERE invoice_id=?", (id,))
    db.execute("DELETE FROM invoice_carryovers WHERE invoice_id=?", (id,))

    # Neu berechnen
    result = calculate_billing(db, period_from, period_to, price_cons, price_gen)

    # Invoice-Kopf aktualisieren
    db.execute("""UPDATE invoices SET total_kwh_traded=?, total_income=?, total_expense=?,
                  total_margin=?, data_status=?, status='draft', finalized_at=NULL WHERE id=?""",
               (result['total_kwh'], result['total_income'], result['total_expense'],
                result['total_margin'], import_status['data_status'], id))

    # Neue Einzelpositionen
    for item in result['items']:
        db.execute("""INSERT INTO invoice_items (invoice_id, member_id, type, kwh, price_per_kwh, amount_eur)
                      VALUES (?, ?, ?, ?, ?, ?)""",
                   (id, item['member_id'], item['type'],
                    item['kwh'], item['price'], item['amount']))
    save_invoice_carryovers(db, id, result['carryovers'])
    db.commit()
    audit_log('invoice_regenerate', f'Abrechnung #{id} neu berechnet: {period_from} - {period_to} '
              f'(Verbrauch: {price_cons} ct, Erzeugung: {price_gen} ct, {result["total_kwh"]:.1f} kWh)')
    if import_status['is_final']:
        flash(f'Abrechnung #{id} wurde mit endgültigen Daten und aktuellen Preisen '
              f'(Verbrauch: {price_cons} ct/kWh, Erzeugung: {price_gen} ct/kWh) neu berechnet.', 'success')
    else:
        flash(f'Abrechnung #{id} wurde vorläufig neu berechnet. Versand und Abschluss bleiben gesperrt, bis endgültige Daten importiert wurden.', 'warning')
    return redirect(url_for('invoice_detail', id=id))


@app.route('/invoices/<int:id>/pdf/<int:member_id>')
@login_required
def invoice_pdf(id, member_id):
    """PDF für ein Mitglied generieren (A4, mehrseitig)."""
    # Members dürfen nur eigene PDFs abrufen
    if not current_user.is_admin and current_user.member_id != member_id:
        audit_log('pdf_access_denied', f'PDF-Zugriff verweigert: Rechnung {id}, Mitglied {member_id}')
        flash('Zugriff verweigert.', 'danger')
        return redirect(url_for('portal_dashboard'))
    audit_log('pdf_download', f'PDF heruntergeladen: Rechnung {id}, Mitglied {member_id}')
    import math
    db = get_db()
    invoice = db.execute("SELECT * FROM invoices WHERE id=?", (id,)).fetchone()
    member = db.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
    items = db.execute("""
        SELECT * FROM invoice_items
        WHERE invoice_id=? AND member_id=?
    """, (id, member_id)).fetchall()
    carryovers = get_invoice_carryovers(db, id, member_id)
    carryover_total = round(sum(row['amount_eur'] for row in carryovers), 2)

    # --- Nettobetrag berechnen (Bezug - Gutschrift) ---
    energy_net_total = 0
    for item in items:
        if item['type'] == 'consumption':
            energy_net_total += item['amount_eur']
        else:
            energy_net_total -= item['amount_eur']
    net_total = round(energy_net_total + carryover_total, 2)

    # --- EPC QR Code für Überweisung ---
    qr_data_uri = ''
    if net_total > 0:
        qr_data_uri = generate_epc_qr(net_total, invoice, member)

    # --- Seite 2: Mitglieder-Statistiken ---
    member_stats = get_member_stats(db, member, invoice['period_from'], invoice['period_to'])

    # --- Seite 3: Community-Statistiken ---
    community_stats = get_community_stats(db, invoice)

    # --- Seite 4: Ersparnis-Berechnung ---
    savings = calculate_member_savings(member_stats, items)
    public_cfg = get_public_config(db)

    # Logo als base64 für PDF-Einbettung
    import base64
    logo_path = os.path.join(BASE_DIR, 'static', 'logo_small.png')
    with open(logo_path, 'rb') as f:
        logo_b64 = 'data:image/png;base64,' + base64.b64encode(f.read()).decode('ascii')

    # --- Pie Chart SVG generieren ---
    def generate_pie_svg(data, colors, size=120):
        """Erzeugt ein SVG-Tortendiagramm."""
        total = sum(d['value'] for d in data)
        if total == 0:
            return ''
        cx, cy, r = size/2, size/2, size/2 - 4
        svg_parts = [f'<svg xmlns="http://www.w3.org/2000/svg" width="{size}" height="{size}" viewBox="0 0 {size} {size}">']
        start_angle = -90
        for i, d in enumerate(data):
            if d['value'] == 0:
                continue
            pct = d['value'] / total
            angle = pct * 360
            end_angle = start_angle + angle
            large_arc = 1 if angle > 180 else 0
            x1 = cx + r * math.cos(math.radians(start_angle))
            y1 = cy + r * math.sin(math.radians(start_angle))
            x2 = cx + r * math.cos(math.radians(end_angle))
            y2 = cy + r * math.sin(math.radians(end_angle))
            color = colors[i % len(colors)]
            if pct >= 0.9999:
                svg_parts.append(f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="{color}"/>')
            else:
                svg_parts.append(f'<path d="M {cx},{cy} L {x1:.2f},{y1:.2f} A {r},{r} 0 {large_arc} 1 {x2:.2f},{y2:.2f} Z" fill="{color}"/>')
            start_angle = end_angle
        svg_parts.append('</svg>')
        return '\n'.join(svg_parts)

    # Pie: Energieverteilung Bezug (pro Mitglied)
    pie_colors = ['#2b7a78', '#3aafa9', '#5cbdb9', '#81cdc6', '#a6ddd6',
                  '#17252a', '#4e8a7a', '#7cc4b5', '#b0e0d6', '#d4f0eb']
    pie_consumption_data = [{'label': m['name'], 'value': m['kwh']}
                            for m in community_stats['member_consumption']]
    pie_consumption_svg = generate_pie_svg(pie_consumption_data, pie_colors, 130)

    # Pie: Erzeugung
    pie_gen_colors = ['#ff9800', '#ffc107', '#ffb74d', '#ffe082', '#fff3e0']
    pie_generation_data = [{'label': m['name'], 'value': m['kwh']}
                           for m in community_stats.get('member_generation', [])]
    pie_generation_svg = generate_pie_svg(pie_generation_data, pie_gen_colors, 130)

    # Pie: Monatlicher Verbrauch des Mitglieds
    pie_monthly_colors = ['#1a535c', '#2b7a78', '#3aafa9', '#5cbdb9', '#7ed6c9',
                          '#a0e8dd', '#c2f5ed', '#17252a', '#4e8a7a', '#81cdc6', '#b0e0d6', '#d4f0eb']
    pie_monthly_data = [{'label': m['label'], 'value': m['consumption']}
                        for m in member_stats['monthly_data']]
    pie_monthly_svg = generate_pie_svg(pie_monthly_data, pie_monthly_colors, 130)

    html = render_template('invoice_pdf.html',
                           invoice=invoice, member=member, items=items,
                           carryovers=carryovers,
                           member_stats=member_stats, community_stats=community_stats,
                           energy_net_total=round(energy_net_total, 2),
                           carryover_total=carryover_total,
                           net_total=net_total, qr_data_uri=qr_data_uri,
                           savings=savings, logo_b64=logo_b64,
                           public_cfg=public_cfg,
                           pie_consumption_svg=pie_consumption_svg,
                           pie_generation_svg=pie_generation_svg,
                           pie_monthly_svg=pie_monthly_svg,
                           pie_consumption_data=pie_consumption_data,
                           pie_generation_data=pie_generation_data,
                           pie_monthly_data=pie_monthly_data,
                           pie_colors=pie_colors,
                           pie_gen_colors=pie_gen_colors,
                           pie_monthly_colors=pie_monthly_colors)

    from weasyprint import HTML
    pdf_filename = safe_invoice_pdf_filename(id, member_id, member['name'])
    pdf_bytes = HTML(string=html, base_url=BASE_DIR).write_pdf()

    preview_pdf = request.args.get('preview') == '1'
    return send_file(io.BytesIO(pdf_bytes), as_attachment=not preview_pdf,
                     download_name=pdf_filename, mimetype='application/pdf')


# === SEPA-Ueberweisung als QR-Code (EPC069-12, GiroCode) ===

def generate_epc_qr(amount, invoice, member):
    """Generiert einen EPC/GiroCode QR-Code als data URI (base64 PNG)."""
    import qrcode
    import base64

    cfg = get_public_config(get_db())
    try:
        payload = build_epc_payload(
            cfg['payment_recipient'], cfg['payment_iban'], amount,
            f'EEG-Abr {invoice["id"]}/{invoice["created_at"][:4]} {member["name"][:30]}',
            cfg['payment_bic'])
    except ValueError as e:
        app.logger.warning('QR-Code für Abrechnung %s nicht erstellt: %s', invoice['id'], e)
        return ''

    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M,
                       box_size=4, border=2)
    qr.add_data(payload)
    qr.make(fit=True)
    buffer = io.BytesIO()
    qr.make_image(fill_color='black', back_color='white').save(buffer, format='PNG')
    return 'data:image/png;base64,' + base64.b64encode(buffer.getvalue()).decode('ascii')


def calculate_member_savings(member_stats, items, price_cons=None, price_gen=None):
    """Berechnet die Ersparnis eines Mitglieds durch EEG-Teilnahme."""
    # Vergleichswerte für lokale EEG ("kleine EEG")
    # Hinweis: Bei lokaler EEG gilt eine höhere Netzentgelt-Reduktion als bei regionaler EEG.
    market_price_ct = 25.0  # Durchschnittlicher Haushaltsstrompreis AT 2026 in ct/kWh
    elabg_ct = 1.5          # Elektrizitätsabgabe (entfällt in EEG)
    eeg_type = 'lokal'
    local_netz_reduction_pct = 57.0
    # Näherung für den energieabhängigen Netzentgelt-Anteil (ct/kWh), auf den die Reduktion wirkt.
    netzentgelt_base_ct = 4.0
    netzentgelt_reduction_ct = netzentgelt_base_ct * (local_netz_reduction_pct / 100.0)

    # Preise aus den tatsächlichen Rechnungspositionen ableiten
    if price_cons is None or price_gen is None:
        # Aus den Items die tatsächlich berechneten Preise lesen
        for item in items:
            if item['type'] == 'consumption' and item['price_per_kwh']:
                price_cons = item['price_per_kwh']
                break
        for item in items:
            if item['type'] == 'generation' and item['price_per_kwh']:
                price_gen = item['price_per_kwh']
                break
    # Fallback nur wenn gar keine Items vorhanden
    eeg_price_ct = price_cons if price_cons else 12.0
    eeg_gen_price_ct = price_gen if price_gen else 10.0

    cons_kwh = member_stats['total_consumption_kwh']
    gen_kwh = member_stats['total_generation_kwh']

    # Berechnung Bezugsseite
    cost_market = cons_kwh * market_price_ct / 100.0  # Was der Strom am Markt kosten würde
    cost_eeg = cons_kwh * eeg_price_ct / 100.0        # Was er in der EEG kostet
    saving_price = cost_market - cost_eeg             # Ersparnis durch günstigen EEG-Preis
    saving_elabg = cons_kwh * elabg_ct / 100.0        # Ersparnis Elektrizitätsabgabe
    saving_netz = cons_kwh * netzentgelt_reduction_ct / 100.0  # Ersparnis Netzentgelt

    # Einspeiseseite (Vergütung)
    market_einspeisetarif_ct = 4.5  # OeMAG Marktpreis-Einspeisung ca. 4-5 ct
    gen_income_eeg = gen_kwh * eeg_gen_price_ct / 100.0
    gen_income_market = gen_kwh * market_einspeisetarif_ct / 100.0
    saving_generation = gen_income_eeg - gen_income_market

    total_saving = saving_price + saving_elabg + saving_netz + saving_generation

    # Kosten in der EEG (was tatsächlich bezahlt wird)
    actual_cost = 0
    actual_income = 0
    for item in items:
        if item['type'] == 'consumption':
            actual_cost += item['amount_eur']
        else:
            actual_income += item['amount_eur']

    return {
        'cons_kwh': cons_kwh,
        'gen_kwh': gen_kwh,
        'eeg_type': eeg_type,
        'market_price_ct': market_price_ct,
        'eeg_price_ct': eeg_price_ct,
        'elabg_ct': elabg_ct,
        'netzentgelt_base_ct': netzentgelt_base_ct,
        'netzentgelt_reduction_pct': local_netz_reduction_pct,
        'netzentgelt_reduction_ct': netzentgelt_reduction_ct,
        'cost_market': round(cost_market, 2),
        'cost_eeg': round(cost_eeg, 2),
        'saving_price': round(saving_price, 2),
        'saving_elabg': round(saving_elabg, 2),
        'saving_netz': round(saving_netz, 2),
        'saving_generation': round(saving_generation, 2),
        'total_saving': round(total_saving, 2),
        'actual_cost': round(actual_cost, 2),
        'actual_income': round(actual_income, 2),
        'eeg_gen_price_ct': eeg_gen_price_ct,
        'market_einspeisetarif_ct': market_einspeisetarif_ct,
        'gen_income_eeg': round(gen_income_eeg, 2),
        'gen_income_market': round(gen_income_market, 2),
    }


def get_member_stats(db, member, period_from, period_to):
    """Berechnet detaillierte Statistiken für ein Mitglied (Seite 2)."""
    ts_from = period_from + "T00:00:00"
    ts_to = period_to + "T23:45:00"

    # Monatliche Daten Bezug (G.03)
    monthly_cons = db.execute("""
        SELECT strftime('%Y-%m', m.timestamp_start) as month,
               ROUND(SUM(m.value_kwh), 2) as kwh
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE mc.code = '1-1:2.9.0 G.03'
          AND m.metering_point_id = ?
          AND m.timestamp_start >= ? AND m.timestamp_start <= ?
        GROUP BY month ORDER BY month
    """, (member['bezug_zp'], ts_from, ts_to)).fetchall()

    # Monatliche Daten Einspeisung (G.01T - P.01T)
    monthly_gen_raw = {}
    if member['einspeiser_zp']:
        rows = db.execute("""
            SELECT strftime('%Y-%m', m.timestamp_start) as month,
                   mc.code,
                   ROUND(SUM(m.value_kwh), 2) as kwh
            FROM measurements m
            JOIN meter_codes mc ON mc.id = m.meter_code_id
            WHERE mc.code IN ('1-1:2.9.0 G.01T', '1-1:2.9.0 P.01T')
              AND m.metering_point_id = ?
              AND m.timestamp_start >= ? AND m.timestamp_start <= ?
            GROUP BY month, mc.code ORDER BY month
        """, (member['einspeiser_zp'], ts_from, ts_to)).fetchall()
        for r in rows:
            if r['month'] not in monthly_gen_raw:
                monthly_gen_raw[r['month']] = {'g01t': 0, 'p01t': 0}
            if 'G.01T' in r['code']:
                monthly_gen_raw[r['month']]['g01t'] = r['kwh']
            else:
                monthly_gen_raw[r['month']]['p01t'] = r['kwh']

    # Gesamter Netz-Bezug für Eigendeckungsgrad
    total_grid = db.execute("""
        SELECT ROUND(SUM(m.value_kwh), 2) as kwh
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE mc.code = '1-1:1.9.0 G.01'
          AND m.metering_point_id = ?
          AND m.timestamp_start >= ? AND m.timestamp_start <= ?
    """, (member['bezug_zp'], ts_from, ts_to)).fetchone()
    total_grid_kwh = total_grid['kwh'] or 0

    # Preise für Berechnung
    price_cons, price_gen = get_price_for_date(db, period_from)

    # Monats-Labels und Daten zusammenführen
    month_names = {'01': 'Jän', '02': 'Feb', '03': 'Mär', '04': 'Apr',
                   '05': 'Mai', '06': 'Jun', '07': 'Jul', '08': 'Aug',
                   '09': 'Sep', '10': 'Okt', '11': 'Nov', '12': 'Dez'}
    monthly_data = []
    total_cons = 0
    total_gen = 0
    for row in monthly_cons:
        month_key = row['month']
        cons = row['kwh'] or 0
        gen_data = monthly_gen_raw.get(month_key, {'g01t': 0, 'p01t': 0})
        gen = max(0, gen_data['g01t'] - gen_data['p01t'])
        net_eur = round(cons * price_cons / 100.0 - gen * price_gen / 100.0, 2)
        label = month_names.get(month_key[-2:], month_key[-2:]) + ' ' + month_key[:4]
        monthly_data.append({
            'month_key': month_key,
            'label': label,
            'consumption': cons,
            'generation': round(gen, 2),
            'net_eur': net_eur
        })
        total_cons += cons
        total_gen += gen

    # Auch Monate mit nur Einspeisung hinzufügen
    existing_months = {row['month'] for row in monthly_cons}
    for month_key in sorted(monthly_gen_raw.keys()):
        if month_key not in existing_months:
            gen_data = monthly_gen_raw[month_key]
            gen = max(0, gen_data['g01t'] - gen_data['p01t'])
            net_eur = round(-gen * price_gen / 100.0, 2)
            label = month_names.get(month_key[-2:], month_key[-2:]) + ' ' + month_key[:4]
            monthly_data.append({
                'month_key': month_key,
                'label': label,
                'consumption': 0,
                'generation': round(gen, 2),
                'net_eur': net_eur
            })
            total_gen += gen
    monthly_data.sort(key=lambda x: x['month_key'])

    # Eigendeckungsgrad: Anteil EEG am Gesamtbezug
    total_member_consumption = total_grid_kwh + total_cons
    self_sufficiency = (total_cons / total_member_consumption * 100) if total_member_consumption > 0 else 0

    monthly_max = max((d['consumption'] for d in monthly_data), default=0)
    monthly_gen_max = max((d['generation'] for d in monthly_data), default=0)

    return {
        'total_consumption_kwh': round(total_cons, 1),
        'total_generation_kwh': round(total_gen, 1),
        'co2_saved_kg': round(total_cons * 0.227, 1),  # 227g CO2/kWh Strommix AT
        'self_sufficiency_pct': round(self_sufficiency, 1),
        'monthly_data': monthly_data,
        'monthly_max': monthly_max,
        'monthly_gen_max': monthly_gen_max,
    }


def get_community_stats(db, invoice):
    """Berechnet EEG-Gesamtstatistiken für Transparenzseite (Seite 3)."""
    ts_from = invoice['period_from'] + "T00:00:00"
    ts_to = invoice['period_to'] + "T23:45:00"
    price_cons, price_gen = get_price_for_date(db, invoice['period_from'])

    # Gehandelte Energie (Summe aller G.03)
    total_traded = db.execute("""
        SELECT ROUND(SUM(m.value_kwh), 1) as kwh
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE mc.code = '1-1:2.9.0 G.03'
          AND m.timestamp_start >= ? AND m.timestamp_start <= ?
    """, (ts_from, ts_to)).fetchone()['kwh'] or 0

    # Erzeugung für Community
    total_generated = db.execute("""
        SELECT
            SUM(CASE WHEN mc.code='1-1:2.9.0 G.01T' THEN m.value_kwh ELSE 0 END) as g01t,
            SUM(CASE WHEN mc.code='1-1:2.9.0 P.01T' THEN m.value_kwh ELSE 0 END) as p01t
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        JOIN members mb ON mb.einspeiser_zp = m.metering_point_id
        WHERE mc.code IN ('1-1:2.9.0 G.01T', '1-1:2.9.0 P.01T')
          AND m.timestamp_start >= ? AND m.timestamp_start <= ?
    """, (ts_from, ts_to)).fetchone()
    total_gen_kwh = round(max(0, (total_generated['g01t'] or 0) - (total_generated['p01t'] or 0)), 1)

    # Mitglieder-Anzahl
    member_count = db.execute("SELECT COUNT(*) FROM members WHERE active=1").fetchone()[0]
    generator_count = db.execute("SELECT COUNT(*) FROM members WHERE active=1 AND einspeiser_zp IS NOT NULL AND einspeiser_zp != ''").fetchone()[0]

    # Pro-Mitglied Verbrauch
    member_cons = db.execute("""
        SELECT mb.name, ROUND(SUM(m.value_kwh), 1) as kwh
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        JOIN members mb ON mb.bezug_zp = m.metering_point_id
        WHERE mc.code = '1-1:2.9.0 G.03'
          AND m.timestamp_start >= ? AND m.timestamp_start <= ?
        GROUP BY mb.id ORDER BY kwh DESC
    """, (ts_from, ts_to)).fetchall()

    member_consumption = [{'name': r['name'], 'kwh': r['kwh']} for r in member_cons]
    max_cons = member_consumption[0]['kwh'] if member_consumption else 0

    # Pro-Mitglied Erzeugung
    member_gen = db.execute("""
        SELECT mb.name,
               ROUND(SUM(CASE WHEN mc.code='1-1:2.9.0 G.01T' THEN m.value_kwh ELSE 0 END) -
                     SUM(CASE WHEN mc.code='1-1:2.9.0 P.01T' THEN m.value_kwh ELSE 0 END), 1) as kwh
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        JOIN members mb ON mb.einspeiser_zp = m.metering_point_id
        WHERE mc.code IN ('1-1:2.9.0 G.01T', '1-1:2.9.0 P.01T')
          AND m.timestamp_start >= ? AND m.timestamp_start <= ?
        GROUP BY mb.id ORDER BY kwh DESC
    """, (ts_from, ts_to)).fetchall()
    member_generation = [{'name': r['name'], 'kwh': max(0, r['kwh'])} for r in member_gen]
    max_gen = member_generation[0]['kwh'] if member_generation else 0

    # Durchschnittlicher Eigendeckungsgrad
    avg_self_suff = 0
    if total_traded > 0:
        total_all_grid = db.execute("""
            SELECT ROUND(SUM(m.value_kwh), 1) as kwh
            FROM measurements m
            JOIN meter_codes mc ON mc.id = m.meter_code_id
            JOIN members mb ON mb.bezug_zp = m.metering_point_id
            WHERE mc.code = '1-1:1.9.0 G.01'
              AND m.timestamp_start >= ? AND m.timestamp_start <= ?
        """, (ts_from, ts_to)).fetchone()['kwh'] or 0
        total_all_consumption = total_all_grid + total_traded
        avg_self_suff = (total_traded / total_all_consumption * 100) if total_all_consumption > 0 else 0

    co2_total = total_traded * 0.227
    trees = int(co2_total / 12.5)  # ~12.5 kg CO2 pro Baum/Jahr

    return {
        'member_count': member_count,
        'generator_count': generator_count,
        'total_traded_kwh': total_traded,
        'total_generated_kwh': total_gen_kwh,
        'avg_self_sufficiency': round(avg_self_suff, 1),
        'total_co2_saved_kg': round(co2_total, 0),
        'trees_equivalent': trees,
        'member_consumption': member_consumption,
        'max_consumption': max_cons,
        'member_generation': member_generation,
        'max_generation': max_gen,
        'price_cons': price_cons,
        'price_gen': price_gen,
    }


@app.route('/invoices/<int:id>/send', methods=['POST'])
@admin_required
def invoice_send(id):
    """E-Mails an alle Mitglieder versenden."""
    db = get_db()
    invoice = db.execute("SELECT * FROM invoices WHERE id=?", (id,)).fetchone()
    blocker = invoice_finalization_blocker(db, invoice)
    if blocker:
        flash(blocker, 'danger')
        audit_log('invoice_send_blocked', f'Rechnung {id}: Sammelversand blockiert ({blocker})')
        return redirect(url_for('invoice_detail', id=id))
    items = invoice_recipient_rows(db, id)

    sent = 0
    failed = 0
    for item in items:
        if not item['email']:
            db.execute("""INSERT INTO email_log (invoice_id, member_id, recipient_email, subject, status, error_message)
                          VALUES (?, ?, ?, ?, 'failed', 'Keine E-Mail-Adresse')""",
                       (id, item['member_id'], '-',
                        f"EEG Abrechnung {invoice['period_from']} - {invoice['period_to']}"))
            failed += 1
            continue

        try:
            send_invoice_email(db, invoice, item)
            db.execute("""INSERT INTO email_log (invoice_id, member_id, recipient_email, subject, status)
                          VALUES (?, ?, ?, ?, 'sent')""",
                       (id, item['member_id'], item['email'],
                        f"EEG Abrechnung {invoice['period_from']} - {invoice['period_to']}"))
            sent += 1
        except Exception as e:
            db.execute("""INSERT INTO email_log (invoice_id, member_id, recipient_email, subject, status, error_message)
                          VALUES (?, ?, ?, ?, 'failed', ?)""",
                       (id, item['member_id'], item['email'],
                        f"EEG Abrechnung {invoice['period_from']} - {invoice['period_to']}",
                        str(e)))
            failed += 1

    db.execute("UPDATE invoices SET status='sent', finalized_at=datetime('now') WHERE id=?", (id,))
    db.commit()
    audit_log('invoice_send_all', f'Rechnung {id}: {sent} E-Mails gesendet, {failed} fehlgeschlagen')
    flash(f'{sent} E-Mails gesendet, {failed} fehlgeschlagen.', 'success' if failed == 0 else 'warning')
    return redirect(url_for('invoice_detail', id=id))


@app.route('/invoices/<int:id>/send/<int:member_id>', methods=['POST'])
@admin_required
def invoice_send_single(id, member_id):
    """E-Mail an ein einzelnes Mitglied senden."""
    db = get_db()
    invoice = db.execute("SELECT * FROM invoices WHERE id=?", (id,)).fetchone()
    member = db.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
    blocker = invoice_finalization_blocker(db, invoice)
    if blocker:
        flash(blocker, 'danger')
        audit_log('invoice_send_blocked', f'Rechnung {id}: Einzelversand blockiert ({blocker})')
        return redirect(url_for('invoice_detail', id=id))

    if not member['email']:
        db.execute("""INSERT INTO email_log (invoice_id, member_id, recipient_email, subject, status, error_message)
                      VALUES (?, ?, ?, ?, 'failed', 'Keine E-Mail-Adresse hinterlegt')""",
                   (id, member_id, '-',
                    f"EEG Abrechnung {invoice['period_from']} - {invoice['period_to']}"))
        db.commit()
        flash(f'Keine E-Mail-Adresse für {member["name"]} hinterlegt.', 'danger')
        return redirect(url_for('invoice_detail', id=id))

    try:
        member_row = {'member_id': member_id, 'name': member['name'], 'email': member['email']}
        send_invoice_email(db, invoice, member_row)
        db.execute("""INSERT INTO email_log (invoice_id, member_id, recipient_email, subject, status)
                      VALUES (?, ?, ?, ?, 'sent')""",
                   (id, member_id, member['email'],
                    f"EEG Abrechnung {invoice['period_from']} - {invoice['period_to']}"))
        # Prüfen ob jetzt alle Mitglieder eine E-Mail erhalten haben → Status auf 'sent' setzen
        total_members = db.execute("""
            SELECT COUNT(*) FROM (
                SELECT m.id
                FROM members m
                WHERE m.id IN (
                    SELECT member_id FROM invoice_items WHERE invoice_id=?
                    UNION
                    SELECT member_id FROM invoice_carryovers WHERE invoice_id=?
                )
                AND m.email IS NOT NULL AND m.email != ''
            )
        """, (id, id)).fetchone()[0]
        sent_members = db.execute("""
            SELECT COUNT(DISTINCT member_id) FROM email_log
            WHERE invoice_id=? AND status='sent'
        """, (id,)).fetchone()[0]
        if sent_members >= total_members and total_members > 0:
            db.execute("UPDATE invoices SET status='sent', finalized_at=datetime('now') WHERE id=?", (id,))
        db.commit()
        audit_log('invoice_send', f'E-Mail gesendet: Rechnung {id} an {member["name"]} ({member["email"]})')
        flash(f'E-Mail an {member["name"]} ({member["email"]}) gesendet.', 'success')
    except Exception as e:
        db.execute("""INSERT INTO email_log (invoice_id, member_id, recipient_email, subject, status, error_message)
                      VALUES (?, ?, ?, ?, 'failed', ?)""",
                   (id, member_id, member['email'],
                    f"EEG Abrechnung {invoice['period_from']} - {invoice['period_to']}",
                    str(e)))
        db.commit()
        flash_exception(e, f'E-Mail an {member["name"]} konnte nicht gesendet werden.')

    return redirect(url_for('invoice_detail', id=id))


@app.route('/invoices/<int:id>/finalize', methods=['POST'])
@admin_required
def invoice_finalize(id):
    """Abrechnung manuell auf 'sent' setzen (z.B. wenn Versand ohne System erfolgte)."""
    db = get_db()
    invoice = db.execute("SELECT * FROM invoices WHERE id=?", (id,)).fetchone()
    blocker = invoice_finalization_blocker(db, invoice)
    if blocker:
        flash(blocker, 'danger')
        audit_log('invoice_finalize_blocked', f'Abrechnung #{id}: Abschluss blockiert ({blocker})')
        return redirect(url_for('invoice_detail', id=id))
    db.execute("UPDATE invoices SET status='sent', finalized_at=datetime('now') WHERE id=?", (id,))
    db.commit()
    audit_log('invoice_finalize', f'Abrechnung #{id} manuell finalisiert')
    flash(f'Abrechnung #{id} wurde als finalisiert markiert.', 'success')
    return redirect(url_for('invoice_detail', id=id))


def send_invoice_email(db, invoice, member_row):
    """Sendet eine Abrechnungs-E-Mail. Konfiguration aus DB-Settings."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication

    mail_cfg = _get_valid_mail_config(db)

    # PDF generieren
    member = db.execute("SELECT * FROM members WHERE id=?", (member_row['member_id'],)).fetchone()
    items = db.execute("SELECT * FROM invoice_items WHERE invoice_id=? AND member_id=?",
                       (invoice['id'], member_row['member_id'])).fetchall()
    carryovers = get_invoice_carryovers(db, invoice['id'], member_row['member_id'])
    carryover_total = round(sum(row['amount_eur'] for row in carryovers), 2)

    energy_net_total = 0
    for it in items:
        if it['type'] == 'consumption':
            energy_net_total += it['amount_eur']
        else:
            energy_net_total -= it['amount_eur']
    net_total = round(energy_net_total + carryover_total, 2)

    qr_data_uri = ''
    if net_total > 0:
        qr_data_uri = generate_epc_qr(net_total, invoice, member)

    member_stats = get_member_stats(db, member, invoice['period_from'], invoice['period_to'])
    community_stats = get_community_stats(db, invoice)
    savings = calculate_member_savings(member_stats, items)
    public_cfg = get_public_config(db)

    import base64 as b64mod
    logo_path = os.path.join(BASE_DIR, 'static', 'logo_small.png')
    with open(logo_path, 'rb') as f:
        logo_b64 = 'data:image/png;base64,' + b64mod.b64encode(f.read()).decode('ascii')

    from weasyprint import HTML
    html_content = render_template('invoice_pdf.html', invoice=invoice, member=member, items=items,
                                   carryovers=carryovers,
                                   member_stats=member_stats, community_stats=community_stats,
                                   energy_net_total=round(energy_net_total, 2),
                                   carryover_total=carryover_total,
                                   net_total=net_total, qr_data_uri=qr_data_uri,
                                   savings=savings, logo_b64=logo_b64, public_cfg=public_cfg)
    pdf_bytes = HTML(string=html_content, base_url=BASE_DIR).write_pdf()

    # E-Mail zusammenbauen
    # Templates aus DB laden
    tpl_rows = db.execute("SELECT key, value FROM settings WHERE key IN ('email_subject', 'email_body')").fetchall()
    tpl = {r['key']: r['value'] for r in tpl_rows}
    replacements = {
        'name': member_row['name'],
        'zeitraum_von': invoice['period_from'],
        'zeitraum_bis': invoice['period_to'],
    }
    subject = tpl.get('email_subject', 'EEG Abrechnung {zeitraum_von} - {zeitraum_bis}').format(**replacements)
    body_text = tpl.get('email_body', 'Hallo {name},\n\nanbei Ihre Abrechnung.').format(**replacements)

    # HTML-Version der E-Mail mit Logo
    import base64 as b64mod2
    logo_email_path = os.path.join(BASE_DIR, 'static', 'logo_small.png')
    with open(logo_email_path, 'rb') as f:
        logo_email_b64 = b64mod2.b64encode(f.read()).decode('ascii')

    body_html = f"""<html><body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6;">
<div style="max-width: 600px; margin: 0 auto;">
    <div style="text-align: center; padding: 15px 0; border-bottom: 2px solid #2b5e3a;">
        <img src="data:image/png;base64,{logo_email_b64}" width="50" height="50" style="border-radius: 8px;">
        <h2 style="color: #2b5e3a; margin: 8px 0 0 0; font-size: 18px;">{public_cfg['org_name']}</h2>
    </div>
    <div style="padding: 20px 0;">
        {''.join(f'<p style="margin: 8px 0;">{line}</p>' if line.strip() else '<br>' for line in body_text.split(chr(10)))}
    </div>
    <div style="border-top: 1px solid #ccc; padding-top: 12px; font-size: 11px; color: #777; text-align: center;">
        <strong>{public_cfg['org_name']}</strong><br>
        {public_cfg['org_address']}<br>
        {public_cfg['org_email']}
    </div>
</div>
</body></html>"""

    msg = MIMEMultipart('mixed')
    msg['From'] = mail_cfg['from_header']
    msg['Reply-To'] = mail_cfg['reply_to_header']
    msg['To'] = member_row['email']
    msg['Subject'] = subject

    # Text + HTML Alternative
    msg_alt = MIMEMultipart('alternative')
    msg_alt.attach(MIMEText(body_text, 'plain', 'utf-8'))
    msg_alt.attach(MIMEText(body_html, 'html', 'utf-8'))
    msg.attach(msg_alt)

    pdf_attachment = MIMEApplication(pdf_bytes, _subtype='pdf')
    pdf_attachment.add_header('Content-Disposition', 'attachment',
                             filename=f"EEG_Abrechnung_{invoice['period_from']}_{invoice['period_to']}.pdf")
    msg.attach(pdf_attachment)

    _log_mail_send(mail_cfg, member_row['email'], subject)
    with smtplib.SMTP(mail_cfg['smtp_host'], mail_cfg['smtp_port']) as server:
        if mail_cfg['smtp_tls']:
            server.starttls()
        server.login(mail_cfg['smtp_user'], mail_cfg['smtp_pass'])
        server.send_message(msg, from_addr=mail_cfg['from_address'], to_addrs=[member_row['email']])


# === Reports ===

REPORT_AGGREGATIONS = {
    'day': {'label': 'Tag', 'strftime': '%Y-%m-%d'},
    'month': {'label': 'Monat', 'strftime': '%Y-%m'},
    'year': {'label': 'Jahr', 'strftime': '%Y'},
}


def _report_float(value):
    return float(value or 0)


def _report_round(value, digits=2):
    return round(_report_float(value), digits)


def _report_pct(part, total):
    total = _report_float(total)
    if total <= 0:
        return 0.0
    return round(_report_float(part) / total * 100.0, 1)


def _report_ts_bounds(db):
    row = db.execute("SELECT MIN(timestamp_start) AS min_ts, MAX(timestamp_start) AS max_ts FROM measurements").fetchone()
    today = local_now().date()
    min_date = (row['min_ts'][:10] if row and row['min_ts'] else today.replace(month=1, day=1).isoformat())
    max_date = (row['max_ts'][:10] if row and row['max_ts'] else today.isoformat())
    return min_date, max_date


def _parse_report_date(value, fallback):
    try:
        return date.fromisoformat((value or '').strip()).isoformat()
    except ValueError:
        return fallback


def _month_iter(start_date, end_date):
    cur = date.fromisoformat(start_date).replace(day=1)
    end = date.fromisoformat(end_date).replace(day=1)
    while cur <= end:
        yield cur
        if cur.month == 12:
            cur = date(cur.year + 1, 1, 1)
        else:
            cur = date(cur.year, cur.month + 1, 1)


def _month_end(day):
    from calendar import monthrange
    return date(day.year, day.month, monthrange(day.year, day.month)[1])


def _dt_ms(value):
    try:
        return int(datetime.fromisoformat(value).timestamp() * 1000)
    except ValueError:
        return 0


def _build_member_report(db, member, period_from, period_to, aggregation):
    ts_from = period_from + 'T00:00:00'
    ts_to = period_to + 'T23:45:00'
    agg = REPORT_AGGREGATIONS.get(aggregation, REPORT_AGGREGATIONS['month'])
    bucket_expr = f"strftime('{agg['strftime']}', m.timestamp_start)"
    price_cons, price_gen = get_price_for_date(db, period_from)
    market_price_ct = 25.0
    market_feed_ct = 4.5

    params = {
        'bezug_zp': member['bezug_zp'] or '',
        'einspeiser_zp': member['einspeiser_zp'] or '',
        'ts_from': ts_from,
        'ts_to': ts_to,
    }
    series_rows = db.execute(f"""
        SELECT {bucket_expr} AS bucket,
               ROUND(SUM(CASE WHEN m.metering_point_id=:bezug_zp AND mc.code='1-1:2.9.0 G.03' THEN m.value_kwh ELSE 0 END), 3) AS eeg,
               ROUND(SUM(CASE WHEN m.metering_point_id=:bezug_zp AND mc.code='1-1:1.9.0 G.01' THEN m.value_kwh ELSE 0 END), 3) AS grid,
               ROUND(SUM(CASE WHEN m.metering_point_id=:einspeiser_zp AND mc.code='1-1:2.9.0 G.01T' THEN m.value_kwh ELSE 0 END), 3) AS generation,
               ROUND(SUM(CASE WHEN m.metering_point_id=:einspeiser_zp AND mc.code='1-1:2.9.0 P.01T' THEN m.value_kwh ELSE 0 END), 3) AS public_feed
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE m.timestamp_start BETWEEN :ts_from AND :ts_to
          AND (
            (m.metering_point_id=:bezug_zp AND mc.code IN ('1-1:2.9.0 G.03', '1-1:1.9.0 G.01'))
            OR
            (m.metering_point_id=:einspeiser_zp AND mc.code IN ('1-1:2.9.0 G.01T', '1-1:2.9.0 P.01T'))
          )
        GROUP BY bucket
        ORDER BY bucket
    """, params).fetchall()

    labels = []
    eeg_values = []
    grid_values = []
    generation_values = []
    public_feed_values = []
    eeg_feed_values = []
    total_consumption_values = []
    eeg_share_values = []
    cost_without_values = []
    cost_actual_values = []
    savings_values = []
    cumulative_savings = []
    running_savings = 0.0
    for row in series_rows:
        eeg = _report_float(row['eeg'])
        grid = _report_float(row['grid'])
        generation = _report_float(row['generation'])
        public_feed = _report_float(row['public_feed'])
        eeg_feed = max(0.0, generation - public_feed)
        total_consumption = eeg + grid
        cost_without = total_consumption * market_price_ct / 100.0 - generation * market_feed_ct / 100.0
        cost_actual = grid * market_price_ct / 100.0 + eeg * price_cons / 100.0 - eeg_feed * price_gen / 100.0 - public_feed * market_feed_ct / 100.0
        savings = cost_without - cost_actual
        running_savings += savings

        labels.append(row['bucket'])
        eeg_values.append(round(eeg, 3))
        grid_values.append(round(grid, 3))
        generation_values.append(round(generation, 3))
        public_feed_values.append(round(public_feed, 3))
        eeg_feed_values.append(round(eeg_feed, 3))
        total_consumption_values.append(round(total_consumption, 3))
        eeg_share_values.append(_report_pct(eeg, total_consumption))
        cost_without_values.append(round(cost_without, 2))
        cost_actual_values.append(round(cost_actual, 2))
        savings_values.append(round(savings, 2))
        cumulative_savings.append(round(running_savings, 2))

    totals = {
        'eeg': round(sum(eeg_values), 2),
        'grid': round(sum(grid_values), 2),
        'consumption': round(sum(total_consumption_values), 2),
        'generation': round(sum(generation_values), 2),
        'public_feed': round(sum(public_feed_values), 2),
        'eeg_feed': round(sum(eeg_feed_values), 2),
        'cost_without': round(sum(cost_without_values), 2),
        'cost_actual': round(sum(cost_actual_values), 2),
        'savings': round(sum(savings_values), 2),
    }
    totals['eeg_share'] = _report_pct(totals['eeg'], totals['consumption'])
    totals['autarky'] = totals['eeg_share']
    totals['avg_savings_per_kwh'] = round(totals['savings'] / totals['consumption'], 4) if totals['consumption'] > 0 else 0
    totals['self_consumption_quote'] = None

    daily_rows = db.execute("""
        SELECT substr(m.timestamp_start, 1, 10) AS day,
               ROUND(SUM(CASE WHEN m.metering_point_id=:bezug_zp AND mc.code='1-1:2.9.0 G.03' THEN m.value_kwh ELSE 0 END), 3) AS eeg,
               ROUND(SUM(CASE WHEN m.metering_point_id=:bezug_zp AND mc.code='1-1:1.9.0 G.01' THEN m.value_kwh ELSE 0 END), 3) AS grid
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE m.timestamp_start BETWEEN :ts_from AND :ts_to
          AND m.metering_point_id=:bezug_zp
          AND mc.code IN ('1-1:2.9.0 G.03', '1-1:1.9.0 G.01')
        GROUP BY day
        ORDER BY day
    """, params).fetchall()
    daily_shares = []
    for row in daily_rows:
        day_total = _report_float(row['eeg']) + _report_float(row['grid'])
        if day_total > 0:
            daily_shares.append({
                'day': row['day'],
                'eeg_share': _report_pct(row['eeg'], day_total),
                'eeg': _report_round(row['eeg'], 2),
                'grid': _report_round(row['grid'], 2),
            })
    best_day = max(daily_shares, key=lambda item: item['eeg_share'], default=None)
    weakest_day = min(daily_shares, key=lambda item: item['eeg_share'], default=None)

    heat_rows = db.execute("""
        SELECT CAST(strftime('%w', m.timestamp_start) AS INTEGER) AS weekday,
               CAST(strftime('%H', m.timestamp_start) AS INTEGER) AS hour,
               ROUND(SUM(m.value_kwh), 3) AS kwh
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE m.timestamp_start BETWEEN :ts_from AND :ts_to
          AND m.metering_point_id=:bezug_zp
          AND mc.code IN ('1-1:2.9.0 G.03', '1-1:1.9.0 G.01')
        GROUP BY weekday, hour
        ORDER BY weekday, hour
    """, params).fetchall()
    heatmap = []
    for row in heat_rows:
        weekday = (int(row['weekday']) + 6) % 7
        heatmap.append([int(row['hour']), weekday, _report_round(row['kwh'], 3)])

    hourly_rows = db.execute("""
        SELECT hour,
               ROUND(AVG(hour_sum), 3) AS kwh
        FROM (
            SELECT substr(m.timestamp_start, 1, 13) AS hour_bucket,
                   CAST(strftime('%H', m.timestamp_start) AS INTEGER) AS hour,
                   SUM(m.value_kwh) AS hour_sum
            FROM measurements m
            JOIN meter_codes mc ON mc.id = m.meter_code_id
            WHERE m.timestamp_start BETWEEN :ts_from AND :ts_to
              AND m.metering_point_id=:bezug_zp
              AND mc.code IN ('1-1:2.9.0 G.03', '1-1:1.9.0 G.01')
            GROUP BY hour_bucket
        ) hourly
        GROUP BY hour
        ORDER BY hour
    """, params).fetchall()
    typical_day = [0.0] * 24
    for row in hourly_rows:
        typical_day[int(row['hour'])] = _report_round(row['kwh'], 3)

    peak_rows = db.execute("""
        SELECT m.timestamp_start AS ts, ROUND(SUM(m.value_kwh), 3) AS kwh
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE m.timestamp_start BETWEEN :ts_from AND :ts_to
          AND m.metering_point_id=:bezug_zp
          AND mc.code IN ('1-1:2.9.0 G.03', '1-1:1.9.0 G.01')
        GROUP BY m.timestamp_start
        ORDER BY kwh DESC
        LIMIT 8
    """, params).fetchall()
    peaks = [{'ts': row['ts'], 'kwh': _report_round(row['kwh'], 3)} for row in peak_rows]

    peak_line_rows = db.execute("""
        SELECT substr(m.timestamp_start, 1, 13) || ':00:00' AS ts, ROUND(SUM(m.value_kwh), 3) AS kwh
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE m.timestamp_start BETWEEN :ts_from AND :ts_to
          AND m.metering_point_id=:bezug_zp
          AND mc.code IN ('1-1:2.9.0 G.03', '1-1:1.9.0 G.01')
        GROUP BY substr(m.timestamp_start, 1, 13)
        ORDER BY ts
    """, params).fetchall()
    peak_line = [[_dt_ms(row['ts']), _report_round(row['kwh'], 3)] for row in peak_line_rows]
    peak_markers = [[_dt_ms(row['ts']), _report_round(row['kwh'], 3)] for row in peaks]

    quality_rows = db.execute("""
        SELECT COALESCE(NULLIF(TRIM(m.quality), ''), 'unbekannt') AS quality, COUNT(*) AS cnt
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE m.timestamp_start BETWEEN :ts_from AND :ts_to
          AND (
            (m.metering_point_id=:bezug_zp AND mc.code IN ('1-1:2.9.0 G.03', '1-1:1.9.0 G.01'))
            OR
            (m.metering_point_id=:einspeiser_zp AND mc.code IN ('1-1:2.9.0 G.01T', '1-1:2.9.0 P.01T'))
          )
        GROUP BY quality
        ORDER BY cnt DESC
    """, params).fetchall()
    quality_total = sum(row['cnt'] for row in quality_rows) or 1
    quality = [{
        'quality': row['quality'],
        'cnt': row['cnt'],
        'pct': round(row['cnt'] / quality_total * 100.0, 1),
    } for row in quality_rows]

    relevant_code_count = 2 + (2 if member['einspeiser_zp'] else 0)
    actual_by_month = {
        row['bucket']: row['cnt']
        for row in db.execute("""
            SELECT strftime('%Y-%m', m.timestamp_start) AS bucket, COUNT(*) AS cnt
            FROM measurements m
            JOIN meter_codes mc ON mc.id = m.meter_code_id
            WHERE m.timestamp_start BETWEEN :ts_from AND :ts_to
              AND (
                (m.metering_point_id=:bezug_zp AND mc.code IN ('1-1:2.9.0 G.03', '1-1:1.9.0 G.01'))
                OR
                (m.metering_point_id=:einspeiser_zp AND mc.code IN ('1-1:2.9.0 G.01T', '1-1:2.9.0 P.01T'))
              )
            GROUP BY bucket
        """, params).fetchall()
    }
    missing_by_month = []
    for month_start in _month_iter(period_from, period_to):
        start = max(month_start, date.fromisoformat(period_from))
        end = min(_month_end(month_start), date.fromisoformat(period_to))
        days = (end - start).days + 1
        expected = max(0, days * 96 * relevant_code_count)
        actual = actual_by_month.get(month_start.strftime('%Y-%m'), 0)
        missing_by_month.append({
            'month': month_start.strftime('%Y-%m'),
            'missing': max(0, expected - actual),
            'expected': expected,
            'actual': actual,
            'completeness': round(actual / expected * 100.0, 1) if expected else 0,
        })
    completeness = round(
        sum(item['actual'] for item in missing_by_month) / max(1, sum(item['expected'] for item in missing_by_month)) * 100.0,
        1
    )

    zeros = db.execute("""
        SELECT COUNT(*)
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE m.timestamp_start BETWEEN :ts_from AND :ts_to
          AND m.value_kwh = 0
          AND (
            (m.metering_point_id=:bezug_zp AND mc.code IN ('1-1:2.9.0 G.03', '1-1:1.9.0 G.01'))
            OR
            (m.metering_point_id=:einspeiser_zp AND mc.code IN ('1-1:2.9.0 G.01T', '1-1:2.9.0 P.01T'))
          )
    """, params).fetchone()[0] or 0
    outliers = db.execute("""
        SELECT COUNT(*)
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE m.timestamp_start BETWEEN :ts_from AND :ts_to
          AND (m.value_kwh < 0 OR m.value_kwh > 100)
          AND (
            (m.metering_point_id=:bezug_zp AND mc.code IN ('1-1:2.9.0 G.03', '1-1:1.9.0 G.01'))
            OR
            (m.metering_point_id=:einspeiser_zp AND mc.code IN ('1-1:2.9.0 G.01T', '1-1:2.9.0 P.01T'))
          )
    """, params).fetchone()[0] or 0

    community_rows = db.execute("""
        SELECT mb.id,
               ROUND(SUM(CASE WHEN m.metering_point_id=mb.bezug_zp AND mc.code='1-1:2.9.0 G.03' THEN m.value_kwh ELSE 0 END), 3) AS eeg,
               ROUND(SUM(CASE WHEN m.metering_point_id=mb.bezug_zp AND mc.code='1-1:1.9.0 G.01' THEN m.value_kwh ELSE 0 END), 3) AS grid,
               ROUND(SUM(CASE WHEN m.metering_point_id=mb.einspeiser_zp AND mc.code='1-1:2.9.0 G.01T' THEN m.value_kwh ELSE 0 END), 3) AS generation,
               ROUND(SUM(CASE WHEN m.metering_point_id=mb.einspeiser_zp AND mc.code='1-1:2.9.0 P.01T' THEN m.value_kwh ELSE 0 END), 3) AS public_feed
        FROM members mb
        LEFT JOIN measurements m
          ON m.timestamp_start BETWEEN :ts_from AND :ts_to
         AND (m.metering_point_id=mb.bezug_zp OR m.metering_point_id=mb.einspeiser_zp)
        LEFT JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE mb.active=1
        GROUP BY mb.id
    """, params).fetchall()
    community_shares = []
    community_generation = 0.0
    community_eeg_feed = 0.0
    for row in community_rows:
        consumption = _report_float(row['eeg']) + _report_float(row['grid'])
        if consumption > 0:
            community_shares.append(_report_float(row['eeg']) / consumption * 100.0)
        generation = _report_float(row['generation'])
        public_feed = _report_float(row['public_feed'])
        community_generation += generation
        community_eeg_feed += max(0.0, generation - public_feed)
    community_avg_eeg_share = round(sum(community_shares) / len(community_shares), 1) if community_shares else 0
    member_generation_share = _report_pct(totals['eeg_feed'], community_eeg_feed)

    evening_grid = db.execute("""
        SELECT ROUND(SUM(m.value_kwh), 3)
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE m.timestamp_start BETWEEN :ts_from AND :ts_to
          AND m.metering_point_id=:bezug_zp
          AND mc.code='1-1:1.9.0 G.01'
          AND CAST(strftime('%H', m.timestamp_start) AS INTEGER) BETWEEN 17 AND 21
    """, params).fetchone()[0] or 0
    noon_eeg = db.execute("""
        SELECT ROUND(SUM(m.value_kwh), 3)
        FROM measurements m
        JOIN meter_codes mc ON mc.id = m.meter_code_id
        WHERE m.timestamp_start BETWEEN :ts_from AND :ts_to
          AND m.metering_point_id=:bezug_zp
          AND mc.code='1-1:2.9.0 G.03'
          AND CAST(strftime('%H', m.timestamp_start) AS INTEGER) BETWEEN 10 AND 15
    """, params).fetchone()[0] or 0
    optimisation_hints = []
    if evening_grid > max(1.0, totals['grid'] * 0.25):
        optimisation_hints.append('Abends wird viel Strom vom öffentlichen Netz benötigt. Geräte wie Waschmaschine, Geschirrspüler oder Boiler könnten wenn möglich früher laufen.')
    if noon_eeg < max(1.0, totals['eeg'] * 0.25):
        optimisation_hints.append('Rund um die Mittagszeit wird noch wenig Strom aus der Energiegemeinschaft genutzt. Verbrauch in diese Zeit zu verschieben kann helfen.')
    if totals['eeg_share'] < community_avg_eeg_share:
        optimisation_hints.append('Der Anteil an Strom aus der Energiegemeinschaft liegt unter dem Durchschnitt der Gemeinschaft.')
    if not optimisation_hints:
        optimisation_hints.append('Im gewählten Zeitraum gibt es keine auffälligen Hinweise. Die Nutzung wirkt bereits ausgewogen.')

    charts = {
        'labels': labels,
        'eeg': eeg_values,
        'grid': grid_values,
        'generation': generation_values,
        'public_feed': public_feed_values,
        'eeg_feed': eeg_feed_values,
        'consumption': total_consumption_values,
        'eeg_share': eeg_share_values,
        'cost_without': cost_without_values,
        'cost_actual': cost_actual_values,
        'savings': savings_values,
        'cumulative_savings': cumulative_savings,
        'daily_eeg_share': [[item['day'], item['eeg_share']] for item in daily_shares],
        'heatmap': heatmap,
        'typical_day': typical_day,
        'peak_line': peak_line,
        'peak_markers': peak_markers,
        'missing_by_month': missing_by_month,
        'quality': quality,
        'sankey': [
            ['Öffentliches Netz', 'Verbrauch', _report_round(totals['grid'], 2)],
            ['EEG', 'Verbrauch', _report_round(totals['eeg'], 2)],
            ['Erzeugung', 'EEG', _report_round(totals['eeg_feed'], 2)],
            ['Erzeugung', 'Öffentliches Netz', _report_round(totals['public_feed'], 2)],
        ],
    }

    return {
        'totals': totals,
        'charts': charts,
        'best_day': best_day,
        'weakest_day': weakest_day,
        'peaks': peaks,
        'quality_summary': {
            'completeness': completeness,
            'zeros': zeros,
            'outliers': outliers,
            'expected_code_count': relevant_code_count,
        },
        'community': {
            'avg_eeg_share': community_avg_eeg_share,
            'member_generation_share': member_generation_share,
            'eeg_feed_total': round(community_eeg_feed, 2),
            'generation_total': round(community_generation, 2),
        },
        'prices': {
            'eeg_consumption_ct': price_cons,
            'eeg_generation_ct': price_gen,
            'market_consumption_ct': market_price_ct,
            'market_feed_ct': market_feed_ct,
            'is_estimate': True,
        },
        'optimisation_hints': optimisation_hints,
        'data_notes': [
            'Strom aus der Energiegemeinschaft ist jener Anteil, den Sie lokal von der EEG beziehen.',
            'Strom aus dem öffentlichen Netz ist jener Anteil, der nicht durch die EEG gedeckt wurde.',
            'Bei Erzeugern wird angezeigt, wie viel Energie an die EEG geliefert und wie viel ins öffentliche Netz abgegeben wurde.',
            'Direkt im Haus verbrauchter PV-Strom ist mit den vorhandenen Netzbetreiber-Daten nicht exakt getrennt sichtbar.',
            'Die Ersparnis ist eine Schätzung. Für eine exakte Berechnung braucht es einen gepflegten Vergleichstarif je Zeitraum.',
        ],
    }


def _reports_response(portal=False):
    db = get_db()
    min_date, max_date = _report_ts_bounds(db)
    aggregation = request.args.get('aggregation', 'month')
    if aggregation not in REPORT_AGGREGATIONS:
        aggregation = 'month'
    period_from = _parse_report_date(request.args.get('date_from'), min_date)
    period_to = _parse_report_date(request.args.get('date_to'), max_date)
    if period_from > period_to:
        period_from, period_to = period_to, period_from

    members = db.execute("""
        SELECT id, name, bezug_zp, einspeiser_zp
        FROM members
        WHERE active=1
        ORDER BY name
    """).fetchall()
    if portal:
        member_id = current_user.member_id
        if not member_id:
            flash('Kein Mitglied zugeordnet.', 'warning')
            return redirect(url_for('portal_dashboard'))
    else:
        member_id = request.args.get('member_id', type=int)
        if not member_id and members:
            member_id = members[0]['id']
    member = db.execute("SELECT * FROM members WHERE id=? AND active=1", (member_id,)).fetchone() if member_id else None
    if not member:
        flash('Kein aktives Mitglied für den Report gefunden.', 'warning')
        return render_template('reports.html', members=members, selected_member=None, report=None,
                               period_from=period_from, period_to=period_to,
                               aggregation=aggregation, aggregations=REPORT_AGGREGATIONS, portal=portal,
                               min_date=min_date, max_date=max_date)

    if portal and current_user.member_id != member['id']:
        abort(403)

    report = _build_member_report(db, member, period_from, period_to, aggregation)
    return render_template(
        'reports.html',
        members=members,
        selected_member=member,
        report=report,
        period_from=period_from,
        period_to=period_to,
        aggregation=aggregation,
        aggregations=REPORT_AGGREGATIONS,
        portal=portal,
        min_date=min_date,
        max_date=max_date,
    )


@app.route('/reports')
@admin_required
def reports():
    return _reports_response(portal=False)


@app.route('/portal/reports')
@login_required
def portal_reports():
    return _reports_response(portal=True)


# === Billing Calculation ===

def calculate_billing(db, period_from, period_to, price_cons, price_gen):
    """Kompatible Fassade für den ausgelagerten Abrechnungsservice."""
    return _calculate_billing(
        db, period_from, period_to, price_cons, price_gen,
        carryover_provider=calculate_carryovers_for_period,
    )


# === Settings ===

@app.route('/release-notes')
@login_required
def release_notes():
    """Release Notes: Liste der letzten Änderungen nach Datum."""
    return render_template('release_notes.html', release_notes=RELEASE_NOTES)


@app.route('/settings', methods=['GET', 'POST'])
@admin_required
def settings():
    db = get_db()
    if request.method == 'POST':
        existing_settings = {
            r['key']: r['value']
            for r in db.execute("SELECT key, value FROM settings").fetchall()
        }
        for key in (
            'smtp_host', 'smtp_port', 'smtp_user', 'smtp_pass', 'smtp_from', 'smtp_tls',
            'mail_from_address', 'mail_from_name', 'mail_reply_to', 'mail_reply_to_name',
            'email_subject', 'email_body',
            'org_name', 'org_email', 'org_website', 'org_address', 'org_legal',
            'payment_recipient', 'payment_iban', 'payment_bic'
        ):
            val = request.form.get(key, '')
            if key == 'smtp_pass' and not val:
                val = existing_settings.get('smtp_pass', '')
            db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, val))
        db.commit()
        audit_log('settings_update', 'SMTP-Einstellungen geändert')
        flash('E-Mail-Einstellungen gespeichert.', 'success')
        return redirect(url_for('settings'))

    # Settings aus DB laden
    rows = db.execute("SELECT key, value FROM settings").fetchall()
    smtp = {r['key']: r['value'] for r in rows}
    smtp_configured, _ = _validate_mail_config(_load_mail_config(db))
    return render_template('settings.html', smtp=smtp, smtp_configured=smtp_configured, db_path=DB_PATH)


# === Backup / Restore ===

def _setting_bool(value):
    return str(value or '').strip().lower() in ('1', 'true', 'yes', 'on')


def _setting_int(value, default, min_value=0, max_value=999):
    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError):
        parsed = default
    return max(min_value, min(parsed, max_value))


def _valid_time_or_default(value, default):
    text = str(value or '').strip()
    match = re.match(r'^([01]\d|2[0-3]):([0-5]\d)$', text)
    return text if match else default


def _backup_week_marker(day):
    year, week, _ = day.isocalendar()
    return f'{year}-W{week:02d}'


def _set_setting(db, key, value):
    db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))


def get_backup_settings(db):
    rows = db.execute("SELECT key, value FROM settings").fetchall()
    raw = dict(BACKUP_SETTING_DEFAULTS)
    raw.update({r['key']: r['value'] for r in rows if r['key'] in BACKUP_SETTING_DEFAULTS})
    public_cfg = get_public_config(db)
    if not raw.get('backup_email_to'):
        raw['backup_email_to'] = public_cfg.get('org_email') or ''

    return {
        'auto_enabled': _setting_bool(raw.get('backup_auto_enabled')),
        'auto_time': _valid_time_or_default(raw.get('backup_auto_time'), BACKUP_SETTING_DEFAULTS['backup_auto_time']),
        'retention_daily': _setting_int(raw.get('backup_retention_daily'), 3, 0, 31),
        'retention_weekly': _setting_int(raw.get('backup_retention_weekly'), 4, 0, 104),
        'retention_monthly': _setting_int(raw.get('backup_retention_monthly'), 6, 0, 120),
        'retention_yearly': _setting_int(raw.get('backup_retention_yearly'), 3, 0, 20),
        'email_enabled': _setting_bool(raw.get('backup_email_enabled')),
        'email_weekday': _setting_int(raw.get('backup_email_weekday'), 6, 0, 6),
        'email_time': _valid_time_or_default(raw.get('backup_email_time'), BACKUP_SETTING_DEFAULTS['backup_email_time']),
        'email_to': (raw.get('backup_email_to') or '').strip(),
        'email_max_mb': _setting_int(raw.get('backup_email_max_mb'), 20, 1, 2000),
        'drive_enabled': _setting_bool(raw.get('backup_drive_enabled')),
        'drive_folder_id': (raw.get('backup_drive_folder_id') or '').strip(),
        'drive_last_upload': raw.get('backup_drive_last_upload') or '',
        'drive_last_check': raw.get('backup_drive_last_check') or '',
        'drive_last_error': raw.get('backup_drive_last_error') or '',
        'auto_last_run_date': raw.get('backup_auto_last_run_date') or '',
        'email_last_attempt_week': raw.get('backup_email_last_attempt_week') or '',
        'email_last_sent_week': raw.get('backup_email_last_sent_week') or '',
    }


def _checkpoint_database():
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError('Datenbankdatei wurde nicht gefunden.')
    with sqlite3.connect(DB_PATH) as checkpoint_db:
        checkpoint_db.execute("PRAGMA wal_checkpoint(TRUNCATE)")


def write_backup_zip(zip_path):
    """Schreibt ein vollstaendiges ZIP-Backup an den angegebenen Pfad."""
    import zipfile

    with BACKUP_JOB_LOCK:
        _checkpoint_database()
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(DB_PATH, 'eeg_data.db')
            zf.writestr(
                'backup_manifest.txt',
                f"created_at={local_now().isoformat(timespec='seconds')}\n"
                f"database=eeg_data.db\n"
                f"invoices_folder=invoices/\n"
            )
            if os.path.isdir(INVOICE_FOLDER):
                for fname in os.listdir(INVOICE_FOLDER):
                    fpath = os.path.join(INVOICE_FOLDER, fname)
                    if os.path.isfile(fpath):
                        zf.write(fpath, f'invoices/{fname}')


def create_local_backup(prefix='eeg_auto'):
    os.makedirs(BACKUP_FOLDER, exist_ok=True)
    timestamp = local_now().strftime('%Y%m%d_%H%M%S')
    zip_filename = f'{prefix}_{timestamp}.zip'
    zip_path = os.path.join(BACKUP_FOLDER, zip_filename)
    write_backup_zip(zip_path)
    return zip_path, zip_filename


def _parse_backup_timestamp(filename):
    match = re.match(r'^eeg_(?:auto|manual)_(\d{8}_\d{6})\.zip$', filename)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), '%Y%m%d_%H%M%S').replace(tzinfo=APP_TIMEZONE)
    except ValueError:
        return None


def list_local_backups():
    backups = []
    if not os.path.isdir(BACKUP_FOLDER):
        return backups
    for fname in os.listdir(BACKUP_FOLDER):
        if not fname.endswith('.zip'):
            continue
        created_at = _parse_backup_timestamp(fname)
        if not created_at:
            continue
        fpath = os.path.join(BACKUP_FOLDER, fname)
        if not os.path.isfile(fpath):
            continue
        backups.append({
            'name': fname,
            'path': fpath,
            'size': os.path.getsize(fpath),
            'created_at': created_at,
            'kind': 'Automatisch' if fname.startswith('eeg_auto_') else 'Manuell',
        })
    backups.sort(key=lambda item: item['created_at'], reverse=True)
    return backups


def local_backup_path_for_delete(filename):
    """Validiert eine lokale Backup-Datei und liefert den sicheren Vollpfad."""
    backup_name = os.path.basename(str(filename or '').strip())
    if backup_name != str(filename or '').strip():
        raise ValueError('Ungültiger Backup-Dateiname.')
    if not _parse_backup_timestamp(backup_name):
        raise ValueError('Ungültiger Backup-Dateiname.')
    backup_path = os.path.abspath(os.path.join(BACKUP_FOLDER, backup_name))
    backup_root = os.path.abspath(BACKUP_FOLDER)
    if os.path.commonpath([backup_root, backup_path]) != backup_root:
        raise ValueError('Ungültiger Backup-Pfad.')
    if not os.path.isfile(backup_path):
        raise FileNotFoundError('Backup-Datei wurde nicht gefunden.')
    return backup_path, backup_name


def _google_libs_available():
    try:
        import google.auth.transport.requests  # noqa: F401
        import google.oauth2.credentials  # noqa: F401
        import google_auth_oauthlib.flow  # noqa: F401
        import googleapiclient.discovery  # noqa: F401
        import googleapiclient.http  # noqa: F401
    except ImportError:
        return False
    return True


def _secret_cipher_for_key(raw_key):
    if not raw_key:
        return None
    try:
        from cryptography.fernet import Fernet
    except ImportError as e:
        raise RuntimeError('cryptography fehlt. Bitte requirements.txt installieren.') from e

    key_text = str(raw_key).strip()
    try:
        return Fernet(key_text.encode('ascii'))
    except Exception:
        derived_key = base64.urlsafe_b64encode(hashlib.sha256(key_text.encode('utf-8')).digest())
        return Fernet(derived_key)


def _secret_cipher_candidates():
    keys = []
    data_key = os.environ.get('EEG_DATA_ENCRYPTION_KEY')
    if data_key:
        keys.append(data_key)
    if _SECRET_KEY and _SECRET_KEY not in keys:
        keys.append(_SECRET_KEY)
    return [_secret_cipher_for_key(key) for key in keys if key]


def _secret_cipher():
    candidates = _secret_cipher_candidates()
    return candidates[0] if candidates else None


def _atomic_write_json_file(path, payload):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp_path = f'{path}.tmp'
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write('\n')
    os.chmod(tmp_path, 0o600)
    os.replace(tmp_path, path)


def _write_private_json_file(path, payload):
    """Speichert lokale Secret-JSONs verschluesselt, wenn ein persistenter Key vorhanden ist."""
    cipher = _secret_cipher()
    if cipher:
        encrypted = cipher.encrypt(json.dumps(payload, ensure_ascii=False).encode('utf-8')).decode('ascii')
        _atomic_write_json_file(path, {
            '_eeg_encrypted': 1,
            'cipher': 'fernet-sha256-v1',
            'payload': encrypted,
        })
        return
    if _IS_PRODUCTION:
        raise RuntimeError('EEG_DATA_ENCRYPTION_KEY oder EEG_SECRET_KEY muss fuer Secret-Speicherung gesetzt sein.')
    app.logger.warning('Secret JSON wird unverschluesselt gespeichert, weil kein persistenter Encryption-Key gesetzt ist.')
    _atomic_write_json_file(path, payload)


def _load_private_json_file(path):
    with open(path, encoding='utf-8') as f:
        payload = json.load(f)
    if not isinstance(payload, dict) or not payload.get('_eeg_encrypted'):
        return payload

    candidates = _secret_cipher_candidates()
    if not candidates:
        raise RuntimeError('Secret-Datei ist verschluesselt, aber EEG_DATA_ENCRYPTION_KEY/EEG_SECRET_KEY fehlt.')
    encrypted = str(payload.get('payload') or '').encode('ascii')
    for cipher in candidates:
        try:
            decrypted = cipher.decrypt(encrypted)
            return json.loads(decrypted.decode('utf-8'))
        except Exception:
            continue
    raise RuntimeError('Secret-Datei konnte nicht entschluesselt werden. Encryption-Key pruefen.')


def _load_json_payload(file_field, text_field, label):
    upload = request.files.get(file_field)
    raw = ''
    if upload and upload.filename:
        raw = upload.read().decode('utf-8')
    else:
        raw = (request.form.get(text_field) or '').strip()
    if not raw:
        raise ValueError(f'{label} fehlt.')
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        raise ValueError(f'{label} ist kein gültiges JSON: {e}') from e


def _manual_google_client_payload():
    client_id = (request.form.get('google_client_id') or '').strip()
    client_secret = (request.form.get('google_client_secret') or '').strip()
    if not client_id and not client_secret:
        return None
    if not client_id or not client_secret:
        raise ValueError('Client-ID und Client-Secret müssen beide ausgefüllt sein.')
    return {
        'web': {
            'client_id': client_id,
            'client_secret': client_secret,
            'auth_uri': (request.form.get('google_auth_uri') or 'https://accounts.google.com/o/oauth2/auth').strip(),
            'token_uri': (request.form.get('google_token_uri') or 'https://oauth2.googleapis.com/token').strip(),
            'redirect_uris': [_google_redirect_uri_for_display()],
        }
    }


def _manual_google_token_payload():
    refresh_token = (request.form.get('google_refresh_token') or '').strip()
    if not refresh_token:
        return None

    client_id = (request.form.get('google_token_client_id') or '').strip()
    client_secret = (request.form.get('google_token_client_secret') or '').strip()
    if (not client_id or not client_secret) and os.path.exists(GOOGLE_CLIENT_SECRETS_FILE):
        client_payload = _load_private_json_file(GOOGLE_CLIENT_SECRETS_FILE)
        section = client_payload.get('web') or client_payload.get('installed') or {}
        client_id = client_id or section.get('client_id', '')
        client_secret = client_secret or section.get('client_secret', '')

    if not client_id or not client_secret:
        raise ValueError('Für einen manuellen Token brauchen wir Refresh-Token, Client-ID und Client-Secret.')

    return {
        'token': (request.form.get('google_access_token') or '').strip(),
        'refresh_token': refresh_token,
        'token_uri': (request.form.get('google_token_token_uri') or 'https://oauth2.googleapis.com/token').strip(),
        'client_id': client_id,
        'client_secret': client_secret,
        'scopes': GOOGLE_DRIVE_SCOPES,
    }


def validate_google_client_config(payload):
    if not isinstance(payload, dict):
        raise ValueError('OAuth Client-Konfiguration muss ein JSON-Objekt sein.')
    section_name = 'web' if isinstance(payload.get('web'), dict) else 'installed'
    section = payload.get(section_name)
    if not isinstance(section, dict):
        raise ValueError('OAuth Client-JSON muss einen Bereich "web" oder "installed" enthalten.')
    missing = [
        key for key in ('client_id', 'client_secret', 'auth_uri', 'token_uri')
        if not section.get(key)
    ]
    if missing:
        raise ValueError(f'OAuth Client-JSON ist unvollständig: {", ".join(missing)} fehlt.')
    return section_name


def validate_google_token_payload(payload):
    if not isinstance(payload, dict):
        raise ValueError('Token muss ein JSON-Objekt sein.')
    missing = [
        key for key in ('refresh_token', 'token_uri', 'client_id', 'client_secret')
        if not payload.get(key)
    ]
    if missing:
        raise ValueError(f'Token-JSON ist unvollständig: {", ".join(missing)} fehlt.')
    try:
        from google.oauth2.credentials import Credentials
    except ImportError as e:
        raise RuntimeError('Google Drive Python-Bibliotheken fehlen. Bitte requirements.txt installieren.') from e
    Credentials.from_authorized_user_info(payload, GOOGLE_DRIVE_SCOPES)
    return True


def _pkce_code_verifier():
    return base64.urlsafe_b64encode(secrets.token_bytes(64)).rstrip(b'=').decode('ascii')


def _pkce_code_challenge(code_verifier):
    digest = hashlib.sha256(code_verifier.encode('ascii')).digest()
    return base64.urlsafe_b64encode(digest).rstrip(b'=').decode('ascii')


def _oauth_verifier_fingerprint(code_verifier):
    return hashlib.sha256(code_verifier.encode('ascii')).hexdigest()[:12]


def _store_google_oauth_pkce(correlation_id, state, code_verifier):
    db = get_db()
    db.execute("DELETE FROM oauth_pkce_sessions WHERE expires_at <= ?", (utc_now_string(),))
    db.execute(
        """INSERT OR REPLACE INTO oauth_pkce_sessions
           (id, user_id, state, code_verifier, created_at, expires_at)
           VALUES (?, ?, ?, ?, ?, datetime(?, '+15 minutes'))""",
        (
            correlation_id,
            current_user.id if current_user and current_user.is_authenticated else None,
            state,
            code_verifier,
            utc_now_string(),
            utc_now_string(),
        )
    )
    db.commit()


def _pop_google_oauth_pkce(correlation_id, state):
    db = get_db()
    row = db.execute(
        """SELECT * FROM oauth_pkce_sessions
           WHERE id=? AND state=? AND expires_at > ?""",
        (correlation_id, state, utc_now_string())
    ).fetchone()
    db.execute("DELETE FROM oauth_pkce_sessions WHERE id=?", (correlation_id,))
    db.commit()
    if not row:
        return None
    if row['user_id'] and current_user and current_user.is_authenticated and row['user_id'] != current_user.id:
        return None
    return row['code_verifier']


def get_google_drive_status():
    token_exists = os.path.exists(GOOGLE_TOKEN_FILE)
    client_exists = os.path.exists(GOOGLE_CLIENT_SECRETS_FILE)
    connected = False
    error = ''
    if token_exists and _google_libs_available():
        try:
            credentials = _load_google_drive_credentials(refresh=True)
            connected = credentials and credentials.valid
        except Exception as e:
            error = str(e)
    return {
        'libs_available': _google_libs_available(),
        'client_file': GOOGLE_CLIENT_SECRETS_FILE,
        'client_configured': client_exists,
        'token_file': GOOGLE_TOKEN_FILE,
        'connected': connected,
        'error': error,
        'redirect_uri': _google_redirect_uri_for_display(),
    }


def _google_redirect_uri():
    return GOOGLE_OAUTH_REDIRECT_URI or public_url_for('admin_backup_google_callback')


def _google_redirect_uri_for_display():
    if GOOGLE_OAUTH_REDIRECT_URI:
        return GOOGLE_OAUTH_REDIRECT_URI
    return public_url_for('admin_backup_google_callback')


def _google_drive_flow():
    if not os.path.exists(GOOGLE_CLIENT_SECRETS_FILE):
        raise RuntimeError(f'Google OAuth Client-Datei fehlt: {GOOGLE_CLIENT_SECRETS_FILE}')
    try:
        from google_auth_oauthlib.flow import Flow
    except ImportError as e:
        raise RuntimeError('Google Drive Python-Bibliotheken fehlen. Bitte requirements.txt installieren.') from e
    client_config = _load_private_json_file(GOOGLE_CLIENT_SECRETS_FILE)
    return Flow.from_client_config(
        client_config,
        scopes=GOOGLE_DRIVE_SCOPES,
        redirect_uri=_google_redirect_uri(),
    )


def _load_google_drive_credentials(refresh=False):
    if not os.path.exists(GOOGLE_TOKEN_FILE):
        return None
    try:
        from google.auth.transport.requests import Request as GoogleAuthRequest
        from google.oauth2.credentials import Credentials
    except ImportError as e:
        raise RuntimeError('Google Drive Python-Bibliotheken fehlen. Bitte requirements.txt installieren.') from e

    token_payload = _load_private_json_file(GOOGLE_TOKEN_FILE)
    credentials = Credentials.from_authorized_user_info(token_payload, GOOGLE_DRIVE_SCOPES)
    if refresh and credentials and credentials.expired and credentials.refresh_token:
        credentials.refresh(GoogleAuthRequest())
        _write_private_json_file(GOOGLE_TOKEN_FILE, json.loads(credentials.to_json()))
    return credentials


def _google_drive_service():
    credentials = _load_google_drive_credentials(refresh=True)
    if not credentials or not credentials.valid:
        raise RuntimeError('Google Drive ist noch nicht verbunden.')
    try:
        from googleapiclient.discovery import build
    except ImportError as e:
        raise RuntimeError('Google Drive Python-Bibliotheken fehlen. Bitte requirements.txt installieren.') from e
    return build('drive', 'v3', credentials=credentials, cache_discovery=False)


def check_google_drive_connection(db):
    service = _google_drive_service()
    service.files().list(pageSize=1, fields='files(id,name)').execute()
    checked_at = local_now().isoformat(timespec='seconds')
    _set_setting(db, 'backup_drive_last_check', checked_at)
    _set_setting(db, 'backup_drive_last_error', '')
    db.commit()
    return checked_at


def upload_backup_to_google_drive(db, backup_path):
    try:
        from googleapiclient.http import MediaFileUpload
    except ImportError as e:
        raise RuntimeError('Google Drive Python-Bibliotheken fehlen. Bitte requirements.txt installieren.') from e

    settings = get_backup_settings(db)
    service = _google_drive_service()
    backup_name = os.path.basename(backup_path)
    metadata = {'name': backup_name}
    if settings['drive_folder_id']:
        metadata['parents'] = [settings['drive_folder_id']]
    media = MediaFileUpload(backup_path, mimetype='application/zip', resumable=True)
    uploaded = service.files().create(
        body=metadata,
        media_body=media,
        fields='id,name,webViewLink',
        supportsAllDrives=True,
    ).execute()
    _set_setting(db, 'backup_drive_last_upload', local_now().isoformat(timespec='seconds'))
    _set_setting(db, 'backup_drive_last_error', '')
    db.commit()
    return uploaded


def _drive_query_literal(value):
    return str(value).replace('\\', '\\\\').replace("'", "\\'")


def _is_google_backup_name(filename):
    return bool(_parse_backup_timestamp(str(filename or '').strip()))


def _normalize_google_drive_backup_file(item):
    size = item.get('size')
    try:
        size = int(size) if size is not None else 0
    except (TypeError, ValueError):
        size = 0
    return {
        'id': item.get('id', ''),
        'name': item.get('name', ''),
        'size': size,
        'created_at': item.get('createdTime') or item.get('modifiedTime') or '',
        'modified_at': item.get('modifiedTime') or '',
        'web_view_link': item.get('webViewLink') or '',
        'mime_type': item.get('mimeType') or '',
    }


def list_google_drive_backups(db, limit=50):
    """Listet von dieser App erreichbare EEG-Backup-ZIP-Dateien in Google Drive."""
    settings = get_backup_settings(db)
    service = _google_drive_service()
    query_parts = [
        "trashed=false",
        "name contains 'eeg_'",
        "name contains '.zip'",
    ]
    if settings['drive_folder_id']:
        folder_id = _drive_query_literal(settings['drive_folder_id'])
        query_parts.append(f"'{folder_id}' in parents")
    response = service.files().list(
        q=' and '.join(query_parts),
        pageSize=max(1, min(int(limit), 100)),
        fields='files(id,name,size,createdTime,modifiedTime,webViewLink,mimeType)',
        orderBy='createdTime desc',
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    backups = []
    for item in response.get('files', []):
        if _is_google_backup_name(item.get('name')):
            backups.append(_normalize_google_drive_backup_file(item))
    return backups


def trash_google_drive_backup(db, file_id):
    """Verschiebt ein von der App verwaltbares Drive-Backup in den Papierkorb."""
    drive_file_id = str(file_id or '').strip()
    if not re.match(r'^[A-Za-z0-9_-]{8,}$', drive_file_id):
        raise ValueError('Ungültige Google Drive Datei-ID.')

    service = _google_drive_service()
    metadata = service.files().get(
        fileId=drive_file_id,
        fields='id,name,mimeType,trashed',
        supportsAllDrives=True,
    ).execute()
    backup_name = metadata.get('name') or ''
    if metadata.get('trashed'):
        raise ValueError('Dieses Google Drive Backup liegt bereits im Papierkorb.')
    if not _is_google_backup_name(backup_name):
        raise ValueError('Aus Sicherheitsgründen können nur EEG-Backup-ZIP-Dateien gelöscht werden.')

    deleted = service.files().update(
        fileId=drive_file_id,
        body={'trashed': True},
        fields='id,name,trashed',
        supportsAllDrives=True,
    ).execute()
    _set_setting(db, 'backup_drive_last_error', '')
    db.commit()
    return deleted


def apply_backup_retention(settings):
    """Wendet eine einfache Grossvater-Vater-Sohn-Aufbewahrung auf Auto-Backups an."""
    auto_backups = [
        item for item in list_local_backups()
        if item['name'].startswith('eeg_auto_')
    ]
    now = local_now()
    kept_buckets = set()
    keep_paths = set()

    for item in auto_backups:
        ts = item['created_at']
        age_days = (now.date() - ts.date()).days
        month_distance = (now.year - ts.year) * 12 + (now.month - ts.month)
        year_distance = now.year - ts.year

        bucket = None
        if age_days < settings['retention_daily']:
            bucket = f"day:{ts.strftime('%Y-%m-%d')}"
        elif age_days < settings['retention_weekly'] * 7:
            if settings['retention_weekly'] > 0:
                bucket = f"week:{ts.strftime('%G-W%V')}"
        elif month_distance < settings['retention_monthly']:
            if settings['retention_monthly'] > 0:
                bucket = f"month:{ts.strftime('%Y-%m')}"
        elif year_distance < settings['retention_yearly']:
            if settings['retention_yearly'] > 0:
                bucket = f"year:{ts.strftime('%Y')}"

        if bucket and bucket not in kept_buckets:
            kept_buckets.add(bucket)
            keep_paths.add(item['path'])

    deleted = 0
    for item in auto_backups:
        if item['path'] in keep_paths:
            continue
        try:
            os.remove(item['path'])
            deleted += 1
        except OSError:
            app.logger.warning('Could not delete old backup %s', item['path'], exc_info=True)
    return deleted


def _time_reached(now, time_text):
    hour, minute = [int(part) for part in _valid_time_or_default(time_text, '00:00').split(':')]
    return (now.hour, now.minute) >= (hour, minute)


def send_backup_email(db, backup_path, recipient, max_mb):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication

    if not _is_valid_email(recipient):
        raise RuntimeError('Empfängeradresse für Backup-Mail ist ungültig.')

    size_mb = os.path.getsize(backup_path) / 1024 / 1024
    if size_mb > max_mb:
        raise RuntimeError(f'Backup ist {size_mb:.1f} MB groß und überschreitet das konfigurierte Mail-Limit von {max_mb} MB.')

    mail_cfg = _get_valid_mail_config(db)
    public_cfg = get_public_config(db)
    backup_name = os.path.basename(backup_path)
    subject = f"EEG Backup {local_now().strftime('%d.%m.%Y')}"
    body_text = (
        f"Automatisches Backup der Webapp {public_cfg['org_name']}.\n\n"
        f"Datei: {backup_name}\n"
        f"Groesse: {size_mb:.1f} MB\n"
        f"Erstellt am: {local_now().strftime('%d.%m.%Y %H:%M')} {getattr(APP_TIMEZONE, 'key', 'Europe/Vienna')}\n\n"
        "Bitte diese Datei geschuetzt aufbewahren, da sie personenbezogene Daten enthalten kann."
    )

    msg = MIMEMultipart('mixed')
    msg['From'] = mail_cfg['from_header']
    msg['Reply-To'] = mail_cfg['reply_to_header']
    msg['To'] = recipient
    msg['Subject'] = subject
    msg.attach(MIMEText(body_text, 'plain', 'utf-8'))

    with open(backup_path, 'rb') as f:
        attachment = MIMEApplication(f.read(), _subtype='zip')
    attachment.add_header('Content-Disposition', 'attachment', filename=backup_name)
    msg.attach(attachment)

    _log_mail_send(mail_cfg, recipient, subject)
    with smtplib.SMTP(mail_cfg['smtp_host'], mail_cfg['smtp_port']) as server:
        if mail_cfg['smtp_tls']:
            server.starttls()
        server.login(mail_cfg['smtp_user'], mail_cfg['smtp_pass'])
        server.send_message(msg, from_addr=mail_cfg['from_address'], to_addrs=[recipient])


def _run_due_backup_jobs():
    with app.app_context():
        db = get_db()
        settings = get_backup_settings(db)
        now = local_now()
        today = now.strftime('%Y-%m-%d')
        week_marker = _backup_week_marker(now.date())

        if settings['auto_enabled'] and _time_reached(now, settings['auto_time']):
            if settings['auto_last_run_date'] != today:
                zip_path, zip_filename = create_local_backup('eeg_auto')
                deleted = apply_backup_retention(settings)
                drive_detail = ''
                if settings['drive_enabled']:
                    try:
                        uploaded = upload_backup_to_google_drive(db, zip_path)
                        drive_detail = f' · Google Drive Upload: {uploaded.get("id")}'
                    except Exception as e:
                        _set_setting(db, 'backup_drive_last_error', str(e))
                        db.commit()
                        app.logger.exception('Google Drive backup upload failed')
                        audit_log('backup_drive_failed', f'Google Drive Upload fehlgeschlagen: {zip_filename} ({e})')
                _set_setting(db, 'backup_auto_last_run_date', today)
                db.commit()
                audit_log('backup_auto', f'Automatisches Backup erstellt: {zip_filename} ({deleted} alte Backups entfernt){drive_detail}')
                app.logger.info('Automatic backup created: %s', zip_path)

        if (settings['email_enabled']
                and now.weekday() == settings['email_weekday']
                and _time_reached(now, settings['email_time'])
                and settings['email_last_attempt_week'] != week_marker):
            email_path = None
            try:
                email_path, zip_filename = create_local_backup('eeg_mail')
                send_backup_email(db, email_path, settings['email_to'], settings['email_max_mb'])
                _set_setting(db, 'backup_email_last_sent_week', week_marker)
                audit_log('backup_email', f'Woechentliches Backup per Mail versendet: {zip_filename} an {settings["email_to"]}')
            except Exception as e:
                app.logger.exception('Weekly backup mail failed')
                audit_log('backup_email_failed', f'Woechentliches Backup-Mail fehlgeschlagen: {e}')
            finally:
                _set_setting(db, 'backup_email_last_attempt_week', week_marker)
                db.commit()
                if email_path and os.path.exists(email_path):
                    try:
                        os.remove(email_path)
                    except OSError:
                        app.logger.warning('Could not remove temporary mail backup %s', email_path, exc_info=True)


def _backup_scheduler_loop():
    last_checkpoint = time.time()
    while True:
        try:
            _run_due_backup_jobs()
        except Exception:
            app.logger.exception('Automatic backup scheduler failed')
        # Stuendlicher WAL-Checkpoint: haelt eeg_data.db-wal klein, damit die
        # Datei nicht unbemerkt zwischen den Backups anwaechst.
        if time.time() - last_checkpoint >= 3600:
            last_checkpoint = time.time()
            try:
                with BACKUP_JOB_LOCK:
                    _checkpoint_database()
            except Exception:
                app.logger.exception('WAL-Checkpoint fehlgeschlagen')
        time.sleep(60)


def start_backup_scheduler():
    global BACKUP_SCHEDULER_STARTED
    with BACKUP_SCHEDULER_LOCK:
        if BACKUP_SCHEDULER_STARTED:
            return
        thread = threading.Thread(target=_backup_scheduler_loop, name='eeg-backup-scheduler', daemon=True)
        thread.start()
        BACKUP_SCHEDULER_STARTED = True


def get_backup_info():
    invoice_count = 0
    invoice_size = 0
    if os.path.isdir(INVOICE_FOLDER):
        for fname in os.listdir(INVOICE_FOLDER):
            fpath = os.path.join(INVOICE_FOLDER, fname)
            if os.path.isfile(fpath):
                invoice_count += 1
                invoice_size += os.path.getsize(fpath)
    wal_path = DB_PATH + '-wal'
    wal_size = os.path.getsize(wal_path) if os.path.exists(wal_path) else 0
    return {
        'db_path': DB_PATH,
        'db_size': os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0,
        'invoice_folder': INVOICE_FOLDER,
        'backup_folder': BACKUP_FOLDER,
        'invoice_count': invoice_count,
        'invoice_size': invoice_size,
        'wal_size': wal_size,
        'wal_warning': wal_size > WAL_WARN_BYTES,
    }


def _backup_redirect_target():
    next_url = request.form.get('next') or request.args.get('next')
    if next_url and is_safe_redirect_url(next_url):
        return next_url
    session_next = session.pop('backup_redirect_next', None)
    if session_next and is_safe_redirect_url(session_next):
        return session_next
    return url_for('admin_backup')


@app.route('/admin/backup')
@admin_required
def admin_backup():
    """Admin-Seite fuer Backup und Restore."""
    db = get_db()
    smtp_configured, _ = _validate_mail_config(_load_mail_config(db))
    google_drive = get_google_drive_status()
    drive_backups = []
    drive_backups_error = ''
    if google_drive['connected']:
        try:
            drive_backups = list_google_drive_backups(db)
        except Exception as e:
            drive_backups_error = str(e)
            app.logger.warning('Could not list Google Drive backups: %s', e, exc_info=True)
    return render_template(
        'admin_backup.html',
        info=get_backup_info(),
        backup_settings=get_backup_settings(db),
        backup_files=list_local_backups()[:20],
        google_drive=google_drive,
        drive_backups=drive_backups,
        drive_backups_error=drive_backups_error,
        smtp_configured=smtp_configured,
        weekdays=[
            (0, 'Montag'),
            (1, 'Dienstag'),
            (2, 'Mittwoch'),
            (3, 'Donnerstag'),
            (4, 'Freitag'),
            (5, 'Samstag'),
            (6, 'Sonntag'),
        ],
    )


@app.route('/admin/backup/settings', methods=['POST'])
@admin_required
def admin_backup_settings():
    """Speichert Zeitplan, Aufbewahrung und Mail-Backup-Konfiguration."""
    db = get_db()
    email_enabled = form_switch_enabled('backup_email_enabled')
    email_to = (request.form.get('backup_email_to') or '').strip()
    if email_enabled and not _is_valid_email(email_to):
        flash('Bitte eine gültige Empfängeradresse für das Mail-Backup eintragen.', 'danger')
        return redirect(_backup_redirect_target())
    drive_enabled = form_switch_enabled('backup_drive_enabled')
    if drive_enabled and not get_google_drive_status()['connected']:
        flash('Google Drive muss zuerst verbunden werden, bevor der automatische Drive-Upload aktiviert werden kann.', 'danger')
        return redirect(_backup_redirect_target())

    values = {
        'backup_auto_enabled': 'true' if form_switch_enabled('backup_auto_enabled') else 'false',
        'backup_auto_time': _valid_time_or_default(request.form.get('backup_auto_time'), BACKUP_SETTING_DEFAULTS['backup_auto_time']),
        'backup_retention_daily': _setting_int(request.form.get('backup_retention_daily'), 3, 0, 31),
        'backup_retention_weekly': _setting_int(request.form.get('backup_retention_weekly'), 4, 0, 104),
        'backup_retention_monthly': _setting_int(request.form.get('backup_retention_monthly'), 6, 0, 120),
        'backup_retention_yearly': _setting_int(request.form.get('backup_retention_yearly'), 3, 0, 20),
        'backup_email_enabled': 'true' if email_enabled else 'false',
        'backup_email_weekday': _setting_int(request.form.get('backup_email_weekday'), 6, 0, 6),
        'backup_email_time': _valid_time_or_default(request.form.get('backup_email_time'), BACKUP_SETTING_DEFAULTS['backup_email_time']),
        'backup_email_to': email_to,
        'backup_email_max_mb': _setting_int(request.form.get('backup_email_max_mb'), 20, 1, 2000),
        'backup_drive_enabled': 'true' if drive_enabled else 'false',
        'backup_drive_folder_id': (request.form.get('backup_drive_folder_id') or '').strip(),
    }
    for key, value in values.items():
        _set_setting(db, key, value)
    db.commit()

    settings = get_backup_settings(db)
    deleted = apply_backup_retention(settings)
    audit_log('backup_settings_update', f'Backup-Konfiguration geändert ({deleted} alte Auto-Backups entfernt)')
    flash('Backup-Konfiguration gespeichert.', 'success')
    return redirect(_backup_redirect_target())


@app.route('/admin/backup/google/connect')
@admin_required
def admin_backup_google_connect():
    """Startet den Google OAuth-Flow fuer Drive-Backups."""
    try:
        next_url = request.args.get('next')
        if next_url and is_safe_redirect_url(next_url):
            session['backup_redirect_next'] = next_url
        flow = _google_drive_flow()
        correlation_id = secrets.token_hex(12)
        oauth_state = secrets.token_urlsafe(32)
        code_verifier = _pkce_code_verifier()
        code_challenge = _pkce_code_challenge(code_verifier)
        _store_google_oauth_pkce(correlation_id, oauth_state, code_verifier)
        session['google_drive_oauth_state'] = oauth_state
        session['google_drive_oauth_correlation_id'] = correlation_id
        session.modified = True
        app.logger.info(
            'Google Drive OAuth start | correlation=%s | verifier_stored=%s | verifier_fp=%s | redirect_uri=%s | secure_cookie=%s | samesite=%s',
            correlation_id,
            True,
            _oauth_verifier_fingerprint(code_verifier),
            _google_redirect_uri(),
            app.config.get('SESSION_COOKIE_SECURE'),
            app.config.get('SESSION_COOKIE_SAMESITE'),
        )
        authorization_url, state = flow.authorization_url(
            state=oauth_state,
            access_type='offline',
            include_granted_scopes='true',
            prompt='consent',
            code_challenge=code_challenge,
            code_challenge_method='S256',
        )
        if state != oauth_state:
            app.logger.warning('Google Drive OAuth state changed by library | correlation=%s', correlation_id)
        return redirect(authorization_url)
    except Exception as e:
        audit_log('backup_drive_connect_failed', f'Google Drive Verbindung fehlgeschlagen: {e}')
        flash_exception(e, 'Google Drive Verbindung konnte nicht gestartet werden.')
        return redirect(_backup_redirect_target())


@app.route('/admin/backup/google/client-config', methods=['POST'])
@admin_required
def admin_backup_google_client_config():
    """Speichert die Google OAuth Client-Konfiguration aus dem Webinterface."""
    try:
        payload = _manual_google_client_payload()
        if payload is None:
            payload = _load_json_payload('google_client_file', 'google_client_json', 'OAuth Client-JSON')
        section_name = validate_google_client_config(payload)
        _write_private_json_file(GOOGLE_CLIENT_SECRETS_FILE, payload)
        if os.path.exists(GOOGLE_TOKEN_FILE):
            os.remove(GOOGLE_TOKEN_FILE)
        db = get_db()
        _set_setting(db, 'backup_drive_enabled', 'false')
        _set_setting(db, 'backup_drive_last_error', '')
        db.commit()
        audit_log('backup_drive_client_config', f'Google OAuth Client-Konfiguration gespeichert ({section_name})')
        flash('Google OAuth Client-Konfiguration gespeichert. Ein vorhandener Token wurde zur Sicherheit entfernt; bitte Google Drive neu verbinden.', 'success')
    except Exception as e:
        audit_log('backup_drive_client_config_failed', f'Google OAuth Client-Konfiguration fehlgeschlagen: {e}')
        flash_exception(e, 'Google OAuth Client-Konfiguration konnte nicht gespeichert werden.')
    return redirect(_backup_redirect_target())


@app.route('/admin/backup/google/token', methods=['POST'])
@admin_required
def admin_backup_google_token():
    """Speichert ein vorhandenes Google OAuth Token-JSON aus dem Webinterface."""
    try:
        payload = _manual_google_token_payload()
        if payload is None:
            payload = _load_json_payload('google_token_file', 'google_token_json', 'Google Token-JSON')
        validate_google_token_payload(payload)
        _write_private_json_file(GOOGLE_TOKEN_FILE, payload)
        db = get_db()
        _set_setting(db, 'backup_drive_last_error', '')
        db.commit()
        audit_log('backup_drive_token', 'Google Drive Token hinterlegt')
        flash('Google Drive Token wurde gespeichert.', 'success')
    except Exception as e:
        audit_log('backup_drive_token_failed', f'Google Drive Token konnte nicht gespeichert werden: {e}')
        flash_exception(e, 'Google Drive Token konnte nicht gespeichert werden.')
    return redirect(_backup_redirect_target())


@app.route('/admin/backup/google/callback')
@admin_required
def admin_backup_google_callback():
    """OAuth Callback fuer Google Drive."""
    correlation_id = session.get('google_drive_oauth_correlation_id')
    code_verifier = None
    try:
        state = session.get('google_drive_oauth_state')
        callback_state = request.args.get('state')
        verifier_found = False
        if state and callback_state and correlation_id:
            code_verifier = _pop_google_oauth_pkce(correlation_id, callback_state)
            verifier_found = bool(code_verifier)
        app.logger.info(
            'Google Drive OAuth callback | correlation=%s | session_state_present=%s | state_matches=%s | verifier_found=%s | scheme=%s | host=%s',
            correlation_id or 'missing',
            bool(state),
            bool(state and state == callback_state),
            verifier_found,
            request.scheme,
            request.host,
        )
        if not state or state != callback_state:
            raise RuntimeError('OAuth-State ist ungültig.')
        if not code_verifier:
            raise RuntimeError('PKCE Code-Verifier wurde in der Session nicht gefunden. Bitte Google Drive erneut verbinden.')
        flow = _google_drive_flow()
        flow.fetch_token(authorization_response=request.url, code_verifier=code_verifier)
        credentials = flow.credentials
        _write_private_json_file(GOOGLE_TOKEN_FILE, json.loads(credentials.to_json()))
        db = get_db()
        _set_setting(db, 'backup_drive_last_error', '')
        db.commit()
        audit_log('backup_drive_connect', 'Google Drive verbunden')
        flash('Google Drive wurde erfolgreich verbunden.', 'success')
    except Exception as e:
        audit_log('backup_drive_connect_failed', f'Google Drive Verbindung fehlgeschlagen: {e}')
        flash_exception(e, 'Google Drive konnte nicht verbunden werden.')
    finally:
        session.pop('google_drive_oauth_state', None)
        session.pop('google_drive_oauth_correlation_id', None)
        session.modified = True
    return redirect(_backup_redirect_target())


@app.route('/admin/backup/google/client-config/delete', methods=['POST'])
@admin_required
def admin_backup_google_client_config_delete():
    """Entfernt lokale Google Client- und Token-Dateien."""
    try:
        for path in (GOOGLE_TOKEN_FILE, GOOGLE_CLIENT_SECRETS_FILE):
            if os.path.exists(path):
                os.remove(path)
        db = get_db()
        _set_setting(db, 'backup_drive_enabled', 'false')
        _set_setting(db, 'backup_drive_last_error', '')
        db.commit()
        audit_log('backup_drive_client_config_delete', 'Google Drive Client-Konfiguration und Token entfernt')
        flash('Google Drive Client-Konfiguration und Token wurden entfernt.', 'success')
    except Exception as e:
        audit_log('backup_drive_client_config_delete_failed', f'Google Drive Client-Konfiguration konnte nicht entfernt werden: {e}')
        flash_exception(e, 'Google Drive Client-Konfiguration konnte nicht entfernt werden.')
    return redirect(_backup_redirect_target())


@app.route('/admin/backup/google/disconnect', methods=['POST'])
@admin_required
def admin_backup_google_disconnect():
    """Entfernt das lokal gespeicherte Google OAuth-Token."""
    try:
        if os.path.exists(GOOGLE_TOKEN_FILE):
            os.remove(GOOGLE_TOKEN_FILE)
        db = get_db()
        _set_setting(db, 'backup_drive_enabled', 'false')
        _set_setting(db, 'backup_drive_last_error', '')
        db.commit()
        audit_log('backup_drive_disconnect', 'Google Drive getrennt')
        flash('Google Drive wurde getrennt. Automatischer Drive-Upload ist deaktiviert.', 'success')
    except Exception as e:
        audit_log('backup_drive_disconnect_failed', f'Google Drive Trennung fehlgeschlagen: {e}')
        flash_exception(e, 'Google Drive konnte nicht getrennt werden.')
    return redirect(_backup_redirect_target())


@app.route('/admin/backup/google/test', methods=['POST'])
@admin_required
def admin_backup_google_test():
    """Prüft Refresh-Token und Google Drive API Zugriff."""
    db = get_db()
    try:
        checked_at = check_google_drive_connection(db)
        audit_log('backup_drive_test', 'Google Drive Verbindung erfolgreich geprüft')
        flash(f'Google Drive Verbindung erfolgreich geprüft: {checked_at}', 'success')
    except Exception as e:
        _set_setting(db, 'backup_drive_last_error', str(e))
        db.commit()
        audit_log('backup_drive_test_failed', f'Google Drive Verbindungstest fehlgeschlagen: {e}')
        flash_exception(e, 'Google Drive Verbindungstest fehlgeschlagen.')
    return redirect(_backup_redirect_target())


@app.route('/admin/backup/run', methods=['POST'])
@admin_required
def admin_backup_run():
    """Erstellt ein lokales Backup im Backup-Ordner."""
    try:
        zip_path, zip_filename = create_local_backup('eeg_manual')
        audit_log('backup_manual', f'Manuelles lokales Backup erstellt: {zip_filename}')
        flash(f'Lokales Backup erstellt: {zip_filename}', 'success')
        app.logger.info('Manual local backup created: %s', zip_path)
    except Exception as e:
        audit_log('backup_manual_failed', f'Manuelles lokales Backup fehlgeschlagen: {e}')
        flash_exception(e, 'Backup konnte nicht erstellt werden.')
    return redirect(_backup_redirect_target())


@app.route('/admin/backup/upload-drive', methods=['POST'])
@admin_required
def admin_backup_upload_drive():
    """Kopiert eine lokale Backup-Datei manuell nach Google Drive."""
    filename = request.form.get('backup_name', '')
    try:
        backup_path, backup_name = local_backup_path_for_delete(filename)
        uploaded = upload_backup_to_google_drive(get_db(), backup_path)
        audit_log('backup_drive_upload', f'Backup nach Google Drive kopiert: {backup_name} ({uploaded.get("id")})')
        flash(f'Backup wurde nach Google Drive kopiert: {backup_name}', 'success')
    except Exception as e:
        db = get_db()
        _set_setting(db, 'backup_drive_last_error', str(e))
        db.commit()
        audit_log('backup_drive_failed', f'Google Drive Upload fehlgeschlagen: {filename} ({e})')
        flash_exception(e, 'Google Drive Upload fehlgeschlagen.')
    return redirect(_backup_redirect_target())


@app.route('/admin/backup/google/delete', methods=['POST'])
@admin_required
def admin_backup_google_delete():
    """Verschiebt eine Google Drive Backup-Datei in den Papierkorb."""
    drive_file_id = request.form.get('drive_file_id', '')
    try:
        deleted = trash_google_drive_backup(get_db(), drive_file_id)
        backup_name = deleted.get('name') or drive_file_id
        audit_log('backup_drive_delete', f'Google Drive Backup in den Papierkorb verschoben: {backup_name} ({deleted.get("id")})')
        flash(f'Google Drive Backup wurde in den Papierkorb verschoben: {backup_name}', 'success')
    except Exception as e:
        db = get_db()
        _set_setting(db, 'backup_drive_last_error', str(e))
        db.commit()
        audit_log('backup_drive_delete_failed', f'Google Drive Backup konnte nicht gelöscht werden: {drive_file_id} ({e})')
        flash_exception(e, 'Google Drive Backup konnte nicht gelöscht werden.')
    return redirect(_backup_redirect_target())


@app.route('/admin/backup/delete', methods=['POST'])
@admin_required
def admin_backup_delete():
    """Löscht eine lokale Backup-Datei nach serverseitiger Sicherheitsprüfung."""
    filename = request.form.get('backup_name', '')
    try:
        backup_path, backup_name = local_backup_path_for_delete(filename)
        size_mb = os.path.getsize(backup_path) / 1024 / 1024
        os.remove(backup_path)
        audit_log('backup_delete', f'Lokales Backup gelöscht: {backup_name} ({size_mb:.1f} MB)')
        flash(f'Backup gelöscht: {backup_name}', 'success')
    except Exception as e:
        audit_log('backup_delete_failed', f'Backup-Löschung fehlgeschlagen: {filename} ({e})')
        flash_exception(e, 'Backup konnte nicht gelöscht werden.')
    return redirect(_backup_redirect_target())


@app.route('/admin/backup/send-mail', methods=['POST'])
@admin_required
def admin_backup_send_mail():
    """Sendet ein Backup sofort per Mail an die konfigurierte Adresse."""
    db = get_db()
    settings = get_backup_settings(db)
    recipient = (request.form.get('backup_email_to') or settings['email_to']).strip()
    max_mb = _setting_int(request.form.get('backup_email_max_mb'), settings['email_max_mb'], 1, 2000)
    email_path = None
    try:
        if not _is_valid_email(recipient):
            raise RuntimeError('Empfängeradresse für Backup-Mail ist ungültig.')
        email_path, zip_filename = create_local_backup('eeg_mail')
        send_backup_email(db, email_path, recipient, max_mb)
        audit_log('backup_email_manual', f'Backup-Mail manuell versendet: {zip_filename} an {recipient}')
        flash(f'Backup-Mail wurde an {recipient} gesendet.', 'success')
    except Exception as e:
        audit_log('backup_email_manual_failed', f'Manuelle Backup-Mail fehlgeschlagen: {e}')
        flash_exception(e, 'Backup-Mail konnte nicht gesendet werden.')
    finally:
        if email_path and os.path.exists(email_path):
            try:
                os.remove(email_path)
            except OSError:
                app.logger.warning('Could not remove temporary mail backup %s', email_path, exc_info=True)
    return redirect(_backup_redirect_target())


@app.route('/backup')
@admin_required
def backup_download():
    """Erstellt ein ZIP-Backup (DB + Rechnungs-PDFs) zum Download."""
    import tempfile

    timestamp = local_now().strftime('%Y%m%d_%H%M%S')
    zip_filename = f"eeg_backup_{timestamp}.zip"

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
    tmp.close()
    write_backup_zip(tmp.name)

    audit_log('backup_download', f'Backup heruntergeladen: {zip_filename}')

    response = send_file(tmp.name, as_attachment=True, download_name=zip_filename,
                         mimetype='application/zip')

    @response.call_on_close
    def cleanup_backup_file():
        try:
            os.unlink(tmp.name)
        except OSError:
            pass

    return response


@app.route('/backup/restore', methods=['POST'])
@admin_required
def backup_restore():
    """Stellt ein Backup aus einem ZIP-File wieder her."""
    import zipfile, tempfile, shutil

    if 'backup_file' not in request.files:
        flash('Keine Datei ausgewählt.', 'danger')
        return redirect(url_for('admin_backup'))

    file = request.files['backup_file']
    if not file.filename.lower().endswith('.zip'):
        flash('Nur ZIP-Dateien sind erlaubt.', 'danger')
        return redirect(url_for('admin_backup'))
    if request.form.get('restore_confirm') != '1':
        flash('Bitte bestätigen Sie die Wiederherstellung ausdrücklich.', 'danger')
        return redirect(url_for('admin_backup'))

    # Temporär speichern
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
    file.save(tmp.name)
    tmp.close()

    try:
        if not zipfile.is_zipfile(tmp.name):
            flash('Ungültiges Backup: Die Datei ist kein lesbares ZIP-Archiv.', 'danger')
            return redirect(url_for('admin_backup'))

        with zipfile.ZipFile(tmp.name, 'r') as zf:
            names = validate_backup_zip(zf)

            # DB schließen
            close_db()

            # DB ersetzen
            target_db = safe_extract_zip_member(zf, 'eeg_data.db', os.path.dirname(DB_PATH))
            if os.path.abspath(target_db) != os.path.abspath(DB_PATH):
                shutil.move(target_db, DB_PATH)

            # PDFs wiederherstellen
            for name in names:
                if name.startswith('invoices/') and name != 'invoices/':
                    safe_extract_zip_member(zf, name, os.path.dirname(INVOICE_FOLDER))

        # WAL-Dateien entfernen falls vorhanden
        for suffix in ('-wal', '-shm'):
            wal_file = DB_PATH + suffix
            if os.path.exists(wal_file):
                os.remove(wal_file)

        audit_log('backup_restore', f'Backup wiederhergestellt aus: {file.filename}')
        flash('Backup erfolgreich wiederhergestellt. Bitte Server neu starten.', 'success')
    except Exception as e:
        audit_log('backup_restore_failed', f'Backup-Wiederherstellung fehlgeschlagen: {e}')
        flash_exception(e, 'Wiederherstellung fehlgeschlagen.')
    finally:
        os.unlink(tmp.name)

    return redirect(url_for('admin_backup'))


# === Datenbank-Wartung ===

def _quote_identifier(name):
    return '"' + str(name).replace('"', '""') + '"'


def _table_exists(conn, table_name):
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,)
    ).fetchone()
    return row is not None


def _column_exists(conn, table_name, column_name):
    if not _table_exists(conn, table_name):
        return False
    return any(row['name'] == column_name for row in conn.execute(f"PRAGMA table_info({_quote_identifier(table_name)})"))


def get_database_stats():
    db_size = os.path.getsize(DB_PATH) if os.path.exists(DB_PATH) else 0
    wal_path = DB_PATH + '-wal'
    shm_path = DB_PATH + '-shm'
    stats = {
        'db_path': DB_PATH,
        'db_size': db_size,
        'wal_size': os.path.getsize(wal_path) if os.path.exists(wal_path) else 0,
        'shm_size': os.path.getsize(shm_path) if os.path.exists(shm_path) else 0,
        'page_count': 0,
        'page_size': 0,
        'freelist_count': 0,
        'fragmentation_mb': 0,
        'tables': [],
    }
    if not os.path.exists(DB_PATH):
        return stats

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        stats['page_count'] = conn.execute("PRAGMA page_count").fetchone()[0]
        stats['page_size'] = conn.execute("PRAGMA page_size").fetchone()[0]
        stats['freelist_count'] = conn.execute("PRAGMA freelist_count").fetchone()[0]
        stats['fragmentation_mb'] = stats['freelist_count'] * stats['page_size'] / 1024 / 1024
        tables = conn.execute("""
            SELECT name FROM sqlite_master
            WHERE type='table' AND name NOT LIKE 'sqlite_%'
            ORDER BY name
        """).fetchall()
        for table in tables:
            name = table['name']
            count = conn.execute(f"SELECT COUNT(*) FROM {_quote_identifier(name)}").fetchone()[0]
            stats['tables'].append({'name': name, 'count': count})
    return stats


def _quality_result(title, status, detail, count=None):
    return {
        'title': title,
        'status': status,
        'detail': detail,
        'count': count,
    }


def _quality_count(conn, title, sql, error_detail, ok_detail='Keine Auffälligkeiten gefunden.'):
    count = conn.execute(sql).fetchone()[0]
    status = 'ok' if count == 0 else 'warning'
    detail = ok_detail if count == 0 else error_detail
    return _quality_result(title, status, detail, count)


def run_database_quality_check():
    results = []
    with BACKUP_JOB_LOCK:
        with sqlite3.connect(DB_PATH, timeout=60) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA foreign_keys=ON")

            integrity_rows = [row[0] for row in conn.execute("PRAGMA integrity_check").fetchall()]
            if integrity_rows == ['ok']:
                results.append(_quality_result('SQLite Integritätsprüfung', 'ok', 'Datenbankdatei ist konsistent.', 0))
            else:
                results.append(_quality_result('SQLite Integritätsprüfung', 'error', '; '.join(integrity_rows[:5]), len(integrity_rows)))

            fk_rows = conn.execute("PRAGMA foreign_key_check").fetchall()
            if fk_rows:
                results.append(_quality_result('Fremdschlüsselprüfung', 'error', f'{len(fk_rows)} verletzte Referenzen gefunden.', len(fk_rows)))
            else:
                results.append(_quality_result('Fremdschlüsselprüfung', 'ok', 'Keine verletzten Fremdschlüssel gefunden.', 0))

            if _table_exists(conn, 'invoice_items'):
                results.append(_quality_count(
                    conn,
                    'Abrechnungspositionen ohne Abrechnung',
                    """SELECT COUNT(*) FROM invoice_items ii
                       LEFT JOIN invoices i ON i.id = ii.invoice_id
                       WHERE i.id IS NULL""",
                    'Abrechnungspositionen verweisen auf gelöschte oder fehlende Abrechnungen.'
                ))
                results.append(_quality_count(
                    conn,
                    'Abrechnungspositionen ohne Mitglied',
                    """SELECT COUNT(*) FROM invoice_items ii
                       LEFT JOIN members m ON m.id = ii.member_id
                       WHERE m.id IS NULL""",
                    'Abrechnungspositionen verweisen auf gelöschte oder fehlende Mitglieder.'
                ))

            if _table_exists(conn, 'email_log'):
                results.append(_quality_count(
                    conn,
                    'E-Mail-Log ohne Abrechnung',
                    """SELECT COUNT(*) FROM email_log el
                       LEFT JOIN invoices i ON i.id = el.invoice_id
                       WHERE el.invoice_id IS NOT NULL AND i.id IS NULL""",
                    'E-Mail-Protokolle verweisen auf fehlende Abrechnungen.'
                ))
                results.append(_quality_count(
                    conn,
                    'E-Mail-Log ohne Mitglied',
                    """SELECT COUNT(*) FROM email_log el
                       LEFT JOIN members m ON m.id = el.member_id
                       WHERE el.member_id IS NOT NULL AND m.id IS NULL""",
                    'E-Mail-Protokolle verweisen auf fehlende Mitglieder.'
                ))

            if _table_exists(conn, 'contracts'):
                results.append(_quality_count(
                    conn,
                    'Verträge ohne Mitglied',
                    """SELECT COUNT(*) FROM contracts c
                       LEFT JOIN members m ON m.id = c.member_id
                       WHERE m.id IS NULL""",
                    'Verträge verweisen auf fehlende Mitglieder.'
                ))

            if _table_exists(conn, 'newsletter_log'):
                results.append(_quality_count(
                    conn,
                    'Newsletter-Log ohne Newsletter',
                    """SELECT COUNT(*) FROM newsletter_log nl
                       LEFT JOIN newsletters n ON n.id = nl.newsletter_id
                       WHERE n.id IS NULL""",
                    'Newsletter-Protokolle verweisen auf fehlende Newsletter.'
                ))
                results.append(_quality_count(
                    conn,
                    'Newsletter-Log ohne Mitglied',
                    """SELECT COUNT(*) FROM newsletter_log nl
                       LEFT JOIN members m ON m.id = nl.member_id
                       WHERE m.id IS NULL""",
                    'Newsletter-Protokolle verweisen auf fehlende Mitglieder.'
                ))

            if _column_exists(conn, 'users', 'member_id'):
                results.append(_quality_count(
                    conn,
                    'Benutzer ohne zugeordnetes Mitglied',
                    """SELECT COUNT(*) FROM users u
                       LEFT JOIN members m ON m.id = u.member_id
                       WHERE u.member_id IS NOT NULL AND m.id IS NULL""",
                    'Benutzerkonten verweisen auf fehlende Mitglieder.'
                ))

            if _table_exists(conn, 'members'):
                results.append(_quality_count(
                    conn,
                    'Aktive Mitglieder ohne Zählpunkt',
                    """SELECT COUNT(*) FROM members
                       WHERE active=1
                         AND COALESCE(TRIM(bezug_zp), '') = ''
                         AND COALESCE(TRIM(einspeiser_zp), '') = ''""",
                    'Aktive Mitglieder ohne Bezugs- oder Einspeise-Zählpunkt gefunden.',
                    ok_detail='Alle aktiven Mitglieder haben mindestens einen Zählpunkt.'
                ))

            if _table_exists(conn, 'measurements'):
                results.append(_quality_count(
                    conn,
                    'Messwerte ohne Import-Batch',
                    """SELECT COUNT(*) FROM measurements m
                       LEFT JOIN import_batches b ON b.id = m.batch_id
                       WHERE b.id IS NULL""",
                    'Messwerte verweisen auf fehlende Import-Batches.'
                ))
                results.append(_quality_count(
                    conn,
                    'Messwerte ohne Meter-Code',
                    """SELECT COUNT(*) FROM measurements m
                       LEFT JOIN meter_codes mc ON mc.id = m.meter_code_id
                       WHERE mc.id IS NULL""",
                    'Messwerte verweisen auf fehlende Meter-Codes.'
                ))
                results.append(_quality_count(
                    conn,
                    'Messwerte mit ungültigem Zeitintervall',
                    """SELECT COUNT(*) FROM measurements
                       WHERE timestamp_start >= timestamp_end OR interval_minutes <= 0""",
                    'Messwerte mit ungültigem Zeitraum oder Intervall gefunden.'
                ))
                results.append(_quality_count(
                    conn,
                    'Messwerte mit negativer Energie',
                    "SELECT COUNT(*) FROM measurements WHERE value_kwh < 0",
                    'Negative kWh-Werte gefunden.'
                ))
                results.append(_quality_count(
                    conn,
                    'Messwerte ohne Qualitätskennzeichen',
                    "SELECT COUNT(*) FROM measurements WHERE COALESCE(TRIM(quality), '') = ''",
                    'Messwerte ohne Qualitätskennzeichen gefunden.'
                ))

            if _table_exists(conn, 'overview_totals'):
                results.append(_quality_count(
                    conn,
                    'Übersichtswerte ohne Import-Batch',
                    """SELECT COUNT(*) FROM overview_totals ot
                       LEFT JOIN import_batches b ON b.id = ot.batch_id
                       WHERE b.id IS NULL""",
                    'Übersichtswerte verweisen auf fehlende Import-Batches.'
                ))
                results.append(_quality_count(
                    conn,
                    'Übersichtswerte ohne Meter-Code',
                    """SELECT COUNT(*) FROM overview_totals ot
                       LEFT JOIN meter_codes mc ON mc.id = ot.meter_code_id
                       WHERE mc.id IS NULL""",
                    'Übersichtswerte verweisen auf fehlende Meter-Codes.'
                ))

    has_error = any(item['status'] == 'error' for item in results)
    has_warning = any(item['status'] == 'warning' for item in results)
    summary = 'Fehler gefunden' if has_error else ('Auffälligkeiten gefunden' if has_warning else 'Keine Fehler gefunden')
    return {
        'checked_at': local_now(),
        'summary': summary,
        'status': 'error' if has_error else ('warning' if has_warning else 'ok'),
        'results': results,
    }


def run_database_maintenance(action):
    action_labels = {
        'checkpoint': 'WAL-Checkpoint',
        'analyze': 'Statistiken aktualisieren',
        'optimize': 'SQLite optimieren',
        'vacuum': 'Defragmentierung',
        'full': 'Komplette Wartung',
    }
    if action not in action_labels:
        raise ValueError('Unbekannte Wartungsaktion.')

    backup_filename = None
    if action in ('vacuum', 'full'):
        _, backup_filename = create_local_backup('eeg_manual')

    before = get_database_stats()
    close_db()
    with BACKUP_JOB_LOCK:
        with sqlite3.connect(DB_PATH, timeout=120, isolation_level=None) as conn:
            conn.execute("PRAGMA foreign_keys=ON")
            if action in ('checkpoint', 'full'):
                conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            if action in ('vacuum', 'full'):
                conn.execute("VACUUM")
            if action in ('analyze', 'full'):
                conn.execute("ANALYZE")
            if action in ('optimize', 'full'):
                conn.execute("PRAGMA optimize")
    after = get_database_stats()
    return {
        'action': action,
        'label': action_labels[action],
        'backup_filename': backup_filename,
        'before_size': before['db_size'],
        'after_size': after['db_size'],
        'before_fragmentation': before['fragmentation_mb'],
        'after_fragmentation': after['fragmentation_mb'],
    }


@app.route('/admin/database')
@admin_required
def admin_database():
    """Admin-Seite fuer Datenbank-Wartung und Qualitaetscheck."""
    return render_template('admin_database.html', stats=get_database_stats())


@app.route('/admin/database/check', methods=['POST'])
@admin_required
def admin_database_check():
    """Führt Integritäts- und Plausibilitätsprüfungen aus."""
    try:
        check_result = run_database_quality_check()
        audit_log('database_quality_check', check_result['summary'])
        flash(f'Datenbank-Qualitätscheck abgeschlossen: {check_result["summary"]}.', 'success' if check_result['status'] == 'ok' else 'warning')
    except Exception as e:
        check_result = None
        audit_log('database_quality_check_failed', f'Datenbank-Qualitätscheck fehlgeschlagen: {e}')
        flash_exception(e, 'Qualitätscheck fehlgeschlagen.')
    return render_template('admin_database.html', stats=get_database_stats(), check_result=check_result)


@app.route('/admin/database/maintenance', methods=['POST'])
@admin_required
def admin_database_maintenance():
    """Führt ausgewählte SQLite-Wartungsaktionen aus."""
    action = request.form.get('maintenance_action', '')
    try:
        result = run_database_maintenance(action)
        detail = f'{result["label"]} ausgeführt'
        if result['backup_filename']:
            detail += f' (Sicherungsbackup: {result["backup_filename"]})'
        audit_log('database_maintenance', detail)
        flash(f'{result["label"]} erfolgreich abgeschlossen.', 'success')
    except Exception as e:
        result = None
        audit_log('database_maintenance_failed', f'Datenbank-Wartung fehlgeschlagen: {e}')
        flash_exception(e, 'Datenbank-Wartung fehlgeschlagen.')
    return render_template('admin_database.html', stats=get_database_stats(), maintenance_result=result)


# === Überweisungsliste / Forderungen ===

def _parse_booking_date(value):
    text = (value or '').strip()
    if not text:
        raise ValueError('Bitte ein Buchungsdatum eintragen.')
    try:
        booking_date = date.fromisoformat(text)
    except ValueError as e:
        raise ValueError('Das Buchungsdatum ist ungültig.') from e
    if booking_date > local_now().date():
        raise ValueError('Das Buchungsdatum darf nicht in der Zukunft liegen.')
    if booking_date < date(2000, 1, 1):
        raise ValueError('Das Buchungsdatum ist zu weit in der Vergangenheit.')
    return booking_date


def _paid_at_from_booking_date(booking_date):
    booked_at = datetime.combine(booking_date, datetime.min.time().replace(hour=12), tzinfo=APP_TIMEZONE)
    return booked_at.isoformat(timespec='seconds')


def _row_reference_date(row):
    for key in ('finalized_at', 'created_at', 'period_to'):
        dt = to_local_datetime(row.get(key) if hasattr(row, 'get') else row[key])
        if dt:
            return dt.date()
    return local_now().date()


def _booking_due_date(row):
    return _row_reference_date(row) + timedelta(days=7)


def _active_payment_bookings(db, member_id=None):
    """Nicht stornierte Buchungen, gruppiert nach Abrechnung und Mitglied.

    Je Abrechnungszeile sind mehrere Buchungen moeglich: Teilzahlungen,
    Nachzahlungen und die Rueckerstattung eines Guthabens.
    """
    params = []
    member_filter = ''
    if member_id is not None:
        member_filter = 'AND member_id=?'
        params.append(member_id)
    rows = db.execute(f"""
        SELECT id, invoice_id, member_id, amount_eur, direction, booking_date,
               note, recorded_at, recorded_by_username
        FROM payment_bookings
        WHERE reversed_at IS NULL {member_filter}
        ORDER BY booking_date, id
    """, params).fetchall()
    changes = _booking_change_map(db, member_id=member_id)
    grouped = {}
    for row in rows:
        booking = dict(row)
        booking['changes'] = changes.get(booking['id'], [])
        grouped.setdefault((row['invoice_id'], row['member_id']), []).append(booking)
    return grouped


def _booking_change_map(db, member_id=None):
    """Aenderungshistorie je Buchung, aelteste zuerst."""
    params = []
    member_filter = ''
    if member_id is not None:
        member_filter = 'WHERE b.member_id=?'
        params.append(member_id)
    rows = db.execute(f"""
        SELECT c.id, c.booking_id, c.changed_at, c.changed_by_username, c.action,
               c.old_amount_eur, c.new_amount_eur, c.old_booking_date,
               c.new_booking_date, c.reason
        FROM payment_booking_changes c
        JOIN payment_bookings b ON b.id = c.booking_id
        {member_filter}
        ORDER BY c.id
    """, params).fetchall()
    grouped = {}
    for row in rows:
        grouped.setdefault(row['booking_id'], []).append(dict(row))
    return grouped


def _record_booking_change(db, booking_id, action, reason, old=None, new=None):
    """Schreibt einen Eintrag in die Aenderungshistorie einer Buchung."""
    old = old or {}
    new = new or {}
    db.execute("""
        INSERT INTO payment_booking_changes (
            booking_id, changed_by_user_id, changed_by_username, action,
            old_amount_eur, new_amount_eur, old_booking_date, new_booking_date, reason
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        booking_id,
        current_user.id if current_user.is_authenticated else None,
        current_user.username if current_user.is_authenticated else None,
        action,
        old.get('amount'), new.get('amount'),
        old.get('date'), new.get('date'),
        reason or '',
    ))


MIN_CHANGE_REASON_LENGTH = 5


def _require_change_reason(value):
    """Der Aenderungsgrund ist Pflicht und wird in der Historie festgehalten."""
    reason = (value or '').strip()
    if len(reason) < MIN_CHANGE_REASON_LENGTH:
        raise ValueError(f'Bitte einen Änderungsgrund mit mindestens '
                         f'{MIN_CHANGE_REASON_LENGTH} Zeichen angeben.')
    return reason[:500]


def _parse_booking_amount(value):
    """Betrag mit Komma oder Punkt; None, wenn nichts eingegeben wurde."""
    text = (value or '').strip().replace('€', '').replace(' ', '')
    if not text:
        return None
    try:
        return round(float(text.replace(',', '.')), 2)
    except ValueError as e:
        raise ValueError('Der Buchungsbetrag ist ungültig.') from e


def _iso_to_date(value):
    try:
        return date.fromisoformat(str(value)[:10])
    except (TypeError, ValueError):
        return None


def _apply_booking_state(db, invoice_id, member_id, booking_date=None):
    """Setzt das Bezahlt-Kennzeichen der Positionen nach jeder Buchungsaenderung."""
    row = get_payment_row(db, invoice_id, member_id)
    if not row:
        return None
    if row['paid']:
        reference = booking_date or _iso_to_date(row['booking_date']) or local_now().date()
        db.execute("""UPDATE invoice_items SET paid=1, paid_at=?
                      WHERE invoice_id=? AND member_id=?""",
                   (_paid_at_from_booking_date(reference), invoice_id, member_id))
    else:
        db.execute("""UPDATE invoice_items SET paid=0, paid_at=NULL
                      WHERE invoice_id=? AND member_id=?""", (invoice_id, member_id))
    return row


def get_payment_rows(db, member_id=None):
    """Liefert Netto-Zahlungszeilen pro Mitglied und Abrechnung."""
    params = []
    member_filter = ''
    if member_id is not None:
        member_filter = 'WHERE ii.member_id=?'
        params.append(member_id)
    items = db.execute(f"""
        SELECT ii.id, ii.invoice_id, ii.member_id, m.name, m.iban, m.bic, m.account_holder,
               i.period_from, i.period_to, i.status AS invoice_status, i.created_at, i.finalized_at,
               ii.type, ii.kwh, ii.amount_eur, COALESCE(ii.paid, 0) AS paid, ii.paid_at
        FROM invoice_items ii
        JOIN members m ON m.id = ii.member_id
        JOIN invoices i ON i.id = ii.invoice_id
        {member_filter}
        ORDER BY i.period_from DESC, m.name
    """, params).fetchall()

    from collections import defaultdict
    grouped = defaultdict(lambda: {'items': [], 'carryovers': [], 'member': None, 'invoice': None})
    latest_period_by_member = {}

    def ensure_group(invoice_id, member_id_value, member_data, invoice_data):
        key = (invoice_id, member_id_value)
        grouped[key]['member'] = member_data
        grouped[key]['invoice'] = invoice_data
        current_latest = latest_period_by_member.get(member_id_value)
        if current_latest is None or invoice_data['period_to'] > current_latest:
            latest_period_by_member[member_id_value] = invoice_data['period_to']
        return key

    for item in items:
        key = ensure_group(
            item['invoice_id'],
            item['member_id'],
            {
                'id': item['member_id'],
                'name': item['name'],
                'iban': item['iban'],
                'bic': item['bic'],
                'account_holder': item['account_holder'],
            },
            {
                'id': item['invoice_id'],
                'period_from': item['period_from'],
                'period_to': item['period_to'],
                'status': item['invoice_status'],
                'created_at': item['created_at'],
                'finalized_at': item['finalized_at'],
            },
        )
        grouped[key]['items'].append(item)

    carryover_params = []
    carryover_member_filter = ''
    if member_id is not None:
        carryover_member_filter = 'WHERE c.member_id=?'
        carryover_params.append(member_id)
    carryover_rows = db.execute(f"""
        SELECT c.*, m.name, m.iban, m.bic, m.account_holder,
               i.period_from, i.period_to, i.status AS invoice_status, i.created_at, i.finalized_at,
               src.period_from AS source_period_from,
               src.period_to AS source_period_to
        FROM invoice_carryovers c
        JOIN members m ON m.id = c.member_id
        JOIN invoices i ON i.id = c.invoice_id
        JOIN invoices src ON src.id = c.source_invoice_id
        {carryover_member_filter}
        ORDER BY i.period_from DESC, m.name, src.period_from
    """, carryover_params).fetchall()
    for carryover in carryover_rows:
        key = ensure_group(
            carryover['invoice_id'],
            carryover['member_id'],
            {
                'id': carryover['member_id'],
                'name': carryover['name'],
                'iban': carryover['iban'],
                'bic': carryover['bic'],
                'account_holder': carryover['account_holder'],
            },
            {
                'id': carryover['invoice_id'],
                'period_from': carryover['period_from'],
                'period_to': carryover['period_to'],
                'status': carryover['invoice_status'],
                'created_at': carryover['created_at'],
                'finalized_at': carryover['finalized_at'],
            },
        )
        grouped[key]['carryovers'].append(carryover)

    payment_rows = []
    today = local_now().date()
    bookings_map = _active_payment_bookings(db, member_id=member_id)
    for key, data in grouped.items():
        energy_net = 0
        for item in data['items']:
            if item['type'] == 'consumption':
                energy_net += item['amount_eur']
            else:
                energy_net -= item['amount_eur']
        carryover_total = round(sum(item['amount_eur'] for item in data['carryovers']), 2)
        net = round(energy_net + carryover_total, 2)
        bookings = bookings_map.get(key, [])
        booked_total = round(sum(booking['amount_eur'] for booking in bookings), 2)
        all_paid = bool(data['items']) and all(item['paid'] for item in data['items'])
        if bookings:
            open_amount = round(net - booked_total, 2)
        elif all_paid:
            # Altbestand: vor Einfuehrung der Buchungssaetze bezahlt.
            booked_total = net
            open_amount = 0.0
        else:
            open_amount = net
        paid = abs(open_amount) < 0.005
        paid_at = data['items'][0]['paid_at'] if all_paid and data['items'] else None
        last_booking = bookings[-1] if bookings else None
        booking_date = last_booking['booking_date'] if last_booking else ''
        if not booking_date and paid_at:
            paid_dt = to_local_datetime(paid_at)
            booking_date = paid_dt.date().isoformat() if paid_dt else ''
        carried_forward = db.execute("""
            SELECT c.invoice_id, i.period_from, i.period_to
            FROM invoice_carryovers c
            JOIN invoices i ON i.id = c.invoice_id
            WHERE c.source_invoice_id=? AND c.member_id=?
            ORDER BY i.period_from DESC, c.invoice_id DESC
            LIMIT 1
        """, (key[0], key[1])).fetchone()
        row = {
            'invoice_id': key[0],
            'member_id': key[1],
            'member_name': data['member']['name'],
            'iban': data['member']['iban'],
            'bic': data['member']['bic'],
            'account_holder': data['member']['account_holder'],
            'period_from': data['invoice']['period_from'],
            'period_to': data['invoice']['period_to'],
            'invoice_status': data['invoice']['status'],
            'created_at': data['invoice']['created_at'],
            'finalized_at': data['invoice']['finalized_at'],
            'reference_date': _row_reference_date(data['invoice']),
            'due_on': _booking_due_date(data['invoice']),
            'net_total': round(net, 2),
            'energy_total': round(energy_net, 2),
            'carryover_total': carryover_total,
            'carryovers': data['carryovers'],
            'bookings': bookings,
            'booked_total': booked_total,
            'open_amount': open_amount,
            'is_partially_booked': bool(bookings) and not paid,
            'paid': paid,
            'paid_at': paid_at,
            'booking_date': booking_date,
            'booking_note': last_booking['note'] if last_booking else '',
            'booking_id': last_booking['id'] if last_booking else None,
            'direction': 'member_to_eeg' if net > 0 else 'eeg_to_member' if net < 0 else 'balanced',
            'open_direction': ('member_to_eeg' if open_amount > 0
                               else 'eeg_to_member' if open_amount < 0 else 'balanced'),
            'is_settled_by_carryover': bool(carried_forward and not paid),
            'carried_forward_to_invoice_id': carried_forward['invoice_id'] if carried_forward and not paid else None,
        }
        row['is_overdue'] = (not row['paid'] and not row['is_settled_by_carryover'] and row['open_amount'] > 0 and today >= row['due_on'])
        row['is_previous_period_open'] = (
            not row['paid']
            and not row['is_settled_by_carryover']
            and latest_period_by_member.get(row['member_id']) is not None
            and row['period_to'] < latest_period_by_member[row['member_id']]
        )
        payment_rows.append(row)

    payment_rows.sort(key=lambda item: (
        item['paid'],
        not item['is_overdue'],
        item['member_name'].lower(),
        item['period_from'],
    ))
    return payment_rows


def get_payment_row(db, invoice_id, member_id):
    for row in get_payment_rows(db, member_id=member_id):
        if row['invoice_id'] == invoice_id:
            return row
    return None


def get_member_account_summary(db, member_id):
    rows = get_payment_rows(db, member_id=member_id)
    active_open_rows = [row for row in rows if not row['paid'] and not row['is_settled_by_carryover']]
    balance = round(sum(row['open_amount'] for row in active_open_rows), 2)
    open_claims = round(sum(row['open_amount'] for row in active_open_rows if row['open_amount'] > 0), 2)
    open_credits = round(sum(row['open_amount'] for row in active_open_rows if row['open_amount'] < 0), 2)
    overdue_claims = round(sum(row['open_amount'] for row in rows if row['is_overdue']), 2)
    previous_open = [row for row in rows if row['is_previous_period_open']]

    events = []
    for row in rows:
        invoice_date = _row_reference_date(row)
        net = row['net_total']
        events.append({
            'sort_date': invoice_date,
            'date': invoice_date,
            'kind': 'invoice',
            'label': f"Abrechnung {row['period_from']} - {row['period_to']}",
            'invoice_id': row['invoice_id'],
            'amount': net,
            'status': 'gebucht' if row['paid'] else 'vorgetragen' if row['is_settled_by_carryover'] else 'offen',
            'is_overdue': row['is_overdue'],
            'is_previous_period_open': row['is_previous_period_open'],
        })
        if row['is_settled_by_carryover']:
            events.append({
                'sort_date': invoice_date,
                'date': invoice_date,
                'kind': 'carryover',
                'label': f"Vortrag in Abrechnung #{row['carried_forward_to_invoice_id']}",
                'invoice_id': row['carried_forward_to_invoice_id'],
                'amount': -row['open_amount'],
                'status': 'vorgetragen',
                'is_overdue': False,
                'is_previous_period_open': False,
            })
        for booking in row['bookings']:
            booking_date = _iso_to_date(booking['booking_date']) or invoice_date
            amount = booking['amount_eur']
            events.append({
                'sort_date': booking_date,
                'date': booking_date,
                'kind': 'booking',
                'label': 'Zahlung gebucht' if amount > 0 else 'Gutschrift ausbezahlt',
                'invoice_id': row['invoice_id'],
                'amount': -amount,
                'status': 'gebucht',
                'is_overdue': False,
                'is_previous_period_open': False,
            })
        if not row['bookings'] and row['paid'] and abs(net) >= 0.005:
            # Altbestand ohne Buchungssatz
            booking_date = _iso_to_date(row['booking_date']) or invoice_date
            events.append({
                'sort_date': booking_date,
                'date': booking_date,
                'kind': 'booking',
                'label': 'Zahlung gebucht' if net > 0 else 'Gutschrift ausbezahlt',
                'invoice_id': row['invoice_id'],
                'amount': -net,
                'status': 'gebucht',
                'is_overdue': False,
                'is_previous_period_open': False,
            })

    running = 0
    for event in sorted(events, key=lambda item: (item['sort_date'], 0 if item['kind'] == 'invoice' else 1, item['invoice_id'])):
        running = round(running + event['amount'], 2)
        event['balance_after'] = running

    return {
        'balance': balance,
        'open_claims': open_claims,
        'open_credits': open_credits,
        'overdue_claims': overdue_claims,
        'previous_open': previous_open,
        'rows': rows,
        'history': list(reversed(events)),
    }


def get_member_account_overview(db):
    """Kontosaldo je Mitglied fuer die Adminuebersicht.

    Positiver Saldo: das Mitglied schuldet der EEG Geld.
    Negativer Saldo: die EEG schuldet dem Mitglied Geld.
    """
    rows = get_payment_rows(db)
    accounts = {}
    for row in rows:
        account = accounts.setdefault(row['member_id'], {
            'member_id': row['member_id'],
            'member_name': row['member_name'],
            'iban': row['iban'],
            'account_holder': row['account_holder'],
            'invoiced_total': 0,
            'booked_total': 0,
            'balance': 0,
            'overdue': 0,
            'open_rows': 0,
            'deviating_rows': 0,
            'carried_rows': 0,
            'last_booking_date': '',
        })
        account['invoiced_total'] += _cents(row['net_total'])
        account['booked_total'] += _cents(row['booked_total'])
        if row['is_settled_by_carryover']:
            account['carried_rows'] += 1
        elif not row['paid']:
            account['balance'] += _cents(row['open_amount'])
            account['open_rows'] += 1
            if row['is_partially_booked']:
                account['deviating_rows'] += 1
        if row['is_overdue']:
            account['overdue'] += _cents(row['open_amount'])
        if row['booking_date'] > account['last_booking_date']:
            account['last_booking_date'] = row['booking_date']

    result = []
    for account in accounts.values():
        for key in ('invoiced_total', 'booked_total', 'balance', 'overdue'):
            account[key] = account[key] / 100
        result.append(account)
    result.sort(key=lambda item: (-abs(item['balance']), item['member_name'].lower()))
    return result


def _payment_redirect_target():
    next_url = request.form.get('next') or request.args.get('next')
    if next_url and is_safe_redirect_url(next_url):
        return next_url
    return url_for('payments')


BOOKED_SORT_KEYS = {
    'member': lambda entry: (entry['row']['member_name'].lower(), entry['booking_date']),
    'invoice': lambda entry: (entry['row']['invoice_id'], entry['row']['member_name'].lower()),
    'period': lambda entry: (entry['row']['period_from'], entry['row']['member_name'].lower()),
    'amount': lambda entry: (entry['amount'], entry['booking_date']),
    'abs_amount': lambda entry: (abs(entry['amount']), entry['booking_date']),
    'date': lambda entry: (entry['booking_date'], entry['row']['member_name'].lower()),
}


def booked_payment_entries(rows, sort='date', direction='desc'):
    """Flache, sortierbare Liste aller gebuchten Zahlungen, eine Zeile je Buchung.

    Altbestaende ohne Buchungssatz erscheinen mit ihrem abgeleiteten Betrag.
    """
    entries = []
    for row in rows:
        for position, booking in enumerate(row['bookings'], start=1):
            entries.append({
                'kind': 'booking',
                'row': row,
                'booking': booking,
                'position': position,
                'total_bookings': len(row['bookings']),
                'amount': booking['amount_eur'],
                'booking_date': booking['booking_date'] or '',
            })
        if not row['bookings'] and row['paid'] and abs(row['net_total']) >= 0.005:
            entries.append({
                'kind': 'legacy',
                'row': row,
                'booking': None,
                'position': 1,
                'total_bookings': 1,
                'amount': row['net_total'],
                'booking_date': row['booking_date'] or '',
            })
    entries.sort(key=BOOKED_SORT_KEYS.get(sort, BOOKED_SORT_KEYS['date']),
                 reverse=(direction == 'desc'))
    return entries


@app.route('/payments')
@admin_required
def payments():
    """Überweisungsliste: offene und bezahlte Forderungen."""
    db = get_db()
    payment_list = get_payment_rows(db)
    booked_sort = request.args.get('booked_sort', 'date')
    booked_dir = request.args.get('booked_dir', 'desc')
    if booked_sort not in BOOKED_SORT_KEYS:
        booked_sort = 'date'
    if booked_dir not in {'asc', 'desc'}:
        booked_dir = 'desc'
    return render_template('payments.html', payments=payment_list,
                           booked_entries=booked_payment_entries(payment_list, booked_sort, booked_dir),
                           booked_sort=booked_sort,
                           booked_dir=booked_dir,
                           reference_for=payment_transfer_reference,
                           sepa_text=_sepa_text,
                           min_change_reason_length=MIN_CHANGE_REASON_LENGTH,
                           today=local_now().date().isoformat())


@app.route('/mitgliederkonten')
@admin_required
def member_accounts():
    """Kontosaldo aller Mitglieder."""
    db = get_db()
    accounts = get_member_account_overview(db)
    totals = {
        'claims': round(sum(a['balance'] for a in accounts if a['balance'] > 0), 2),
        'credits': round(sum(a['balance'] for a in accounts if a['balance'] < 0), 2),
        'balance': round(sum(a['balance'] for a in accounts), 2),
        'overdue': round(sum(a['overdue'] for a in accounts), 2),
    }
    return render_template('member_accounts.html', accounts=accounts, totals=totals)


@app.route('/mitgliederkonten/<int:id>')
@admin_required
def member_account_detail(id):
    """Kontoauszug eines Mitglieds mit allen Buchungen."""
    db = get_db()
    member = db.execute("SELECT * FROM members WHERE id=?", (id,)).fetchone()
    if not member:
        flash('Mitglied nicht gefunden.', 'danger')
        return redirect(url_for('member_accounts'))
    return render_template('member_account_detail.html',
                           member=member,
                           account=get_member_account_summary(db, id),
                           today=local_now().date().isoformat())


@app.route('/payments/mark_paid', methods=['POST'])
@admin_required
def payment_mark_paid():
    """Bucht eine Forderung oder Gutschrift mit Bank-Buchungsdatum.

    Ohne Betragsangabe wird der offene Restbetrag gebucht. Weicht die
    Ueberweisung davon ab, kann der tatsaechliche Betrag eingetragen werden;
    die Differenz bleibt als Restforderung oder Guthaben am Mitgliedskonto.
    """
    db = get_db()
    invoice_id = request.form.get('invoice_id', type=int)
    member_id = request.form.get('member_id', type=int)
    try:
        booking_date = _parse_booking_date(request.form.get('booking_date'))
        note = (request.form.get('booking_note') or '').strip()[:500]
        amount = _parse_booking_amount(request.form.get('amount_eur'))
        row = get_payment_row(db, invoice_id, member_id)
        if not row:
            raise ValueError('Die Buchung wurde nicht gefunden.')
        if row['is_settled_by_carryover']:
            raise ValueError(f'Diese offene Buchung wurde bereits in Abrechnung #{row["carried_forward_to_invoice_id"]} vorgetragen.')
        if amount is None:
            if row['paid']:
                raise ValueError('Diese Buchung ist bereits vollständig gebucht.')
            amount = row['open_amount']
        if abs(amount) < 0.005:
            raise ValueError('Der Buchungsbetrag darf nicht null sein.')
        # Weicht der gebuchte Betrag vom offenen Betrag ab, ist eine Begruendung Pflicht.
        deviates = abs(amount - row['open_amount']) >= 0.005
        reason = (request.form.get('change_reason') or '').strip()[:500]
        if deviates:
            reason = _require_change_reason(request.form.get('change_reason'))

        cursor = db.execute("""
            INSERT INTO payment_bookings (
                invoice_id, member_id, amount_eur, direction, booking_date,
                recorded_by_user_id, recorded_by_username, note
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            invoice_id,
            member_id,
            amount,
            'member_to_eeg' if amount > 0 else 'eeg_to_member',
            booking_date.isoformat(),
            current_user.id if current_user.is_authenticated else None,
            current_user.username if current_user.is_authenticated else None,
            note,
        ))
        _record_booking_change(db, cursor.lastrowid, 'create', reason,
                               new={'amount': amount, 'date': booking_date.isoformat()})
        updated = _apply_booking_state(db, invoice_id, member_id, booking_date)
        db.commit()
        action_label = 'Zahlung' if amount > 0 else 'Gutschrift'
        rest = updated['open_amount'] if updated else 0.0
        audit_log('payment_paid', f'{action_label} gebucht: {row["member_name"]}, Rechnung {invoice_id}, '
                                  f'Betrag {amount:.2f} EUR, Buchungsdatum {booking_date.isoformat()}, '
                                  f'Rest {rest:.2f} EUR'
                                  + (f', Grund: {reason}' if reason else ''))
        message = (f'{action_label} über {abs(amount):.2f} € für {row["member_name"]} '
                   f'mit Buchungsdatum {booking_date.strftime("%d.%m.%Y")} gebucht.')
        if abs(rest) >= 0.005:
            message += (f' Offen bleiben {abs(rest):.2f} € '
                        f'({"Restforderung" if rest > 0 else "Guthaben des Mitglieds"}).')
            flash(message, 'warning')
        else:
            flash(message, 'success')
    except Exception as e:
        db.rollback()
        audit_log('payment_paid_failed', f'Zahlungsbuchung fehlgeschlagen: Mitglied {member_id}, Rechnung {invoice_id} ({e})')
        flash_exception(e, 'Buchung konnte nicht gespeichert werden.')
    return redirect(_payment_redirect_target())


def payment_transfer_reference(row):
    """Verwendungszweck fuer die Auszahlung einer Gutschrift."""
    year = (row['period_from'] or '')[:4]
    return f'EEG-Abr {row["invoice_id"]}/{year} Gutschrift {row["member_name"]}'


@app.route('/payments/<int:invoice_id>/<int:member_id>/qr.svg')
@admin_required
def payment_transfer_qr(invoice_id, member_id):
    """SEPA-QR-Code (GiroCode) zum Auszahlen einer Gutschrift."""
    db = get_db()
    row = get_payment_row(db, invoice_id, member_id)
    if not row:
        abort(404)
    try:
        amount = _parse_booking_amount(request.args.get('amount'))
        if amount is None:
            amount = row['open_amount']
        payload = build_epc_payload(
            recipient=row['account_holder'] or row['member_name'],
            iban=row['iban'],
            amount=abs(amount),
            remittance=payment_transfer_reference(row),
            bic=row['bic'] or '',
        )
    except ValueError as e:
        return app.response_class(str(e), status=400, mimetype='text/plain; charset=utf-8')

    response = app.response_class(render_epc_qr_svg(payload), mimetype='image/svg+xml')
    # Die Daten haengen am Betrag, deshalb nicht zwischenspeichern.
    response.headers['Cache-Control'] = 'no-store'
    return response


@app.route('/payments/legacy_booking', methods=['POST'])
@admin_required
def payment_legacy_booking():
    """Traegt fuer einen Altbestand den tatsaechlich gebuchten Betrag nach.

    Aeltere Abrechnungen wurden ohne Buchungssatz bezahlt; ihr Betrag ist nur
    aus den Positionen abgeleitet und damit nicht korrigierbar. Hier wird der
    Buchungssatz mit dem echten Betrag vom Kontoauszug nachgetragen.
    """
    db = get_db()
    invoice_id = request.form.get('invoice_id', type=int)
    member_id = request.form.get('member_id', type=int)
    try:
        row = get_payment_row(db, invoice_id, member_id)
        if not row:
            raise ValueError('Die Buchung wurde nicht gefunden.')
        if row['bookings']:
            raise ValueError('Für diese Zeile gibt es bereits einen Buchungssatz. '
                             'Bitte den Betrag dort korrigieren.')
        if not row['paid']:
            raise ValueError('Diese Zeile ist offen und kann normal gebucht werden.')
        amount = _parse_booking_amount(request.form.get('amount_eur'))
        if amount is None:
            raise ValueError('Bitte einen Buchungsbetrag angeben.')
        if abs(amount) < 0.005:
            raise ValueError('Der Buchungsbetrag darf nicht null sein.')
        booking_date = _parse_booking_date(
            request.form.get('booking_date')
            or (row['booking_date'] or local_now().date().isoformat()))
        reason = _require_change_reason(request.form.get('change_reason'))

        cursor = db.execute("""
            INSERT INTO payment_bookings (
                invoice_id, member_id, amount_eur, direction, booking_date,
                recorded_by_user_id, recorded_by_username, note
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            invoice_id,
            member_id,
            amount,
            'member_to_eeg' if amount > 0 else 'eeg_to_member',
            booking_date.isoformat(),
            current_user.id if current_user.is_authenticated else None,
            current_user.username if current_user.is_authenticated else None,
            (request.form.get('booking_note') or '').strip()[:500],
        ))
        _record_booking_change(db, cursor.lastrowid, 'migrate', reason,
                               old={'amount': row['net_total'], 'date': row['booking_date']},
                               new={'amount': amount, 'date': booking_date.isoformat()})
        updated = _apply_booking_state(db, invoice_id, member_id, booking_date)
        db.commit()
        rest = updated['open_amount'] if updated else 0.0
        audit_log('payment_legacy_booking',
                  f'Altbestand nachgetragen: {row["member_name"]}, Rechnung {invoice_id}, '
                  f'abgerechnet {row["net_total"]:.2f} EUR, gebucht {amount:.2f} EUR, '
                  f'Buchungsdatum {booking_date.isoformat()}, Rest {rest:.2f} EUR, Grund: {reason}')
        message = (f'Buchung für {row["member_name"]} mit {amount:.2f} € nachgetragen.')
        if abs(rest) >= 0.005:
            message += (f' Offen bleiben {abs(rest):.2f} € '
                        f'({"Restforderung" if rest > 0 else "Guthaben des Mitglieds"}).')
            flash(message, 'warning')
        else:
            flash(message, 'success')
    except Exception as e:
        db.rollback()
        audit_log('payment_legacy_booking_failed',
                  f'Nachtrag fehlgeschlagen: Mitglied {member_id}, Rechnung {invoice_id} ({e})')
        flash_exception(e, 'Buchung konnte nicht nachgetragen werden.')
    return redirect(_payment_redirect_target())


@app.route('/payments/bookings/<int:id>/edit', methods=['POST'])
@admin_required
def payment_booking_edit(id):
    """Korrigiert Betrag, Buchungsdatum oder Notiz einer bestehenden Buchung."""
    db = get_db()
    try:
        booking = db.execute("""
            SELECT b.*, m.name AS member_name
            FROM payment_bookings b JOIN members m ON m.id = b.member_id
            WHERE b.id=?
        """, (id,)).fetchone()
        if not booking:
            raise ValueError('Die Buchung wurde nicht gefunden.')
        if booking['reversed_at']:
            raise ValueError('Eine stornierte Buchung kann nicht mehr bearbeitet werden.')
        amount = _parse_booking_amount(request.form.get('amount_eur'))
        if amount is None:
            raise ValueError('Bitte einen Buchungsbetrag angeben.')
        if abs(amount) < 0.005:
            raise ValueError('Der Buchungsbetrag darf nicht null sein.')
        booking_date = _parse_booking_date(request.form.get('booking_date') or booking['booking_date'])
        note = (request.form.get('booking_note') or '').strip()[:500]
        reason = _require_change_reason(request.form.get('change_reason'))
        if (abs(amount - booking['amount_eur']) < 0.005
                and booking_date.isoformat() == booking['booking_date']):
            raise ValueError('Betrag und Buchungsdatum sind unverändert.')

        db.execute("""UPDATE payment_bookings
                      SET amount_eur=?, direction=?, booking_date=?, note=?
                      WHERE id=?""",
                   (amount, 'member_to_eeg' if amount > 0 else 'eeg_to_member',
                    booking_date.isoformat(), note, id))
        _record_booking_change(db, id, 'edit', reason,
                               old={'amount': booking['amount_eur'], 'date': booking['booking_date']},
                               new={'amount': amount, 'date': booking_date.isoformat()})
        updated = _apply_booking_state(db, booking['invoice_id'], booking['member_id'], booking_date)
        db.commit()
        rest = updated['open_amount'] if updated else 0.0
        audit_log('payment_booking_edit',
                  f'Buchung #{id} geändert: {booking["member_name"]}, Rechnung {booking["invoice_id"]}, '
                  f'{booking["amount_eur"]:.2f} → {amount:.2f} EUR, '
                  f'Buchungsdatum {booking["booking_date"]} → {booking_date.isoformat()}, '
                  f'Rest {rest:.2f} EUR, Grund: {reason}')
        message = f'Buchung für {booking["member_name"]} auf {amount:.2f} € geändert.'
        if abs(rest) >= 0.005:
            message += (f' Offen bleiben {abs(rest):.2f} € '
                        f'({"Restforderung" if rest > 0 else "Guthaben des Mitglieds"}).')
            flash(message, 'warning')
        else:
            flash(message, 'success')
    except Exception as e:
        db.rollback()
        audit_log('payment_booking_edit_failed', f'Buchungsänderung fehlgeschlagen: Buchung {id} ({e})')
        flash_exception(e, 'Buchung konnte nicht geändert werden.')
    return redirect(_payment_redirect_target())


@app.route('/payments/bookings/<int:id>/reverse', methods=['POST'])
@admin_required
def payment_booking_reverse(id):
    """Storniert eine einzelne Buchung, andere Buchungen bleiben bestehen."""
    db = get_db()
    try:
        booking = db.execute("""
            SELECT b.*, m.name AS member_name
            FROM payment_bookings b JOIN members m ON m.id = b.member_id
            WHERE b.id=?
        """, (id,)).fetchone()
        if not booking:
            raise ValueError('Die Buchung wurde nicht gefunden.')
        if booking['reversed_at']:
            raise ValueError('Diese Buchung ist bereits storniert.')
        reason = _require_change_reason(request.form.get('change_reason'))
        db.execute("""
            UPDATE payment_bookings
            SET reversed_at=datetime('now'), reversed_by_user_id=?,
                reversed_by_username=?, reverse_note=?
            WHERE id=?
        """, (
            current_user.id if current_user.is_authenticated else None,
            current_user.username if current_user.is_authenticated else None,
            reason,
            id,
        ))
        _record_booking_change(db, id, 'reverse', reason,
                               old={'amount': booking['amount_eur'], 'date': booking['booking_date']})
        _apply_booking_state(db, booking['invoice_id'], booking['member_id'])
        db.commit()
        audit_log('payment_booking_reverse',
                  f'Buchung #{id} storniert: {booking["member_name"]}, Rechnung {booking["invoice_id"]}, '
                  f'{booking["amount_eur"]:.2f} EUR, Grund: {reason}')
        flash(f'Buchung über {booking["amount_eur"]:.2f} € wurde storniert.', 'info')
    except Exception as e:
        db.rollback()
        audit_log('payment_booking_reverse_failed', f'Buchungsstorno fehlgeschlagen: Buchung {id} ({e})')
        flash_exception(e, 'Buchung konnte nicht storniert werden.')
    return redirect(_payment_redirect_target())


@app.route('/payments/mark_unpaid', methods=['POST'])
@admin_required
def payment_mark_unpaid():
    """Markiert eine Forderung als offen (Storno)."""
    db = get_db()
    invoice_id = request.form.get('invoice_id', type=int)
    member_id = request.form.get('member_id', type=int)
    try:
        row = get_payment_row(db, invoice_id, member_id)
        if not row:
            raise ValueError('Die Buchung wurde nicht gefunden.')
        if not row['paid']:
            raise ValueError('Diese Buchung ist bereits offen.')
        reason = _require_change_reason(request.form.get('change_reason'))
        for booking in row['bookings']:
            _record_booking_change(db, booking['id'], 'reverse', reason,
                                   old={'amount': booking['amount_eur'], 'date': booking['booking_date']})
        db.execute("""
            UPDATE payment_bookings
            SET reversed_at=datetime('now'),
                reversed_by_user_id=?,
                reversed_by_username=?,
                reverse_note=?
            WHERE invoice_id=? AND member_id=? AND reversed_at IS NULL
        """, (
            current_user.id if current_user.is_authenticated else None,
            current_user.username if current_user.is_authenticated else None,
            reason,
            invoice_id,
            member_id,
        ))
        db.execute("""UPDATE invoice_items SET paid=0, paid_at=NULL
                      WHERE invoice_id=? AND member_id=?""", (invoice_id, member_id))
        db.commit()
        audit_log('payment_unpaid', f'Buchung storniert: {row["member_name"]}, Rechnung {invoice_id}, '
                                    f'Grund: {reason}')
        flash('Buchung wurde storniert und wieder als offen markiert.', 'info')
    except Exception as e:
        db.rollback()
        audit_log('payment_unpaid_failed', f'Buchungsstorno fehlgeschlagen: Mitglied {member_id}, Rechnung {invoice_id} ({e})')
        flash_exception(e, 'Buchung konnte nicht zurückgesetzt werden.')
    return redirect(_payment_redirect_target())


# ═══════════════════════════════════════════════════════
# VEREINSKASSABUCH
# ═══════════════════════════════════════════════════════

# Startkategorien beim ersten Start; Admins koennen weitere anlegen.
CASHBOOK_DEFAULT_CATEGORIES = [
    ('Mitgliedsbeiträge', 'income', 10),
    ('Förderungen und Zuschüsse', 'income', 20),
    ('Zinserträge', 'income', 30),
    ('Sonstige Einnahmen', 'income', 40),
    ('Bewirtung', 'expense', 50),
    ('Verwaltungskosten', 'expense', 60),
    ('Bankspesen', 'expense', 70),
    ('Versicherung', 'expense', 80),
    ('Miete und Betriebskosten', 'expense', 90),
    ('Sonstige Ausgaben', 'expense', 100),
    ('Rundungsdifferenz', 'both', 110),
]
CASHBOOK_DIRECTIONS = {'income': 'Einnahme', 'expense': 'Ausgabe'}
CASHBOOK_PAYMENT_METHODS = {'cash': 'Bar', 'transfer': 'Überweisung'}
# Strombewegungen stammen aus payment_bookings und erhalten feste Kategorien.
CASHBOOK_ENERGY_CATEGORIES = {
    'income': 'Stromverkauf an Mitglieder',
    'expense': 'Stromeinkauf von Mitgliedern',
}
CASHBOOK_RECEIPT_MAX_BYTES = 10 * 1024 * 1024
# Dateiendung -> (MIME-Typ, erwartete Signatur am Dateianfang)
CASHBOOK_RECEIPT_TYPES = {
    '.pdf': ('application/pdf', b'%PDF-'),
    '.jpg': ('image/jpeg', b'\xff\xd8\xff'),
    '.jpeg': ('image/jpeg', b'\xff\xd8\xff'),
    '.png': ('image/png', b'\x89PNG\r\n\x1a\n'),
}


def _date_de(value):
    """ISO-Datum als 31.12.2026; unbekannte Werte bleiben unveraendert."""
    try:
        return date.fromisoformat(str(value)[:10]).strftime('%d.%m.%Y')
    except (TypeError, ValueError):
        return str(value or '')


def _cents(value):
    """Betrag in ganze Cent. Summen und Salden werden so exakt gerechnet."""
    return int(round(float(value or 0) * 100))


def _logo_data_uri():
    """Vereinslogo als data URI, damit es in PDF-Berichten eingebettet ist."""
    try:
        with open(os.path.join(BASE_DIR, 'static', 'logo.png'), 'rb') as handle:
            return 'data:image/png;base64,' + base64.b64encode(handle.read()).decode('ascii')
    except OSError as e:
        app.logger.warning('Logo für den Bericht nicht lesbar: %s', e)
        return ''


def get_cashbook_categories(db, only_active=True):
    query = "SELECT id, name, direction, active FROM cashbook_categories"
    if only_active:
        query += " WHERE active=1"
    query += " ORDER BY sort_order, name"
    return [dict(row) for row in db.execute(query).fetchall()]


def _cashbook_manual_rows(db):
    """Manuell erfasste Buchungen wie Bewirtung oder Verwaltungskosten."""
    rows = db.execute("""
        SELECT e.id, e.entry_date, e.direction, e.amount_eur, e.payment_method,
               e.description, e.counterparty, e.document_number, e.created_by,
               e.receipt_filename, e.category_id, c.name AS category_name
        FROM cashbook_entries e
        LEFT JOIN cashbook_categories c ON c.id = e.category_id
    """).fetchall()
    entries = []
    for row in rows:
        amount = round(abs(row['amount_eur'] or 0), 2)
        entries.append({
            'source': 'manual',
            'id': row['id'],
            'entry_date': row['entry_date'],
            'direction': row['direction'],
            'amount_eur': amount,
            'signed_amount': amount if row['direction'] == 'income' else -amount,
            'category': row['category_name'] or 'Ohne Kategorie',
            'category_id': row['category_id'],
            'payment_method': row['payment_method'],
            'description': row['description'],
            'counterparty': row['counterparty'] or '',
            'document_number': row['document_number'] or '',
            'has_receipt': bool(row['receipt_filename']),
            'receipt_filename': row['receipt_filename'] or '',
            'member_id': None,
            'invoice_id': None,
            'recorded_by': row['created_by'] or '',
            'deletable': True,
        })
    return entries


def _cashbook_energy_rows(db):
    """Stromverkauf und -einkauf aus allen gebuchten Abrechnungen.

    Massgeblich ist der tatsaechlich gebuchte Betrag: jede Zahlungsbuchung ist
    eine echte Kontobewegung. Aeltere Abrechnungen ohne Buchungssatz werden aus
    den bezahlten Positionen abgeleitet. Offene Restbetraege stehen nicht im
    Kassabuch, sondern am Mitgliedskonto.
    """
    entries = []
    for row in get_payment_rows(db):
        period = f'({_date_de(row["period_from"])} bis {_date_de(row["period_to"])})'
        movements = [{
            'key': booking['id'],
            'amount': booking['amount_eur'],
            'date': booking['booking_date'],
            'note': booking['note'],
            'user': booking['recorded_by_username'],
        } for booking in row['bookings']]
        if not movements and row['paid'] and abs(row['net_total']) >= 0.005:
            movements.append({
                'key': 'alt',
                'amount': row['net_total'],
                'date': row['booking_date'] or row['reference_date'].isoformat(),
                'note': '',
                'user': '',
            })
        for movement in movements:
            amount = round(movement['amount'] or 0, 2)
            if abs(amount) < 0.005:
                continue
            direction = 'income' if amount > 0 else 'expense'
            description = f'Abrechnung #{row["invoice_id"]} {period}'
            if row['carryover_total']:
                description += f', inkl. Vortrag {row["carryover_total"]:.2f} €'
            if len(movements) > 1:
                description += ', Teilbuchung'
            if movement['note']:
                description += f' – {movement["note"]}'
            entries.append({
                'source': 'energy',
                'id': f'{row["invoice_id"]}-{row["member_id"]}-{movement["key"]}',
                'entry_date': movement['date'] or row['reference_date'].isoformat(),
                'direction': direction,
                'amount_eur': abs(amount),
                'signed_amount': amount,
                'category': CASHBOOK_ENERGY_CATEGORIES[direction],
                'category_id': None,
                # Abrechnungen werden immer ueber das Bankkonto beglichen.
                'payment_method': 'transfer',
                'description': description,
                'counterparty': row['member_name'] or '',
                'document_number': f'A{row["invoice_id"]}-{row["member_id"]:03d}',
                'has_receipt': False,
                'receipt_filename': '',
                'member_id': row['member_id'],
                'invoice_id': row['invoice_id'],
                'recorded_by': movement['user'] or '',
                'deletable': False,
            })
    return entries


def resolve_cashbook_period(year='', date_from='', date_to=''):
    """Loest Jahr oder Von-Bis in einen konkreten Berichtszeitraum auf.

    Ein gesetzter Von-Bis-Zeitraum hat Vorrang vor dem Jahr. Leere Werte
    bedeuten offenes Ende. Liefert (start, ende, Beschriftung) als ISO-Strings.
    """
    start = _iso_to_date(date_from)
    end = _iso_to_date(date_to)
    if not start and not end and year:
        if re.fullmatch(r'\d{4}', year):
            start = date(int(year), 1, 1)
            end = date(int(year), 12, 31)
    if start and end and start > end:
        start, end = end, start

    start_iso = start.isoformat() if start else ''
    end_iso = end.isoformat() if end else ''
    if start and end:
        label = f'{_date_de(start_iso)} bis {_date_de(end_iso)}'
    elif start:
        label = f'ab {_date_de(start_iso)}'
    elif end:
        label = f'bis {_date_de(end_iso)}'
    else:
        label = 'gesamter Zeitraum'
    return start_iso, end_iso, label


def build_cashbook(db, year='', category='', direction='', method='', search='',
                   date_from='', date_to=''):
    """Fuehrt manuelle Buchungen und Strombewegungen zu einem Kassabuch zusammen.

    Der laufende Saldo wird ueber alle Bewegungen berechnet und erst danach
    gefiltert. Dadurch stimmen Anfangs- und Endsaldo eines Berichtszeitraums
    auch dann, wenn davor schon Buchungen liegen.
    """
    period_from, period_to, period_label = resolve_cashbook_period(year, date_from, date_to)

    all_rows = _cashbook_manual_rows(db) + _cashbook_energy_rows(db)
    # Innerhalb eines Tages nach Belegnummer, damit die Reihenfolge stabil und
    # nachvollziehbar bleibt.
    all_rows.sort(key=lambda row: (row['entry_date'] or '', row['document_number'], str(row['id'])))
    # In Cent rechnen, damit sich keine Fliesskomma-Abweichung aufsummiert.
    balance = cash_balance = bank_balance = 0
    opening = opening_cash = opening_bank = 0
    # Fortlaufende Belegnummer je Jahr in chronologischer Reihenfolge. Sie wird
    # ueber alle Bewegungen vergeben, damit sie von Filtern unabhaengig bleibt.
    counters = {}
    for row in all_rows:
        amount_cents = _cents(row['signed_amount'])
        if period_from and (row['entry_date'] or '') < period_from:
            opening += amount_cents
            if row['payment_method'] == 'cash':
                opening_cash += amount_cents
            else:
                opening_bank += amount_cents
        balance += amount_cents
        if row['payment_method'] == 'cash':
            cash_balance += amount_cents
        else:
            bank_balance += amount_cents
        row['balance'] = balance / 100
        row['entry_date_de'] = _date_de(row['entry_date'])
        year_key = (row['entry_date'] or '')[:4]
        counters[year_key] = counters.get(year_key, 0) + 1
        row['sequence_number'] = f'{year_key}/{counters[year_key]:03d}' if year_key else ''
        # Herkunftsnachweis: Belegnummer der manuellen Buchung bzw. Abrechnung
        row['reference'] = row['document_number']
    years = sorted({row['entry_date'][:4] for row in all_rows if row['entry_date']}, reverse=True)

    needle = (search or '').strip().lower()

    def in_period(row):
        entry_date = row['entry_date'] or ''
        if period_from and entry_date < period_from:
            return False
        if period_to and entry_date > period_to:
            return False
        return True

    def matches(row):
        if not in_period(row):
            return False
        if category and row['category'] != category:
            return False
        if direction and row['direction'] != direction:
            return False
        if method and row['payment_method'] != method:
            return False
        if needle:
            haystack = ' '.join((row['description'], row['counterparty'],
                                 row['category'], row['document_number'],
                                 row['sequence_number'])).lower()
            if needle not in haystack:
                return False
        return True

    rows = [row for row in all_rows if matches(row)]
    income_total = sum(_cents(row['amount_eur']) for row in rows if row['direction'] == 'income')
    expense_total = sum(_cents(row['amount_eur']) for row in rows if row['direction'] == 'expense')

    # Endsaldo: alle Bewegungen bis zum Ende des Zeitraums, unabhaengig von den
    # uebrigen Filtern. Sonst wuerde eine Kategorieauswahl den Saldo verfaelschen.
    closing = opening
    closing_cash, closing_bank = opening_cash, opening_bank
    for row in all_rows:
        if not in_period(row):
            continue
        amount_cents = _cents(row['signed_amount'])
        closing += amount_cents
        if row['payment_method'] == 'cash':
            closing_cash += amount_cents
        else:
            closing_bank += amount_cents

    def aggregate(source_rows, key_name, key_func):
        buckets = {}
        for row in source_rows:
            key = key_func(row)
            bucket = buckets.setdefault(key, {key_name: key, 'income': 0, 'expense': 0, 'count': 0})
            bucket['income' if row['direction'] == 'income' else 'expense'] += _cents(row['amount_eur'])
            bucket['count'] += 1
        result = []
        for bucket in buckets.values():
            income_cents, expense_cents = bucket['income'], bucket['expense']
            bucket['income'] = income_cents / 100
            bucket['expense'] = expense_cents / 100
            bucket['result'] = (income_cents - expense_cents) / 100
            result.append(bucket)
        return result

    by_category = aggregate(rows, 'category', lambda row: row['category'])
    by_category.sort(key=lambda item: item['income'] + item['expense'], reverse=True)
    by_year = aggregate(all_rows, 'year', lambda row: (row['entry_date'] or '')[:4])
    by_year.sort(key=lambda item: item['year'], reverse=True)

    return {
        'rows': list(reversed(rows)),   # neueste zuerst fuer die Anzeige
        'rows_chronological': rows,     # aelteste zuerst fuer Export und PDF
        'categories': get_cashbook_categories(db),
        'years': years,
        'by_category': by_category,
        'by_year': by_year,
        'period': {
            'from': period_from,
            'to': period_to,
            'label': period_label,
            'from_de': _date_de(period_from) if period_from else '',
            'to_de': _date_de(period_to) if period_to else '',
        },
        'summary': {
            'income_total': income_total / 100,
            'expense_total': expense_total / 100,
            'result': (income_total - expense_total) / 100,
            'entry_count': len(rows),
            'opening_balance': opening / 100,
            'opening_cash': opening_cash / 100,
            'opening_bank': opening_bank / 100,
            'closing_balance': closing / 100,
            'closing_cash': closing_cash / 100,
            'closing_bank': closing_bank / 100,
            # Gesamtstand heute, unabhaengig vom Berichtszeitraum
            'cash_balance': cash_balance / 100,
            'bank_balance': bank_balance / 100,
            'balance': balance / 100,
            'missing_receipts': len([row for row in rows
                                     if row['source'] == 'manual' and not row['has_receipt']]),
        },
    }


def _cashbook_filters():
    return {
        'year': (request.args.get('year') or '').strip(),
        'date_from': (request.args.get('date_from') or '').strip(),
        'date_to': (request.args.get('date_to') or '').strip(),
        'category': (request.args.get('category') or '').strip(),
        'direction': (request.args.get('direction') or '').strip(),
        'method': (request.args.get('method') or '').strip(),
        'search': (request.args.get('search') or '').strip(),
    }


def _cashbook_redirect_target():
    next_url = request.form.get('next') or request.args.get('next')
    if next_url and is_safe_redirect_url(next_url):
        return next_url
    return url_for('cashbook')


def _read_cashbook_receipt(file):
    """Prueft und liest einen Beleg; gibt None zurueck, wenn keiner gewaehlt wurde."""
    if not file or not file.filename:
        return None
    filename = secure_filename(file.filename)
    extension = os.path.splitext(filename)[1].lower()
    if extension not in CASHBOOK_RECEIPT_TYPES:
        raise ValueError('Als Beleg sind nur PDF-, JPG- und PNG-Dateien erlaubt.')
    mimetype, signature = CASHBOOK_RECEIPT_TYPES[extension]
    data = file.read()
    if not data:
        raise ValueError('Die Belegdatei ist leer.')
    if len(data) > CASHBOOK_RECEIPT_MAX_BYTES:
        raise ValueError(f'Der Beleg ist zu groß (max. {CASHBOOK_RECEIPT_MAX_BYTES // 1024 // 1024} MB).')
    if not data.startswith(signature):
        raise ValueError('Der Inhalt der Belegdatei passt nicht zur Dateiendung.')
    return {'filename': filename, 'mimetype': mimetype, 'data': data}


def _next_cashbook_document_number(db, entry_date):
    """Fortlaufende Belegnummer je Jahr, z.B. 2026-0007."""
    prefix = f'{entry_date.year}-'
    row = db.execute("""SELECT document_number FROM cashbook_entries
                        WHERE document_number LIKE ?
                        ORDER BY document_number DESC LIMIT 1""", (prefix + '%',)).fetchone()
    counter = 1
    if row and row['document_number']:
        try:
            counter = int(row['document_number'].split('-', 1)[1]) + 1
        except (IndexError, ValueError):
            counter = 1
    return f'{prefix}{counter:04d}'


def _save_cashbook_entry_from_request(db):
    """Legt eine manuelle Kassabuch-Buchung an und liefert die Belegnummer."""
    try:
        entry_date = date.fromisoformat((request.form.get('entry_date') or '').strip())
    except ValueError:
        raise ValueError('Bitte ein gültiges Buchungsdatum angeben.')
    if entry_date > local_now().date():
        raise ValueError('Das Buchungsdatum darf nicht in der Zukunft liegen.')

    direction = request.form.get('direction', '')
    if direction not in CASHBOOK_DIRECTIONS:
        raise ValueError('Bitte Einnahme oder Ausgabe auswählen.')
    payment_method = request.form.get('payment_method', '')
    if payment_method not in CASHBOOK_PAYMENT_METHODS:
        raise ValueError('Bitte Bar oder Überweisung auswählen.')

    try:
        amount = round(abs(float((request.form.get('amount_eur') or '').replace(',', '.').strip())), 2)
    except ValueError:
        raise ValueError('Bitte einen gültigen Betrag angeben.')
    if amount <= 0:
        raise ValueError('Der Betrag muss größer als null sein.')

    description = (request.form.get('description') or '').strip()
    if not description:
        raise ValueError('Bitte eine Begründung zur Buchung angeben.')

    category_id = request.form.get('category_id', type=int)
    if category_id and not db.execute("SELECT 1 FROM cashbook_categories WHERE id=?",
                                      (category_id,)).fetchone():
        raise ValueError('Die gewählte Kategorie existiert nicht.')

    receipt = _read_cashbook_receipt(request.files.get('receipt'))
    document_number = _next_cashbook_document_number(db, entry_date)
    db.execute("""INSERT INTO cashbook_entries
                  (entry_date, direction, amount_eur, category_id, payment_method,
                   description, counterparty, document_number,
                   receipt_filename, receipt_mimetype, receipt_data, created_by)
                  VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
               (entry_date.isoformat(), direction, amount, category_id or None, payment_method,
                description, (request.form.get('counterparty') or '').strip(), document_number,
                receipt['filename'] if receipt else None,
                receipt['mimetype'] if receipt else None,
                receipt['data'] if receipt else None,
                current_user.username))
    db.commit()
    audit_log('cashbook_create',
              f'Kassabuch {document_number}: {CASHBOOK_DIRECTIONS[direction]} '
              f'{amount:.2f} EUR ({CASHBOOK_PAYMENT_METHODS[payment_method]}) – {description}')
    return document_number


@app.route('/kassabuch')
@admin_required
def cashbook():
    """Vereinskassabuch mit allen Einnahmen und Ausgaben."""
    db = get_db()
    filters = _cashbook_filters()
    book = build_cashbook(db, **filters)
    today = local_now().date()
    quarter_start = date(today.year, ((today.month - 1) // 3) * 3 + 1, 1)
    return render_template('kassabuch.html',
                           book=book,
                           filters=filters,
                           export_args={key: value for key, value in filters.items() if value},
                           directions=CASHBOOK_DIRECTIONS,
                           methods=CASHBOOK_PAYMENT_METHODS,
                           year_start=date(today.year, 1, 1).isoformat(),
                           last_year_start=date(today.year - 1, 1, 1).isoformat(),
                           last_year_end=date(today.year - 1, 12, 31).isoformat(),
                           quarter_start=quarter_start.isoformat(),
                           today=today.isoformat())


@app.route('/kassabuch/new', methods=['POST'])
@admin_required
def cashbook_create():
    try:
        document_number = _save_cashbook_entry_from_request(get_db())
        flash(f'Buchung {document_number} gespeichert.', 'success')
    except Exception as e:
        flash_exception(e, 'Buchung konnte nicht gespeichert werden.')
    return redirect(_cashbook_redirect_target())


@app.route('/kassabuch/<int:id>/delete', methods=['POST'])
@admin_required
def cashbook_delete(id):
    db = get_db()
    row = db.execute("""SELECT document_number, amount_eur, direction, description
                        FROM cashbook_entries WHERE id=?""", (id,)).fetchone()
    if not row:
        flash('Buchung nicht gefunden.', 'danger')
        return redirect(_cashbook_redirect_target())
    db.execute("DELETE FROM cashbook_entries WHERE id=?", (id,))
    db.commit()
    audit_log('cashbook_delete',
              f'Kassabuch {row["document_number"]} gelöscht: '
              f'{CASHBOOK_DIRECTIONS.get(row["direction"], row["direction"])} '
              f'{row["amount_eur"]:.2f} EUR – {row["description"]}')
    flash(f'Buchung {row["document_number"]} wurde gelöscht.', 'success')
    return redirect(_cashbook_redirect_target())


@app.route('/kassabuch/<int:id>/beleg')
@admin_required
def cashbook_receipt(id):
    db = get_db()
    row = db.execute("""SELECT document_number, receipt_filename, receipt_mimetype, receipt_data
                        FROM cashbook_entries WHERE id=?""", (id,)).fetchone()
    if not row or not row['receipt_data']:
        flash('Zu dieser Buchung ist kein Beleg hinterlegt.', 'warning')
        return redirect(_cashbook_redirect_target())
    audit_log('cashbook_receipt_download', f'Beleg zu Buchung {row["document_number"]} geöffnet')
    return send_file(io.BytesIO(row['receipt_data']),
                     mimetype=row['receipt_mimetype'] or 'application/octet-stream',
                     as_attachment=request.args.get('preview') != '1',
                     download_name=row['receipt_filename'] or f'beleg-{id}')


def _cashbook_export_name(book, suffix):
    period = book['period']
    if period['from'] and period['to']:
        scope = f"{period['from']}_bis_{period['to']}"
    elif period['from']:
        scope = f"ab_{period['from']}"
    elif period['to']:
        scope = f"bis_{period['to']}"
    else:
        scope = 'gesamt'
    return f'kassabuch-{scope}-{local_now().strftime("%Y%m%d")}.{suffix}'


# Spalten der Exporte: Titel, Breite in Excel, Ausrichtung
CASHBOOK_EXPORT_COLUMNS = [
    ('Beleg-Nr', 12, 'left'),
    ('Referenz', 12, 'left'),
    ('Datum', 12, 'left'),
    ('Art', 10, 'left'),
    ('Kategorie', 28, 'left'),
    ('Zahlungsart', 13, 'left'),
    ('Begründung', 46, 'left'),
    ('Zahler/Empfänger', 26, 'left'),
    ('Einnahme', 12, 'right'),
    ('Ausgabe', 12, 'right'),
    ('Saldo', 12, 'right'),
    ('Beleg', 8, 'left'),
    ('Erfasst von', 16, 'left'),
]
# Spaltennummern der Betragsspalten in den Exporten
CASHBOOK_EXPORT_AMOUNT_COLUMNS = (9, 10, 11)
CASHBOOK_EXPORT_BALANCE_COLUMN = 11


@app.route('/kassabuch/export.csv')
@admin_required
def cashbook_export_csv():
    import csv

    filters = _cashbook_filters()
    book = build_cashbook(get_db(), **filters)
    summary = book['summary']

    def eur(value):
        return f'{value:.2f}'.replace('.', ',')

    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=';')
    writer.writerow(['Kassabuch', book['period']['label']])
    writer.writerow(['Erstellt am', local_now().strftime('%d.%m.%Y %H:%M')])
    writer.writerow([])
    writer.writerow([title for title, _, _ in CASHBOOK_EXPORT_COLUMNS])
    writer.writerow(['', '', '', '', 'Anfangssaldo', '', '', '', '', '',
                     eur(summary['opening_balance']), '', ''])
    for row in book['rows_chronological']:
        writer.writerow([
            row['sequence_number'],
            row['reference'],
            _date_de(row['entry_date']),
            CASHBOOK_DIRECTIONS.get(row['direction'], row['direction']),
            row['category'],
            CASHBOOK_PAYMENT_METHODS.get(row['payment_method'], row['payment_method']),
            row['description'],
            row['counterparty'],
            eur(row['amount_eur']) if row['direction'] == 'income' else '',
            eur(row['amount_eur']) if row['direction'] == 'expense' else '',
            eur(row['balance']),
            'ja' if row['has_receipt'] else '',
            row['recorded_by'],
        ])
    writer.writerow(['', '', '', '', f'Summe ({summary["entry_count"]} Buchungen)', '', '', '',
                     eur(summary['income_total']), eur(summary['expense_total']),
                     eur(summary['result']), '', ''])
    writer.writerow(['', '', '', '', 'Endsaldo', '', '', '', '', '',
                     eur(summary['closing_balance']), '', ''])
    writer.writerow([])
    writer.writerow(['Kassastand bar', eur(summary['closing_cash'])])
    writer.writerow(['Kontostand Bank', eur(summary['closing_bank'])])
    audit_log('cashbook_export', f'Kassabuch als CSV exportiert ({summary["entry_count"]} Buchungen)')
    # BOM, damit Excel die Umlaute korrekt erkennt
    return send_file(io.BytesIO(buffer.getvalue().encode('utf-8-sig')),
                     mimetype='text/csv', as_attachment=True,
                     download_name=_cashbook_export_name(book, 'csv'))


@app.route('/kassabuch/export.xlsx')
@admin_required
def cashbook_export_xlsx():
    """Kassabuch als Excel-Datei mit echten Zahlen- und Datumswerten."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    filters = _cashbook_filters()
    db = get_db()
    book = build_cashbook(db, **filters)
    summary = book['summary']
    org = get_public_config(db)

    money = '#,##0.00 €'
    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2B7A78')
    total_fill = PatternFill('solid', fgColor='E8F1F0')
    thin_top = Border(top=Side(style='thin', color='999999'))

    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Kassabuch'

    sheet['A1'] = f'Kassabuch – {org["org_name"]}'
    sheet['A1'].font = title_font
    sheet['A2'] = 'Berichtszeitraum'
    sheet['A2'].font = label_font
    sheet['B2'] = book['period']['label']
    sheet['A3'] = 'Erstellt am'
    sheet['A3'].font = label_font
    sheet['B3'] = local_now().strftime('%d.%m.%Y %H:%M')
    for index, (label, value) in enumerate((
            ('Anfangssaldo', summary['opening_balance']),
            ('Einnahmen', summary['income_total']),
            ('Ausgaben', summary['expense_total']),
            ('Endsaldo', summary['closing_balance']),
            ('davon bar', summary['closing_cash']),
            ('davon Bank', summary['closing_bank']))):
        cell_label = sheet.cell(row=2 + index, column=4, value=label)
        cell_label.font = label_font
        cell_value = sheet.cell(row=2 + index, column=5, value=value)
        cell_value.number_format = money

    header_row = 9
    for index, (title, width, align) in enumerate(CASHBOOK_EXPORT_COLUMNS, start=1):
        cell = sheet.cell(row=header_row, column=index, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal=align)
        sheet.column_dimensions[get_column_letter(index)].width = width

    current = header_row + 1
    sheet.cell(row=current, column=5, value='Anfangssaldo').font = label_font
    opening_cell = sheet.cell(row=current, column=CASHBOOK_EXPORT_BALANCE_COLUMN,
                              value=summary['opening_balance'])
    opening_cell.number_format = money
    opening_cell.font = label_font

    for row in book['rows_chronological']:
        current += 1
        values = [
            row['sequence_number'],
            row['reference'],
            _iso_to_date(row['entry_date']),
            CASHBOOK_DIRECTIONS.get(row['direction'], row['direction']),
            row['category'],
            CASHBOOK_PAYMENT_METHODS.get(row['payment_method'], row['payment_method']),
            row['description'],
            row['counterparty'],
            row['amount_eur'] if row['direction'] == 'income' else None,
            row['amount_eur'] if row['direction'] == 'expense' else None,
            row['balance'],
            'ja' if row['has_receipt'] else '',
            row['recorded_by'],
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=current, column=index, value=value)
            if index == 3:
                cell.number_format = 'DD.MM.YYYY'
            elif index in CASHBOOK_EXPORT_AMOUNT_COLUMNS:
                cell.number_format = money

    current += 1
    sheet.cell(row=current, column=5,
               value=f'Summe ({summary["entry_count"]} Buchungen)').font = label_font
    for column, value in zip(CASHBOOK_EXPORT_AMOUNT_COLUMNS,
                             (summary['income_total'], summary['expense_total'], summary['result'])):
        cell = sheet.cell(row=current, column=column, value=value)
        cell.number_format = money
        cell.font = label_font
    for column in range(1, len(CASHBOOK_EXPORT_COLUMNS) + 1):
        sheet.cell(row=current, column=column).fill = total_fill
        sheet.cell(row=current, column=column).border = thin_top

    current += 1
    sheet.cell(row=current, column=5, value='Endsaldo').font = label_font
    closing_cell = sheet.cell(row=current, column=CASHBOOK_EXPORT_BALANCE_COLUMN,
                              value=summary['closing_balance'])
    closing_cell.number_format = money
    closing_cell.font = label_font
    for column in range(1, len(CASHBOOK_EXPORT_COLUMNS) + 1):
        sheet.cell(row=current, column=column).fill = total_fill

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    sheet.auto_filter.ref = f'A{header_row}:{get_column_letter(len(CASHBOOK_EXPORT_COLUMNS))}{current - 2}'

    report = wb.create_sheet('Auswertung')
    report_row = 1
    for heading, key, entries in (('Auswertung nach Kategorie', 'category', book['by_category']),
                                  ('Auswertung nach Jahr', 'year', book['by_year'])):
        report.cell(row=report_row, column=1, value=heading).font = title_font
        report_row += 1
        for index, title in enumerate(('Bezeichnung', 'Buchungen', 'Einnahmen', 'Ausgaben', 'Ergebnis'),
                                      start=1):
            cell = report.cell(row=report_row, column=index, value=title)
            cell.font = header_font
            cell.fill = header_fill
        report_row += 1
        for entry in entries:
            report.cell(row=report_row, column=1, value=entry[key])
            report.cell(row=report_row, column=2, value=entry['count'])
            for column, field in ((3, 'income'), (4, 'expense'), (5, 'result')):
                cell = report.cell(row=report_row, column=column, value=entry[field])
                cell.number_format = money
            report_row += 1
        report_row += 2
    for index, width in enumerate((32, 12, 14, 14, 14), start=1):
        report.column_dimensions[get_column_letter(index)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    audit_log('cashbook_export', f'Kassabuch als Excel exportiert ({summary["entry_count"]} Buchungen)')
    return send_file(buffer, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=_cashbook_export_name(book, 'xlsx'))


@app.route('/kassabuch/export.pdf')
@admin_required
def cashbook_export_pdf():
    from weasyprint import HTML

    filters = _cashbook_filters()
    db = get_db()
    book = build_cashbook(db, **filters)
    html = render_template('kassabuch_pdf.html',
                           book=book,
                           filters=filters,
                           directions=CASHBOOK_DIRECTIONS,
                           methods=CASHBOOK_PAYMENT_METHODS,
                           org=get_public_config(db),
                           logo_b64=_logo_data_uri(),
                           created_by=current_user.username if current_user.is_authenticated else '',
                           created_at=local_now())
    pdf_bytes = HTML(string=html, base_url=BASE_DIR).write_pdf()
    audit_log('cashbook_export', f'Kassabuch als PDF exportiert ({book["summary"]["entry_count"]} Buchungen)')
    return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf',
                     as_attachment=request.args.get('preview') != '1',
                     download_name=_cashbook_export_name(book, 'pdf'))


@app.route('/kassabuch/kategorien', methods=['POST'])
@admin_required
def cashbook_category_create():
    db = get_db()
    name = (request.form.get('name') or '').strip()
    direction = request.form.get('direction', 'both')
    if not name:
        flash('Bitte einen Namen für die Kategorie angeben.', 'danger')
    elif direction not in {'income', 'expense', 'both'}:
        flash('Ungültige Zuordnung für die Kategorie.', 'danger')
    elif db.execute("SELECT 1 FROM cashbook_categories WHERE LOWER(name)=?", (name.lower(),)).fetchone():
        flash(f'Die Kategorie "{name}" existiert bereits.', 'warning')
    else:
        db.execute("INSERT INTO cashbook_categories (name, direction) VALUES (?, ?)", (name, direction))
        db.commit()
        audit_log('cashbook_category_create', f'Kassabuch-Kategorie angelegt: {name}')
        flash(f'Kategorie "{name}" angelegt.', 'success')
    return redirect(_cashbook_redirect_target())


@app.route('/kassabuch/kategorien/<int:id>/delete', methods=['POST'])
@admin_required
def cashbook_category_delete(id):
    db = get_db()
    row = db.execute("SELECT name FROM cashbook_categories WHERE id=?", (id,)).fetchone()
    if not row:
        flash('Kategorie nicht gefunden.', 'danger')
        return redirect(_cashbook_redirect_target())
    in_use = db.execute("SELECT COUNT(*) FROM cashbook_entries WHERE category_id=?", (id,)).fetchone()[0]
    if in_use:
        # Bereits gebuchte Kategorien bleiben erhalten, damit alte Buchungen lesbar bleiben.
        db.execute("UPDATE cashbook_categories SET active=0 WHERE id=?", (id,))
        flash(f'Kategorie "{row["name"]}" wird nicht mehr angeboten, '
              f'bleibt aber bei {in_use} bestehenden Buchungen sichtbar.', 'info')
    else:
        db.execute("DELETE FROM cashbook_categories WHERE id=?", (id,))
        flash(f'Kategorie "{row["name"]}" gelöscht.', 'success')
    db.commit()
    audit_log('cashbook_category_delete', f'Kassabuch-Kategorie entfernt: {row["name"]}')
    return redirect(_cashbook_redirect_target())


# ═══════════════════════════════════════════════════════
# BENUTZERVERWALTUNG (Admin)
# ═══════════════════════════════════════════════════════

CONTRACTS_FOLDER = os.path.join(BASE_DIR, '..', 'contracts')
os.makedirs(CONTRACTS_FOLDER, exist_ok=True)

CONTRACT_TYPE_LABELS = {'bezieher': 'Bezieher', 'einspeiser': 'Einspeiser'}
_CONTRACT_UMLAUT_MAP = {
    'ä': 'ae', 'ö': 'oe', 'ü': 'ue', 'Ä': 'Ae', 'Ö': 'Oe', 'Ü': 'Ue', 'ß': 'ss',
}


def _contract_name_slug(name):
    """Wandelt einen Mitgliedsnamen in einen dateinamentauglichen Slug um (Umlaute transliteriert)."""
    text = ''.join(_CONTRACT_UMLAUT_MAP.get(ch, ch) for ch in str(name or ''))
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^A-Za-z0-9]+', '_', text).strip('_')
    return text or 'Mitglied'


def generate_contract_filename(member_name, member_id, contract_type, upload_date=None):
    """Erzeugt einen einheitlichen, sprechenden Dateinamen fuer einen Vertrags-Upload,
    z.B. Vertrag_Bezieher_Max_Mustermann_M42_2026-07-29.pdf"""
    typ_label = CONTRACT_TYPE_LABELS.get(contract_type, contract_type.capitalize())
    date_str = upload_date or datetime.now().strftime('%Y-%m-%d')
    return f'Vertrag_{typ_label}_{_contract_name_slug(member_name)}_M{member_id}_{date_str}.pdf'


def _admin_users_redirect_target():
    next_url = request.form.get('next') or request.args.get('next')
    if next_url and is_safe_redirect_url(next_url):
        return next_url
    return url_for('admin_users')


@app.route('/admin/users')
@admin_required
def admin_users():
    """Benutzerverwaltung – alle User anzeigen."""
    db = get_db()
    users = db.execute("""
        SELECT u.*, m.name as member_name, m.email as member_email
        FROM users u LEFT JOIN members m ON u.member_id = m.id
        WHERE NOT EXISTS (
            SELECT 1
            FROM users other
            WHERE other.id != u.id
              AND (
                  (u.member_id IS NOT NULL AND other.member_id = u.member_id)
                  OR (
                      u.email IS NOT NULL AND u.email != ''
                      AND other.email IS NOT NULL AND other.email != ''
                      AND LOWER(other.email) = LOWER(u.email)
                  )
              )
              AND (
                  (other.invite_token IS NULL AND u.invite_token IS NOT NULL)
                  OR (
                      (other.invite_token IS NULL) = (u.invite_token IS NULL)
                      AND other.id < u.id
                  )
              )
        )
        ORDER BY u.is_admin DESC, u.username
    """).fetchall()
    members = db.execute("SELECT id, name, email FROM members WHERE active=1 ORDER BY name").fetchall()

    active_members = db.execute(
        "SELECT id, name, bezug_zp, einspeiser_zp FROM members WHERE active=1 ORDER BY name"
    ).fetchall()
    # "Aktive Nutzer" = Mitglieder mit einem bereits akzeptierten Benutzerkonto
    # (kein offener Einladungslink). Mitglieder ohne oder mit nur offener Einladung
    # sollen nicht als "fehlender Vertrag" auftauchen.
    active_user_member_ids = {
        row['member_id'] for row in db.execute(
            "SELECT DISTINCT member_id FROM users WHERE member_id IS NOT NULL AND invite_token IS NULL"
        ).fetchall()
    }
    existing_types = {}
    for row in db.execute("SELECT member_id, type FROM contracts").fetchall():
        existing_types.setdefault(row['member_id'], set()).add(row['type'])
    missing_contracts = []
    for m in active_members:
        if m['id'] not in active_user_member_ids:
            continue
        have = existing_types.get(m['id'], set())
        if m['bezug_zp'] and 'bezieher' not in have:
            missing_contracts.append({'member_id': m['id'], 'member_name': m['name'], 'type': 'bezieher'})
        if m['einspeiser_zp'] and 'einspeiser' not in have:
            missing_contracts.append({'member_id': m['id'], 'member_name': m['name'], 'type': 'einspeiser'})

    return render_template('admin_users.html', users=users, members=members,
                           missing_contracts=missing_contracts)


@app.route('/admin/users/create', methods=['POST'])
@admin_required
def admin_user_create():
    """Neuen Benutzer für ein Mitglied anlegen."""
    db = get_db()
    member_id = request.form.get('member_id', type=int)
    role = request.form.get('role', 'member')
    if role not in ('admin', 'member'):
        role = 'member'

    member = db.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
    if not member:
        flash('Mitglied nicht gefunden.', 'danger')
        return redirect(_admin_users_redirect_target())

    # Username: email oder vorname+nachname lowercase
    if member['email']:
        username = member['email'].lower().strip()
    else:
        username = member['name'].lower().replace(' ', '').replace('&', '')
        # Umlaute normalisieren
        for old, new in [('ä', 'ae'), ('ö', 'oe'), ('ü', 'ue'), ('ß', 'ss')]:
            username = username.replace(old, new)

    existing = db.execute("""
        SELECT id, username FROM users
        WHERE LOWER(username)=?
           OR member_id=?
           OR (? != '' AND email IS NOT NULL AND email != '' AND LOWER(email)=?)
        ORDER BY CASE WHEN invite_token IS NULL THEN 0 ELSE 1 END, id
        LIMIT 1
    """, (username.lower(), member_id, (member['email'] or '').strip(), (member['email'] or '').strip().lower())).fetchone()
    if existing:
        flash(f'Für dieses Mitglied existiert bereits der Benutzer "{existing["username"]}".', 'warning')
        return redirect(_admin_users_redirect_target())

    # Einladungs-Token generieren
    invite_token = secrets.token_urlsafe(32)
    invite_expires = (datetime.now().replace(hour=23, minute=59) +
                      __import__('datetime').timedelta(days=14)).isoformat()
    # Temporäres Passwort (wird beim ersten Login über Invite-Link gesetzt)
    temp_hash = generate_password_hash(secrets.token_hex(16))

    db.execute("""INSERT INTO users (username, password_hash, email, is_admin, role, member_id,
                  invite_token, invite_expires) VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
               (username, temp_hash, member['email'], 1 if role == 'admin' else 0,
                role, member_id, invite_token, invite_expires))
    db.commit()

    invite_url = public_url_for('v2_invite_accept', token=invite_token)
    audit_log('user_create', f'Benutzer angelegt: {username} (Rolle: {role}, Mitglied-ID: {member_id})')
    if member['email']:
        invite_user = {
            'username': username,
            'email': member['email'],
            'role': role,
            'member_name': member['name'],
        }
        try:
            send_invitation_email(db, invite_user, invite_url, invite_expires)
            flash(f'Benutzer "{username}" angelegt und Einladung an {member["email"]} gesendet.', 'success')
        except Exception as e:
            app.logger.exception('Invitation mail failed for user %s', username)
            flash_exception(e, f'Benutzer "{username}" angelegt, aber die Einladung konnte nicht per E-Mail gesendet werden.', 'warning')
            flash(f'Einladungslink: {invite_url}', 'info')
    else:
        flash(f'Benutzer "{username}" angelegt. Keine E-Mail-Adresse hinterlegt; Einladungslink: {invite_url}', 'warning')
    return redirect(_admin_users_redirect_target())


@app.route('/admin/users/<int:id>/invite', methods=['POST'])
@admin_required
def admin_user_reinvite(id):
    """Neuen Einladungslink generieren."""
    db = get_db()
    invite_action = request.form.get('invite_action', 'send')
    if invite_action not in ('show', 'send'):
        invite_action = 'send'
    invite_token = secrets.token_urlsafe(32)
    invite_expires = (datetime.now().replace(hour=23, minute=59) +
                      __import__('datetime').timedelta(days=14)).isoformat()
    db.execute("UPDATE users SET invite_token=?, invite_expires=? WHERE id=?",
               (invite_token, invite_expires, id))
    db.commit()
    user = db.execute("""
        SELECT u.*, m.name as member_name
        FROM users u LEFT JOIN members m ON u.member_id = m.id
        WHERE u.id=?
    """, (id,)).fetchone()
    invite_url = public_url_for('v2_invite_accept', token=invite_token)
    audit_log('user_reinvite', f'Neuer Einladungslink für: {user["username"]}' if user else f'Reinvite User-ID {id}')
    if invite_action == 'show':
        flash(f'Neuer Einladungslink generiert: {invite_url}', 'info')
    elif not user:
        flash(f'Neuer Einladungslink generiert: {invite_url}', 'success')
    elif user['email']:
        try:
            send_invitation_email(db, user, invite_url, invite_expires)
            flash(f'Neuer Einladungslink für "{user["username"]}" generiert und an {user["email"]} gesendet.', 'success')
        except Exception as e:
            app.logger.exception('Invitation mail failed for user %s', user['username'])
            flash_exception(e, 'Neuer Einladungslink generiert, aber die Einladung konnte nicht per E-Mail gesendet werden.', 'warning')
            flash(f'Einladungslink: {invite_url}', 'info')
    else:
        flash(f'Neuer Einladungslink generiert. Keine E-Mail-Adresse hinterlegt; Link: {invite_url}', 'warning')
    return redirect(_admin_users_redirect_target())


@app.route('/admin/users/<int:id>/toggle-role', methods=['POST'])
@admin_required
def admin_user_toggle_role(id):
    """Rolle umschalten admin <-> member."""
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (id,)).fetchone()
    if not user:
        flash('Benutzer nicht gefunden.', 'danger')
        return redirect(_admin_users_redirect_target())
    new_role = 'member' if user['role'] == 'admin' else 'admin'
    new_admin = 1 if new_role == 'admin' else 0
    db.execute("UPDATE users SET role=?, is_admin=? WHERE id=?", (new_role, new_admin, id))
    db.commit()
    audit_log('user_role_change', f'Rolle geändert: {user["username"]} → {new_role}')
    flash(f'Rolle auf "{new_role}" geändert.', 'success')
    return redirect(_admin_users_redirect_target())


@app.route('/admin/users/<int:id>/delete', methods=['POST'])
@admin_required
def admin_user_delete(id):
    """Benutzer löschen."""
    if id == current_user.id:
        flash('Sie können sich nicht selbst löschen.', 'danger')
        return redirect(_admin_users_redirect_target())
    db = get_db()
    user = db.execute("SELECT username FROM users WHERE id=?", (id,)).fetchone()
    db.execute("DELETE FROM users WHERE id=?", (id,))
    db.commit()
    audit_log('user_delete', f'Benutzer gelöscht: {user["username"]}' if user else f'User-ID {id} gelöscht')
    flash('Benutzer gelöscht.', 'success')
    return redirect(_admin_users_redirect_target())


@app.route('/admin/contracts/upload', methods=['POST'])
@admin_required
def admin_contract_upload():
    """Vertrag hochladen für ein Mitglied."""
    db = get_db()
    member_id = request.form.get('member_id', type=int)
    contract_type = request.form.get('type', '')
    if contract_type not in ('bezieher', 'einspeiser'):
        flash('Ungültiger Vertragstyp.', 'danger')
        return redirect(_admin_users_redirect_target())
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('Keine Datei ausgewählt.', 'danger')
        return redirect(_admin_users_redirect_target())
    original_name = secure_filename(file.filename)
    if not original_name.lower().endswith('.pdf'):
        flash('Nur PDF-Dateien sind als Vertrag erlaubt.', 'danger')
        return redirect(_admin_users_redirect_target())
    file_data = file.read()
    if len(file_data) > 10 * 1024 * 1024:
        flash('Datei zu groß (max. 10 MB).', 'danger')
        return redirect(_admin_users_redirect_target())
    if not file_data.startswith(b'%PDF-'):
        flash('Die hochgeladene Datei ist keine gültige PDF-Datei.', 'danger')
        return redirect(_admin_users_redirect_target())
    member = db.execute("SELECT name FROM members WHERE id=?", (member_id,)).fetchone()
    if not member:
        flash('Mitglied nicht gefunden.', 'danger')
        return redirect(_admin_users_redirect_target())
    filename = generate_contract_filename(member['name'], member_id, contract_type)
    db.execute("""INSERT INTO contracts (member_id, type, filename, file_data, uploaded_by)
                  VALUES (?, ?, ?, ?, ?)""",
               (member_id, contract_type, filename, file_data, current_user.username))
    db.commit()
    audit_log('contract_upload', f'Vertrag hochgeladen: {filename} ({contract_type}) für {member["name"]}')
    flash(f'Vertrag "{filename}" hochgeladen.', 'success')
    return redirect(_admin_users_redirect_target())


@app.route('/contracts/<int:id>/download')
@login_required
def contract_download(id):
    """Vertrag herunterladen (Admins alle, Members nur eigene)."""
    db = get_db()
    contract = db.execute("SELECT * FROM contracts WHERE id=?", (id,)).fetchone()
    if not contract:
        flash('Vertrag nicht gefunden.', 'danger')
        return redirect(url_for('admin_users'))
    if not current_user.is_admin and current_user.member_id != contract['member_id']:
        flash('Zugriff verweigert.', 'danger')
        return redirect(url_for('portal_dashboard'))
    audit_log('contract_download', f'Vertrag heruntergeladen: {contract["filename"]} (ID {id})')
    import io
    is_preview = request.args.get('preview') == '1'
    filename = contract['filename'] or ''
    mimetype = mimetypes.guess_type(filename)[0] or 'application/octet-stream'
    return send_file(
        io.BytesIO(contract['file_data']),
        mimetype=mimetype,
        as_attachment=not is_preview,
        download_name=filename
    )


@app.route('/contracts/<int:id>/delete', methods=['POST'])
@admin_required
def contract_delete(id):
    """Vertrag löschen."""
    db = get_db()
    contract = db.execute("SELECT filename, member_id FROM contracts WHERE id=?", (id,)).fetchone()
    db.execute("DELETE FROM contracts WHERE id=?", (id,))
    db.commit()
    audit_log('contract_delete', f'Vertrag gelöscht: {contract["filename"]}' if contract else f'Vertrag ID {id} gelöscht')
    flash('Vertrag gelöscht.', 'success')
    return redirect(_admin_users_redirect_target())


# ═══════════════════════════════════════════════════════
# EINLADUNG / PASSWORT SETZEN
# ═══════════════════════════════════════════════════════

@app.route('/invite/<token>', methods=['GET', 'POST'])
def invite_accept(token):
    """Einladungslink – Passwort setzen."""
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE invite_token=?", (token,)).fetchone()
    if not user:
        flash('Ungültiger Einladungslink.', 'danger')
        return redirect(url_for('login'))
    if user['invite_expires'] and user['invite_expires'] < datetime.now().isoformat():
        flash('Einladungslink abgelaufen. Bitte Admin kontaktieren.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        policy_error = validate_password(password, user['username'])
        if policy_error:
            flash(policy_error, 'danger')
            return render_template('invite.html', token=token, username=user['username'])
        if password != confirm:
            flash('Passwörter stimmen nicht überein.', 'danger')
            return render_template('invite.html', token=token, username=user['username'])
        db.execute("""UPDATE users
                      SET password_hash=?, invite_token=NULL, invite_expires=NULL,
                          password_change_required=0
                      WHERE id=?""",
                   (generate_password_hash(password), user['id']))
        db.commit()
        audit_log('invite_accept', f'Einladung angenommen, Passwort gesetzt', user_id=user['id'], username=user['username'])
        flash('Passwort erfolgreich gesetzt. Sie können sich jetzt einloggen.', 'success')
        return redirect(url_for('login'))

    return render_template('invite.html', token=token, username=user['username'])


@app.route('/v2/invite/<token>', methods=['GET', 'POST'])
def v2_invite_accept(token):
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE invite_token=?", (token,)).fetchone()
    if not user:
        flash('Ungültiger Einladungslink.', 'danger')
        return redirect(url_for('v2_login'))
    if user['invite_expires'] and user['invite_expires'] < datetime.now().isoformat():
        flash('Einladungslink abgelaufen. Bitte Admin kontaktieren.', 'danger')
        return redirect(url_for('v2_login'))

    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        policy_error = validate_password(password, user['username'])
        if policy_error:
            flash(policy_error, 'danger')
            return render_template('v2_public.html', page='invite', token=token, username=user['username'])
        if password != confirm:
            flash('Passwörter stimmen nicht überein.', 'danger')
            return render_template('v2_public.html', page='invite', token=token, username=user['username'])
        db.execute("""UPDATE users
                      SET password_hash=?, invite_token=NULL, invite_expires=NULL,
                          password_change_required=0
                      WHERE id=?""",
                   (generate_password_hash(password), user['id']))
        db.commit()
        audit_log('invite_accept', 'Einladung angenommen, Passwort gesetzt', user_id=user['id'], username=user['username'])
        flash('Passwort erfolgreich gesetzt. Sie können sich jetzt einloggen.', 'success')
        return redirect(url_for('v2_login'))

    return render_template('v2_public.html', page='invite', token=token, username=user['username'])


@app.route('/api/contracts')
@admin_required
def api_contracts():
    """JSON-API: Alle Verträge auflisten."""
    db = get_db()
    rows = db.execute("""
        SELECT c.id, c.member_id, c.type, c.filename, c.uploaded_at, c.uploaded_by, m.name as member_name
        FROM contracts c JOIN members m ON m.id = c.member_id
        ORDER BY m.name, c.type
    """).fetchall()
    data = []
    for row in rows:
        item = dict(row)
        item['uploaded_at'] = format_local_date(item.get('uploaded_at'))
        data.append(item)
    return jsonify(data)


@app.route('/admin/audit')
@admin_required
def admin_audit():
    """Audit-Log anzeigen."""
    db = get_db()
    page = request.args.get('page', 1, type=int)
    per_page = 50
    offset = (page - 1) * per_page

    # Filter
    action_filter = request.args.get('action', '')
    user_filter = request.args.get('user', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    where_clauses = []
    params = []
    if action_filter:
        where_clauses.append("a.action = ?")
        params.append(action_filter)
    if user_filter:
        where_clauses.append("a.username LIKE ?")
        params.append(f'%{user_filter}%')
    if date_from:
        date_from_utc, _ = local_day_bounds_as_utc_strings(date_from)
        where_clauses.append("a.timestamp >= ?")
        params.append(date_from_utc)
    if date_to:
        _, date_to_utc = local_day_bounds_as_utc_strings(date_to)
        where_clauses.append("a.timestamp <= ?")
        params.append(date_to_utc)

    where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    total = db.execute(f"SELECT COUNT(*) FROM audit_log a{where_sql}", params).fetchone()[0]
    logs = db.execute(f"""
        SELECT a.* FROM audit_log a{where_sql}
        ORDER BY a.timestamp DESC LIMIT ? OFFSET ?
    """, params + [per_page, offset]).fetchall()
    logs = [dict(row) for row in logs]
    for log in logs:
        log['timestamp_display'] = format_local_datetime(log.get('timestamp'))

    # Alle vorhandenen Aktionstypen für Filter-Dropdown
    actions = db.execute("SELECT DISTINCT action FROM audit_log ORDER BY action").fetchall()
    action_list = [r['action'] for r in actions]

    # Statistiken
    today_from_utc, today_to_utc = local_day_bounds_as_utc_strings()
    stats = {
        'total_entries': db.execute("SELECT COUNT(*) FROM audit_log").fetchone()[0],
        'today_entries': db.execute(
            "SELECT COUNT(*) FROM audit_log WHERE timestamp >= ? AND timestamp <= ?",
            (today_from_utc, today_to_utc)
        ).fetchone()[0],
        'active_users': db.execute(
            "SELECT COUNT(DISTINCT username) FROM audit_log WHERE timestamp >= ? AND timestamp <= ?",
            (today_from_utc, today_to_utc)
        ).fetchone()[0],
    }

    total_pages = (total + per_page - 1) // per_page

    return render_template('admin_audit.html',
                           logs=logs, page=page, total_pages=total_pages, total=total,
                           action_filter=action_filter, user_filter=user_filter,
                           date_from=date_from, date_to=date_to,
                           action_list=action_list, stats=stats)


# ═══════════════════════════════════════════════════════
# MITGLIEDER-PORTAL
# ═══════════════════════════════════════════════════════

@app.route('/portal')
@login_required
def portal_dashboard():
    """Teilnehmer-Dashboard."""
    if current_user.is_admin and not current_user.member_id:
        return redirect(url_for('dashboard'))
    db = get_db()
    member_id = current_user.member_id
    if not member_id:
        flash('Kein Mitglied zugeordnet.', 'warning')
        return render_template('portal_dashboard.html', member=None, invoices=[], stats=None, account=None)

    member = db.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
    account = get_member_account_summary(db, member_id)
    # Abrechnungen des Mitglieds
    invoices = db.execute("""
        SELECT DISTINCT i.* FROM invoices i
        WHERE i.id IN (
            SELECT invoice_id FROM invoice_items WHERE member_id=?
            UNION
            SELECT invoice_id FROM invoice_carryovers WHERE member_id=?
        )
        ORDER BY i.period_from DESC
    """, (member_id, member_id)).fetchall()

    # Letzte Abrechnung: Stats berechnen
    stats = None
    if invoices:
        latest = invoices[0]
        stats = get_member_stats(db, member, latest['period_from'], latest['period_to'])
        # Net total
        items = db.execute("SELECT * FROM invoice_items WHERE invoice_id=? AND member_id=?",
                           (latest['id'], member_id)).fetchall()
        carryovers = get_invoice_carryovers(db, latest['id'], member_id)
        net = sum(i['amount_eur'] if i['type'] == 'consumption' else -i['amount_eur'] for i in items)
        net += sum(c['amount_eur'] for c in carryovers)
        stats['net_total'] = round(net, 2)
        stats['invoice_id'] = latest['id']

    return render_template('portal_dashboard.html', member=member, invoices=invoices, stats=stats, account=account)


@app.route('/portal/data', methods=['GET', 'POST'])
@login_required
def portal_data():
    """Teilnehmer kann eigene Stammdaten bearbeiten."""
    if not current_user.member_id:
        flash('Kein Mitglied zugeordnet.', 'warning')
        return redirect(url_for('portal_dashboard'))
    db = get_db()
    member = db.execute("SELECT * FROM members WHERE id=?", (current_user.member_id,)).fetchone()

    if request.method == 'POST':
        newsletter_optout = 0 if form_switch_enabled('newsletter_enabled') else 1
        db.execute("""UPDATE members SET
            name=?, email=?, phone=?,
            address_street=?, address_zip=?, address_city=?,
            iban=?, bic=?, account_holder=?, newsletter_optout=?,
            updated_at=datetime('now')
            WHERE id=?""", (
            request.form.get('name', member['name']),
            request.form.get('email', member['email']),
            request.form.get('phone', member['phone']),
            request.form.get('address_street', member['address_street']),
            request.form.get('address_zip', member['address_zip']),
            request.form.get('address_city', member['address_city']),
            request.form.get('iban', member['iban']),
            request.form.get('bic', member['bic']),
            request.form.get('account_holder', member['account_holder']),
            newsletter_optout,
            current_user.member_id))
        db.commit()
        audit_log('portal_data_update', f'Eigene Stammdaten aktualisiert')
        flash('Daten aktualisiert.', 'success')
        return redirect(url_for('portal_data'))

    return render_template('portal_data.html', member=member)


@app.route('/portal/invoices')
@login_required
def portal_invoices():
    """Teilnehmer: Eigene Abrechnungen."""
    if not current_user.member_id:
        flash('Kein Mitglied zugeordnet.', 'warning')
        return redirect(url_for('portal_dashboard'))
    db = get_db()
    account = get_member_account_summary(db, current_user.member_id)
    rows = db.execute("""
        SELECT i.id, i.period_from, i.period_to, i.status, i.created_at,
               COALESCE(SUM(CASE WHEN ii.type='consumption' THEN ii.amount_eur ELSE 0 END), 0) as total_cons,
               COALESCE(SUM(CASE WHEN ii.type='generation' THEN ii.amount_eur ELSE 0 END), 0) as total_gen,
               COALESCE(SUM(ii.kwh), 0) as total_kwh
        FROM invoices i
        LEFT JOIN invoice_items ii ON ii.invoice_id = i.id AND ii.member_id = ?
        WHERE i.id IN (
            SELECT invoice_id FROM invoice_items WHERE member_id=?
            UNION
            SELECT invoice_id FROM invoice_carryovers WHERE member_id=?
        )
        GROUP BY i.id
        ORDER BY i.period_from DESC
    """, (current_user.member_id, current_user.member_id, current_user.member_id)).fetchall()
    payment_by_invoice = {row['invoice_id']: row for row in account['rows']}
    return render_template(
        'portal_invoices.html',
        invoices=rows,
        member_id=current_user.member_id,
        payment_by_invoice=payment_by_invoice,
    )


@app.route('/portal/contracts')
@login_required
def portal_contracts():
    """Teilnehmer: Eigene Verträge."""
    if not current_user.member_id:
        flash('Kein Mitglied zugeordnet.', 'warning')
        return redirect(url_for('portal_dashboard'))
    db = get_db()
    contracts = db.execute("SELECT * FROM contracts WHERE member_id=? ORDER BY uploaded_at DESC",
                           (current_user.member_id,)).fetchall()
    return render_template('portal_contracts.html', contracts=contracts)


@app.route('/portal/newsletter', methods=['POST'])
@login_required
def portal_newsletter_toggle():
    """Teilnehmer: Newsletter an/abbestellen."""
    db = get_db()
    if not current_user.member_id:
        flash('Kein Mitglied zugeordnet.', 'warning')
        return redirect(url_for('portal_data'))
    if 'newsletter_enabled' in request.form:
        optout = 0 if form_switch_enabled('newsletter_enabled') else 1
    else:
        optout = 1 if request.form.get('optout') == '1' else 0
    db.execute("UPDATE members SET newsletter_optout=? WHERE id=?", (optout, current_user.member_id))
    db.commit()
    if optout:
        audit_log('newsletter_optout', f'Newsletter abbestellt (Mitglied {current_user.member_id})')
        flash('Newsletter abbestellt.', 'info')
    else:
        audit_log('newsletter_optin', f'Newsletter wieder abonniert (Mitglied {current_user.member_id})')
        flash('Newsletter abonniert.', 'success')
    return redirect(url_for('portal_data'))


@app.route('/newsletter/unsubscribe/<token>')
def newsletter_unsubscribe(token):
    """Öffentliche Abmeldung per Link aus E-Mail."""
    if not token or len(token) < 16:
        return render_template('newsletter_unsubscribe.html', status='invalid', member=None), 400
    db = get_db()
    member = db.execute("SELECT id, name FROM members WHERE unsubscribe_token=?", (token,)).fetchone()
    if not member:
        return render_template('newsletter_unsubscribe.html', status='invalid', member=None), 404
    db.execute("UPDATE members SET newsletter_optout=1 WHERE id=?", (member['id'],))
    db.commit()
    audit_log('newsletter_optout', f'Newsletter per Link abbestellt: {member["name"]} (ID {member["id"]})',
              user_id=None, username='system')
    return render_template('newsletter_unsubscribe.html', status='success', member=member)


@app.route('/v2/newsletter/unsubscribe/<token>')
def v2_newsletter_unsubscribe(token):
    status = 'invalid'
    member = None
    http_status = 200
    if not token or len(token) < 16:
        http_status = 400
    else:
        db = get_db()
        member = db.execute("SELECT id, name FROM members WHERE unsubscribe_token=?", (token,)).fetchone()
        if not member:
            http_status = 404
        else:
            db.execute("UPDATE members SET newsletter_optout=1 WHERE id=?", (member['id'],))
            db.commit()
            audit_log('newsletter_optout', f'Newsletter per Link abbestellt: {member["name"]} (ID {member["id"]})',
                      user_id=None, username='system')
            status = 'success'
    return render_template('v2_public.html', page='unsubscribe', status=status, member=member), http_status


# === Newsletter Admin ===

def _newsletter_redirect_target():
    next_url = request.form.get('next') or request.args.get('next')
    if next_url and is_safe_redirect_url(next_url):
        return next_url
    return url_for('newsletter_list')


@app.route('/newsletter')
@admin_required
def newsletter_list():
    """Alle Newsletter anzeigen."""
    db = get_db()
    newsletters = db.execute("SELECT * FROM newsletters ORDER BY created_at DESC").fetchall()
    return render_template('newsletter_list.html', newsletters=newsletters)


@app.route('/newsletter/new', methods=['GET', 'POST'])
@admin_required
def newsletter_new():
    """Neuen Newsletter erstellen."""
    if request.method == 'POST':
        subject = request.form.get('subject', '').strip()
        body_html = sanitize_newsletter_html(request.form.get('body_html', '').strip())
        if not subject or not body_html:
            flash('Betreff und Inhalt sind erforderlich.', 'danger')
            return render_template('newsletter_edit.html', newsletter=None,
                                   subject=subject, body_html=body_html)
        db = get_db()
        db.execute("INSERT INTO newsletters (subject, body_html, created_by) VALUES (?,?,?)",
                   (subject, body_html, current_user.username))
        db.commit()
        audit_log('newsletter_create', f'Newsletter erstellt: {subject}')
        flash('Newsletter gespeichert.', 'success')
        return redirect(_newsletter_redirect_target())
    return render_template('newsletter_edit.html', newsletter=None, subject='', body_html='')


@app.route('/newsletter/<int:id>/edit', methods=['GET', 'POST'])
@admin_required
def newsletter_edit(id):
    """Newsletter bearbeiten."""
    db = get_db()
    nl = db.execute("SELECT * FROM newsletters WHERE id=?", (id,)).fetchone()
    if not nl:
        flash('Newsletter nicht gefunden.', 'danger')
        return redirect(url_for('newsletter_list'))
    if nl['sent_at']:
        flash('Bereits versendeter Newsletter kann nicht bearbeitet werden.', 'warning')
        return redirect(url_for('newsletter_list'))
    if request.method == 'POST':
        subject = request.form.get('subject', '').strip()
        body_html = sanitize_newsletter_html(request.form.get('body_html', '').strip())
        if not subject or not body_html:
            flash('Betreff und Inhalt sind erforderlich.', 'danger')
            return render_template('newsletter_edit.html', newsletter=nl,
                                   subject=subject, body_html=body_html)
        db.execute("UPDATE newsletters SET subject=?, body_html=? WHERE id=?", (subject, body_html, id))
        db.commit()
        audit_log('newsletter_edit', f'Newsletter bearbeitet: {subject} (ID {id})')
        flash('Newsletter aktualisiert.', 'success')
        return redirect(url_for('newsletter_list'))
    return render_template('newsletter_edit.html', newsletter=nl,
                           subject=nl['subject'], body_html=sanitize_newsletter_html(nl['body_html']))


@app.route('/newsletter/<int:id>/preview')
@admin_required
def newsletter_preview(id):
    """Vorschau des Newsletters im E-Mail-Template."""
    db = get_db()
    nl = db.execute("SELECT * FROM newsletters WHERE id=?", (id,)).fetchone()
    if not nl:
        flash('Newsletter nicht gefunden.', 'danger')
        return redirect(url_for('newsletter_list'))
    logo_url = public_url_for('static', filename='logo.png')
    html = render_template('newsletter_email.html',
        subject=nl['subject'],
        preview_text=nl['subject'],
        logo_url=logo_url,
        edition_label=nl['subject'].split('–')[0].strip() if '\u2013' in nl['subject'] else nl['subject'],
        headline=nl['subject'],
        subtitle='',
        body_html=sanitize_newsletter_html(nl['body_html']),
        unsubscribe_url='#',
    )
    return html


@app.route('/newsletter/<int:id>/test', methods=['POST'])
@admin_required
def newsletter_test(id):
    """Test-E-Mail an eine einzelne Adresse senden."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    test_email = request.form.get('test_email', '').strip()
    if not _is_valid_email(test_email):
        flash('Bitte eine gültige Test-E-Mail-Adresse eingeben.', 'danger')
        return redirect(_newsletter_redirect_target())

    db = get_db()
    nl = db.execute("SELECT * FROM newsletters WHERE id=?", (id,)).fetchone()
    if not nl:
        flash('Newsletter nicht gefunden.', 'danger')
        return redirect(_newsletter_redirect_target())

    try:
        mail_cfg = _get_valid_mail_config(db)
    except RuntimeError as e:
        flash(f'E-Mail-Konfiguration ungültig: {e}', 'danger')
        return redirect(_newsletter_redirect_target())

    base_url = public_base_url()
    logo_url = public_url_for('static', filename='logo.png')

    full_html = render_template('newsletter_email.html',
        subject=nl['subject'],
        preview_text=nl['subject'],
        logo_url=logo_url,
        edition_label=nl['subject'].split('–')[0].strip() if '\u2013' in nl['subject'] else nl['subject'],
        headline=nl['subject'],
        subtitle='',
        body_html=sanitize_newsletter_html(nl['body_html']),
        unsubscribe_url=f"{base_url}/newsletter/unsubscribe/test-preview",
    )

    try:
        with smtplib.SMTP(mail_cfg['smtp_host'], mail_cfg['smtp_port']) as server:
            if mail_cfg['smtp_tls']:
                server.starttls()
            server.login(mail_cfg['smtp_user'], mail_cfg['smtp_pass'])

            msg = MIMEMultipart('alternative')
            msg['From'] = mail_cfg['from_header']
            msg['Reply-To'] = mail_cfg['reply_to_header']
            msg['To'] = test_email
            msg['Subject'] = f"[TEST] {nl['subject']}"
            msg.attach(MIMEText(full_html, 'html', 'utf-8'))

            _log_mail_send(mail_cfg, test_email, msg['Subject'])
            server.sendmail(mail_cfg['from_address'], [test_email], msg.as_string())
        flash(f'Test-E-Mail erfolgreich an {test_email} gesendet.', 'success')
    except Exception as e:
        flash_exception(e, 'Test-E-Mail konnte nicht gesendet werden.')

    audit_log('newsletter_test', f'Test-E-Mail für "{nl["subject"]}" an {test_email}')
    return redirect(_newsletter_redirect_target())


@app.route('/newsletter/<int:id>/send', methods=['POST'])
@admin_required
def newsletter_send(id):
    """Newsletter an alle aktiven Mitglieder mit E-Mail senden (die nicht abbestellt haben)."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    import secrets

    db = get_db()
    nl = db.execute("SELECT * FROM newsletters WHERE id=?", (id,)).fetchone()
    if not nl:
        flash('Newsletter nicht gefunden.', 'danger')
        return redirect(_newsletter_redirect_target())

    try:
        mail_cfg = _get_valid_mail_config(db)
    except RuntimeError as e:
        flash(f'E-Mail-Konfiguration ungültig: {e}', 'danger')
        return redirect(_newsletter_redirect_target())

    # Empfänger: aktive Mitglieder mit E-Mail, die nicht abbestellt haben
    members = db.execute("""
        SELECT id, name, email, unsubscribe_token FROM members
        WHERE active=1 AND email IS NOT NULL AND email != ''
          AND (newsletter_optout IS NULL OR newsletter_optout=0)
    """).fetchall()

    if not members:
        flash('Keine Empfänger gefunden (alle abbestellt oder keine E-Mail hinterlegt).', 'warning')
        return redirect(_newsletter_redirect_target())

    sent = 0
    failed = 0
    # Logo-URL für E-Mail
    logo_url = public_url_for('static', filename='logo.png')

    try:
        server = smtplib.SMTP(mail_cfg['smtp_host'], mail_cfg['smtp_port'])
        if mail_cfg['smtp_tls']:
            server.starttls()
        server.login(mail_cfg['smtp_user'], mail_cfg['smtp_pass'])

        for member in members:
            # Unsubscribe-Token generieren falls nicht vorhanden
            unsub_token = member['unsubscribe_token']
            if not unsub_token:
                unsub_token = secrets.token_urlsafe(32)
                db.execute("UPDATE members SET unsubscribe_token=? WHERE id=?", (unsub_token, member['id']))
                db.commit()

            unsub_url = public_url_for('v2_newsletter_unsubscribe', token=unsub_token)

            # HTML aus Template rendern
            full_html = render_template('newsletter_email.html',
                subject=nl['subject'],
                preview_text=nl['subject'],
                logo_url=logo_url,
                edition_label=nl['subject'].split('–')[0].strip() if '–' in nl['subject'] else nl['subject'],
                headline=nl['subject'],
                subtitle='',
                body_html=sanitize_newsletter_html(nl['body_html']),
                unsubscribe_url=unsub_url,
            )

            msg = MIMEMultipart('alternative')
            msg['From'] = mail_cfg['from_header']
            msg['Reply-To'] = mail_cfg['reply_to_header']
            msg['To'] = member['email']
            msg['Subject'] = nl['subject']
            msg['List-Unsubscribe'] = f'<{unsub_url}>'
            msg.attach(MIMEText(full_html, 'html', 'utf-8'))

            try:
                _log_mail_send(mail_cfg, member['email'], nl['subject'])
                server.sendmail(mail_cfg['from_address'], [member['email']], msg.as_string())
                db.execute("""INSERT INTO newsletter_log (newsletter_id, member_id, email, status)
                              VALUES (?,?,?,?)""", (id, member['id'], member['email'], 'sent'))
                sent += 1
            except Exception as e:
                db.execute("""INSERT INTO newsletter_log (newsletter_id, member_id, email, status, error_message)
                              VALUES (?,?,?,?,?)""", (id, member['id'], member['email'], 'failed', str(e)))
                failed += 1

        server.quit()
    except Exception as e:
        flash_exception(e, 'SMTP-Verbindung fehlgeschlagen.')
        return redirect(_newsletter_redirect_target())

    db.execute("UPDATE newsletters SET sent_at=datetime('now'), recipients_count=? WHERE id=?", (sent, id))
    db.commit()
    audit_log('newsletter_send', f'Newsletter "{nl["subject"]}" versendet: {sent} gesendet, {failed} fehlgeschlagen')
    flash(f'Newsletter versendet: {sent} erfolgreich, {failed} fehlgeschlagen.', 'success')
    return redirect(_newsletter_redirect_target())


@app.route('/newsletter/<int:id>/delete', methods=['POST'])
@admin_required
def newsletter_delete(id):
    """Newsletter löschen."""
    db = get_db()
    nl = db.execute("SELECT subject FROM newsletters WHERE id=?", (id,)).fetchone()
    if nl:
        db.execute("DELETE FROM newsletter_log WHERE newsletter_id=?", (id,))
        db.execute("DELETE FROM newsletters WHERE id=?", (id,))
        db.commit()
        audit_log('newsletter_delete', f'Newsletter gelöscht: {nl["subject"]} (ID {id})')
        flash('Newsletter gelöscht.', 'success')
    return redirect(_newsletter_redirect_target())


# === Entry Point ===

if __name__ == '__main__':
    init_db()
    _startup_mail_config_check()
    start_backup_scheduler()
    app.run(
        host=os.environ.get('EEG_HOST', '127.0.0.1'),
        port=int(os.environ.get('EEG_PORT', '5000')),
        debug=os.environ.get('FLASK_DEBUG') == '1'
    )
