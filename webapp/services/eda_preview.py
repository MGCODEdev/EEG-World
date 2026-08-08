"""Read-only preview and validation for EDA XLSX imports."""

import hashlib
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


KNOWN_LABELS = {
    "Gesamtverbrauch lt. Messung (bei Teilnahme gem. Erzeugung) [KWH]",
    "Verbrauch lt. Messung entsprechend dem Teilnahmefaktor je ZP und EC-ID [KWH]",
    "Anteil gemeinschaftliche Erzeugung [KWH]",
    "Eigendeckung gemeinschaftliche Erzeugung [KWH]",
    "Eigendeckung aus erneuerbarer Energie [KWH]",
    "Gesamte gemeinschaftliche Erzeugung [KWH]",
    "Erzeugung lt. Messung entsprechend dem Teilnahmefaktor und EC-ID [KWH]",
    "Gesamt/Überschusserzeugung, Gemeinschaftsüberschuss [KWH]",
    "Restüberschuss bei EG und je ZP [KWH]",
}


class EDAPreviewError(ValueError):
    pass


@dataclass(frozen=True)
class PreviewWarning:
    code: str
    severity: str
    message: str


@dataclass(frozen=True)
class EDAPreview:
    filename: str
    sha256: str
    size_bytes: int
    sheet_name: str
    metering_point_count: int
    series_count: int
    interval_count: int
    measurement_count: int
    data_available_from: str | None
    data_available_until: str | None
    interval_minutes: int | None
    duplicate_timestamps: int
    missing_timestamp_intervals: int
    empty_measurement_cells: int
    quality_counts: dict[str, int]
    directions: dict[str, int]
    warnings: tuple[PreviewWarning, ...]

    def to_dict(self) -> dict:
        return asdict(self)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _parse_timestamp(value) -> datetime:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    text = str(value or "").strip()
    for fmt in (
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise EDAPreviewError(f"Unbekannter Zeitstempel: {text[:40]}")


def _cell(row: tuple, index: int):
    return row[index] if index < len(row) else None


def preview_eda_xlsx(filepath: str) -> EDAPreview:
    """Validiert eine EDA-Datei ohne Datenbank- oder Dateisystemaenderungen."""
    path = Path(filepath)
    if path.suffix.lower() != ".xlsx":
        raise EDAPreviewError("Nur XLSX-Dateien werden unterstützt.")
    if not path.is_file():
        raise EDAPreviewError("Die Importdatei wurde nicht gefunden.")

    warnings: list[PreviewWarning] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = next(
            (name for name in workbook.sheetnames if "energiedaten" in name.lower()),
            None,
        )
        if not sheet_name:
            raise EDAPreviewError("Das Tabellenblatt 'Energiedaten' fehlt.")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header: dict[int, tuple] = {}
        data_rows: list[tuple] = []
        for row_number, row in enumerate(rows, 1):
            if row_number <= 16:
                header[row_number] = row
            elif row and row[0] not in (None, ""):
                data_rows.append(row)

        if not data_rows:
            raise EDAPreviewError("Die Datei enthält keine Messintervalle.")

        row2 = header.get(2, ())
        row4 = header.get(4, ())
        row7 = header.get(7, ())
        row8 = header.get(8, ())
        row12 = header.get(12, ())
        row14 = header.get(14, ())
        max_columns = max(len(row2), len(row4), len(row14))
        series = []
        metering_points = set()
        directions = Counter()
        all_series_are_quarter_hour = True
        col = 1
        while col < max_columns:
            point = str(_cell(row2, col) or "").strip()
            direction = str(_cell(row4, col) or "").strip().upper()
            label = str(_cell(row14, col) or "").strip()
            if not point or point in {"MM", "TOTAL"}:
                col += 1
                continue
            quality_col = None
            next_point = str(_cell(row2, col + 1) or "").strip()
            if next_point in {"", "MM"}:
                quality_col = col + 1
            try:
                series_start = _parse_timestamp(_cell(row7, col)) if _cell(row7, col) else None
                series_end = _parse_timestamp(_cell(row8, col)) if _cell(row8, col) else None
            except EDAPreviewError:
                series_start = None
                series_end = None
                warnings.append(PreviewWarning(
                    "INVALID_SERIES_PERIOD", "error",
                    f"Ungültiger Datenzeitraum in Spalte {col + 1}.",
                ))
            series.append((col, quality_col, point, direction, label, series_start, series_end))
            metering_points.add(point)
            if direction:
                directions[direction] += 1
            if label not in KNOWN_LABELS:
                warnings.append(PreviewWarning(
                    "UNKNOWN_SERIES", "error",
                    f"Nicht unterstützte Messreihe in Spalte {col + 1}.",
                ))
            interval_label = str(_cell(row12, col) or "").strip().lower()
            if interval_label and "viertel" not in interval_label and "qh" not in interval_label:
                all_series_are_quarter_hour = False
                warnings.append(PreviewWarning(
                    "UNSUPPORTED_INTERVAL", "error",
                    f"Nicht unterstütztes Messintervall in Spalte {col + 1}.",
                ))
            col += 2 if quality_col is not None else 1

        if not series:
            raise EDAPreviewError("Es wurden keine Messreihen erkannt.")

        timestamps = [_parse_timestamp(row[0]) for row in data_rows]
        timestamp_counts = Counter(timestamps)
        duplicate_timestamps = sum(count - 1 for count in timestamp_counts.values() if count > 1)
        ordered_unique = sorted(timestamp_counts)
        positive_deltas = [
            int((current - previous).total_seconds() // 60)
            for previous, current in zip(ordered_unique, ordered_unique[1:])
            if current > previous
        ]
        interval_minutes = (
            15 if all_series_are_quarter_hour
            else Counter(positive_deltas).most_common(1)[0][0] if positive_deltas else None
        )
        missing_intervals = 0
        if interval_minutes:
            for delta in positive_deltas:
                if delta > interval_minutes and delta % interval_minutes == 0:
                    missing_intervals += delta // interval_minutes - 1
        if duplicate_timestamps:
            warnings.append(PreviewWarning(
                "DUPLICATE_TIMESTAMPS", "error",
                f"{duplicate_timestamps} doppelte Zeitstempel erkannt.",
            ))
        if missing_intervals:
            warnings.append(PreviewWarning(
                "MISSING_TIMESTAMPS", "warning",
                f"{missing_intervals} fehlende Zeitintervalle erkannt.",
            ))
        if interval_minutes != 15:
            warnings.append(PreviewWarning(
                "INTERVAL_NOT_QUARTER_HOUR", "error",
                f"Erkanntes Zeitraster: {interval_minutes or 0} Minuten.",
            ))

        quality_counts = Counter()
        empty_cells = 0
        measurement_count = 0
        for row, row_timestamp in zip(data_rows, timestamps):
            for value_col, quality_col, _, _, _, series_start, series_end in series:
                value = _cell(row, value_col)
                if value in (None, ""):
                    if ((series_start is None or row_timestamp >= series_start)
                            and (series_end is None or row_timestamp <= series_end)):
                        empty_cells += 1
                    continue
                try:
                    numeric = float(value)
                except (TypeError, ValueError):
                    warnings.append(PreviewWarning(
                        "INVALID_VALUE", "error",
                        f"Nicht numerischer Messwert in Datenzeile {measurement_count + 17}.",
                    ))
                    continue
                if numeric < 0:
                    warnings.append(PreviewWarning(
                        "NEGATIVE_VALUE", "error", "Negativer Energiewert erkannt.",
                    ))
                measurement_count += 1
                quality = str(_cell(row, quality_col) or "L1").strip().upper() if quality_col else "L1"
                quality_counts[quality] += 1

        if empty_cells:
            warnings.append(PreviewWarning(
                "EMPTY_MEASUREMENT_CELLS", "warning",
                f"{empty_cells} leere Messwertzellen erkannt.",
            ))

        return EDAPreview(
            filename=path.name,
            sha256=_sha256(path),
            size_bytes=path.stat().st_size,
            sheet_name=sheet_name,
            metering_point_count=len(metering_points),
            series_count=len(series),
            interval_count=len(data_rows),
            measurement_count=measurement_count,
            data_available_from=ordered_unique[0].isoformat() if ordered_unique else None,
            data_available_until=(ordered_unique[-1] + timedelta(minutes=interval_minutes or 0)).isoformat()
            if ordered_unique else None,
            interval_minutes=interval_minutes,
            duplicate_timestamps=duplicate_timestamps,
            missing_timestamp_intervals=missing_intervals,
            empty_measurement_cells=empty_cells,
            quality_counts=dict(quality_counts),
            directions=dict(directions),
            warnings=tuple(warnings),
        )
    finally:
        workbook.close()
