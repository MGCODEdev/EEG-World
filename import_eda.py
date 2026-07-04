#!/usr/bin/env python3
"""
EDA Energiedatenreport Importer v2
Importiert Excel-Reports aus dem EDA-Anwenderportal in SQLite.
Exakt zugeschnitten auf das EDA-Report-Format für Energiegemeinschaften.

Blattstruktur "Energiedaten":
  Zeile 1:  Disclaimer-Text
  Zeile 2:  MeteringPointId  | ZP1 | MM | ZP1 | MM | ... | TOTAL | TOTAL | ...
  Zeile 3:  Name             | Name1 | | Name1 | | ...
  Zeile 4:  Energydirection  | CONSUMPTION | | ... | GENERATION | | ...
  Zeile 5-13: Metadaten (Filter, Period, Interval, etc.)
  Zeile 14: MeterCode        | Label1 | | Label2 | | ...
  Zeile 15: Data Completeness
  Zeile 16: Spaltensumme / minimale Qualität
  Ab Zeile 17: Zeitstempel | Wert | Qualität | Wert | Qualität | ...

Spalten kommen immer paarweise: Wert-Spalte (gerade), MM-Spalte (ungerade).
Ausnahme: TOTAL-Spalten am Ende ohne MM.
"""

import sqlite3
import sys
import os
import re
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl nicht installiert. Bitte: pip install openpyxl")
    sys.exit(1)


DB_PATH = os.environ.get("EEG_DB_PATH", "eeg_data.db")
DATA_DIR = os.environ.get("EEG_DATA_DIR", "data")

# Mapping: Excel MeterCode-Label -> OBIS-Code
LABEL_TO_METERCODE = {
    # CONSUMPTION
    "Gesamtverbrauch lt. Messung (bei Teilnahme gem. Erzeugung) [KWH]": "1-1:1.9.0 G.01",
    "Verbrauch lt. Messung entsprechend dem Teilnahmefaktor je ZP und EC-ID [KWH]": "1-1:1.9.0 G.01T",
    "Anteil gemeinschaftliche Erzeugung [KWH]": "1-1:1.9.0 G.02",
    "Eigendeckung gemeinschaftliche Erzeugung [KWH]": "1-1:2.9.0 G.03",
    "Eigendeckung aus erneuerbarer Energie [KWH]": "1-1:2.9.0 G.03R",
    # GENERATION
    "Gesamte gemeinschaftliche Erzeugung [KWH]": "1-1:2.9.0 G.01",
    "Erzeugung lt. Messung entsprechend dem Teilnahmefaktor und EC-ID [KWH]": "1-1:2.9.0 G.01T",
    "Gesamt/Überschusserzeugung, Gemeinschaftsüberschuss [KWH]": "1-1:2.9.0 G.02",
    "Restüberschuss bei EG und je ZP [KWH]": "1-1:2.9.0 P.01T",
}


def init_db(db_path: str) -> sqlite3.Connection:
    """Initialisiert DB mit Schema."""
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA synchronous=NORMAL")
    schema_path = Path(__file__).parent / "schema.sql"
    if schema_path.exists():
        conn.executescript(schema_path.read_text())
    return conn


def parse_filename(filename: str) -> dict:
    """Extrahiert Metadaten aus dem Dateinamen."""
    info = {"report_code": None, "period_start": None, "period_end": None}
    m = re.match(
        r"([A-Z]+\d+)_(\d{4}-\d{2}-\d{2})T(\d{2})_(\d{2})-(\d{4}-\d{2}-\d{2})T(\d{2})_(\d{2})",
        filename,
    )
    if m:
        info["report_code"] = m.group(1)
        info["period_start"] = f"{m.group(2)}T{m.group(3)}:{m.group(4)}:00"
        info["period_end"] = f"{m.group(5)}T{m.group(6)}:{m.group(7)}:00"
    return info


def resolve_meter_code(label: str, conn: sqlite3.Connection) -> int:
    """Loest ein Excel-Label zu einer meter_code_id auf."""
    code = LABEL_TO_METERCODE.get(label)

    if not code:
        label_lower = label.lower()
        for key, val in LABEL_TO_METERCODE.items():
            if key.lower()[:30] in label_lower or label_lower[:30] in key.lower():
                code = val
                break

    if not code:
        code = f"UNKNOWN:{label[:50]}"
        conn.execute(
            "INSERT OR IGNORE INTO meter_codes (code, label_de, direction) VALUES (?, ?, 'UNKNOWN')",
            (code, label),
        )
        conn.commit()

    row = conn.execute("SELECT id FROM meter_codes WHERE code = ?", (code,)).fetchone()
    if row:
        return row[0]

    conn.execute(
        "INSERT OR IGNORE INTO meter_codes (code, label_de, direction) VALUES (?, ?, 'UNKNOWN')",
        (code, label),
    )
    conn.commit()
    return conn.execute("SELECT id FROM meter_codes WHERE code = ?", (code,)).fetchone()[0]


def parse_timestamp(val) -> datetime:
    """Parst Zeitstempel aus Excel-Zelle."""
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"Unbekanntes Zeitformat: {val}")


def parse_columns(ws) -> list:
    """
    Parst die Header-Zeilen und gibt Spalten-Definitionen zurueck.
    """
    columns = []

    # Header-Zeilen einlesen
    rows_data = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=16, values_only=True), 1):
        rows_data[row_idx] = list(row)

    row2 = rows_data.get(2, [])   # MeteringPointId
    row4 = rows_data.get(4, [])   # Energydirection
    row14 = rows_data.get(14, []) # MeterCode

    max_col = max(len(row2), len(row4), len(row14))

    col_idx = 1  # Skip col A (labels)
    while col_idx < max_col:
        mp_id = str(row2[col_idx]).strip() if col_idx < len(row2) and row2[col_idx] else ""
        direction = str(row4[col_idx]).strip().upper() if col_idx < len(row4) and row4[col_idx] else ""
        meter_label = str(row14[col_idx]).strip() if col_idx < len(row14) and row14[col_idx] else ""

        if not mp_id or mp_id == "None" or mp_id == "MM":
            col_idx += 1
            continue

        is_total = (mp_id == "TOTAL")

        # Check if next column is MM (quality column)
        mm_col = None
        if not is_total and col_idx + 1 < max_col:
            next_mp = str(row2[col_idx + 1]).strip() if col_idx + 1 < len(row2) and row2[col_idx + 1] else ""
            if next_mp == "MM" or next_mp == "" or next_mp == "None":
                mm_col = col_idx + 1

        columns.append({
            "col_idx": col_idx,
            "metering_point_id": mp_id,
            "energy_direction": direction if direction in ("CONSUMPTION", "GENERATION") else "",
            "meter_label": meter_label,
            "mm_col_idx": mm_col,
            "is_total": is_total,
        })

        if mm_col is not None:
            col_idx += 2
        else:
            col_idx += 1

    return columns


def import_file(filepath: str, conn: sqlite3.Connection, allow_duplicate: bool = False) -> int:
    """Importiert eine EDA-Excel-Datei."""
    filename = os.path.basename(filepath)
    print(f"\n{'='*60}")
    print(f"Import: {filename}")
    print(f"{'='*60}")

    # Duplikat-Check
    if not allow_duplicate:
        try:
            existing = conn.execute(
                "SELECT id FROM import_batches WHERE source_file = ? AND replaced_at IS NULL",
                (filename,),
            ).fetchone()
        except sqlite3.OperationalError:
            existing = conn.execute(
                "SELECT id FROM import_batches WHERE source_file = ?", (filename,)
            ).fetchone()
        if existing:
            print(f"  SKIP: Bereits importiert (batch_id={existing[0]})")
            return 0

    info = parse_filename(filename)
    print(f"  Report: {info['report_code']}, Zeitraum: {info['period_start']} - {info['period_end']}")

    print(f"  Lade Workbook...")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    print(f"  Blaetter: {wb.sheetnames}")

    # Batch anlegen
    cursor = conn.execute("""
        INSERT INTO import_batches (source_file, report_code, period_start, period_end)
        VALUES (?, ?, ?, ?)
    """, (filename, info["report_code"], info["period_start"], info["period_end"]))
    batch_id = cursor.lastrowid
    conn.commit()

    # Energiedaten-Blatt finden
    energy_sheet = None
    for name in wb.sheetnames:
        if "energiedaten" in name.lower() or "energy" in name.lower():
            energy_sheet = name
            break
    if not energy_sheet:
        energy_sheet = wb.sheetnames[-1]

    print(f"  Blatt: '{energy_sheet}'")
    ws = wb[energy_sheet]

    # Spalten parsen
    columns = parse_columns(ws)
    data_columns = [c for c in columns if not c["is_total"]]

    if not data_columns:
        print("  FEHLER: Keine Datenspalten erkannt!")
        wb.close()
        return 0

    unique_mps = set(c["metering_point_id"] for c in data_columns)
    print(f"  {len(data_columns)} Datenspalten, {len(unique_mps)} Zaehlpunkte")

    # Metering Points in DB
    for col in data_columns:
        conn.execute("""
            INSERT OR IGNORE INTO metering_points (metering_point_id, energy_direction, first_seen)
            VALUES (?, ?, datetime('now'))
        """, (col["metering_point_id"], col["energy_direction"]))
    conn.commit()

    # MeterCode IDs cachen
    mc_cache = {}
    for col in data_columns:
        label = col["meter_label"]
        if label and label not in mc_cache:
            mc_cache[label] = resolve_meter_code(label, conn)

    # Daten lesen (ab Zeile 17)
    print(f"  Lese Viertelstundenwerte ab Zeile 17...")
    count = 0
    batch_data = []
    interval_minutes = 15

    for row_idx, row in enumerate(ws.iter_rows(min_row=17, values_only=True), 17):
        row_list = list(row)
        if not row_list or not row_list[0]:
            continue

        try:
            ts_start = parse_timestamp(row_list[0])
        except (ValueError, TypeError):
            continue

        ts_end = ts_start + timedelta(minutes=interval_minutes)
        ts_start_iso = ts_start.strftime("%Y-%m-%dT%H:%M:%S")
        ts_end_iso = ts_end.strftime("%Y-%m-%dT%H:%M:%S")

        for col in data_columns:
            cidx = col["col_idx"]
            label = col["meter_label"]
            if not label or label not in mc_cache:
                continue
            if cidx >= len(row_list):
                continue

            value = row_list[cidx]
            if value is None:
                continue

            try:
                value_kwh = float(value)
            except (ValueError, TypeError):
                continue

            # Quality
            quality = "L1"
            if col["mm_col_idx"] is not None and col["mm_col_idx"] < len(row_list):
                q_val = row_list[col["mm_col_idx"]]
                if q_val:
                    quality = str(q_val).strip()

            is_estimated = 1 if quality in ("L2", "L3", "03", "04") else 0

            batch_data.append((
                batch_id,
                col["metering_point_id"],
                ts_start_iso,
                ts_end_iso,
                interval_minutes,
                mc_cache[label],
                value_kwh,
                quality,
                is_estimated,
            ))
            count += 1

        # Batch-Insert
        if len(batch_data) >= 10000:
            conn.executemany("""
                INSERT INTO measurements
                (batch_id, metering_point_id, timestamp_start, timestamp_end,
                 interval_minutes, meter_code_id, value_kwh, quality, is_estimated)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, batch_data)
            conn.commit()
            batch_data = []
            if count % 100000 == 0:
                print(f"    ... {count:,} Werte")

    if batch_data:
        conn.executemany("""
            INSERT INTO measurements
            (batch_id, metering_point_id, timestamp_start, timestamp_end,
             interval_minutes, meter_code_id, value_kwh, quality, is_estimated)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, batch_data)
        conn.commit()

    wb.close()

    print(f"  -> {count:,} Messwerte importiert")
    return count


def main():
    """Hauptprogramm."""
    print("=" * 60)
    print("  EDA Energiedatenreport Importer v2.0")
    print("=" * 60)
    print(f"Datenbank: {DB_PATH}")
    print(f"Datenverzeichnis: {DATA_DIR}")

    conn = init_db(DB_PATH)

    data_path = Path(DATA_DIR)
    if not data_path.exists():
        data_path.mkdir(parents=True)
        print(f"\nVerzeichnis '{DATA_DIR}' erstellt.")
        conn.close()
        sys.exit(0)

    xlsx_files = sorted(data_path.glob("*.xlsx"))
    if not xlsx_files:
        xlsx_files = sorted(Path(".").glob("RC*.xlsx"))

    if not xlsx_files:
        print(f"\nKeine XLSX-Dateien gefunden.")
        conn.close()
        sys.exit(0)

    print(f"\nDateien: {len(xlsx_files)}")
    for f in xlsx_files:
        print(f"  - {f.name} ({f.stat().st_size / 1024:.0f} KB)")

    total = 0
    for f in xlsx_files:
        try:
            count = import_file(str(f), conn)
            total += count
        except Exception as e:
            print(f"\n  FEHLER bei {f.name}: {e}")
            import traceback
            traceback.print_exc()

    # Abschluss
    print(f"\n{'='*60}")
    print(f"  IMPORT ABGESCHLOSSEN: {total:,} Messwerte")
    print(f"{'='*60}")

    stats = conn.execute("""
        SELECT
            COUNT(*) as total,
            COUNT(DISTINCT metering_point_id) as mps,
            COUNT(DISTINCT batch_id) as batches,
            MIN(timestamp_start) as ts_min,
            MAX(timestamp_end) as ts_max
        FROM measurements
    """).fetchone()

    print(f"  Messwerte gesamt:  {stats[0]:,}")
    print(f"  Zaehlpunkte:       {stats[1]}")
    print(f"  Import-Batches:    {stats[2]}")
    print(f"  Zeitraum:          {stats[3]} bis {stats[4]}")
    print(f"  DB-Groesse:        {os.path.getsize(DB_PATH) / 1024 / 1024:.1f} MB")

    conn.close()


if __name__ == "__main__":
    main()
